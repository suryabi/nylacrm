"""
Distributor Management Routes
CRUD operations for distributors, operating coverage, and locations
"""
from fastapi import APIRouter, HTTPException, Depends, Query, Response, BackgroundTasks
from datetime import datetime, timezone, timedelta
from typing import Optional, List, Dict, Any
import asyncio
import logging
import re
import uuid
import bcrypt

from database import db, get_tenant_db
from deps import get_current_user
from core.tenant import get_current_tenant_id
from models.distributor import (
    DistributorCreate, DistributorUpdate, Distributor,
    OperatingCoverageCreate, OperatingCoverageUpdate,
    DistributorLocationCreate, DistributorLocationUpdate,
    MarginMatrixCreate, MarginMatrixUpdate,
    AccountDistributorCreate, AccountDistributorUpdate,
    PrimaryShipmentCreate, PrimaryShipmentUpdate, ShipmentItemCreate,
    AccountDeliveryCreate, AccountDeliveryUpdate, DeliveryItemCreate,
    DistributorSettlementCreate, DistributorSettlementUpdate,
    BillingConfigCreate, BillingConfigUpdate,
    ProvisionalInvoiceCreate, ReconciliationCreate, DebitCreditNoteCreate
)
from utils.pdf_generator import generate_debit_credit_note_pdf, generate_customer_invoice_pdf
from utils.storage import upload_pdf, download_pdf
from utils.object_storage import init_storage

router = APIRouter()
logger = logging.getLogger(__name__)

from core.distributor_auth import DISTRIBUTOR_DEFAULT_PASSWORD  # noqa: E402


def hash_password(password: str) -> str:
    """Hash password using bcrypt (same as server.py)"""
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')


async def _record_deletion_audit(tenant_id: str, entity_type: str, entity: dict,
                                 distributor_id: str, current_user: dict,
                                 entity_number: str = None, extra: dict = None):
    """Persist a deletion audit record so who/when/what is queryable later.

    Stores a snapshot of the deleted document. Best-effort: never blocks the
    delete if the audit write fails.
    """
    try:
        rec = {
            'id': str(uuid.uuid4()),
            'tenant_id': tenant_id,
            'entity_type': entity_type,
            'entity_id': entity.get('id'),
            'entity_number': entity_number or entity.get('delivery_number') or entity.get('shipment_number'),
            'distributor_id': distributor_id,
            'status_at_deletion': entity.get('status'),
            'deleted_by': current_user.get('id'),
            'deleted_by_email': current_user.get('email'),
            'deleted_by_name': current_user.get('name') or current_user.get('full_name'),
            'deleted_by_role': current_user.get('role'),
            'deleted_at': datetime.now(timezone.utc).isoformat(),
            'snapshot': {k: v for k, v in entity.items() if k != '_id'},
        }
        if extra:
            rec.update(extra)
        await db.deletion_audit.insert_one(rec)
    except Exception as e:
        logger.error(f"Failed to write deletion audit for {entity_type} {entity.get('id')}: {e}")


def is_distributor_admin(user: dict) -> bool:
    """Check if user can manage distributors"""
    return user.get('role') in ['CEO', 'Director', 'Admin', 'System Admin', 'Vice President', 'National Sales Head']


# Common Indian city name aliases (old/anglicized names <-> official local names).
# Used to normalize city comparisons between accounts and distributor operating coverage,
# so that an account in "Bangalore" can be assigned to a distributor servicing "Bengaluru".
CITY_ALIASES = {
    "bangalore": "bengaluru",
    "bengaluru": "bengaluru",
    "bombay": "mumbai",
    "mumbai": "mumbai",
    "madras": "chennai",
    "chennai": "chennai",
    "calcutta": "kolkata",
    "kolkata": "kolkata",
    "poona": "pune",
    "pune": "pune",
    "trivandrum": "thiruvananthapuram",
    "thiruvananthapuram": "thiruvananthapuram",
    "cochin": "kochi",
    "kochi": "kochi",
    "baroda": "vadodara",
    "vadodara": "vadodara",
    "mysore": "mysuru",
    "mysuru": "mysuru",
    "mangalore": "mangaluru",
    "mangaluru": "mangaluru",
    "belgaum": "belagavi",
    "belagavi": "belagavi",
    "gurgaon": "gurugram",
    "gurugram": "gurugram",
    "allahabad": "prayagraj",
    "prayagraj": "prayagraj",
}


def normalize_city(city: str) -> str:
    """Normalize a city name for alias-aware comparison."""
    if not city:
        return ""
    key = city.strip().lower()
    return CITY_ALIASES.get(key, key)


def cities_match(city_a: str, city_b: str) -> bool:
    """Compare two city names treating common aliases (e.g. Bangalore/Bengaluru) as equivalent."""
    return normalize_city(city_a) == normalize_city(city_b)


def is_delete_authorized(user: dict) -> bool:
    """Check if user can delete distributors/warehouses (CEO and System Admin only)"""
    return user.get('role') in ['CEO', 'System Admin']


def is_distributor_user(user: dict) -> bool:
    """Check if user is a distributor"""
    return user.get('role') == 'Distributor'


def can_manage_distributor_data(user: dict, distributor_id: str) -> bool:
    """True if user is an admin OR a distributor user operating on their own org."""
    if is_distributor_admin(user):
        return True
    if is_distributor_user(user) and user.get('distributor_id') == distributor_id:
        return True
    return False


def get_user_distributor_id(user: dict) -> Optional[str]:
    """Get distributor_id for a distributor user"""
    if is_distributor_user(user):
        return user.get('distributor_id')
    return None


async def generate_distributor_code(tenant_id: str) -> str:
    """Generate unique distributor code"""
    count = await db.distributors.count_documents({"tenant_id": tenant_id})
    return f"DIST-{count + 1:04d}"


async def generate_location_code(distributor_id: str, tenant_id: str) -> str:
    """Generate unique location code"""
    count = await db.distributor_locations.count_documents({
        "tenant_id": tenant_id,
        "distributor_id": distributor_id
    })
    return f"LOC-{count + 1:03d}"


# ============ Distributor Master CRUD ============

@router.get("")
async def list_distributors(
    status: Optional[str] = None,
    search: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    current_user: dict = Depends(get_current_user)
):
    """List all distributors for current tenant (filtered for distributor users)"""
    tenant_id = get_current_tenant_id()
    
    query = {"tenant_id": tenant_id}
    
    # If user is a Distributor, only show their own distributor
    if is_distributor_user(current_user):
        user_distributor_id = get_user_distributor_id(current_user)
        if user_distributor_id:
            query["id"] = user_distributor_id
        else:
            # No distributor linked - return empty
            return {
                "distributors": [],
                "total": 0,
                "page": page,
                "page_size": page_size,
                "total_pages": 0
            }
    
    if status and status != 'all':
        query["status"] = status
    
    if search:
        query["$or"] = [
            {"distributor_name": {"$regex": search, "$options": "i"}},
            {"distributor_code": {"$regex": search, "$options": "i"}},
            {"primary_contact_name": {"$regex": search, "$options": "i"}},
            {"primary_contact_mobile": {"$regex": search, "$options": "i"}}
        ]
    
    total = await db.distributors.count_documents(query)
    
    distributors = await db.distributors.find(
        query,
        {"_id": 0}
    ).sort("created_at", -1).skip((page - 1) * page_size).limit(page_size).to_list(page_size)
    
    # Enrich with counts
    for dist in distributors:
        dist['coverage_count'] = await db.distributor_operating_coverage.count_documents({
            "tenant_id": tenant_id,
            "distributor_id": dist['id'],
            "status": "active"
        })
        dist['locations_count'] = await db.distributor_locations.count_documents({
            "tenant_id": tenant_id,
            "distributor_id": dist['id'],
            "status": "active"
        })
    
    return {
        "distributors": distributors,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size
    }


@router.get("/summary")
async def get_distributors_summary(current_user: dict = Depends(get_current_user)):
    """Get distributors summary stats"""
    tenant_id = get_current_tenant_id()
    
    total = await db.distributors.count_documents({"tenant_id": tenant_id})
    active = await db.distributors.count_documents({"tenant_id": tenant_id, "status": "active"})
    inactive = await db.distributors.count_documents({"tenant_id": tenant_id, "status": "inactive"})
    suspended = await db.distributors.count_documents({"tenant_id": tenant_id, "status": "suspended"})
    
    total_locations = await db.distributor_locations.count_documents({
        "tenant_id": tenant_id,
        "status": "active"
    })
    
    return {
        "total": total,
        "active": active,
        "inactive": inactive,
        "suspended": suspended,
        "total_locations": total_locations
    }


@router.get("/{distributor_id}")
async def get_distributor(distributor_id: str, current_user: dict = Depends(get_current_user)):
    """Get a specific distributor by ID"""
    tenant_id = get_current_tenant_id()
    
    # If user is a Distributor, verify they can only access their own distributor
    if is_distributor_user(current_user):
        user_distributor_id = get_user_distributor_id(current_user)
        if user_distributor_id != distributor_id:
            raise HTTPException(status_code=403, detail="Access denied - you can only view your own distributor profile")
    
    distributor = await db.distributors.find_one(
        {"id": distributor_id, "tenant_id": tenant_id},
        {"_id": 0}
    )
    
    if not distributor:
        raise HTTPException(status_code=404, detail="Distributor not found")
    
    # Get related data
    distributor['operating_coverage'] = await db.distributor_operating_coverage.find(
        {"tenant_id": tenant_id, "distributor_id": distributor_id},
        {"_id": 0}
    ).sort("city", 1).to_list(500)
    
    distributor['locations'] = await db.distributor_locations.find(
        {"tenant_id": tenant_id, "distributor_id": distributor_id},
        {"_id": 0}
    ).sort("location_name", 1).to_list(100)
    
    return distributor


@router.post("")
async def create_distributor(
    data: DistributorCreate,
    current_user: dict = Depends(get_current_user)
):
    """Create a new distributor and auto-create a user account for them"""
    if not is_distributor_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin access required to create distributors")
    
    tenant_id = get_current_tenant_id()
    now = datetime.now(timezone.utc).isoformat()
    
    # Generate distributor code if not provided
    distributor_code = data.distributor_code or await generate_distributor_code(tenant_id)
    
    # Check if code already exists
    existing = await db.distributors.find_one({
        "tenant_id": tenant_id,
        "distributor_code": distributor_code
    })
    if existing:
        raise HTTPException(status_code=400, detail="Distributor code already exists")
    
    distributor_id = str(uuid.uuid4())
    
    distributor_doc = {
        "id": distributor_id,
        "tenant_id": tenant_id,
        "distributor_name": data.distributor_name,
        "legal_entity_name": data.legal_entity_name,
        "distributor_code": distributor_code,
        "gstin": data.gstin,
        "pan": data.pan,
        "billing_address": data.billing_address,
        "registered_address": data.registered_address,
        "primary_contact_name": data.primary_contact_name,
        "primary_contact_mobile": data.primary_contact_mobile,
        "primary_contact_email": data.primary_contact_email,
        "secondary_contact_name": data.secondary_contact_name,
        "secondary_contact_mobile": data.secondary_contact_mobile,
        "secondary_contact_email": data.secondary_contact_email,
        "payment_terms": data.payment_terms or "net_30",
        "credit_days": data.credit_days or 30,
        "credit_limit": data.credit_limit or 0,
        "security_deposit": data.security_deposit or 0,
        "is_self_managed": data.is_self_managed or False,
        "billing_approach": data.billing_approach or "margin_upfront",
        "status": data.status or "active",
        "notes": data.notes,
        "created_at": now,
        "updated_at": now,
        "created_by": current_user.get('id')
    }
    
    await db.distributors.insert_one(distributor_doc)
    distributor_doc.pop('_id', None)
    
    # Auto-create user for distributor if primary contact email is provided
    user_created = False
    if data.primary_contact_email:
        # Check if user already exists with this email
        existing_user = await db.users.find_one({'email': data.primary_contact_email}, {'_id': 0})
        
        if not existing_user:
            # Create new user for distributor
            user_doc = {
                'id': str(uuid.uuid4()),
                'email': data.primary_contact_email,
                'name': data.primary_contact_name or data.distributor_name,
                'password': hash_password(DISTRIBUTOR_DEFAULT_PASSWORD),
                'role': 'Distributor',
                'designation': 'Distributor',
                'department': 'Distribution',
                'phone': data.primary_contact_mobile,
                'is_active': True,
                'distributor_id': distributor_id,
                'force_password_change': True,
                'created_at': now,
                'tenant_id': tenant_id
            }
            await db.users.insert_one(user_doc)
            user_created = True
            logger.info(f"User account created for distributor '{data.distributor_name}' with email '{data.primary_contact_email}'")
        else:
            # Update existing user to link to this distributor if they don't have one
            if not existing_user.get('distributor_id'):
                await db.users.update_one(
                    {'email': data.primary_contact_email},
                    {'$set': {
                        'distributor_id': distributor_id,
                        'role': 'Distributor',
                        'department': 'Distribution'
                    }}
                )
                logger.info(f"Existing user '{data.primary_contact_email}' linked to distributor '{data.distributor_name}'")
    
    logger.info(f"Distributor '{data.distributor_name}' created by {current_user['email']}")
    
    # Add user_created flag to response
    distributor_doc['user_created'] = user_created
    
    return distributor_doc


@router.put("/{distributor_id}")
async def update_distributor(
    distributor_id: str,
    data: DistributorUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Update a distributor"""
    if not is_distributor_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin access required to update distributors")
    
    tenant_id = get_current_tenant_id()
    
    distributor = await db.distributors.find_one(
        {"id": distributor_id, "tenant_id": tenant_id},
        {"_id": 0}
    )
    
    if not distributor:
        raise HTTPException(status_code=404, detail="Distributor not found")
    
    update_data = {"updated_at": datetime.now(timezone.utc).isoformat()}
    
    # Update only provided fields
    for field in ['distributor_name', 'legal_entity_name', 'distributor_code', 'gstin', 'pan',
                  'billing_address', 'registered_address', 'primary_contact_name', 'primary_contact_mobile',
                  'primary_contact_email', 'secondary_contact_name', 'secondary_contact_mobile',
                  'secondary_contact_email', 'payment_terms', 'credit_days', 'credit_limit',
                  'security_deposit', 'is_self_managed', 'billing_approach', 'status', 'notes']:
        value = getattr(data, field, None)
        if value is not None:
            update_data[field] = value
    
    await db.distributors.update_one(
        {"id": distributor_id, "tenant_id": tenant_id},
        {"$set": update_data}
    )
    
    updated = await db.distributors.find_one(
        {"id": distributor_id, "tenant_id": tenant_id},
        {"_id": 0}
    )
    
    logger.info(f"Distributor '{distributor_id}' updated by {current_user['email']}")
    
    return updated


@router.delete("/{distributor_id}")
async def delete_distributor(distributor_id: str, current_user: dict = Depends(get_current_user)):
    """Hard delete a distributor and all child data (CEO/System Admin only)"""
    if not is_delete_authorized(current_user):
        raise HTTPException(status_code=403, detail="Only CEO and System Admin can delete distributors")
    
    tenant_id = get_current_tenant_id()
    
    distributor = await db.distributors.find_one(
        {"id": distributor_id, "tenant_id": tenant_id},
        {"_id": 0, "distributor_name": 1}
    )
    
    if not distributor:
        raise HTTPException(status_code=404, detail="Distributor not found")
    
    dist_name = distributor.get('distributor_name', distributor_id)
    
    # Cascading delete all child data
    child_filter = {"tenant_id": tenant_id, "distributor_id": distributor_id}
    
    # 1. Get shipment IDs for item cleanup
    shipment_ids = [s["id"] async for s in db.distributor_shipments.find(child_filter, {"id": 1})]
    if shipment_ids:
        await db.distributor_shipment_items.delete_many({"tenant_id": tenant_id, "shipment_id": {"$in": shipment_ids}})
    
    # 2. Get delivery IDs for item cleanup
    delivery_ids = [d["id"] async for d in db.distributor_deliveries.find(child_filter, {"id": 1})]
    if delivery_ids:
        await db.distributor_delivery_items.delete_many({"tenant_id": tenant_id, "delivery_id": {"$in": delivery_ids}})
    
    # 3. Get settlement IDs for item cleanup
    settlement_ids = [s["id"] async for s in db.distributor_settlements.find(child_filter, {"id": 1})]
    if settlement_ids:
        await db.distributor_settlement_items.delete_many({"tenant_id": tenant_id, "settlement_id": {"$in": settlement_ids}})
    
    # 4. Get reconciliation IDs for line item + note cleanup
    recon_ids = [r["id"] async for r in db.distributor_reconciliations.find(child_filter, {"id": 1})]
    if recon_ids:
        await db.distributor_reconciliation_items.delete_many({"tenant_id": tenant_id, "reconciliation_id": {"$in": recon_ids}})
        await db.distributor_debit_credit_notes.delete_many({"tenant_id": tenant_id, "reconciliation_id": {"$in": recon_ids}})
    
    # 5. Delete top-level child collections
    del_results = {}
    for coll_name, collection in [
        ("operating_coverage", db.distributor_operating_coverage),
        ("locations", db.distributor_locations),
        ("margin_matrix", db.distributor_margin_matrix),
        ("account_assignments", db.account_distributor_assignments),
        ("shipments", db.distributor_shipments),
        ("deliveries", db.distributor_deliveries),
        ("settlements", db.distributor_settlements),
        ("billing_configs", db.distributor_billing_configs),
        ("provisional_invoices", db.distributor_provisional_invoices),
        ("reconciliations", db.distributor_reconciliations),
    ]:
        result = await collection.delete_many(child_filter)
        del_results[coll_name] = result.deleted_count
    
    # 6. Delete linked user accounts
    user_result = await db.users.delete_many({"tenant_id": tenant_id, "distributor_id": distributor_id})
    del_results["users"] = user_result.deleted_count
    
    # 7. Delete the distributor itself
    await db.distributors.delete_one({"id": distributor_id, "tenant_id": tenant_id})
    
    logger.info(f"Distributor '{dist_name}' hard-deleted with all child data by {current_user['email']}. Counts: {del_results}")
    
    return {
        "message": f"Distributor '{dist_name}' and all related data deleted permanently",
        "deleted_counts": del_results
    }


@router.post("/cleanup/orphaned-data")
async def cleanup_orphaned_distributor_data(current_user: dict = Depends(get_current_user)):
    """Remove all transactional data for distributors that no longer exist. CEO/System Admin only."""
    if not is_delete_authorized(current_user):
        raise HTTPException(status_code=403, detail="Only CEO and System Admin can run cleanup")
    
    tenant_id = get_current_tenant_id()
    tdb = get_tenant_db()
    
    # Get all existing distributor IDs
    existing_ids = await db.distributors.distinct("id", {"tenant_id": tenant_id})
    existing_set = set(existing_ids)
    
    results = {}
    
    # Collections with distributor_id field
    dist_collections = [
        ("operating_coverage", db.distributor_operating_coverage),
        ("locations", db.distributor_locations),
        ("margin_matrix", db.distributor_margin_matrix),
        ("account_assignments", db.account_distributor_assignments),
        ("shipments", db.distributor_shipments),
        ("deliveries", db.distributor_deliveries),
        ("settlements", db.distributor_settlements),
        ("billing_configs", db.distributor_billing_configs),
        ("provisional_invoices", db.distributor_provisional_invoices),
        ("reconciliations", db.distributor_reconciliations),
        ("users", db.users),
    ]
    
    for coll_name, collection in dist_collections:
        # Find orphaned distributor_ids in this collection
        all_dist_ids = await collection.distinct("distributor_id", {"tenant_id": tenant_id})
        orphaned = [did for did in all_dist_ids if did and did not in existing_set]
        if orphaned:
            r = await collection.delete_many({"tenant_id": tenant_id, "distributor_id": {"$in": orphaned}})
            results[coll_name] = r.deleted_count
    
    # Clean orphaned shipment items (shipment no longer exists)
    existing_shipment_ids = await db.distributor_shipments.distinct("id", {"tenant_id": tenant_id})
    all_item_shipment_ids = await db.distributor_shipment_items.distinct("shipment_id", {"tenant_id": tenant_id})
    orphaned_ship_items = [sid for sid in all_item_shipment_ids if sid not in set(existing_shipment_ids)]
    if orphaned_ship_items:
        r = await db.distributor_shipment_items.delete_many({"tenant_id": tenant_id, "shipment_id": {"$in": orphaned_ship_items}})
        results["shipment_items"] = r.deleted_count
    
    # Clean orphaned delivery items
    existing_delivery_ids = await db.distributor_deliveries.distinct("id", {"tenant_id": tenant_id})
    all_del_item_ids = await db.distributor_delivery_items.distinct("delivery_id", {"tenant_id": tenant_id})
    orphaned_del_items = [did for did in all_del_item_ids if did not in set(existing_delivery_ids)]
    if orphaned_del_items:
        r = await db.distributor_delivery_items.delete_many({"tenant_id": tenant_id, "delivery_id": {"$in": orphaned_del_items}})
        results["delivery_items"] = r.deleted_count
    
    # Clean orphaned settlement items
    existing_settlement_ids = await db.distributor_settlements.distinct("id", {"tenant_id": tenant_id})
    all_sett_item_ids = await db.distributor_settlement_items.distinct("settlement_id", {"tenant_id": tenant_id})
    orphaned_sett_items = [sid for sid in all_sett_item_ids if sid not in set(existing_settlement_ids)]
    if orphaned_sett_items:
        r = await db.distributor_settlement_items.delete_many({"tenant_id": tenant_id, "settlement_id": {"$in": orphaned_sett_items}})
        results["settlement_items"] = r.deleted_count
    
    # Clean orphaned reconciliation items and debit/credit notes
    existing_recon_ids = await db.distributor_reconciliations.distinct("id", {"tenant_id": tenant_id})
    all_recon_item_ids = await db.distributor_reconciliation_items.distinct("reconciliation_id", {"tenant_id": tenant_id})
    orphaned_recon = [rid for rid in all_recon_item_ids if rid not in set(existing_recon_ids)]
    if orphaned_recon:
        r1 = await db.distributor_reconciliation_items.delete_many({"tenant_id": tenant_id, "reconciliation_id": {"$in": orphaned_recon}})
        r2 = await db.distributor_debit_credit_notes.delete_many({"tenant_id": tenant_id, "reconciliation_id": {"$in": orphaned_recon}})
        results["reconciliation_items"] = r1.deleted_count
        results["debit_credit_notes"] = r2.deleted_count
    
    # Clean factory warehouse stock for orphaned locations
    existing_loc_ids = await db.distributor_locations.distinct("id", {"tenant_id": tenant_id})
    all_fws_loc_ids = await tdb.factory_warehouse_stock.distinct("warehouse_location_id", {"tenant_id": tenant_id})
    orphaned_fws = [lid for lid in all_fws_loc_ids if lid not in set(existing_loc_ids)]
    if orphaned_fws:
        r = await tdb.factory_warehouse_stock.delete_many({"tenant_id": tenant_id, "warehouse_location_id": {"$in": orphaned_fws}})
        results["factory_warehouse_stock"] = r.deleted_count
    
    logger.info(f"Orphaned distributor data cleanup by {current_user['email']}: {results}")
    
    return {"message": "Orphaned distributor data cleaned up", "deleted_counts": results}


@router.post("/admin/wipe-transactional-data")
async def wipe_distributor_transactional_data(
    payload: dict,
    current_user: dict = Depends(get_current_user)
):
    """Wipe ALL distributor transactional data for the current tenant.

    Wipes (per tenant):
      • distributor_deliveries + distributor_delivery_items   (Stock Out)
      • distributor_settlements + distributor_settlement_items (Settlements)
      • distributor_debit_credit_notes + distributor_reconciliations + items (Billing)
      • customer_returns + credit_notes + credit_note_issuances (Returns)
      • distributor_factory_returns                            (Factory Returns)

    Optionally wipes when payload includes `"include_stock_in": true`:
      • distributor_shipments + distributor_shipment_items     (Stock In)
      • distributor_stock                                       (current on-hand)

    Keeps untouched:
      • distributor_shipments + items (Stock In) — unless include_stock_in is true
      • distributors, locations, coverage, margin_matrix, billing_config, contacts
      • account_distributor_assignments
      • accounts, leads, invoices, master data, etc.

    After deletion `distributor_stock` is rebuilt from delivered shipments so
    on-hand quantities reflect a clean baseline (everything received, nothing
    dispatched or returned). When `include_stock_in` is true, stock is left
    fully empty since no shipments remain.

    Guards:
      • CEO / System Admin only
      • Body must contain {"confirm": "DELETE_TRANSACTIONS"} to avoid accidents
    """
    if not is_delete_authorized(current_user):
        raise HTTPException(status_code=403, detail="Only CEO and System Admin can run this operation")

    if (payload or {}).get("confirm") != "DELETE_TRANSACTIONS":
        raise HTTPException(
            status_code=400,
            detail='Missing or invalid confirm token. Provide {"confirm": "DELETE_TRANSACTIONS"} to proceed.'
        )

    include_stock_in = bool((payload or {}).get("include_stock_in"))

    tenant_id = get_current_tenant_id()
    now = datetime.now(timezone.utc).isoformat()
    deleted: dict = {}

    targets = [
        # Stock Out
        ("distributor_deliveries", db.distributor_deliveries),
        ("distributor_delivery_items", db.distributor_delivery_items),
        # Settlements
        ("distributor_settlements", db.distributor_settlements),
        ("distributor_settlement_items", db.distributor_settlement_items),
        # Billing & Reconciliation
        ("distributor_debit_credit_notes", db.distributor_debit_credit_notes),
        ("distributor_reconciliations", db.distributor_reconciliations),
        ("distributor_reconciliation_items", db.distributor_reconciliation_items),
        # Returns + linked credits
        ("customer_returns", db.customer_returns),
        ("credit_notes", db.credit_notes),
        ("credit_note_issuances", db.credit_note_issuances),
        # Factory Returns
        ("distributor_factory_returns", db.distributor_factory_returns),
    ]
    if include_stock_in:
        targets.extend([
            # Stock In
            ("distributor_shipments", db.distributor_shipments),
            ("distributor_shipment_items", db.distributor_shipment_items),
        ])

    for coll_name, collection in targets:
        try:
            res = await collection.delete_many({"tenant_id": tenant_id})
            deleted[coll_name] = res.deleted_count
        except Exception as e:
            logger.error(f"wipe-transactional: failed to delete {coll_name}: {e}")
            deleted[coll_name] = f"error: {e}"

    # Rebuild distributor_stock from delivered shipments so on-hand reflects
    # "everything received, nothing dispatched/returned".
    # If Stock In was also wiped, simply zero out stock (no shipments left to rebuild from).
    stock_rebuilt = 0
    try:
        if include_stock_in:
            sres = await db.distributor_stock.delete_many({"tenant_id": tenant_id})
            deleted["distributor_stock"] = sres.deleted_count
            stock_rebuilt = 0
        else:
            # Collect aggregate qty per (distributor_id, location_id, sku_id) from
            # delivered shipments only.
            delivered_shipments = await db.distributor_shipments.find(
                {"tenant_id": tenant_id, "status": "delivered"},
                {"_id": 0, "id": 1, "distributor_id": 1, "distributor_location_id": 1,
                 "distributor_name": 1, "distributor_location_name": 1}
            ).to_list(10000)
            ship_meta = {s['id']: s for s in delivered_shipments}
            ship_ids = list(ship_meta.keys())

            # Wipe existing stock for the tenant first
            sres = await db.distributor_stock.delete_many({"tenant_id": tenant_id})
            deleted["distributor_stock"] = sres.deleted_count

            if ship_ids:
                items = await db.distributor_shipment_items.find(
                    {"tenant_id": tenant_id, "shipment_id": {"$in": ship_ids}},
                    {"_id": 0}
                ).to_list(50000)
                agg: dict = {}
                for it in items:
                    s = ship_meta.get(it.get('shipment_id')) or {}
                    key = (
                        s.get('distributor_id'),
                        s.get('distributor_location_id'),
                        it.get('sku_id'),
                    )
                    if not all(key):
                        continue
                    row = agg.setdefault(key, {
                        "tenant_id": tenant_id,
                        "distributor_id": key[0],
                        "distributor_location_id": key[1],
                        "sku_id": key[2],
                        "sku_name": it.get('sku_name'),
                        "sku_code": it.get('sku_code'),
                        "distributor_name": s.get('distributor_name'),
                        "location_name": s.get('distributor_location_name'),
                        "quantity": 0,
                        "id": str(uuid.uuid4()),
                        "created_at": now,
                        "updated_at": now,
                    })
                    row["quantity"] += it.get('quantity', 0) or 0
                if agg:
                    await db.distributor_stock.insert_many(list(agg.values()))
                    stock_rebuilt = len(agg)
    except Exception as e:
        logger.error(f"wipe-transactional: stock rebuild failed: {e}")
        deleted["distributor_stock_rebuild_error"] = str(e)

    deleted["distributor_stock_rows_rebuilt"] = stock_rebuilt

    logger.warning(
        f"WIPE distributor transactional data — tenant={tenant_id} "
        f"by {current_user.get('email')} (role={current_user.get('role')}) "
        f"results={deleted}"
    )

    return {
        "message": "Distributor transactional data wiped. Stock rebuilt from delivered shipments.",
        "tenant_id": tenant_id,
        "deleted_counts": deleted,
        "executed_by": current_user.get('email'),
        "executed_at": now,
    }



# ============ Operating Coverage CRUD ============

@router.get("/{distributor_id}/coverage")
async def list_operating_coverage(
    distributor_id: str,
    current_user: dict = Depends(get_current_user)
):
    """List operating coverage for a distributor"""
    tenant_id = get_current_tenant_id()
    
    coverage = await db.distributor_operating_coverage.find(
        {"tenant_id": tenant_id, "distributor_id": distributor_id},
        {"_id": 0}
    ).sort([("state", 1), ("city", 1)]).to_list(500)
    
    return {"coverage": coverage, "total": len(coverage)}


@router.post("/{distributor_id}/coverage")
async def add_operating_coverage(
    distributor_id: str,
    data: OperatingCoverageCreate,
    current_user: dict = Depends(get_current_user)
):
    """Add operating coverage for a distributor"""
    if not is_distributor_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    tenant_id = get_current_tenant_id()
    now = datetime.now(timezone.utc).isoformat()
    
    # Check if distributor exists
    distributor = await db.distributors.find_one({"id": distributor_id, "tenant_id": tenant_id})
    if not distributor:
        raise HTTPException(status_code=404, detail="Distributor not found")
    
    # Check for duplicate
    existing = await db.distributor_operating_coverage.find_one({
        "tenant_id": tenant_id,
        "distributor_id": distributor_id,
        "state": data.state,
        "city": data.city,
        "status": "active"
    })
    if existing:
        raise HTTPException(status_code=400, detail="Coverage for this city already exists")
    
    coverage_doc = {
        "id": str(uuid.uuid4()),
        "tenant_id": tenant_id,
        "distributor_id": distributor_id,
        "state": data.state,
        "city": data.city,
        "effective_from": data.effective_from or now,
        "effective_to": data.effective_to,
        "status": data.status or "active",
        "created_at": now,
        "updated_at": now
    }
    
    await db.distributor_operating_coverage.insert_one(coverage_doc)
    coverage_doc.pop('_id', None)
    
    return coverage_doc


@router.post("/{distributor_id}/coverage/bulk")
async def add_bulk_operating_coverage(
    distributor_id: str,
    data: List[OperatingCoverageCreate],
    current_user: dict = Depends(get_current_user)
):
    """Add multiple operating coverages for a distributor"""
    if not is_distributor_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    tenant_id = get_current_tenant_id()
    now = datetime.now(timezone.utc).isoformat()
    
    # Check if distributor exists
    distributor = await db.distributors.find_one({"id": distributor_id, "tenant_id": tenant_id})
    if not distributor:
        raise HTTPException(status_code=404, detail="Distributor not found")
    
    added = []
    skipped = []
    
    for item in data:
        # Check for duplicate
        existing = await db.distributor_operating_coverage.find_one({
            "tenant_id": tenant_id,
            "distributor_id": distributor_id,
            "state": item.state,
            "city": item.city,
            "status": "active"
        })
        
        if existing:
            skipped.append({"state": item.state, "city": item.city, "reason": "Already exists"})
            continue
        
        coverage_doc = {
            "id": str(uuid.uuid4()),
            "tenant_id": tenant_id,
            "distributor_id": distributor_id,
            "state": item.state,
            "city": item.city,
            "effective_from": item.effective_from or now,
            "effective_to": item.effective_to,
            "status": item.status or "active",
            "created_at": now,
            "updated_at": now
        }
        
        await db.distributor_operating_coverage.insert_one(coverage_doc)
        coverage_doc.pop('_id', None)
        added.append(coverage_doc)
    
    return {
        "added": added,
        "added_count": len(added),
        "skipped": skipped,
        "skipped_count": len(skipped)
    }


@router.delete("/{distributor_id}/coverage/{coverage_id}")
async def delete_operating_coverage(
    distributor_id: str,
    coverage_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Delete operating coverage"""
    if not is_distributor_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    tenant_id = get_current_tenant_id()
    
    result = await db.distributor_operating_coverage.delete_one({
        "id": coverage_id,
        "tenant_id": tenant_id,
        "distributor_id": distributor_id
    })
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Coverage not found")
    
    return {"message": "Coverage deleted successfully"}


# ============ Distributor Locations CRUD ============

@router.get("/{distributor_id}/locations")
async def list_distributor_locations(
    distributor_id: str,
    current_user: dict = Depends(get_current_user)
):
    """List locations for a distributor"""
    tenant_id = get_current_tenant_id()
    
    locations = await db.distributor_locations.find(
        {"tenant_id": tenant_id, "distributor_id": distributor_id},
        {"_id": 0}
    ).sort("location_name", 1).to_list(100)
    
    return {"locations": locations, "total": len(locations)}


@router.post("/{distributor_id}/locations")
async def create_distributor_location(
    distributor_id: str,
    data: DistributorLocationCreate,
    current_user: dict = Depends(get_current_user)
):
    """Create a new distributor location/warehouse"""
    if not is_distributor_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    tenant_id = get_current_tenant_id()
    now = datetime.now(timezone.utc).isoformat()
    
    # Check if distributor exists
    distributor = await db.distributors.find_one({"id": distributor_id, "tenant_id": tenant_id})
    if not distributor:
        raise HTTPException(status_code=404, detail="Distributor not found")
    
    # Check if city is in operating coverage
    coverage = await db.distributor_operating_coverage.find_one({
        "tenant_id": tenant_id,
        "distributor_id": distributor_id,
        "city": data.city,
        "status": "active"
    })
    if not coverage:
        raise HTTPException(
            status_code=400, 
            detail=f"City '{data.city}' is not in distributor's operating coverage. Add coverage first."
        )
    
    # Generate location code if not provided
    location_code = data.location_code or await generate_location_code(distributor_id, tenant_id)
    
    # If this is marked as default, unset other defaults
    if data.is_default:
        await db.distributor_locations.update_many(
            {"tenant_id": tenant_id, "distributor_id": distributor_id},
            {"$set": {"is_default": False}}
        )
    
    location_doc = {
        "id": str(uuid.uuid4()),
        "tenant_id": tenant_id,
        "distributor_id": distributor_id,
        "location_name": data.location_name,
        "location_code": location_code,
        "address_line_1": data.address_line_1,
        "address_line_2": data.address_line_2,
        "state": data.state,
        "city": data.city,
        "pincode": data.pincode,
        "contact_person": data.contact_person,
        "contact_number": data.contact_number,
        "email": data.email,
        "is_default": data.is_default or False,
        "is_factory": data.is_factory or False,
        "track_batches": bool(data.track_batches) if data.track_batches is not None else False,
        "status": data.status or "active",
        "gstin": data.gstin,
        "zoho_branch_id": data.zoho_branch_id,
        "zoho_branch_name": data.zoho_branch_name,
        "lat": data.lat,
        "lng": data.lng,
        "formatted_address": data.formatted_address,
        "created_at": now,
        "updated_at": now
    }
    
    await db.distributor_locations.insert_one(location_doc)
    location_doc.pop('_id', None)
    
    logger.info(f"Distributor location '{data.location_name}' created by {current_user['email']}")
    
    return location_doc


@router.put("/{distributor_id}/locations/{location_id}")
async def update_distributor_location(
    distributor_id: str,
    location_id: str,
    data: DistributorLocationUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Update a distributor location"""
    if not is_distributor_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    tenant_id = get_current_tenant_id()
    
    location = await db.distributor_locations.find_one({
        "id": location_id,
        "tenant_id": tenant_id,
        "distributor_id": distributor_id
    })
    
    if not location:
        raise HTTPException(status_code=404, detail="Location not found")
    
    update_data = {"updated_at": datetime.now(timezone.utc).isoformat()}
    
    # If setting as default, unset other defaults
    if data.is_default:
        await db.distributor_locations.update_many(
            {"tenant_id": tenant_id, "distributor_id": distributor_id, "id": {"$ne": location_id}},
            {"$set": {"is_default": False}}
        )
    
    for field in ['location_name', 'location_code', 'address_line_1', 'address_line_2',
                  'state', 'city', 'pincode', 'contact_person', 'contact_number',
                  'email', 'is_default', 'is_factory', 'track_batches', 'status',
                  'gstin', 'zoho_branch_id', 'zoho_branch_name',
                  'lat', 'lng', 'formatted_address']:
        value = getattr(data, field, None)
        if value is not None:
            update_data[field] = value
    
    await db.distributor_locations.update_one(
        {"id": location_id, "tenant_id": tenant_id},
        {"$set": update_data}
    )
    
    updated = await db.distributor_locations.find_one(
        {"id": location_id, "tenant_id": tenant_id},
        {"_id": 0}
    )
    
    return updated


@router.post("/{distributor_id}/locations/{location_id}/sync-to-zoho")
async def sync_location_to_zoho_branch(
    distributor_id: str,
    location_id: str,
    current_user: dict = Depends(get_current_user),
):
    """Push the CRM warehouse's address + GSTIN to the linked Zoho branch.

    This is the one-click "Sync to Zoho" action on each warehouse card. It
    requires the warehouse to already have a `zoho_branch_id` populated —
    we do NOT create branches automatically because Zoho enforces
    tax-settings prerequisites that must be configured in the Zoho UI first.
    """
    if not is_distributor_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin access required")

    tenant_id = get_current_tenant_id()
    location = await db.distributor_locations.find_one(
        {"id": location_id, "tenant_id": tenant_id, "distributor_id": distributor_id},
        {"_id": 0},
    )
    if not location:
        raise HTTPException(status_code=404, detail="Location not found")

    try:
        from services.zoho_service import (
            sync_warehouse_to_zoho_branch,
            ZohoPushSkippedError,
            ZohoApiError,
        )
        # Bound the total runtime so we always respond well before the gateway
        # timeout. A slow/unreachable Zoho would otherwise let the request hang
        # until Cloudflare returns an opaque 502 origin error to the user.
        branch = await asyncio.wait_for(
            sync_warehouse_to_zoho_branch(tenant_id=tenant_id, location=location),
            timeout=25.0,
        )
    except ZohoPushSkippedError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except ZohoApiError as e:
        raise HTTPException(status_code=502, detail=f"Zoho rejected the sync: {e}")
    except asyncio.TimeoutError:
        logger.warning("Zoho branch sync timed out for location %s", location_id)
        raise HTTPException(
            status_code=504,
            detail="Zoho Books didn't respond in time. This is usually a temporary Zoho slowdown — please try the sync again in a moment.",
        )
    except Exception as e:
        logger.exception("Zoho branch sync failed for location %s", location_id)
        raise HTTPException(
            status_code=502,
            detail=f"Couldn't reach Zoho Books to sync this warehouse: {str(e)[:200]}. Please verify Zoho is connected and try again.",
        )

    # Record when the last successful sync happened so the UI can show it.
    now_iso = datetime.now(timezone.utc).isoformat()
    await db.distributor_locations.update_one(
        {"id": location_id, "tenant_id": tenant_id},
        {"$set": {
            "zoho_branch_synced_at": now_iso,
            # Keep the cached branch name in sync with whatever Zoho returned.
            "zoho_branch_name": branch.get("branch_name") or location.get("zoho_branch_name"),
            "updated_at": now_iso,
        }},
    )

    return {
        "ok": True,
        "synced_at": now_iso,
        "zoho_branch_id": location.get("zoho_branch_id"),
        "zoho_branch_name": branch.get("branch_name") or location.get("zoho_branch_name"),
        "message": "Warehouse address and GSTIN pushed to Zoho. Future invoices and challans from this warehouse will print the updated header.",
    }


@router.delete("/{distributor_id}/locations/{location_id}")
async def delete_distributor_location(
    distributor_id: str,
    location_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Hard delete a distributor location/warehouse and related data (CEO/System Admin only)"""
    if not is_delete_authorized(current_user):
        raise HTTPException(status_code=403, detail="Only CEO and System Admin can delete warehouses")
    
    tenant_id = get_current_tenant_id()
    
    location = await db.distributor_locations.find_one({
        "id": location_id,
        "tenant_id": tenant_id,
        "distributor_id": distributor_id
    })
    
    if not location:
        raise HTTPException(status_code=404, detail="Location not found")
    
    loc_name = location.get('location_name', location_id)
    loc_filter = {"tenant_id": tenant_id, "distributor_location_id": location_id}
    
    # Cascade: delete shipments and their items that target this location
    shipment_ids = [s["id"] async for s in db.distributor_shipments.find(loc_filter, {"id": 1})]
    if shipment_ids:
        await db.distributor_shipment_items.delete_many({"tenant_id": tenant_id, "shipment_id": {"$in": shipment_ids}})
    await db.distributor_shipments.delete_many(loc_filter)
    
    # Cascade: delete deliveries and their items from this location
    delivery_ids = [d["id"] async for d in db.distributor_deliveries.find(loc_filter, {"id": 1})]
    if delivery_ids:
        await db.distributor_delivery_items.delete_many({"tenant_id": tenant_id, "delivery_id": {"$in": delivery_ids}})
    await db.distributor_deliveries.delete_many(loc_filter)
    
    # Hard delete the location
    await db.distributor_locations.delete_one({"id": location_id, "tenant_id": tenant_id})
    
    logger.info(f"Location '{loc_name}' hard-deleted with related data by {current_user['email']}")
    
    return {"message": f"Warehouse '{loc_name}' and all related data deleted permanently"}


# ============ Dropdown Data ============

@router.get("/dropdown/active")
async def get_active_distributors_dropdown(current_user: dict = Depends(get_current_user)):
    """Get active distributors for dropdown"""
    tenant_id = get_current_tenant_id()
    
    distributors = await db.distributors.find(
        {"tenant_id": tenant_id, "status": "active"},
        {"_id": 0, "id": 1, "distributor_name": 1, "distributor_code": 1}
    ).sort("distributor_name", 1).to_list(500)
    
    return {"distributors": distributors}


@router.get("/{distributor_id}/locations/dropdown")
async def get_distributor_locations_dropdown(
    distributor_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get active locations for a distributor (for dropdown)"""
    tenant_id = get_current_tenant_id()
    
    locations = await db.distributor_locations.find(
        {"tenant_id": tenant_id, "distributor_id": distributor_id, "status": "active"},
        {"_id": 0, "id": 1, "location_name": 1, "location_code": 1, "city": 1, "is_default": 1, "is_factory": 1}
    ).sort("location_name", 1).to_list(100)
    
    return {"locations": locations}


# ============ Margin Matrix CRUD ============

MARGIN_TYPES = {
    "percentage": "Percentage on Account Invoice Value",
    "fixed_per_bottle": "Fixed Amount per Bottle",
    "fixed_per_case": "Fixed Amount per Case/Crate"
}


@router.get("/{distributor_id}/margins")
async def list_margin_matrix(
    distributor_id: str,
    city: Optional[str] = None,
    sku_id: Optional[str] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """List margin matrix entries for a distributor"""
    tenant_id = get_current_tenant_id()
    
    query = {"tenant_id": tenant_id, "distributor_id": distributor_id}
    
    if city and city != 'all':
        query["city"] = city
    if sku_id and sku_id != 'all':
        query["sku_id"] = sku_id
    if status and status != 'all':
        query["status"] = status
    
    margins = await db.distributor_margin_matrix.find(
        query,
        {"_id": 0}
    ).sort([("city", 1), ("sku_name", 1)]).to_list(1000)
    
    # Get summary stats
    total = len(margins)
    active = len([m for m in margins if m.get('status') == 'active'])
    by_type = {}
    for m in margins:
        mt = m.get('margin_type', 'unknown')
        by_type[mt] = by_type.get(mt, 0) + 1
    
    return {
        "margins": margins,
        "total": total,
        "active": active,
        "by_type": by_type
    }


@router.get("/{distributor_id}/margins/{margin_id}")
async def get_margin_entry(
    distributor_id: str,
    margin_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get a specific margin matrix entry"""
    tenant_id = get_current_tenant_id()
    
    margin = await db.distributor_margin_matrix.find_one(
        {"id": margin_id, "tenant_id": tenant_id, "distributor_id": distributor_id},
        {"_id": 0}
    )
    
    if not margin:
        raise HTTPException(status_code=404, detail="Margin entry not found")
    
    return margin


@router.post("/{distributor_id}/margins")
async def create_margin_entry(
    distributor_id: str,
    data: MarginMatrixCreate,
    current_user: dict = Depends(get_current_user)
):
    """Create a new margin matrix entry"""
    if not is_distributor_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    tenant_id = get_current_tenant_id()
    now = datetime.now(timezone.utc).isoformat()
    
    # Check if distributor exists
    distributor = await db.distributors.find_one({"id": distributor_id, "tenant_id": tenant_id})
    if not distributor:
        raise HTTPException(status_code=404, detail="Distributor not found")
    
    # Check if city is in operating coverage
    coverage = await db.distributor_operating_coverage.find_one({
        "tenant_id": tenant_id,
        "distributor_id": distributor_id,
        "city": data.city,
        "status": "active"
    })
    if not coverage:
        raise HTTPException(
            status_code=400,
            detail=f"City '{data.city}' is not in distributor's operating coverage"
        )
    
    # Validate margin type
    if data.margin_type not in MARGIN_TYPES:
        raise HTTPException(status_code=400, detail=f"Invalid margin type. Must be one of: {list(MARGIN_TYPES.keys())}")
    
    # Check for date overlap with existing active entries for same city+SKU
    # Two date ranges overlap if: start1 <= end2 AND start2 <= end1
    new_start = data.active_from or now[:10]
    new_end = data.active_to or "9999-12-31"  # Far future date for open-ended entries
    
    existing_entries = await db.distributor_margin_matrix.find({
        "tenant_id": tenant_id,
        "distributor_id": distributor_id,
        "city": data.city,
        "sku_id": data.sku_id,
        "status": "active"
    }).to_list(100)
    
    for existing in existing_entries:
        exist_start = existing.get('active_from') or "1900-01-01"
        exist_end = existing.get('active_to') or "9999-12-31"
        
        # Check for overlap: new_start <= exist_end AND exist_start <= new_end
        if new_start <= exist_end and exist_start <= new_end:
            raise HTTPException(
                status_code=400,
                detail=f"Date range overlaps with existing entry (ID: {existing.get('id')[:8]}..., Active: {exist_start} to {exist_end if exist_end != '9999-12-31' else 'ongoing'}). Please adjust dates to avoid overlap."
            )
    
    # Get SKU name if not provided
    sku_name = data.sku_name
    if not sku_name and data.sku_id:
        sku = await db.skus.find_one({"id": data.sku_id}, {"_id": 0, "name": 1})
        if sku:
            sku_name = sku.get('name')
    
    # Calculate transfer price based on billing approach
    distributor = await db.distributors.find_one(
        {"id": distributor_id, "tenant_id": tenant_id},
        {"_id": 0, "billing_approach": 1}
    )
    billing_approach = (distributor or {}).get('billing_approach', 'margin_upfront')
    
    transfer_price = None
    if data.base_price:
        if billing_approach == 'cost_based':
            transfer_price = data.base_price
        elif data.margin_type == 'percentage':
            transfer_price = data.base_price * (1 - data.margin_value / 100)
    
    margin_doc = {
        "id": str(uuid.uuid4()),
        "tenant_id": tenant_id,
        "distributor_id": distributor_id,
        "state": data.state,
        "city": data.city,
        "sku_id": data.sku_id,
        "sku_name": sku_name,
        "base_price": data.base_price,
        "margin_type": data.margin_type,
        "margin_value": data.margin_value,
        "transfer_price": round(transfer_price, 2) if transfer_price else None,
        "min_quantity": data.min_quantity,
        "max_quantity": data.max_quantity,
        "active_from": data.active_from or now[:10],
        "active_to": data.active_to,
        "remarks": data.remarks,
        "status": data.status or "active",
        "created_at": now,
        "updated_at": now,
        "created_by": current_user.get('id')
    }
    
    await db.distributor_margin_matrix.insert_one(margin_doc)
    margin_doc.pop('_id', None)
    
    logger.info(f"Margin entry created for distributor {distributor_id}, city {data.city}, SKU {sku_name} by {current_user['email']}")
    
    return margin_doc


@router.post("/{distributor_id}/margins/bulk")
async def create_bulk_margin_entries(
    distributor_id: str,
    data: List[MarginMatrixCreate],
    current_user: dict = Depends(get_current_user)
):
    """Create multiple margin matrix entries"""
    if not is_distributor_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    tenant_id = get_current_tenant_id()
    now = datetime.now(timezone.utc).isoformat()
    
    # Check if distributor exists
    distributor = await db.distributors.find_one({"id": distributor_id, "tenant_id": tenant_id})
    if not distributor:
        raise HTTPException(status_code=404, detail="Distributor not found")
    
    billing_approach = distributor.get('billing_approach', 'margin_upfront')
    added = []
    skipped = []
    
    for item in data:
        # Check if city is in operating coverage
        coverage = await db.distributor_operating_coverage.find_one({
            "tenant_id": tenant_id,
            "distributor_id": distributor_id,
            "city": item.city,
            "status": "active"
        })
        if not coverage:
            skipped.append({"city": item.city, "sku_id": item.sku_id, "reason": "City not in coverage"})
            continue
        
        # Check for duplicate
        existing = await db.distributor_margin_matrix.find_one({
            "tenant_id": tenant_id,
            "distributor_id": distributor_id,
            "city": item.city,
            "sku_id": item.sku_id,
            "status": "active"
        })
        if existing:
            skipped.append({"city": item.city, "sku_id": item.sku_id, "reason": "Already exists"})
            continue
        
        # Get SKU name
        sku_name = item.sku_name
        if not sku_name and item.sku_id:
            sku = await db.skus.find_one({"id": item.sku_id}, {"_id": 0, "name": 1})
            if sku:
                sku_name = sku.get('name')
        
        # Calculate transfer price based on billing approach
        transfer_price = None
        if item.base_price:
            if billing_approach == 'cost_based':
                transfer_price = item.base_price
            elif item.margin_type == 'percentage':
                transfer_price = item.base_price * (1 - item.margin_value / 100)
        
        margin_doc = {
            "id": str(uuid.uuid4()),
            "tenant_id": tenant_id,
            "distributor_id": distributor_id,
            "state": item.state,
            "city": item.city,
            "sku_id": item.sku_id,
            "sku_name": sku_name,
            "base_price": item.base_price,
            "margin_type": item.margin_type,
            "margin_value": item.margin_value,
            "transfer_price": round(transfer_price, 2) if transfer_price else None,
            "min_quantity": item.min_quantity,
            "max_quantity": item.max_quantity,
            "active_from": item.active_from or now[:10],
            "active_to": item.active_to,
            "remarks": item.remarks,
            "status": item.status or "active",
            "created_at": now,
            "updated_at": now,
            "created_by": current_user.get('id')
        }
        
        await db.distributor_margin_matrix.insert_one(margin_doc)
        margin_doc.pop('_id', None)
        added.append(margin_doc)
    
    return {
        "added": added,
        "added_count": len(added),
        "skipped": skipped,
        "skipped_count": len(skipped)
    }


@router.put("/{distributor_id}/margins/{margin_id}")
async def update_margin_entry(
    distributor_id: str,
    margin_id: str,
    data: MarginMatrixUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Update a margin matrix entry"""
    if not is_distributor_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    tenant_id = get_current_tenant_id()
    
    margin = await db.distributor_margin_matrix.find_one({
        "id": margin_id,
        "tenant_id": tenant_id,
        "distributor_id": distributor_id
    })
    
    if not margin:
        raise HTTPException(status_code=404, detail="Margin entry not found")
    
    update_data = {"updated_at": datetime.now(timezone.utc).isoformat()}
    
    for field in ['state', 'city', 'sku_id', 'sku_name', 'base_price', 'margin_type', 'margin_value',
                  'min_quantity', 'max_quantity', 'active_from', 'active_to',
                  'remarks', 'status']:
        value = getattr(data, field, None)
        if value is not None:
            update_data[field] = value
    
    # Validate margin type if being updated
    if data.margin_type and data.margin_type not in MARGIN_TYPES:
        raise HTTPException(status_code=400, detail=f"Invalid margin type. Must be one of: {list(MARGIN_TYPES.keys())}")
    
    # Check for date overlap if dates are being changed
    if data.active_from is not None or data.active_to is not None:
        new_start = update_data.get('active_from', margin.get('active_from')) or "1900-01-01"
        new_end = update_data.get('active_to', margin.get('active_to')) or "9999-12-31"
        
        # Get other entries for same city+SKU (excluding current one)
        other_entries = await db.distributor_margin_matrix.find({
            "tenant_id": tenant_id,
            "distributor_id": distributor_id,
            "city": margin.get('city'),
            "sku_id": margin.get('sku_id'),
            "status": "active",
            "id": {"$ne": margin_id}  # Exclude current entry
        }).to_list(100)
        
        for existing in other_entries:
            exist_start = existing.get('active_from') or "1900-01-01"
            exist_end = existing.get('active_to') or "9999-12-31"
            
            # Check for overlap
            if new_start <= exist_end and exist_start <= new_end:
                raise HTTPException(
                    status_code=400,
                    detail=f"Date range overlaps with existing entry (ID: {existing.get('id')[:8]}..., Active: {exist_start} to {exist_end if exist_end != '9999-12-31' else 'ongoing'}). Please adjust dates to avoid overlap."
                )
    
    # Recalculate transfer price based on billing approach
    base_price = update_data.get('base_price', margin.get('base_price'))
    margin_type = update_data.get('margin_type', margin.get('margin_type'))
    margin_value = update_data.get('margin_value', margin.get('margin_value'))
    
    dist = await db.distributors.find_one(
        {"id": distributor_id, "tenant_id": tenant_id},
        {"_id": 0, "billing_approach": 1}
    )
    billing_approach = (dist or {}).get('billing_approach', 'margin_upfront')
    
    if base_price:
        if billing_approach == 'cost_based':
            update_data['transfer_price'] = round(base_price, 2)
        elif margin_type == 'percentage':
            update_data['transfer_price'] = round(base_price * (1 - margin_value / 100), 2)
    
    await db.distributor_margin_matrix.update_one(
        {"id": margin_id, "tenant_id": tenant_id},
        {"$set": update_data}
    )
    
    updated = await db.distributor_margin_matrix.find_one(
        {"id": margin_id, "tenant_id": tenant_id},
        {"_id": 0}
    )
    
    logger.info(f"Margin entry {margin_id} updated by {current_user['email']}")
    
    return updated


@router.delete("/{distributor_id}/margins/{margin_id}")
async def delete_margin_entry(
    distributor_id: str,
    margin_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Delete a margin matrix entry"""
    if not is_distributor_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    tenant_id = get_current_tenant_id()
    
    margin = await db.distributor_margin_matrix.find_one({
        "id": margin_id,
        "tenant_id": tenant_id,
        "distributor_id": distributor_id
    })
    
    if not margin:
        raise HTTPException(status_code=404, detail="Margin entry not found")
    
    await db.distributor_margin_matrix.delete_one({
        "id": margin_id,
        "tenant_id": tenant_id
    })
    
    logger.info(f"Margin entry {margin_id} deleted by {current_user['email']}")
    
    return {"message": "Margin entry deleted successfully"}


@router.get("/{distributor_id}/margins/cities/list")
async def get_margin_cities(
    distributor_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get list of cities with margin entries for this distributor"""
    tenant_id = get_current_tenant_id()
    
    margins = await db.distributor_margin_matrix.find(
        {"tenant_id": tenant_id, "distributor_id": distributor_id},
        {"_id": 0, "city": 1}
    ).to_list(1000)
    
    cities = list(set([m['city'] for m in margins]))
    cities.sort()
    
    return {"cities": cities}


@router.get("/{distributor_id}/margins/skus/list")
async def get_margin_skus(
    distributor_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get list of SKUs with margin entries for this distributor"""
    tenant_id = get_current_tenant_id()
    
    margins = await db.distributor_margin_matrix.find(
        {"tenant_id": tenant_id, "distributor_id": distributor_id},
        {"_id": 0, "sku_id": 1, "sku_name": 1}
    ).to_list(1000)
    
    # Deduplicate by sku_id
    skus = {}
    for m in margins:
        if m['sku_id'] not in skus:
            skus[m['sku_id']] = {"id": m['sku_id'], "name": m.get('sku_name', m['sku_id'])}
    
    return {"skus": list(skus.values())}



# ============ Account-Distributor Assignment CRUD ============

@router.get("/assignments/all")
async def list_all_account_assignments(
    distributor_id: Optional[str] = None,
    account_id: Optional[str] = None,
    city: Optional[str] = None,
    status: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    current_user: dict = Depends(get_current_user)
):
    """List all account-distributor assignments"""
    tenant_id = get_current_tenant_id()
    
    query = {"tenant_id": tenant_id}
    
    if distributor_id:
        query["distributor_id"] = distributor_id
    if account_id:
        query["account_id"] = account_id
    if city and city != 'all':
        query["servicing_city"] = city
    if status and status != 'all':
        query["status"] = status
    
    total = await db.account_distributor_assignments.count_documents(query)
    
    assignments = await db.account_distributor_assignments.find(
        query,
        {"_id": 0}
    ).sort([("servicing_city", 1), ("account_name", 1)]).skip((page - 1) * page_size).limit(page_size).to_list(page_size)
    
    return {
        "assignments": assignments,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size
    }


@router.get("/{distributor_id}/assignments")
async def list_distributor_account_assignments(
    distributor_id: str,
    city: Optional[str] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """List account assignments for a specific distributor"""
    tenant_id = get_current_tenant_id()
    
    query = {"tenant_id": tenant_id, "distributor_id": distributor_id}
    
    if city and city != 'all':
        query["servicing_city"] = city
    if status and status != 'all':
        query["status"] = status
    
    assignments = await db.account_distributor_assignments.find(
        query,
        {"_id": 0}
    ).sort([("servicing_city", 1), ("account_name", 1)]).to_list(1000)

    # Enrich each assignment with the account's GST / PAN / billing details so the
    # distributor's Account Assignments view can surface compliance info without
    # an extra round-trip. Read-only — distributors don't edit account-level data.
    account_ids = list({a.get('account_id') for a in assignments if a.get('account_id')})
    accounts_map: Dict[str, Dict[str, Any]] = {}
    if account_ids:
        async for ad in db.accounts.find(
            {"id": {"$in": account_ids}, "tenant_id": tenant_id},
            {"_id": 0, "id": 1, "gst_number": 1, "pan_number": 1, "billing_address": 1,
             "gst_legal_name": 1, "gst_trade_name": 1, "contact_name": 1, "contact_number": 1,
             "delivery_address": 1, "delivery_contact_name": 1, "delivery_contact_phone": 1}
        ):
            accounts_map[ad['id']] = ad
    for a in assignments:
        acc = accounts_map.get(a.get('account_id')) or {}
        a['gst_number'] = acc.get('gst_number')
        a['pan_number'] = acc.get('pan_number')
        a['billing_address'] = acc.get('billing_address')
        a['gst_legal_name'] = acc.get('gst_legal_name')
        a['gst_trade_name'] = acc.get('gst_trade_name')
        a['account_contact_name'] = acc.get('contact_name')
        a['account_contact_number'] = acc.get('contact_number')
        a['delivery_address'] = acc.get('delivery_address')
        a['delivery_contact_name'] = acc.get('delivery_contact_name')
        a['delivery_contact_phone'] = acc.get('delivery_contact_phone')
    
    # Group by city
    by_city = {}
    for a in assignments:
        city_name = a.get('servicing_city', 'Unknown')
        if city_name not in by_city:
            by_city[city_name] = []
        by_city[city_name].append(a)
    
    return {
        "assignments": assignments,
        "by_city": by_city,
        "total": len(assignments)
    }


@router.get("/{distributor_id}/assignments/{assignment_id}")
async def get_assignment(
    distributor_id: str,
    assignment_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get a specific account-distributor assignment"""
    tenant_id = get_current_tenant_id()
    
    assignment = await db.account_distributor_assignments.find_one(
        {"id": assignment_id, "tenant_id": tenant_id, "distributor_id": distributor_id},
        {"_id": 0}
    )
    
    if not assignment:
        raise HTTPException(status_code=404, detail="Assignment not found")
    
    return assignment


@router.post("/{distributor_id}/assignments")
async def create_account_assignment(
    distributor_id: str,
    data: AccountDistributorCreate,
    current_user: dict = Depends(get_current_user)
):
    """Create a new account-distributor assignment"""
    if not is_distributor_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    tenant_id = get_current_tenant_id()
    now = datetime.now(timezone.utc).isoformat()
    
    # Validate distributor exists
    distributor = await db.distributors.find_one({"id": distributor_id, "tenant_id": tenant_id})
    if not distributor:
        raise HTTPException(status_code=404, detail="Distributor not found")
    
    # Validate account exists
    account = await db.accounts.find_one({"id": data.account_id, "tenant_id": tenant_id})
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    
    # Validate account's city matches servicing city (alias-aware: Bangalore <-> Bengaluru, etc.)
    account_city = account.get('city', '')
    if account_city and data.servicing_city and not cities_match(account_city, data.servicing_city):
        raise HTTPException(
            status_code=400,
            detail=f"Account is located in '{account_city}' but servicing city is '{data.servicing_city}'. Account can only be assigned to distributors serving its city."
        )

    # Validate city is in distributor's operating coverage (alias-aware)
    normalized_servicing = normalize_city(data.servicing_city)
    coverage = None
    if normalized_servicing:
        # Try exact match first (case-insensitive)
        coverage = await db.distributor_operating_coverage.find_one({
            "tenant_id": tenant_id,
            "distributor_id": distributor_id,
            "city": {"$regex": f"^{data.servicing_city}$", "$options": "i"},
            "status": "active"
        })
        if not coverage:
            # Fallback: scan all active coverage rows for this distributor and alias-match
            async for cov in db.distributor_operating_coverage.find({
                "tenant_id": tenant_id,
                "distributor_id": distributor_id,
                "status": "active"
            }):
                if normalize_city(cov.get('city', '')) == normalized_servicing:
                    coverage = cov
                    break
    if not coverage:
        raise HTTPException(
            status_code=400,
            detail=f"City '{data.servicing_city}' is not in distributor's operating coverage"
        )
    
    # Validate distributor location if provided
    if data.distributor_location_id:
        location = await db.distributor_locations.find_one({
            "id": data.distributor_location_id,
            "tenant_id": tenant_id,
            "distributor_id": distributor_id,
            "status": "active"
        })
        if not location:
            raise HTTPException(status_code=400, detail="Invalid distributor location")
    
    # ── DEDUP: don't allow the same (account, distributor, city) to be assigned twice. ──
    # An active assignment for this exact triple already exists → return 409 with the
    # existing assignment so the caller can surface a clear message and link the user
    # to it instead of inserting a duplicate row.
    existing_same = await db.account_distributor_assignments.find_one({
        "tenant_id": tenant_id,
        "account_id": data.account_id,
        "distributor_id": distributor_id,
        "status": "active",
    }, {"_id": 0})
    if existing_same:
        # Match if the city aliases match (Bangalore <-> Bengaluru, etc.)
        if cities_match(existing_same.get('servicing_city', ''), data.servicing_city):
            raise HTTPException(
                status_code=409,
                detail=(
                    f"Account '{existing_same.get('account_name')}' is already assigned to "
                    f"distributor '{existing_same.get('distributor_name')}' for "
                    f"'{existing_same.get('servicing_city')}'. Edit the existing assignment "
                    "instead of creating a new one."
                ),
            )

    # Check for existing primary assignment for same account + city (alias-aware)
    if data.is_primary:
        existing_primaries = db.account_distributor_assignments.find({
            "tenant_id": tenant_id,
            "account_id": data.account_id,
            "is_primary": True,
            "status": "active"
        })
        async for existing_primary in existing_primaries:
            if cities_match(existing_primary.get('servicing_city', ''), data.servicing_city) \
               and existing_primary.get('distributor_id') != distributor_id:
                raise HTTPException(
                    status_code=400,
                    detail=f"Account already has a primary distributor for {data.servicing_city}. Remove or change existing assignment first."
                )
    
    # Get names for denormalization
    account_name = data.account_name or account.get('account_name') or account.get('company') or account.get('name')
    distributor_name = data.distributor_name or distributor.get('distributor_name')
    location_name = data.distributor_location_name
    if data.distributor_location_id and not location_name:
        loc = await db.distributor_locations.find_one({"id": data.distributor_location_id}, {"_id": 0, "location_name": 1})
        location_name = loc.get('location_name') if loc else None
    
    assignment_doc = {
        "id": str(uuid.uuid4()),
        "tenant_id": tenant_id,
        "account_id": data.account_id,
        "account_name": account_name,
        "distributor_id": distributor_id,
        "distributor_name": distributor_name,
        "servicing_state": data.servicing_state,
        "servicing_city": data.servicing_city,
        "distributor_location_id": data.distributor_location_id,
        "distributor_location_name": location_name,
        "is_primary": data.is_primary if data.is_primary is not None else True,
        "is_backup": data.is_backup if data.is_backup is not None else False,
        "has_special_override": data.has_special_override or False,
        "override_type": data.override_type,
        "override_value": data.override_value,
        "effective_from": data.effective_from or now,
        "effective_to": data.effective_to,
        "remarks": data.remarks,
        "status": data.status or "active",
        "created_at": now,
        "updated_at": now,
        "created_by": current_user.get('id')
    }
    
    await db.account_distributor_assignments.insert_one(assignment_doc)
    assignment_doc.pop('_id', None)
    
    logger.info(f"Account '{account_name}' assigned to distributor '{distributor_name}' in {data.servicing_city} by {current_user['email']}")
    
    return assignment_doc


@router.put("/{distributor_id}/assignments/{assignment_id}")
async def update_account_assignment(
    distributor_id: str,
    assignment_id: str,
    data: AccountDistributorUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Update an account-distributor assignment"""
    if not is_distributor_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    tenant_id = get_current_tenant_id()
    
    assignment = await db.account_distributor_assignments.find_one({
        "id": assignment_id,
        "tenant_id": tenant_id,
        "distributor_id": distributor_id
    })
    
    if not assignment:
        raise HTTPException(status_code=404, detail="Assignment not found")
    
    update_data = {"updated_at": datetime.now(timezone.utc).isoformat()}
    
    # Validate distributor location if being updated
    if data.distributor_location_id:
        location = await db.distributor_locations.find_one({
            "id": data.distributor_location_id,
            "tenant_id": tenant_id,
            "distributor_id": distributor_id,
            "status": "active"
        })
        if not location:
            raise HTTPException(status_code=400, detail="Invalid distributor location")
        update_data["distributor_location_name"] = location.get('location_name')
    
    for field in ['distributor_id', 'distributor_name', 'servicing_state', 'servicing_city',
                  'distributor_location_id', 'distributor_location_name', 'is_primary', 'is_backup',
                  'has_special_override', 'override_type', 'override_value',
                  'effective_from', 'effective_to', 'remarks', 'status']:
        value = getattr(data, field, None)
        if value is not None:
            update_data[field] = value
    
    await db.account_distributor_assignments.update_one(
        {"id": assignment_id, "tenant_id": tenant_id},
        {"$set": update_data}
    )
    
    updated = await db.account_distributor_assignments.find_one(
        {"id": assignment_id, "tenant_id": tenant_id},
        {"_id": 0}
    )
    
    logger.info(f"Assignment {assignment_id} updated by {current_user['email']}")
    
    return updated


@router.delete("/{distributor_id}/assignments/{assignment_id}")
async def delete_account_assignment(
    distributor_id: str,
    assignment_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Delete an account-distributor assignment"""
    if not is_distributor_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    tenant_id = get_current_tenant_id()
    
    assignment = await db.account_distributor_assignments.find_one({
        "id": assignment_id,
        "tenant_id": tenant_id,
        "distributor_id": distributor_id
    })
    
    if not assignment:
        raise HTTPException(status_code=404, detail="Assignment not found")
    
    await db.account_distributor_assignments.delete_one({
        "id": assignment_id,
        "tenant_id": tenant_id
    })
    
    logger.info(f"Assignment for account '{assignment.get('account_name')}' deleted by {current_user['email']}")
    
    return {"message": "Assignment deleted successfully"}


@router.get("/{distributor_id}/assignments/cities/list")
async def get_assignment_cities(
    distributor_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get list of cities with account assignments for this distributor"""
    tenant_id = get_current_tenant_id()
    
    assignments = await db.account_distributor_assignments.find(
        {"tenant_id": tenant_id, "distributor_id": distributor_id},
        {"_id": 0, "servicing_city": 1}
    ).to_list(1000)
    
    cities = list(set([a['servicing_city'] for a in assignments]))
    cities.sort()
    
    return {"cities": cities}


# Get account's distributor assignments (for account detail page)
@router.get("/account/{account_id}/distributors")
async def get_account_distributors(
    account_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get distributor assignments for a specific account"""
    tenant_id = get_current_tenant_id()
    
    assignments = await db.account_distributor_assignments.find(
        {"tenant_id": tenant_id, "account_id": account_id},
        {"_id": 0}
    ).sort([("is_primary", -1), ("servicing_city", 1)]).to_list(100)
    
    return {
        "assignments": assignments,
        "total": len(assignments)
    }


# Search accounts for assignment
@router.get("/accounts/search")
async def search_accounts_for_assignment(
    q: str = Query(..., min_length=2),
    city: Optional[str] = None,
    limit: int = Query(20, le=50),
    current_user: dict = Depends(get_current_user)
):
    """Search accounts for distributor assignment"""
    tenant_id = get_current_tenant_id()
    
    query = {
        "tenant_id": tenant_id,
        "$or": [
            {"account_name": {"$regex": q, "$options": "i"}},
            {"contact_name": {"$regex": q, "$options": "i"}},
            {"account_id": {"$regex": q, "$options": "i"}},
            {"city": {"$regex": q, "$options": "i"}}
        ]
    }
    
    if city:
        query["city"] = city
    
    accounts = await db.accounts.find(
        query,
        {"_id": 0, "id": 1, "account_name": 1, "contact_name": 1, "city": 1, "state": 1, "account_id": 1, "territory": 1, "contact_number": 1, "delivery_address": 1}
    ).limit(limit).to_list(limit)
    
    return {"accounts": accounts}


@router.get("/{distributor_id}/search-assignable-accounts")
async def search_assignable_accounts_for_distributor(
    distributor_id: str,
    q: str = Query(..., min_length=2),
    limit: int = Query(20, le=50),
    current_user: dict = Depends(get_current_user)
):
    """Search accounts that can be assigned to this distributor (only accounts in covered cities)"""
    tenant_id = get_current_tenant_id()
    
    # Get distributor's covered cities
    coverage = await db.distributor_operating_coverage.find(
        {"tenant_id": tenant_id, "distributor_id": distributor_id, "status": "active"},
        {"_id": 0, "city": 1}
    ).to_list(100)
    
    covered_cities = [c.get('city') for c in coverage if c.get('city')]
    
    if not covered_cities:
        return {"accounts": [], "message": "Distributor has no operating coverage configured. Please add coverage first."}

    # Expand covered cities with their aliases so e.g. "Bengaluru" coverage also matches accounts in "Bangalore"
    expanded_cities = set()
    for city in covered_cities:
        expanded_cities.add(city)
        normalized = normalize_city(city)
        for alias_key, alias_canonical in CITY_ALIASES.items():
            if alias_canonical == normalized:
                expanded_cities.add(alias_key.title())
                expanded_cities.add(alias_key)
    # Build case-insensitive city match
    city_regex_list = [{"city": {"$regex": f"^{c}$", "$options": "i"}} for c in expanded_cities]

    # Search accounts only in covered cities (alias-aware)
    query = {
        "tenant_id": tenant_id,
        "$and": [
            {"$or": city_regex_list},
            {"$or": [
                {"account_name": {"$regex": q, "$options": "i"}},
                {"contact_name": {"$regex": q, "$options": "i"}},
                {"account_id": {"$regex": q, "$options": "i"}},
                {"city": {"$regex": q, "$options": "i"}}
            ]}
        ]
    }
    
    accounts = await db.accounts.find(
        query,
        {"_id": 0, "id": 1, "account_name": 1, "contact_name": 1, "city": 1, "state": 1, "account_id": 1, "territory": 1, "contact_number": 1, "delivery_address": 1}
    ).limit(limit).to_list(limit)
    
    return {
        "accounts": accounts,
        "covered_cities": covered_cities
    }


# ============ Primary Shipment / Stock Receipt CRUD ============

SHIPMENT_STATUSES = {
    "draft": "Draft - Not yet confirmed",
    "confirmed": "Confirmed - Ready for dispatch",
    "in_transit": "In Transit - On the way to distributor",
    "delivered": "Delivered - Received by distributor",
    "partially_delivered": "Partially Delivered - Some items received",
    "discrepancy_pending": "Discrepancy - Awaiting supplier approval",
    "cancelled": "Cancelled"
}


async def generate_shipment_number(tenant_id: str) -> str:
    """Generate unique shipment number like SHP-2026-0001"""
    year = datetime.now().year
    count = await db.distributor_shipments.count_documents({
        "tenant_id": tenant_id,
        "shipment_number": {"$regex": f"^SHP-{year}-"}
    })
    return f"SHP-{year}-{count + 1:04d}"


def calculate_item_amounts(item: dict) -> dict:
    """Calculate amounts for a shipment item"""
    quantity = item.get('quantity', 0)
    unit_price = item.get('unit_price', 0)
    discount_percent = item.get('discount_percent', 0) or 0
    tax_percent = item.get('tax_percent', 0) or 0
    
    gross_amount = quantity * unit_price
    discount_amount = round(gross_amount * discount_percent / 100, 2)
    taxable_amount = gross_amount - discount_amount
    tax_amount = round(taxable_amount * tax_percent / 100, 2)
    net_amount = taxable_amount + tax_amount
    
    return {
        **item,
        'gross_amount': round(gross_amount, 2),
        'discount_amount': discount_amount,
        'taxable_amount': round(taxable_amount, 2),
        'tax_amount': tax_amount,
        'net_amount': round(net_amount, 2)
    }


@router.get("/shipments/all")
async def list_all_shipments(
    distributor_id: Optional[str] = None,
    location_id: Optional[str] = None,
    status: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    current_user: dict = Depends(get_current_user)
):
    """List all primary shipments with filters"""
    tenant_id = get_current_tenant_id()
    
    query = {"tenant_id": tenant_id}
    
    if distributor_id:
        query["distributor_id"] = distributor_id
    if location_id:
        query["distributor_location_id"] = location_id
    if status and status != 'all':
        query["status"] = status
    if from_date:
        query["shipment_date"] = {"$gte": from_date}
    if to_date:
        if "shipment_date" in query:
            query["shipment_date"]["$lte"] = to_date
        else:
            query["shipment_date"] = {"$lte": to_date}
    
    total = await db.distributor_shipments.count_documents(query)
    
    shipments = await db.distributor_shipments.find(
        query,
        {"_id": 0}
    ).sort("shipment_date", -1).skip((page - 1) * page_size).limit(page_size).to_list(page_size)
    
    return {
        "shipments": shipments,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size
    }


@router.get("/shipments/summary")
async def get_shipments_summary(
    distributor_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get shipments summary stats"""
    tenant_id = get_current_tenant_id()
    
    query = {"tenant_id": tenant_id}
    if distributor_id:
        query["distributor_id"] = distributor_id
    
    total = await db.distributor_shipments.count_documents(query)
    
    # Count by status
    draft = await db.distributor_shipments.count_documents({**query, "status": "draft"})
    confirmed = await db.distributor_shipments.count_documents({**query, "status": "confirmed"})
    in_transit = await db.distributor_shipments.count_documents({**query, "status": "in_transit"})
    delivered = await db.distributor_shipments.count_documents({**query, "status": "delivered"})
    cancelled = await db.distributor_shipments.count_documents({**query, "status": "cancelled"})
    
    # Calculate totals
    pipeline = [
        {"$match": {**query, "status": {"$ne": "cancelled"}}},
        {"$group": {
            "_id": None,
            "total_quantity": {"$sum": "$total_quantity"},
            "total_amount": {"$sum": "$total_net_amount"}
        }}
    ]
    
    totals = await db.distributor_shipments.aggregate(pipeline).to_list(1)
    total_quantity = totals[0]["total_quantity"] if totals else 0
    total_amount = totals[0]["total_amount"] if totals else 0
    
    return {
        "total": total,
        "by_status": {
            "draft": draft,
            "confirmed": confirmed,
            "in_transit": in_transit,
            "delivered": delivered,
            "cancelled": cancelled
        },
        "total_quantity": total_quantity,
        "total_amount": round(total_amount, 2)
    }


@router.get("/{distributor_id}/shipments")
async def list_distributor_shipments(
    distributor_id: str,
    status: Optional[str] = None,
    location_id: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    current_user: dict = Depends(get_current_user)
):
    """List shipments for a specific distributor with item aggregates"""
    tenant_id = get_current_tenant_id()
    
    query = {"tenant_id": tenant_id, "distributor_id": distributor_id}
    
    if status and status != 'all':
        query["status"] = status

    if location_id and location_id != 'all':
        query["distributor_location_id"] = location_id
    
    total = await db.distributor_shipments.count_documents(query)
    
    shipments = await db.distributor_shipments.find(
        query,
        {"_id": 0}
    ).sort("shipment_date", -1).skip((page - 1) * page_size).limit(page_size).to_list(page_size)
    
    # For each shipment, get aggregated item data
    for shipment in shipments:
        items = await db.distributor_shipment_items.find(
            {"shipment_id": shipment['id'], "tenant_id": tenant_id},
            {"_id": 0, "base_price": 1, "distributor_margin": 1, "unit_price": 1, "tax_percent": 1, "quantity": 1}
        ).to_list(500)
        
        if items:
            # Calculate weighted averages
            total_qty = sum(item.get('quantity', 0) for item in items)
            if total_qty > 0:
                # Weighted average base price
                weighted_base = sum((item.get('base_price') or 0) * item.get('quantity', 0) for item in items)
                shipment['avg_base_price'] = round(weighted_base / total_qty, 2) if weighted_base else None
                
                # Weighted average margin
                weighted_margin = sum((item.get('distributor_margin') or 0) * item.get('quantity', 0) for item in items)
                shipment['avg_distributor_margin'] = round(weighted_margin / total_qty, 2) if weighted_margin else None
                
                # Weighted average transfer price
                weighted_transfer = sum((item.get('unit_price') or 0) * item.get('quantity', 0) for item in items)
                shipment['avg_transfer_price'] = round(weighted_transfer / total_qty, 2) if weighted_transfer else None
                
                # Weighted average GST
                weighted_gst = sum((item.get('tax_percent') or 0) * item.get('quantity', 0) for item in items)
                shipment['avg_gst_percent'] = round(weighted_gst / total_qty, 2) if weighted_gst else None
    
    return {
        "shipments": shipments,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size
    }


@router.get("/{distributor_id}/shipments/export")
async def export_distributor_shipments(
    distributor_id: str,
    status: Optional[str] = None,
    location_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Export the (filtered) Stock In shipments list to an Excel (.xlsx) file."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse

    tenant_id = get_current_tenant_id()

    distributor = await db.distributors.find_one(
        {"id": distributor_id, "tenant_id": tenant_id}, {"_id": 0, "distributor_name": 1}
    )
    if not distributor:
        raise HTTPException(status_code=404, detail="Distributor not found")

    query = {"tenant_id": tenant_id, "distributor_id": distributor_id}
    if status and status != 'all':
        query["status"] = status
    if location_id and location_id != 'all':
        query["distributor_location_id"] = location_id

    shipments = await db.distributor_shipments.find(
        query, {"_id": 0}
    ).sort("shipment_date", -1).to_list(5000)

    # Resolve weighted-average GST% per shipment (mirrors the list endpoint)
    for shipment in shipments:
        items = await db.distributor_shipment_items.find(
            {"shipment_id": shipment['id'], "tenant_id": tenant_id},
            {"_id": 0, "tax_percent": 1, "quantity": 1}
        ).to_list(500)
        total_qty = sum(item.get('quantity', 0) for item in items)
        if total_qty > 0:
            weighted_gst = sum((item.get('tax_percent') or 0) * item.get('quantity', 0) for item in items)
            shipment['avg_gst_percent'] = round(weighted_gst / total_qty, 2) if weighted_gst else 0

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock In"

    headers = ['Shipment #', 'Reference / PO', 'Date', 'Destination Location', 'From Warehouse',
               'Status', 'Total Qty', 'Subtotal', 'Discount', 'GST %', 'GST Amount', 'Total']
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="166534", end_color="166534", fill_type="solid")
    thin = Side(style="thin", color="d0d5dd")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c_idx, name in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    widths = [18, 18, 14, 26, 24, 14, 12, 16, 14, 10, 14, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    for r_idx, s in enumerate(shipments, 2):
        gst_pct = s.get('gst_percent')
        if gst_pct is None:
            gst_pct = s.get('avg_gst_percent') or 0
        row = [
            s.get('shipment_number') or '-',
            s.get('reference_number') or '-',
            (s.get('shipment_date') or '')[:10] if isinstance(s.get('shipment_date'), str) else (s.get('shipment_date') or '-'),
            s.get('distributor_location_name') or '-',
            s.get('source_warehouse_name') or '-',
            s.get('status') or '-',
            s.get('total_quantity') or 0,
            round(s.get('total_gross_amount') or 0, 2),
            round(s.get('total_discount_amount') or 0, 2),
            round(gst_pct or 0, 2),
            round(s.get('total_tax_amount') or 0, 2),
            round(s.get('total_net_amount') or 0, 2),
        ]
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border
            if c_idx == 7:
                cell.number_format = '#,##0'
            elif c_idx in (8, 9, 11, 12):
                cell.number_format = '#,##0.00'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"stock_in_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{distributor_id}/shipments/{shipment_id}")
async def get_shipment(
    distributor_id: str,
    shipment_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get a specific shipment with items"""
    tenant_id = get_current_tenant_id()
    
    shipment = await db.distributor_shipments.find_one(
        {"id": shipment_id, "tenant_id": tenant_id, "distributor_id": distributor_id},
        {"_id": 0}
    )
    
    if not shipment:
        raise HTTPException(status_code=404, detail="Shipment not found")
    
    # Get shipment items
    items = await db.distributor_shipment_items.find(
        {"shipment_id": shipment_id, "tenant_id": tenant_id},
        {"_id": 0}
    ).to_list(500)
    
    shipment['items'] = items
    
    return shipment


@router.post("/{distributor_id}/shipments")
async def create_shipment(
    distributor_id: str,
    data: PrimaryShipmentCreate,
    current_user: dict = Depends(get_current_user)
):
    """Create a new primary shipment"""
    if not is_distributor_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    tenant_id = get_current_tenant_id()
    now = datetime.now(timezone.utc).isoformat()
    
    # Validate distributor exists
    distributor = await db.distributors.find_one(
        {"id": distributor_id, "tenant_id": tenant_id},
        {"_id": 0, "distributor_name": 1, "distributor_code": 1}
    )
    if not distributor:
        raise HTTPException(status_code=404, detail="Distributor not found")
    
    # Validate location exists and belongs to distributor
    location = await db.distributor_locations.find_one(
        {"id": data.distributor_location_id, "tenant_id": tenant_id, "distributor_id": distributor_id},
        {"_id": 0, "location_name": 1, "address_line_1": 1, "city": 1, "state": 1, "pincode": 1, "track_batches": 1}
    )
    if not location:
        raise HTTPException(status_code=400, detail="Invalid distributor location")
    dest_tracks_batches = bool(location.get("track_batches"))
    
    # Validate source factory warehouse if provided
    source_warehouse_name = None
    source_tracks_batches = False
    if data.source_warehouse_id:
        source_warehouse = await db.distributor_locations.find_one(
            {"id": data.source_warehouse_id, "tenant_id": tenant_id, "is_factory": True, "status": "active"},
            {"_id": 0, "location_name": 1, "track_batches": 1}
        )
        if not source_warehouse:
            raise HTTPException(status_code=400, detail="Invalid source factory warehouse")
        source_warehouse_name = source_warehouse.get('location_name')
        source_tracks_batches = bool(source_warehouse.get('track_batches'))

    # Validate items
    if not data.items or len(data.items) == 0:
        raise HTTPException(status_code=400, detail="At least one item is required")

    # Phase 2 batch tracking: when EITHER the source factory warehouse has
    # `track_batches=True` OR the destination distributor location does,
    # every line MUST carry a `batch_id`. Reject up-front with a friendly
    # error if any line is missing one.
    if source_tracks_batches or dest_tracks_batches:
        missing_batch = [
            (item.sku_name or item.sku_id)
            for item in data.items
            if not item.batch_id
        ]
        if missing_batch:
            who = (
                f"Source warehouse '{source_warehouse_name}'" if source_tracks_batches
                else f"Destination location '{location.get('location_name')}'"
            )
            raise HTTPException(
                status_code=400,
                detail=(
                    f"{who} tracks batches — "
                    f"please pick a production batch for: {', '.join(missing_batch[:5])}"
                    + (f" and {len(missing_batch) - 5} more" if len(missing_batch) > 5 else "")
                ),
            )
    
    # Generate shipment number
    shipment_number = await generate_shipment_number(tenant_id)
    shipment_id = str(uuid.uuid4())
    
    # Build shipping address from location if not provided
    shipping_address = data.shipping_address
    if not shipping_address:
        addr_parts = [location.get('location_name'), location.get('address_line_1'), 
                      location.get('city'), location.get('state'), location.get('pincode')]
        shipping_address = ', '.join([p for p in addr_parts if p])
    
    # Process items and calculate totals
    items_to_insert = []
    total_quantity = 0
    total_gross_amount = 0
    total_discount_amount = 0
    total_tax_amount = 0
    total_net_amount = 0
    
    for item_data in data.items:
        # Get SKU info if not provided
        sku_name = item_data.sku_name
        sku_code = item_data.sku_code
        if not sku_name or not sku_code:
            sku = await db.skus.find_one({"id": item_data.sku_id}, {"_id": 0, "name": 1, "sku_code": 1})
            if sku:
                sku_name = sku_name or sku.get('name')
                sku_code = sku_code or sku.get('sku_code')
        
        item_dict = {
            'id': str(uuid.uuid4()),
            'tenant_id': tenant_id,
            'shipment_id': shipment_id,
            'sku_id': item_data.sku_id,
            'sku_name': sku_name,
            'sku_code': sku_code,
            'quantity': item_data.quantity,
            'base_price': item_data.base_price,
            'distributor_margin': item_data.distributor_margin,
            'unit_price': item_data.unit_price,
            'discount_percent': item_data.discount_percent or 0,
            'tax_percent': item_data.tax_percent or 0,
            'remarks': item_data.remarks,
            # Batch identity (Phase 2). Always written so downstream code can
            # rely on the field existing — None when source doesn't track batches.
            'batch_id': item_data.batch_id,
            'batch_code': item_data.batch_code,
            # Packaging breakdown for stock-in detail view + invoice description.
            'packaging_type_name': item_data.packaging_type_name,
            'packaging_units': item_data.packaging_units,
            'packages': item_data.packages,
        }
        
        # Calculate amounts
        item_dict = calculate_item_amounts(item_dict)
        items_to_insert.append(item_dict)
        
        total_quantity += item_data.quantity
        total_gross_amount += item_dict['gross_amount']
        total_discount_amount += item_dict['discount_amount']
    
    # Calculate GST at shipment level (not per-item)
    subtotal_after_discount = total_gross_amount - total_discount_amount
    gst_pct = data.gst_percent or 0
    total_tax_amount = round(subtotal_after_discount * (gst_pct / 100), 2)
    total_net_amount = round(subtotal_after_discount + total_tax_amount, 2)
    
    # Create shipment document
    shipment_doc = {
        "id": shipment_id,
        "tenant_id": tenant_id,
        "shipment_number": shipment_number,
        "distributor_id": distributor_id,
        "distributor_name": distributor.get('distributor_name'),
        "distributor_code": distributor.get('distributor_code'),
        "distributor_location_id": data.distributor_location_id,
        "distributor_location_name": location.get('location_name'),
        "source_warehouse_id": data.source_warehouse_id,
        "source_warehouse_name": source_warehouse_name,
        "shipment_date": data.shipment_date,
        "expected_delivery_date": data.expected_delivery_date,
        "actual_delivery_date": None,
        "reference_number": data.reference_number,
        "vehicle_number": data.vehicle_number,
        "driver_name": data.driver_name,
        "driver_contact": data.driver_contact,
        "shipping_address": shipping_address,
        "status": "draft",
        "total_quantity": total_quantity,
        "total_gross_amount": round(total_gross_amount, 2),
        "total_discount_amount": round(total_discount_amount, 2),
        "total_tax_amount": round(total_tax_amount, 2),
        "total_net_amount": round(total_net_amount, 2),
        "gst_percent": data.gst_percent or 0,
        "remarks": data.remarks,
        "created_at": now,
        "updated_at": now,
        "created_by": current_user.get('id'),
        "confirmed_at": None,
        "confirmed_by": None,
        "delivered_at": None,
        "delivered_by": None
    }
    
    # Insert shipment and items
    await db.distributor_shipments.insert_one(shipment_doc)
    if items_to_insert:
        await db.distributor_shipment_items.insert_many(items_to_insert)
    
    shipment_doc.pop('_id', None)
    shipment_doc['items'] = [
        {k: v for k, v in item.items() if k not in ['_id', 'tenant_id']} 
        for item in items_to_insert
    ]
    
    logger.info(f"Shipment {shipment_number} created for distributor {distributor_id} by {current_user['email']}")
    
    return shipment_doc


@router.put("/{distributor_id}/shipments/{shipment_id}")
async def update_shipment(
    distributor_id: str,
    shipment_id: str,
    data: PrimaryShipmentUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Update a shipment (only draft shipments can be fully edited)"""
    if not is_distributor_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    tenant_id = get_current_tenant_id()
    
    shipment = await db.distributor_shipments.find_one(
        {"id": shipment_id, "tenant_id": tenant_id, "distributor_id": distributor_id}
    )
    
    if not shipment:
        raise HTTPException(status_code=404, detail="Shipment not found")
    
    current_status = shipment.get('status')
    
    # Only allow limited updates for non-draft shipments
    if current_status != 'draft':
        allowed_fields = ['remarks', 'vehicle_number', 'driver_name', 'driver_contact', 'expected_delivery_date']
        update_data = {"updated_at": datetime.now(timezone.utc).isoformat()}
        
        for field in allowed_fields:
            value = getattr(data, field, None)
            if value is not None:
                update_data[field] = value
        
        # Status transitions
        if data.status:
            valid_transitions = {
                'confirmed': ['in_transit', 'cancelled'],
                'in_transit': ['delivered', 'partially_delivered'],
                'partially_delivered': ['delivered']
            }
            if data.status not in valid_transitions.get(current_status, []):
                raise HTTPException(
                    status_code=400, 
                    detail=f"Cannot change status from '{current_status}' to '{data.status}'"
                )
            update_data['status'] = data.status
            
            # Set delivery date when marked as delivered
            if data.status in ['delivered', 'partially_delivered']:
                update_data['actual_delivery_date'] = datetime.now(timezone.utc).isoformat()[:10]
                update_data['delivered_at'] = datetime.now(timezone.utc).isoformat()
                update_data['delivered_by'] = current_user.get('id')
    else:
        # Full update for draft shipments
        update_data = {"updated_at": datetime.now(timezone.utc).isoformat()}
        
        for field in ['shipment_date', 'expected_delivery_date', 'reference_number', 
                      'vehicle_number', 'driver_name', 'driver_contact', 'shipping_address', 'remarks']:
            value = getattr(data, field, None)
            if value is not None:
                update_data[field] = value
    
    await db.distributor_shipments.update_one(
        {"id": shipment_id, "tenant_id": tenant_id},
        {"$set": update_data}
    )
    
    updated = await db.distributor_shipments.find_one(
        {"id": shipment_id, "tenant_id": tenant_id},
        {"_id": 0}
    )
    
    logger.info(f"Shipment {shipment_id} updated by {current_user['email']}")
    
    return updated


async def _safe_sync_shipment_to_zoho(tenant_id: str, distributor_id: str, shipment_id: str) -> None:
    """Background wrapper around sync_shipment_to_zoho that swallows the
    expected skip case (Zoho not configured/connected) so it doesn't surface as
    a noisy task error. Real failures are already recorded by the orchestrator."""
    from services.zoho_service import sync_shipment_to_zoho, ZohoPushSkippedError
    try:
        await sync_shipment_to_zoho(tenant_id, distributor_id, shipment_id)
    except ZohoPushSkippedError as e:
        logger.info(f"Zoho shipment push skipped for {shipment_id}: {e}")
    except Exception:
        logger.exception(f"Unexpected error during Zoho shipment push for {shipment_id}")


@router.post("/{distributor_id}/shipments/{shipment_id}/confirm")
async def confirm_shipment(
    distributor_id: str,
    shipment_id: str,
    background_tasks: BackgroundTasks,
    current_user: dict = Depends(get_current_user)
):
    """Confirm a draft shipment - marks it ready for dispatch"""
    if not is_distributor_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    tenant_id = get_current_tenant_id()
    
    shipment = await db.distributor_shipments.find_one(
        {"id": shipment_id, "tenant_id": tenant_id, "distributor_id": distributor_id}
    )
    
    if not shipment:
        raise HTTPException(status_code=404, detail="Shipment not found")
    
    if shipment.get('status') != 'draft':
        raise HTTPException(status_code=400, detail="Only draft shipments can be confirmed")
    
    # Validate items exist
    items_count = await db.distributor_shipment_items.count_documents({"shipment_id": shipment_id})
    if items_count == 0:
        raise HTTPException(status_code=400, detail="Cannot confirm shipment without items")
    
    now = datetime.now(timezone.utc).isoformat()
    
    # Deduct stock from source factory warehouse if specified
    source_warehouse_id = shipment.get('source_warehouse_id')
    if source_warehouse_id:
        tdb = get_tenant_db()
        # Check whether the source warehouse tracks batches — drives whether
        # the per-batch stock check applies (batch-aware deduction) vs the
        # legacy per-SKU aggregate.
        src_loc = await db.distributor_locations.find_one(
            {"id": source_warehouse_id, "tenant_id": tenant_id},
            {"_id": 0, "track_batches": 1, "location_name": 1},
        )
        src_tracks_batches = bool(src_loc and src_loc.get("track_batches"))

        # Get shipment items (now also batch_id / batch_code)
        items = await db.distributor_shipment_items.find(
            {"shipment_id": shipment_id, "tenant_id": tenant_id},
            {"_id": 0, "sku_id": 1, "sku_name": 1, "quantity": 1, "batch_id": 1, "batch_code": 1}
        ).to_list(500)

        # Validate stock availability before deducting.
        # When the source warehouse tracks batches, every item must carry a
        # batch_id (we already enforced this on create) and we validate per
        # (sku_id, batch_id) — not aggregated against the SKU's total.
        insufficient = []
        for item in items:
            stock_query = {
                "tenant_id": tenant_id,
                "warehouse_location_id": source_warehouse_id,
                "sku_id": item["sku_id"],
            }
            if src_tracks_batches and item.get("batch_id"):
                stock_query["batch_id"] = item["batch_id"]
            stock = await tdb.factory_warehouse_stock.find_one(stock_query)
            available = stock.get("quantity", 0) if stock else 0
            if available < item["quantity"]:
                label = item.get("sku_name") or item["sku_id"]
                if item.get("batch_code"):
                    label = f"{label} (batch {item['batch_code']})"
                insufficient.append(f"{label}: need {item['quantity']}, have {available}")

        if insufficient:
            raise HTTPException(
                status_code=400,
                detail=f"Insufficient stock in factory warehouse: {'; '.join(insufficient)}"
            )

        # Deduct stock for each line. Same key shape as the validation — when
        # tracking batches, deduct from the specific batch row only.
        for item in items:
            update_query = {
                "tenant_id": tenant_id,
                "warehouse_location_id": source_warehouse_id,
                "sku_id": item["sku_id"],
            }
            if src_tracks_batches and item.get("batch_id"):
                update_query["batch_id"] = item["batch_id"]
            await tdb.factory_warehouse_stock.update_one(
                update_query,
                {"$inc": {"quantity": -item["quantity"]}, "$set": {"updated_at": now}}
            )

        logger.info(f"Deducted stock from factory warehouse {source_warehouse_id} for shipment {shipment['shipment_number']}")
    
    await db.distributor_shipments.update_one(
        {"id": shipment_id, "tenant_id": tenant_id},
        {"$set": {
            "status": "confirmed",
            "confirmed_at": now,
            "confirmed_by": current_user.get('id'),
            "updated_at": now
        }}
    )
    
    logger.info(f"Shipment {shipment['shipment_number']} confirmed by {current_user['email']}")
    
    # On confirm, push a Zoho Books Tax Invoice billing the distributor (best-effort,
    # with retry). The shipment is flagged zoho_push_pending on failure so the UI can
    # offer a manual Retry. Skipped cleanly when Zoho is not configured/connected.
    background_tasks.add_task(_safe_sync_shipment_to_zoho, tenant_id, distributor_id, shipment_id)
    
    return {"message": f"Shipment {shipment['shipment_number']} confirmed successfully", "status": "confirmed"}


@router.post("/{distributor_id}/shipments/{shipment_id}/dispatch")
async def dispatch_shipment(
    distributor_id: str,
    shipment_id: str,
    vehicle_number: Optional[str] = None,
    driver_name: Optional[str] = None,
    driver_contact: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Mark shipment as dispatched/in-transit"""
    if not is_distributor_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    tenant_id = get_current_tenant_id()
    
    shipment = await db.distributor_shipments.find_one(
        {"id": shipment_id, "tenant_id": tenant_id, "distributor_id": distributor_id}
    )
    
    if not shipment:
        raise HTTPException(status_code=404, detail="Shipment not found")
    
    if shipment.get('status') != 'confirmed':
        raise HTTPException(status_code=400, detail="Only confirmed shipments can be dispatched")
    
    now = datetime.now(timezone.utc).isoformat()
    
    update_data = {
        "status": "in_transit",
        "updated_at": now
    }
    
    if vehicle_number:
        update_data['vehicle_number'] = vehicle_number
    if driver_name:
        update_data['driver_name'] = driver_name
    if driver_contact:
        update_data['driver_contact'] = driver_contact
    
    await db.distributor_shipments.update_one(
        {"id": shipment_id, "tenant_id": tenant_id},
        {"$set": update_data}
    )
    
    logger.info(f"Shipment {shipment['shipment_number']} dispatched by {current_user['email']}")
    
    return {"message": f"Shipment {shipment['shipment_number']} dispatched", "status": "in_transit"}


@router.post("/{distributor_id}/shipments/{shipment_id}/deliver")
async def mark_shipment_delivered(
    distributor_id: str,
    shipment_id: str,
    delivery_date: Optional[str] = None,
    remarks: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Mark shipment as delivered"""
    if not is_distributor_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    tenant_id = get_current_tenant_id()
    
    shipment = await db.distributor_shipments.find_one(
        {"id": shipment_id, "tenant_id": tenant_id, "distributor_id": distributor_id}
    )
    
    if not shipment:
        raise HTTPException(status_code=404, detail="Shipment not found")
    
    if shipment.get('status') not in ['confirmed', 'in_transit', 'partially_delivered']:
        raise HTTPException(status_code=400, detail="Shipment cannot be marked as delivered in current status")
    
    now = datetime.now(timezone.utc).isoformat()
    actual_date = delivery_date or now[:10]
    
    update_data = {
        "status": "delivered",
        "actual_delivery_date": actual_date,
        "delivered_at": now,
        "delivered_by": current_user.get('id'),
        "updated_at": now
    }
    
    if remarks:
        update_data['remarks'] = (shipment.get('remarks', '') + '\n' + remarks).strip()
    
    await db.distributor_shipments.update_one(
        {"id": shipment_id, "tenant_id": tenant_id},
        {"$set": update_data}
    )
    
    # Update distributor stock (add to location inventory). When the source
    # factory tracked batches, the line carries the batch_id forward so the
    # destination's `distributor_stock` row is keyed by (loc, sku, batch).
    items = await db.distributor_shipment_items.find(
        {"shipment_id": shipment_id, "tenant_id": tenant_id},
        {"_id": 0}
    ).to_list(500)

    for item in items:
        batch_id = item.get('batch_id')
        stock_key = {
            "tenant_id": tenant_id,
            "distributor_id": distributor_id,
            "distributor_location_id": shipment.get('distributor_location_id'),
            "sku_id": item.get('sku_id'),
        }
        # Add batch to the key so batched and legacy aggregate rows stay separate.
        # `{"$in": [None]}` matches a stored None OR missing field, which is
        # what aggregate-only legacy rows have.
        stock_key["batch_id"] = batch_id if batch_id else {"$in": [None]}

        set_on_insert = {
            "id": str(uuid.uuid4()),
            "tenant_id": tenant_id,
            "distributor_id": distributor_id,
            "distributor_location_id": shipment.get('distributor_location_id'),
            "sku_id": item.get('sku_id'),
            "batch_id": batch_id,
            "created_at": now,
        }
        set_fields = {
            "sku_name": item.get('sku_name'),
            "sku_code": item.get('sku_code'),
            "distributor_name": shipment.get('distributor_name'),
            "location_name": shipment.get('distributor_location_name'),
            "updated_at": now,
        }
        if item.get('batch_code'):
            set_fields["batch_code"] = item.get('batch_code')

        await db.distributor_stock.update_one(
            stock_key,
            {
                "$inc": {"quantity": item.get('quantity', 0)},
                "$set": set_fields,
                "$setOnInsert": set_on_insert,
            },
            upsert=True
        )
    
    logger.info(f"Shipment {shipment['shipment_number']} delivered by {current_user['email']}")
    


# ============ Stock-In Receipt Acknowledgement Flow ============
# Workflow:
#   1) Admin dispatches shipment (status: in_transit)
#   2) Distributor reviews & acknowledges actual received quantity per item
#        - If all received qty == sent qty -> status: delivered (stock added)
#        - If any discrepancy             -> status: discrepancy_pending
#   3) Supplier admin reviews discrepancy and either approves or rejects:
#        - approve  -> shipment item quantity becomes received_quantity, status: delivered,
#                      stock added at received qty, factory stock refunded the delta
#        - reject   -> received_quantity cleared, status back to in_transit; distributor re-acknowledges


def _apply_stock_on_delivery(tenant_id, distributor_id, shipment, items, qty_field, now):
    """Helper: upsert distributor stock for delivered items using given quantity field.
    Batch-aware: keys on `(loc, sku, batch_id)` when the line carries a batch,
    otherwise falls back to the aggregate per-SKU row used by legacy data."""
    ops = []
    for item in items:
        batch_id = item.get('batch_id')
        stock_key = {
            "tenant_id": tenant_id,
            "distributor_id": distributor_id,
            "distributor_location_id": shipment.get('distributor_location_id'),
            "sku_id": item.get('sku_id'),
        }
        stock_key["batch_id"] = batch_id if batch_id else {"$in": [None]}
        set_fields = {
            "sku_name": item.get('sku_name'),
            "sku_code": item.get('sku_code'),
            "distributor_name": shipment.get('distributor_name'),
            "location_name": shipment.get('distributor_location_name'),
            "updated_at": now,
        }
        if item.get('batch_code'):
            set_fields["batch_code"] = item.get('batch_code')
        ops.append(db.distributor_stock.update_one(
            stock_key,
            {
                "$inc": {"quantity": int(item.get(qty_field, 0) or 0)},
                "$set": set_fields,
                "$setOnInsert": {
                    "id": str(uuid.uuid4()),
                    "tenant_id": tenant_id,
                    "distributor_id": distributor_id,
                    "distributor_location_id": shipment.get('distributor_location_id'),
                    "sku_id": item.get('sku_id'),
                    "batch_id": batch_id,
                    "created_at": now,
                },
            },
            upsert=True,
        ))
    return ops


@router.post("/{distributor_id}/shipments/{shipment_id}/acknowledge")
async def acknowledge_shipment_receipt(
    distributor_id: str,
    shipment_id: str,
    payload: dict,
    current_user: dict = Depends(get_current_user)
):
    """
    Distributor acknowledges receipt of a shipment, entering actual received qty per item.

    payload = {
       "items": [{"item_id": "...", "received_quantity": 10, "discrepancy_remark": "..."}],
       "acknowledgement_note": "Optional overall note"
    }

    - If all received qty == sent qty -> shipment marked delivered, stock added
    - If any discrepancy -> shipment marked discrepancy_pending for supplier review
    """
    tenant_id = get_current_tenant_id()

    # Allow either the supplier-admin (override) or the distributor owner
    if is_distributor_user(current_user):
        if current_user.get('distributor_id') != distributor_id:
            raise HTTPException(status_code=403, detail="You can only acknowledge your own shipments")
    elif not is_distributor_admin(current_user):
        raise HTTPException(status_code=403, detail="Not authorised to acknowledge shipments")

    shipment = await db.distributor_shipments.find_one(
        {"id": shipment_id, "tenant_id": tenant_id, "distributor_id": distributor_id}
    )
    if not shipment:
        raise HTTPException(status_code=404, detail="Shipment not found")

    if shipment.get('status') not in ['in_transit', 'partially_delivered']:
        raise HTTPException(
            status_code=400,
            detail=f"Shipment cannot be acknowledged in status '{shipment.get('status')}' — it must be dispatched (in_transit) first."
        )

    received_items = payload.get('items') or []
    if not received_items:
        raise HTTPException(status_code=400, detail="Received items required")

    # Load existing items keyed by id
    db_items = await db.distributor_shipment_items.find(
        {"shipment_id": shipment_id, "tenant_id": tenant_id},
        {"_id": 0}
    ).to_list(500)
    items_by_id = {it['id']: it for it in db_items}

    # Validate & determine if there is any discrepancy
    has_discrepancy = False
    parsed = []  # (item_doc, received_qty, remark)
    for entry in received_items:
        item_id = entry.get('item_id')
        if not item_id or item_id not in items_by_id:
            raise HTTPException(status_code=400, detail=f"Unknown shipment item: {item_id}")
        received_qty = int(entry.get('received_quantity', 0) or 0)
        if received_qty < 0:
            raise HTTPException(status_code=400, detail="Received quantity must be >= 0")
        sent_qty = int(items_by_id[item_id].get('quantity', 0) or 0)
        if received_qty > sent_qty:
            raise HTTPException(
                status_code=400,
                detail=f"Received qty ({received_qty}) cannot exceed sent qty ({sent_qty}) for {items_by_id[item_id].get('sku_name', item_id)}"
            )
        if received_qty != sent_qty:
            has_discrepancy = True
        parsed.append((items_by_id[item_id], received_qty, (entry.get('discrepancy_remark') or '').strip()))

    # Persist received quantities on each item
    now = datetime.now(timezone.utc).isoformat()
    for item_doc, received_qty, remark in parsed:
        await db.distributor_shipment_items.update_one(
            {"id": item_doc['id'], "tenant_id": tenant_id},
            {"$set": {
                "received_quantity": received_qty,
                "discrepancy_remark": remark or None,
                "acknowledged_at": now,
                "acknowledged_by": current_user.get('id')
            }}
        )

    new_status = 'discrepancy_pending' if has_discrepancy else 'delivered'
    update_data = {
        "status": new_status,
        "acknowledged_at": now,
        "acknowledged_by": current_user.get('id'),
        "acknowledgement_note": payload.get('acknowledgement_note'),
        "has_discrepancy": has_discrepancy,
        "updated_at": now,
    }
    if not has_discrepancy:
        update_data["actual_delivery_date"] = now[:10]
        update_data["delivered_at"] = now
        update_data["delivered_by"] = current_user.get('id')

    await db.distributor_shipments.update_one(
        {"id": shipment_id, "tenant_id": tenant_id},
        {"$set": update_data}
    )

    # If no discrepancy, add stock immediately
    if not has_discrepancy:
        refreshed_items = await db.distributor_shipment_items.find(
            {"shipment_id": shipment_id, "tenant_id": tenant_id},
            {"_id": 0}
        ).to_list(500)
        for op in _apply_stock_on_delivery(tenant_id, distributor_id, shipment, refreshed_items, 'received_quantity', now):
            await op

    logger.info(
        f"Shipment {shipment['shipment_number']} acknowledged by {current_user['email']} - "
        f"status={new_status}, discrepancy={has_discrepancy}"
    )

    return {
        "message": (
            f"Receipt acknowledged. Shipment marked {'delivered' if not has_discrepancy else 'pending supplier approval due to discrepancy'}."
        ),
        "status": new_status,
        "has_discrepancy": has_discrepancy,
    }


@router.post("/{distributor_id}/shipments/{shipment_id}/approve-receipt")
async def approve_shipment_receipt(
    distributor_id: str,
    shipment_id: str,
    payload: Optional[dict] = None,
    current_user: dict = Depends(get_current_user)
):
    """
    Supplier admin approves a discrepancy-acknowledged shipment.
    The received_quantity becomes final. Stock added to distributor at received_qty.
    Factory warehouse stock refunded for the (sent - received) delta.
    """
    if not is_distributor_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin access required")

    tenant_id = get_current_tenant_id()

    shipment = await db.distributor_shipments.find_one(
        {"id": shipment_id, "tenant_id": tenant_id, "distributor_id": distributor_id}
    )
    if not shipment:
        raise HTTPException(status_code=404, detail="Shipment not found")
    if shipment.get('status') != 'discrepancy_pending':
        raise HTTPException(status_code=400, detail="Only shipments in 'discrepancy_pending' status can be approved")

    now = datetime.now(timezone.utc).isoformat()
    items = await db.distributor_shipment_items.find(
        {"shipment_id": shipment_id, "tenant_id": tenant_id},
        {"_id": 0}
    ).to_list(500)

    # Refund factory stock for the under-received delta
    source_warehouse_id = shipment.get('source_warehouse_id')
    if source_warehouse_id:
        tdb = get_tenant_db()
        for it in items:
            sent_qty = int(it.get('quantity', 0) or 0)
            recv_qty = int(it.get('received_quantity', sent_qty) or 0)
            delta = sent_qty - recv_qty
            if delta > 0:
                await tdb.factory_warehouse_stock.update_one(
                    {
                        "tenant_id": tenant_id,
                        "warehouse_location_id": source_warehouse_id,
                        "sku_id": it["sku_id"]
                    },
                    {"$inc": {"quantity": delta}, "$set": {"updated_at": now}}
                )

    # Lock in received_quantity as the final quantity on items (and recompute amounts)
    new_total_qty = 0
    new_total_gross = 0.0
    new_total_discount = 0.0
    new_total_taxable = 0.0
    for it in items:
        recv_qty = int(it.get('received_quantity', it.get('quantity', 0)) or 0)
        unit_price = float(it.get('unit_price') or 0)
        discount_percent = float(it.get('discount_percent') or 0)
        gross = recv_qty * unit_price
        discount_amount = gross * (discount_percent / 100.0)
        taxable = gross - discount_amount
        await db.distributor_shipment_items.update_one(
            {"id": it['id'], "tenant_id": tenant_id},
            {"$set": {
                "original_sent_quantity": it.get('original_sent_quantity', int(it.get('quantity', 0) or 0)),
                "quantity": recv_qty,
                "gross_amount": round(gross, 2),
                "discount_amount": round(discount_amount, 2),
                "taxable_amount": round(taxable, 2),
                # tax_amount/net_amount re-computed at shipment-level GST below
                "updated_at": now,
            }}
        )
        new_total_qty += recv_qty
        new_total_gross += gross
        new_total_discount += discount_amount
        new_total_taxable += taxable

    gst_percent = float(shipment.get('gst_percent') or 0)
    new_tax_amount = new_total_taxable * (gst_percent / 100.0)
    new_net_amount = new_total_taxable + new_tax_amount

    update_data = {
        "status": "delivered",
        "actual_delivery_date": now[:10],
        "delivered_at": now,
        "delivered_by": current_user.get('id'),
        "discrepancy_approved_at": now,
        "discrepancy_approved_by": current_user.get('id'),
        "discrepancy_approval_note": (payload or {}).get('note'),
        "total_quantity": new_total_qty,
        "total_gross_amount": round(new_total_gross, 2),
        "total_discount_amount": round(new_total_discount, 2),
        "total_tax_amount": round(new_tax_amount, 2),
        "total_net_amount": round(new_net_amount, 2),
        "updated_at": now,
    }
    await db.distributor_shipments.update_one(
        {"id": shipment_id, "tenant_id": tenant_id},
        {"$set": update_data}
    )

    # Add stock to distributor at received quantity
    for op in _apply_stock_on_delivery(tenant_id, distributor_id, shipment, items, 'received_quantity', now):
        await op

    logger.info(f"Shipment {shipment['shipment_number']} discrepancy approved by {current_user['email']}")

    return {
        "message": f"Shipment {shipment['shipment_number']} approved. Quantities reconciled.",
        "status": "delivered",
    }


@router.post("/{distributor_id}/shipments/{shipment_id}/reject-receipt")
async def reject_shipment_receipt(
    distributor_id: str,
    shipment_id: str,
    payload: Optional[dict] = None,
    current_user: dict = Depends(get_current_user)
):
    """
    Supplier admin rejects a discrepancy submission, asking distributor to re-verify.
    Resets received_quantity and reverts status to in_transit.
    """
    if not is_distributor_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin access required")

    tenant_id = get_current_tenant_id()

    shipment = await db.distributor_shipments.find_one(
        {"id": shipment_id, "tenant_id": tenant_id, "distributor_id": distributor_id}
    )
    if not shipment:
        raise HTTPException(status_code=404, detail="Shipment not found")
    if shipment.get('status') != 'discrepancy_pending':
        raise HTTPException(status_code=400, detail="Only shipments in 'discrepancy_pending' status can be rejected")

    now = datetime.now(timezone.utc).isoformat()
    reason = ((payload or {}).get('reason') or '').strip()

    await db.distributor_shipment_items.update_many(
        {"shipment_id": shipment_id, "tenant_id": tenant_id},
        {"$unset": {"received_quantity": "", "discrepancy_remark": "", "acknowledged_at": "", "acknowledged_by": ""}}
    )
    await db.distributor_shipments.update_one(
        {"id": shipment_id, "tenant_id": tenant_id},
        {"$set": {
            "status": "in_transit",
            "discrepancy_rejected_at": now,
            "discrepancy_rejected_by": current_user.get('id'),
            "discrepancy_rejection_reason": reason or None,
            "acknowledged_at": None,
            "acknowledged_by": None,
            "has_discrepancy": False,
            "updated_at": now,
        }}
    )

    logger.info(f"Shipment {shipment['shipment_number']} discrepancy rejected by {current_user['email']}: {reason}")

    return {
        "message": f"Shipment {shipment['shipment_number']} sent back to distributor for re-verification.",
        "status": "in_transit",
        "rejection_reason": reason,
    }


    return {"message": f"Shipment {shipment['shipment_number']} delivered", "status": "delivered"}


@router.delete("/{distributor_id}/shipments/{shipment_id}")
async def delete_shipment(
    distributor_id: str,
    shipment_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Delete a draft shipment"""
    if not is_distributor_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    tenant_id = get_current_tenant_id()
    
    shipment = await db.distributor_shipments.find_one(
        {"id": shipment_id, "tenant_id": tenant_id, "distributor_id": distributor_id}
    )
    
    if not shipment:
        raise HTTPException(status_code=404, detail="Shipment not found")
    
    if shipment.get('status') != 'draft':
        raise HTTPException(status_code=400, detail="Only draft shipments can be deleted")
    
    # Delete items first
    ship_items = await db.distributor_shipment_items.find({"shipment_id": shipment_id, "tenant_id": tenant_id}).to_list(1000)
    await db.distributor_shipment_items.delete_many({"shipment_id": shipment_id, "tenant_id": tenant_id})
    
    # Record deletion audit BEFORE removing the shipment
    await _record_deletion_audit(
        tenant_id, 'shipment', shipment, distributor_id, current_user,
        entity_number=shipment.get('shipment_number'),
        extra={'item_count': len(ship_items), 'items_snapshot': [{k: v for k, v in i.items() if k != '_id'} for i in ship_items]},
    )
    
    # Delete shipment
    await db.distributor_shipments.delete_one({"id": shipment_id, "tenant_id": tenant_id})
    
    logger.info(f"Shipment {shipment['shipment_number']} deleted by {current_user['email']}")
    
    return {"message": f"Shipment {shipment['shipment_number']} deleted"}


@router.post("/{distributor_id}/shipments/{shipment_id}/cancel")
async def cancel_shipment(
    distributor_id: str,
    shipment_id: str,
    reason: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Cancel a shipment (only draft or confirmed can be cancelled)"""
    if not is_distributor_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    tenant_id = get_current_tenant_id()
    
    shipment = await db.distributor_shipments.find_one(
        {"id": shipment_id, "tenant_id": tenant_id, "distributor_id": distributor_id}
    )
    
    if not shipment:
        raise HTTPException(status_code=404, detail="Shipment not found")
    
    prior_status = shipment.get('status')
    if prior_status not in ['draft', 'confirmed']:
        raise HTTPException(status_code=400, detail="Only draft or confirmed shipments can be cancelled")
    
    now = datetime.now(timezone.utc).isoformat()
    
    # When a CONFIRMED shipment is cancelled, the source factory warehouse stock
    # was already deducted at confirm time — add it back so units are not left
    # stuck/"reserved". A DRAFT shipment never deducted anything, so this is a
    # no-op for drafts.
    source_restored = False
    if prior_status == 'confirmed':
        try:
            await _readd_shipment_source_stock(tenant_id, shipment)
            source_restored = True
        except Exception:
            logger.exception(f"Failed to restore source stock while cancelling shipment {shipment_id}")
    
    remarks = shipment.get('remarks', '') or ''
    if reason:
        remarks = f"{remarks}\nCancelled: {reason}".strip()
    
    await db.distributor_shipments.update_one(
        {"id": shipment_id, "tenant_id": tenant_id},
        {"$set": {
            "status": "cancelled",
            "remarks": remarks,
            "cancelled_at": now,
            "cancelled_by": current_user.get('id'),
            "cancelled_from_status": prior_status,
            "source_stock_restored": source_restored,
            "updated_at": now
        }}
    )
    
    logger.info(f"Shipment {shipment['shipment_number']} cancelled by {current_user['email']} (source_restored={source_restored})")
    
    return {"message": f"Shipment {shipment['shipment_number']} cancelled", "status": "cancelled"}


@router.post("/{distributor_id}/shipments/{shipment_id}/retry-zoho-push")
async def retry_shipment_zoho_push(
    distributor_id: str,
    shipment_id: str,
    current_user: dict = Depends(get_current_user),
):
    """Manually re-attempt the Zoho Books invoice push for a confirmed shipment."""
    if not is_distributor_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin access required")

    tenant_id = get_current_tenant_id()
    shipment = await db.distributor_shipments.find_one(
        {"id": shipment_id, "tenant_id": tenant_id, "distributor_id": distributor_id}, {"_id": 0}
    )
    if not shipment:
        raise HTTPException(status_code=404, detail="Shipment not found")
    if shipment.get("status") in ("draft", "cancelled", "reversed"):
        raise HTTPException(status_code=400, detail="Only confirmed/dispatched/delivered shipments can be invoiced.")

    from services.zoho_service import sync_shipment_to_zoho, ZohoPushSkippedError
    try:
        await sync_shipment_to_zoho(tenant_id, distributor_id, shipment_id)
    except ZohoPushSkippedError as e:
        raise HTTPException(status_code=400, detail=str(e))

    fresh = await db.distributor_shipments.find_one(
        {"id": shipment_id, "tenant_id": tenant_id},
        {"_id": 0, "zoho_invoice_url": 1, "zoho_invoice_number": 1, "zoho_invoice_id": 1, "zoho_push_error": 1},
    )
    if not fresh.get("zoho_invoice_url"):
        raise HTTPException(
            status_code=400,
            detail=fresh.get("zoho_push_error") or "Zoho push completed but no invoice URL was produced. Check that the source warehouse is mapped to a Zoho Branch and SKUs have Zoho item mappings.",
        )
    return {
        "message": "Zoho invoice generated.",
        "zoho_invoice_url": fresh.get("zoho_invoice_url"),
        "zoho_invoice_number": fresh.get("zoho_invoice_number"),
        "zoho_invoice_id": fresh.get("zoho_invoice_id"),
    }


async def _readd_shipment_source_stock(tenant_id: str, shipment: dict) -> None:
    """Inverse of confirm's deduction: add a shipment's quantities back to the
    SOURCE factory warehouse (`factory_warehouse_stock`). Batch-aware when the
    source warehouse tracks batches; otherwise keyed on the per-SKU row."""
    source_warehouse_id = shipment.get('source_warehouse_id')
    if not source_warehouse_id:
        return
    now = datetime.now(timezone.utc).isoformat()
    src_loc = await db.distributor_locations.find_one(
        {"id": source_warehouse_id, "tenant_id": tenant_id},
        {"_id": 0, "track_batches": 1},
    )
    src_tracks_batches = bool(src_loc and src_loc.get("track_batches"))
    items = await db.distributor_shipment_items.find(
        {"shipment_id": shipment.get("id"), "tenant_id": tenant_id}, {"_id": 0},
    ).to_list(500)
    for item in items:
        qty = int(item.get("quantity", 0) or 0)
        if not qty:
            continue
        stock_query = {
            "tenant_id": tenant_id,
            "warehouse_location_id": source_warehouse_id,
            "sku_id": item.get("sku_id"),
        }
        if src_tracks_batches and item.get("batch_id"):
            stock_query["batch_id"] = item.get("batch_id")
        await db.factory_warehouse_stock.update_one(
            stock_query,
            {"$inc": {"quantity": qty}, "$set": {"updated_at": now}},
            upsert=True,
        )


async def _remove_shipment_destination_stock(tenant_id: str, distributor_id: str, shipment: dict) -> None:
    """Inverse of deliver's addition: subtract a delivered shipment's quantities
    from the DESTINATION `distributor_stock`. Batch-aware, matching the same
    key convention used when the stock was added on delivery."""
    now = datetime.now(timezone.utc).isoformat()
    items = await db.distributor_shipment_items.find(
        {"shipment_id": shipment.get("id"), "tenant_id": tenant_id}, {"_id": 0},
    ).to_list(500)
    for item in items:
        qty = int(item.get("quantity", 0) or 0)
        if not qty:
            continue
        batch_id = item.get("batch_id")
        stock_key = {
            "tenant_id": tenant_id,
            "distributor_id": distributor_id,
            "distributor_location_id": shipment.get("distributor_location_id"),
            "sku_id": item.get("sku_id"),
        }
        stock_key["batch_id"] = batch_id if batch_id else {"$in": [None]}
        await db.distributor_stock.update_one(
            stock_key,
            {"$inc": {"quantity": -qty}, "$set": {"updated_at": now}},
        )


@router.post("/{distributor_id}/shipments/{shipment_id}/reverse")
async def reverse_shipment(
    distributor_id: str,
    shipment_id: str,
    reason: Optional[str] = None,
    acknowledge: bool = False,
    current_user: dict = Depends(get_current_user),
):
    """Reverse a Stock In (factory → distributor) shipment at ANY stage and undo
    all of its stock effects, keeping the record for audit (status 'reversed'):
      • draft: nothing was committed → just mark reversed.
      • confirmed / in_transit / discrepancy_pending: source factory stock was
        deducted at confirm → ADD it back.
      • delivered / partially_delivered: source stock was deducted AND
        destination distributor stock was added → restore source AND remove the
        destination units.
    Non-draft reversals require `acknowledge=true` (server-side double-confirm)."""
    if not is_distributor_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin access required")

    tenant_id = get_current_tenant_id()
    shipment = await db.distributor_shipments.find_one(
        {"id": shipment_id, "tenant_id": tenant_id, "distributor_id": distributor_id}
    )
    if not shipment:
        raise HTTPException(status_code=404, detail="Shipment not found")

    status = (shipment.get("status") or "").lower()
    if status in {"cancelled", "reversed"}:
        raise HTTPException(status_code=400, detail="This shipment is already cancelled or reversed.")

    is_draft = status == "draft"
    if not is_draft and not acknowledge:
        raise HTTPException(
            status_code=400,
            detail="This shipment is past draft. Confirm the reversal explicitly (acknowledge=true).",
        )

    now = datetime.now(timezone.utc).isoformat()
    source_deducted = status in {"confirmed", "in_transit", "discrepancy_pending", "delivered", "partially_delivered"}
    dest_added = status in {"delivered", "partially_delivered"}

    # 1) Remove destination units first (only if they were added on delivery).
    if dest_added:
        try:
            await _remove_shipment_destination_stock(tenant_id, distributor_id, shipment)
        except Exception:
            logger.exception(f"Failed to remove destination stock during reverse for shipment {shipment_id}")

    # 2) Restore source factory warehouse stock (deducted at confirm).
    if source_deducted:
        try:
            await _readd_shipment_source_stock(tenant_id, shipment)
        except Exception:
            logger.exception(f"Failed to restore source stock during reverse for shipment {shipment_id}")

    remarks = shipment.get("remarks", "") or ""
    if reason:
        remarks = f"{remarks}\nReversed: {reason}".strip()

    await db.distributor_shipments.update_one(
        {"id": shipment_id, "tenant_id": tenant_id},
        {"$set": {
            "status": "reversed",
            "remarks": remarks,
            "reversed_at": now,
            "reversed_by": current_user.get("id"),
            "reversed_by_name": current_user.get("name"),
            "reversed_from_status": status,
            "source_stock_restored": source_deducted,
            "destination_stock_removed": dest_added,
            "updated_at": now,
        }}
    )

    if dest_added:
        stock_msg = "destination stock removed and source warehouse stock restored"
    elif source_deducted:
        stock_msg = "source warehouse stock restored"
    else:
        stock_msg = "no stock was committed"
    logger.info(f"Shipment {shipment['shipment_number']} reversed by {current_user['email']} — {stock_msg}")

    return {
        "message": f"Shipment {shipment['shipment_number']} reversed — {stock_msg}",
        "status": "reversed",
    }


# ============ Shipment Items Management ============

@router.post("/{distributor_id}/shipments/{shipment_id}/items")
async def add_shipment_item(
    distributor_id: str,
    shipment_id: str,
    data: ShipmentItemCreate,
    current_user: dict = Depends(get_current_user)
):
    """Add an item to a draft shipment"""
    if not is_distributor_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    tenant_id = get_current_tenant_id()
    
    shipment = await db.distributor_shipments.find_one(
        {"id": shipment_id, "tenant_id": tenant_id, "distributor_id": distributor_id}
    )
    
    if not shipment:
        raise HTTPException(status_code=404, detail="Shipment not found")
    
    if shipment.get('status') != 'draft':
        raise HTTPException(status_code=400, detail="Can only add items to draft shipments")
    
    # Get SKU info
    sku_name = data.sku_name
    sku_code = data.sku_code
    if not sku_name or not sku_code:
        sku = await db.skus.find_one({"id": data.sku_id}, {"_id": 0, "name": 1, "sku_code": 1})
        if sku:
            sku_name = sku_name or sku.get('name')
            sku_code = sku_code or sku.get('sku_code')
    
    item_dict = {
        'id': str(uuid.uuid4()),
        'tenant_id': tenant_id,
        'shipment_id': shipment_id,
        'sku_id': data.sku_id,
        'sku_name': sku_name,
        'sku_code': sku_code,
        'quantity': data.quantity,
        'unit_price': data.unit_price,
        'discount_percent': data.discount_percent or 0,
        'tax_percent': data.tax_percent or 0,
        'remarks': data.remarks
    }
    
    item_dict = calculate_item_amounts(item_dict)
    
    await db.distributor_shipment_items.insert_one(item_dict)
    
    # Update shipment totals
    await recalculate_shipment_totals(shipment_id, tenant_id)
    
    item_dict.pop('_id', None)
    
    return item_dict


@router.delete("/{distributor_id}/shipments/{shipment_id}/items/{item_id}")
async def remove_shipment_item(
    distributor_id: str,
    shipment_id: str,
    item_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Remove an item from a draft shipment"""
    if not is_distributor_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    tenant_id = get_current_tenant_id()
    
    shipment = await db.distributor_shipments.find_one(
        {"id": shipment_id, "tenant_id": tenant_id, "distributor_id": distributor_id}
    )
    
    if not shipment:
        raise HTTPException(status_code=404, detail="Shipment not found")
    
    if shipment.get('status') != 'draft':
        raise HTTPException(status_code=400, detail="Can only remove items from draft shipments")
    
    result = await db.distributor_shipment_items.delete_one({
        "id": item_id, 
        "shipment_id": shipment_id, 
        "tenant_id": tenant_id
    })
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Item not found")
    
    # Recalculate totals
    await recalculate_shipment_totals(shipment_id, tenant_id)
    
    return {"message": "Item removed"}


async def recalculate_shipment_totals(shipment_id: str, tenant_id: str):
    """Recalculate and update shipment totals"""
    items = await db.distributor_shipment_items.find(
        {"shipment_id": shipment_id, "tenant_id": tenant_id},
        {"_id": 0}
    ).to_list(500)
    
    total_quantity = sum(item.get('quantity', 0) for item in items)
    total_gross_amount = sum(item.get('gross_amount', 0) for item in items)
    total_discount_amount = sum(item.get('discount_amount', 0) for item in items)
    total_tax_amount = sum(item.get('tax_amount', 0) for item in items)
    total_net_amount = sum(item.get('net_amount', 0) for item in items)
    
    await db.distributor_shipments.update_one(
        {"id": shipment_id, "tenant_id": tenant_id},
        {"$set": {
            "total_quantity": total_quantity,
            "total_gross_amount": round(total_gross_amount, 2),
            "total_discount_amount": round(total_discount_amount, 2),
            "total_tax_amount": round(total_tax_amount, 2),
            "total_net_amount": round(total_net_amount, 2),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )


# ============ Stock Dashboard ============

@router.get("/dashboard/stock-summary")
async def get_stock_dashboard_summary(
    city: Optional[str] = None,
    distributor_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Tenant-wide stock dashboard.

    Pulls from BOTH `distributor_stock` (regular distributor warehouses) AND
    `factory_warehouse_stock` (factory / master warehouses) and reports
    **everything in crates** (numbers in the raw collections are bottles —
    converted per-SKU via `bottles_per_crate`).

    Bug fixes vs. earlier:
      • Factory warehouses were entirely missing.
      • Numbers were bottles, labelled as units; the UI then showed inflated
        counts (e.g. 27,613 instead of ~1,150 crates).
      • Negative rows polluted totals — floored to zero per row (the negative
        is surfaced as a warehouse-overview anomaly elsewhere).
    """
    tenant_id = get_current_tenant_id()

    # ── 1. Pull raw stock from both collections ───────────────────────────
    dist_query: dict = {"tenant_id": tenant_id}
    if distributor_id:
        dist_query["distributor_id"] = distributor_id
    distributor_rows = await db.distributor_stock.find(dist_query, {"_id": 0}).to_list(5000)

    fwh_query: dict = {"tenant_id": tenant_id}
    factory_rows = await db.factory_warehouse_stock.find(fwh_query, {"_id": 0}).to_list(5000)

    # ── 2. Apply optional city filter (works for both collections) ────────
    if city:
        loc_ids_in_city = {
            loc["id"] for loc in await db.distributor_locations.find(
                {"tenant_id": tenant_id, "city": city}, {"_id": 0, "id": 1}
            ).to_list(2000)
        }
        distributor_rows = [s for s in distributor_rows if s.get("distributor_location_id") in loc_ids_in_city]
        factory_rows = [s for s in factory_rows if s.get("warehouse_location_id") in loc_ids_in_city]

    # If a specific distributor was requested, factory warehouses owned by
    # that distributor (self-managed factories) should still be included; the
    # link is via `distributor_locations.distributor_id`.
    if distributor_id:
        loc_owned = {
            loc["id"] for loc in await db.distributor_locations.find(
                {"tenant_id": tenant_id, "distributor_id": distributor_id, "is_factory": True},
                {"_id": 0, "id": 1},
            ).to_list(500)
        }
        factory_rows = [s for s in factory_rows if s.get("warehouse_location_id") in loc_owned]

    # ── 3. Stock quantities are stored **already in crates** in both
    #       `distributor_stock` and `factory_warehouse_stock` (the per-SKU
    #       bottles-per-crate conversion happens at the moment stock is
    #       received / dispatched, never on read). We therefore do NOT divide
    #       again here — doing so previously made small crate counts collapse
    #       to 0 (e.g. 12 crates ÷ 24 bpc = 0). We only floor negatives so
    #       any reconciliation glitch can't pollute totals.
    def crates(_sku_id: str, qty) -> int:
        try:
            return max(0, int(qty))
        except Exception:
            try:
                return max(0, int(qty or 0))
            except Exception:
                return 0

    # ── 4. Resolve location metadata for factory rows ─────────────────────
    fwh_loc_ids = list({s.get("warehouse_location_id") for s in factory_rows if s.get("warehouse_location_id")})
    fwh_locs: dict = {}
    if fwh_loc_ids:
        fwh_locs = {loc["id"]: loc for loc in await db.distributor_locations.find(
            {"tenant_id": tenant_id, "id": {"$in": fwh_loc_ids}},
            {"_id": 0, "id": 1, "distributor_id": 1, "location_name": 1, "city": 1, "is_factory": 1},
        ).to_list(len(fwh_loc_ids) + 1)}
    parent_dist_ids = list({loc.get("distributor_id") for loc in fwh_locs.values() if loc.get("distributor_id")})
    parent_dists: dict = {}
    if parent_dist_ids:
        parent_dists = {d["id"]: d for d in await db.distributors.find(
            {"tenant_id": tenant_id, "id": {"$in": parent_dist_ids}},
            {"_id": 0, "id": 1, "distributor_name": 1},
        ).to_list(len(parent_dist_ids) + 1)}

    # ── 5. Build a unified "stock_rows" list of crate-denominated entries ─
    unified: list[dict] = []
    for s in distributor_rows:
        unified.append({
            "kind": "distributor",
            "distributor_id": s.get("distributor_id"),
            "distributor_name": s.get("distributor_name") or "Unknown",
            "location_id": s.get("distributor_location_id"),
            "location_name": s.get("location_name") or "Unknown",
            "sku_id": s.get("sku_id"),
            "sku_name": s.get("sku_name") or "Unknown",
            "quantity": crates(s.get("sku_id", ""), s.get("quantity", 0)),
            "is_factory": False,
        })
    for s in factory_rows:
        loc = fwh_locs.get(s.get("warehouse_location_id"), {})
        parent = parent_dists.get(loc.get("distributor_id"), {})
        unified.append({
            "kind": "factory",
            "distributor_id": loc.get("distributor_id"),
            "distributor_name": parent.get("distributor_name") or s.get("warehouse_name") or "Factory",
            "location_id": s.get("warehouse_location_id"),
            "location_name": s.get("warehouse_name") or loc.get("location_name") or "Factory Warehouse",
            "sku_id": s.get("sku_id"),
            "sku_name": s.get("sku_name") or "Unknown",
            "quantity": crates(s.get("sku_id", ""), s.get("quantity", 0)),
            "is_factory": True,
        })

    # ── 6. Top-level summary ──────────────────────────────────────────────
    total_quantity = sum(r["quantity"] for r in unified)
    total_skus = len({r["sku_id"] for r in unified if r.get("sku_id")})
    total_locations = len({r["location_id"] for r in unified if r.get("location_id")})
    total_distributors = len({r["distributor_id"] for r in unified if r.get("distributor_id")})

    # ── 7. Groupings (by_distributor / by_sku / by_location) ──────────────
    by_dist: dict = {}
    for r in unified:
        did = r.get("distributor_id") or "__factory__"
        bucket = by_dist.setdefault(did, {
            "distributor_id": did,
            "distributor_name": r.get("distributor_name"),
            "total_quantity": 0,
            "sku_count": set(),
            "locations": set(),
        })
        bucket["total_quantity"] += r["quantity"]
        if r.get("sku_id"):
            bucket["sku_count"].add(r["sku_id"])
        if r.get("location_id"):
            bucket["locations"].add(r["location_id"])
    distributor_summary = sorted([
        {"distributor_id": v["distributor_id"], "distributor_name": v["distributor_name"],
         "total_quantity": v["total_quantity"], "sku_count": len(v["sku_count"]),
         "location_count": len(v["locations"])}
        for v in by_dist.values()
    ], key=lambda x: x["total_quantity"], reverse=True)

    by_sku: dict = {}
    for r in unified:
        sid = r.get("sku_id")
        if not sid:
            continue
        bucket = by_sku.setdefault(sid, {
            "sku_id": sid, "sku_name": r.get("sku_name"),
            "total_quantity": 0, "locations": set(),
        })
        bucket["total_quantity"] += r["quantity"]
        if r.get("location_id"):
            bucket["locations"].add(r["location_id"])
    sku_summary = sorted([
        {"sku_id": v["sku_id"], "sku_name": v["sku_name"],
         "total_quantity": v["total_quantity"], "location_count": len(v["locations"])}
        for v in by_sku.values()
    ], key=lambda x: x["total_quantity"], reverse=True)

    by_loc: dict = {}
    for r in unified:
        lid = r.get("location_id")
        if not lid:
            continue
        bucket = by_loc.setdefault(lid, {
            "location_id": lid, "location_name": r.get("location_name"),
            "distributor_name": r.get("distributor_name"),
            "is_factory": r.get("is_factory", False),
            "total_quantity": 0, "sku_count": set(), "items": [],
        })
        # If the same location surfaces from both collections (rare) mark as
        # factory if either row says so — factories are the more specific kind.
        if r.get("is_factory"):
            bucket["is_factory"] = True
        bucket["total_quantity"] += r["quantity"]
        bucket["sku_count"].add(r.get("sku_id"))
        bucket["items"].append({
            "sku_id": r.get("sku_id"),
            "sku_name": r.get("sku_name"),
            "quantity": r["quantity"],
        })
    # Aggregate duplicate SKU lines within a location (when the same SKU
    # appears in both distributor_stock and factory_warehouse_stock — yes,
    # this happens for self-managed factories — or across multiple batches).
    location_summary = []
    for v in by_loc.values():
        items_by_sku: dict = {}
        for it in v["items"]:
            sid = it.get("sku_id")
            if not sid:
                continue
            items_by_sku.setdefault(sid, {"sku_id": sid, "sku_name": it.get("sku_name"), "quantity": 0})
            items_by_sku[sid]["quantity"] += int(it.get("quantity") or 0)
        location_summary.append({
            "location_id": v["location_id"],
            "location_name": v["location_name"],
            "distributor_name": v["distributor_name"],
            "is_factory": v["is_factory"],
            "total_quantity": v["total_quantity"],
            "sku_count": len([s for s in v["sku_count"] if s]),
            "items": sorted(items_by_sku.values(), key=lambda x: (x.get("sku_name") or "").lower()),
        })
    location_summary.sort(key=lambda x: x["total_quantity"], reverse=True)

    # ── 8. Enrich each SKU with its DEFAULT crate packaging ────────────────
    # The rep can then toggle the dashboard from raw bottles to that SKU's
    # default crate (e.g. "12 Bottle Crate" for 1L) on the frontend. We use
    # `packaging_config.stock_out` (customer-facing) as the canonical default
    # because that's the unit reps think in when reading the dashboard. Falls
    # back to `packaging_config.stock_in`, then to any first packaging found.
    sku_ids_in_play = list({r.get("sku_id") for r in unified if r.get("sku_id")})
    sku_pkg: dict = {}
    if sku_ids_in_play:
        # Master SKUs may pre-date the tenant_id field — accept rows without it
        # too, mirroring how the rest of the dashboard treats legacy masters.
        async for sku in db.master_skus.find(
            {
                "id": {"$in": sku_ids_in_play},
                "$or": [{"tenant_id": tenant_id}, {"tenant_id": {"$exists": False}}],
            },
            {"_id": 0, "id": 1, "packaging_config": 1},
        ):
            cfg = sku.get("packaging_config") or {}
            chosen = None
            for ctx in ("stock_out", "stock_in", "production", "promo_stock_out"):
                rows = cfg.get(ctx) or []
                if not rows:
                    continue
                pick = next((p for p in rows if p.get("is_default")), None) or rows[0]
                if pick and (pick.get("units_per_package") or 0) > 1:
                    chosen = pick
                    break
                if pick and not chosen:
                    chosen = pick  # fallback (e.g. "Bottle (1)") — used if no crate exists
            if chosen:
                sku_pkg[sku["id"]] = {
                    "default_packaging_name": chosen.get("packaging_type_name"),
                    "default_units_per_package": int(chosen.get("units_per_package") or 1),
                }

    def _attach(item: dict):
        """Add default-packaging fields + the converted crate count to a row."""
        pkg = sku_pkg.get(item.get("sku_id"))
        item["default_packaging_name"] = (pkg or {}).get("default_packaging_name")
        item["default_units_per_package"] = (pkg or {}).get("default_units_per_package") or 1
        upp = item["default_units_per_package"]
        qty = int(item.get("total_quantity") if "total_quantity" in item else item.get("quantity", 0))
        item["default_packaging_quantity"] = qty // upp if upp > 0 else qty
        return item

    for row in sku_summary:
        _attach(row)
    for loc in location_summary:
        loc["items"] = [_attach(dict(it)) for it in loc.get("items", [])]

    # Cities (just the ones whose locations actually have stock)
    cities_with_stock = await db.distributor_locations.distinct(
        "city",
        {"tenant_id": tenant_id, "id": {"$in": list({r.get("location_id") for r in unified if r.get("location_id")})}}
    )

    return {
        "summary": {
            "total_quantity": total_quantity,
            "total_skus": total_skus,
            "total_locations": total_locations,
            "total_distributors": total_distributors,
            "unit": "crates",
        },
        "by_distributor": distributor_summary,
        "by_sku": sku_summary,
        "by_location": location_summary,
        "cities": sorted(cities_with_stock) if cities_with_stock else [],
    }


# ============ Distributor Stock ============

@router.get("/{distributor_id}/stock")
async def get_distributor_stock(
    distributor_id: str,
    location_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get current stock levels at distributor locations"""
    tenant_id = get_current_tenant_id()
    
    query = {"tenant_id": tenant_id, "distributor_id": distributor_id}
    
    if location_id:
        query["distributor_location_id"] = location_id
    
    stock = await db.distributor_stock.find(
        query,
        {"_id": 0}
    ).sort([("location_name", 1), ("sku_name", 1)]).to_list(1000)
    
    # Group by location
    by_location = {}
    for item in stock:
        loc_id = item.get('distributor_location_id')
        loc_name = item.get('location_name', 'Unknown')
        if loc_id not in by_location:
            by_location[loc_id] = {
                'location_id': loc_id,
                'location_name': loc_name,
                'items': [],
                'total_quantity': 0
            }
        by_location[loc_id]['items'].append(item)
        by_location[loc_id]['total_quantity'] += item.get('quantity', 0)
    
    return {
        "stock": stock,
        "by_location": list(by_location.values()),
        "total_items": len(stock)
    }


# ============ Distributor-to-Account Delivery CRUD ============

DELIVERY_STATUSES = {
    "draft": "Draft - Not yet confirmed",
    "confirmed": "Confirmed - Ready for delivery",
    "delivery_assigned": "Delivery Assigned - Attached to a schedule",
    "delivery_scheduled": "Delivery Scheduled - Schedule approved",
    "on_the_way": "On the Way - Driver started the run",
    "complete": "Complete - Delivered to customer",
    # Legacy values kept for backward compatibility (older rows):
    "scheduled": "Scheduled - On a confirmed delivery schedule",
    "in_transit": "In Transit - On the way",
    "delivered": "Delivered - Completed",
    "partially_delivered": "Partially Delivered",
    "returned": "Returned",
    "cancelled": "Cancelled"
}


async def generate_delivery_number(tenant_id: str) -> str:
    """Generate unique delivery number like DEL-2026-0001"""
    year = datetime.now().year
    count = await db.distributor_deliveries.count_documents({
        "tenant_id": tenant_id,
        "delivery_number": {"$regex": f"^DEL-{year}-"}
    })
    return f"DEL-{year}-{count + 1:04d}"


def calculate_delivery_item_amounts(item: dict, margin_type: str = None, margin_value: float = None, transfer_price: float = None, base_price: float = None, billing_approach: str = 'margin_upfront') -> dict:
    """Calculate amounts for a delivery item including margin and adjustment calculations
    
    Columns (in display order):
    - Base Price: The original/standard price of the product
    - Transfer Price: base_price × (1 - margin%) — price at which factory sells to distributor
    - Billed to Dist: qty × transfer_price — amount initially billed based on base price
    - Customer Price: What the customer actually pays
    - New Transfer Price: customer_price × (1 - margin%) — adjusted transfer price based on actual sale
    - Actual Billable to Dist: qty × new_transfer_price — amount that should be billed based on customer price
    - Adjustment (Dist → Factory): Actual Billable - Billed to Dist = qty × (1 - margin%) × (customer_price - base_price)
      Positive = distributor owes factory (sold higher), Negative = factory owes distributor (sold lower)
    - Customer Invoice: qty × customer_price — what the customer pays
    """
    quantity = item.get('quantity', 0)
    unit_price = item.get('unit_price', 0)  # Customer Selling Price
    customer_selling_price = item.get('customer_selling_price') or unit_price
    discount_percent = item.get('discount_percent', 0) or 0
    tax_percent = item.get('tax_percent', 0) or 0
    
    # Base price from margin matrix
    item_base_price = item.get('base_price') or base_price or transfer_price or 0
    
    # Get commission/margin percentage from item or margin_value
    commission_percent = item.get('distributor_commission_percent') or margin_value or 0
    
    # Transfer Price: for cost_based, equals base price; for margin_upfront, base × (1 - margin%)
    is_cost_based = billing_approach == 'cost_based'
    if is_cost_based:
        item_transfer_price = item_base_price  # Cost-based: always use base price (no margin deducted at transfer)
        # Factory's due = customer price × (1 - margin%) — always deduct margin for factory's share
        new_transfer_price = round(customer_selling_price * (1 - commission_percent / 100), 2) if customer_selling_price and commission_percent else customer_selling_price
    else:
        item_transfer_price = item.get('transfer_price') or (round(item_base_price * (1 - commission_percent / 100), 2) if item_base_price and commission_percent else item_base_price)
        new_transfer_price = round(customer_selling_price * (1 - commission_percent / 100), 2) if customer_selling_price and commission_percent else customer_selling_price
    
    # Billed to Distributor = qty × transfer_price (initial billing based on base price)
    billed_to_dist = round(quantity * item_transfer_price, 2) if item_transfer_price else 0
    
    # Actual Billable to Distributor = qty × new_transfer_price (actual billing based on customer price)
    actual_billable_to_dist = round(quantity * new_transfer_price, 2) if new_transfer_price else 0
    
    # Calculate billing amounts
    gross_amount = quantity * customer_selling_price  # Customer Invoice = qty × customer_price
    discount_amount = round(gross_amount * discount_percent / 100, 2)
    taxable_amount = gross_amount - discount_amount
    tax_amount = round(taxable_amount * tax_percent / 100, 2)
    net_amount = taxable_amount + tax_amount
    
    # Customer Invoice = qty × customer_price
    customer_billing = round(gross_amount, 2)
    
    # NEW FORMULA: Adjustment (Dist → Factory) = Actual Billable - Billed to Dist
    # = qty × new_transfer_price - qty × transfer_price
    # = qty × (1 - margin%) × (customer_price - base_price)
    # Positive when customer_price > base_price (distributor owes factory)
    # Negative when customer_price < base_price (factory owes distributor)
    factory_distributor_adjustment = round(actual_billable_to_dist - billed_to_dist, 2)
    
    # Margin per unit = commission% of customer_price
    margin_per_unit = round(customer_selling_price * commission_percent / 100, 2) if commission_percent else 0
    
    # Legacy calculations (kept for backward compatibility)
    distributor_earnings = round(gross_amount * commission_percent / 100, 2) if commission_percent else 0
    margin_at_transfer_price = 0 if is_cost_based else (round(quantity * item_base_price * commission_percent / 100, 2) if item_base_price and commission_percent else 0)
    adjustment_payable = round(distributor_earnings - margin_at_transfer_price, 2)
    price_premium_payable = round(quantity * (customer_selling_price - item_base_price), 2) if customer_selling_price > item_base_price and item_base_price > 0 else 0
    
    # Legacy margin calculation
    margin_amount = 0
    if margin_type and margin_value:
        if margin_type == 'percentage':
            margin_amount = round(net_amount * margin_value / 100, 2)
        elif margin_type == 'fixed_per_bottle':
            margin_amount = round(quantity * margin_value, 2)
        elif margin_type == 'fixed_per_case':
            cases = quantity / 12
            margin_amount = round(cases * margin_value, 2)
    
    return {
        **item,
        'base_price': item_base_price,
        'transfer_price': item_transfer_price,
        'customer_selling_price': customer_selling_price,
        'new_transfer_price': new_transfer_price,
        'billed_to_dist': billed_to_dist,
        'actual_billable_to_dist': actual_billable_to_dist,
        'distributor_commission_percent': commission_percent,
        'margin_per_unit': margin_per_unit,
        'gross_amount': round(gross_amount, 2),
        'customer_billing': customer_billing,
        'factory_distributor_adjustment': factory_distributor_adjustment,
        'discount_amount': discount_amount,
        'taxable_amount': round(taxable_amount, 2),
        'tax_amount': tax_amount,
        'net_amount': round(net_amount, 2),
        # Legacy fields
        'distributor_earnings': distributor_earnings,
        'margin_at_transfer_price': margin_at_transfer_price,
        'adjustment_payable': adjustment_payable,
        'price_premium_payable': price_premium_payable,
        'margin_type': margin_type,
        'margin_value': margin_value,
        'margin_amount': margin_amount or distributor_earnings
    }


@router.get("/deliveries/all")
async def list_all_deliveries(
    distributor_id: Optional[str] = None,
    account_id: Optional[str] = None,
    status: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    current_user: dict = Depends(get_current_user)
):
    """List all deliveries with filters"""
    tenant_id = get_current_tenant_id()
    
    query = {"tenant_id": tenant_id}
    
    if distributor_id:
        query["distributor_id"] = distributor_id
    if account_id:
        query["account_id"] = account_id
    if status and status != 'all':
        query["status"] = status
    if from_date:
        query["delivery_date"] = {"$gte": from_date}
    if to_date:
        if "delivery_date" in query:
            query["delivery_date"]["$lte"] = to_date
        else:
            query["delivery_date"] = {"$lte": to_date}
    
    total = await db.distributor_deliveries.count_documents(query)
    
    deliveries = await db.distributor_deliveries.find(
        query,
        {"_id": 0}
    ).sort([("delivery_date", -1), ("created_at", -1)]).skip((page - 1) * page_size).limit(page_size).to_list(page_size)
    
    return {
        "deliveries": deliveries,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size
    }


@router.get("/deliveries/summary")
async def get_deliveries_summary(
    distributor_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get deliveries summary stats"""
    tenant_id = get_current_tenant_id()
    
    query = {"tenant_id": tenant_id}
    if distributor_id:
        query["distributor_id"] = distributor_id
    
    total = await db.distributor_deliveries.count_documents(query)
    
    # Count by status
    draft = await db.distributor_deliveries.count_documents({**query, "status": "draft"})
    confirmed = await db.distributor_deliveries.count_documents({**query, "status": "confirmed"})
    delivered = await db.distributor_deliveries.count_documents({**query, "status": {"$in": ["delivered", "complete"]}})
    cancelled = await db.distributor_deliveries.count_documents({**query, "status": "cancelled"})
    
    # Calculate totals
    pipeline = [
        {"$match": {**query, "status": {"$nin": ["cancelled", "returned"]}}},
        {"$group": {
            "_id": None,
            "total_quantity": {"$sum": "$total_quantity"},
            "total_amount": {"$sum": "$total_net_amount"},
            "total_margin": {"$sum": "$total_margin_amount"}
        }}
    ]
    
    totals = await db.distributor_deliveries.aggregate(pipeline).to_list(1)
    total_quantity = totals[0]["total_quantity"] if totals else 0
    total_amount = totals[0]["total_amount"] if totals else 0
    total_margin = totals[0]["total_margin"] if totals else 0
    
    return {
        "total": total,
        "by_status": {
            "draft": draft,
            "confirmed": confirmed,
            "delivered": delivered,
            "cancelled": cancelled
        },
        "total_quantity": total_quantity,
        "total_amount": round(total_amount, 2),
        "total_margin": round(total_margin, 2)
    }


@router.get("/{distributor_id}/deliveries")
async def list_distributor_deliveries(
    distributor_id: str,
    status: Optional[str] = None,
    account_id: Optional[str] = None,
    account_ids: Optional[str] = None,
    location_id: Optional[str] = None,
    time_filter: Optional[str] = 'this_month',
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    current_user: dict = Depends(get_current_user)
):
    """List deliveries for a specific distributor with items, pagination, and time filter"""
    tenant_id = get_current_tenant_id()
    
    query = {"tenant_id": tenant_id, "distributor_id": distributor_id}
    
    if status and status != 'all':
        query["status"] = status
    
    # Account filter — supports a single account_id (legacy) or a
    # comma-separated account_ids list (multi-select). When both are present
    # they are unioned.
    selected_accounts = set()
    if account_id and account_id != 'all':
        selected_accounts.add(account_id)
    if account_ids:
        selected_accounts.update(a for a in account_ids.split(',') if a and a != 'all')
    if selected_accounts:
        query["account_id"] = {"$in": list(selected_accounts)}
    
    if location_id and location_id != 'all':
        query["distributor_location_id"] = location_id
    
    # Apply time filter
    now = datetime.now(timezone.utc)
    if time_filter and time_filter != 'lifetime':
        if time_filter == 'this_week':
            start_date = (now - timedelta(days=now.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = now
        elif time_filter == 'last_week':
            start_date = (now - timedelta(days=now.weekday() + 7)).replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = (now - timedelta(days=now.weekday() + 1)).replace(hour=23, minute=59, second=59, microsecond=999999)
        elif time_filter == 'this_month':
            start_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            end_date = now
        elif time_filter == 'last_month':
            first_of_current = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            last_day_of_prev = first_of_current - timedelta(days=1)
            start_date = last_day_of_prev.replace(day=1)
            end_date = last_day_of_prev.replace(hour=23, minute=59, second=59, microsecond=999999)
        elif time_filter == 'last_3_months':
            start_date = (now - timedelta(days=90)).replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = now
        elif time_filter == 'last_6_months':
            start_date = (now - timedelta(days=180)).replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = now
        elif time_filter == 'this_year':
            start_date = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
            end_date = now
        else:
            start_date = None
            end_date = None
        
        if start_date and end_date:
            query["delivery_date"] = {
                "$gte": start_date.strftime("%Y-%m-%d"),
                "$lte": end_date.strftime("%Y-%m-%d")
            }
    
    total = await db.distributor_deliveries.count_documents(query)
    
    deliveries = await db.distributor_deliveries.find(
        query,
        {"_id": 0}
    ).sort([("delivery_date", -1), ("created_at", -1)]).skip((page - 1) * page_size).limit(page_size).to_list(page_size)
    
    # Fetch items for each delivery
    for delivery in deliveries:
        items = await db.distributor_delivery_items.find(
            {"delivery_id": delivery['id'], "tenant_id": tenant_id},
            {"_id": 0}
        ).to_list(500)
        delivery['items'] = items
    
    return {
        "deliveries": deliveries,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size
    }


@router.get("/{distributor_id}/deliveries/export")
async def export_distributor_deliveries(
    distributor_id: str,
    status: Optional[str] = None,
    account_id: Optional[str] = None,
    account_ids: Optional[str] = None,
    location_id: Optional[str] = None,
    time_filter: Optional[str] = 'this_month',
    current_user: dict = Depends(get_current_user)
):
    """Export the (filtered) Stock Out deliveries to an Excel (.xlsx) file with
    a per-delivery summary row and collapsible per-SKU detail rows (drill-down)."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse

    tenant_id = get_current_tenant_id()

    query = {"tenant_id": tenant_id, "distributor_id": distributor_id}
    if status and status != 'all':
        query["status"] = status

    selected_accounts = set()
    if account_id and account_id != 'all':
        selected_accounts.add(account_id)
    if account_ids:
        selected_accounts.update(a for a in account_ids.split(',') if a and a != 'all')
    if selected_accounts:
        query["account_id"] = {"$in": list(selected_accounts)}

    if location_id and location_id != 'all':
        query["distributor_location_id"] = location_id

    now = datetime.now(timezone.utc)
    if time_filter and time_filter != 'lifetime':
        start_date = end_date = None
        if time_filter == 'this_week':
            start_date = (now - timedelta(days=now.weekday())).replace(hour=0, minute=0, second=0, microsecond=0); end_date = now
        elif time_filter == 'last_week':
            start_date = (now - timedelta(days=now.weekday() + 7)).replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = (now - timedelta(days=now.weekday() + 1)).replace(hour=23, minute=59, second=59, microsecond=999999)
        elif time_filter == 'this_month':
            start_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0); end_date = now
        elif time_filter == 'last_month':
            first_of_current = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            last_day_of_prev = first_of_current - timedelta(days=1)
            start_date = last_day_of_prev.replace(day=1); end_date = last_day_of_prev.replace(hour=23, minute=59, second=59, microsecond=999999)
        elif time_filter == 'last_3_months':
            start_date = (now - timedelta(days=90)).replace(hour=0, minute=0, second=0, microsecond=0); end_date = now
        elif time_filter == 'last_6_months':
            start_date = (now - timedelta(days=180)).replace(hour=0, minute=0, second=0, microsecond=0); end_date = now
        elif time_filter == 'this_year':
            start_date = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0); end_date = now
        if start_date and end_date:
            query["delivery_date"] = {"$gte": start_date.strftime("%Y-%m-%d"), "$lte": end_date.strftime("%Y-%m-%d")}

    deliveries = await db.distributor_deliveries.find(query, {"_id": 0}) \
        .sort([("delivery_date", -1), ("created_at", -1)]).to_list(5000)
    for delivery in deliveries:
        delivery['items'] = await db.distributor_delivery_items.find(
            {"delivery_id": delivery['id'], "tenant_id": tenant_id}, {"_id": 0}
        ).to_list(500)

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Out"
    ws.sheet_properties.outlinePr.summaryBelow = False

    headers = ['Delivery #', 'Date', 'Account', 'City', 'SKU', 'Quantity', 'Margin %',
               'Base Price', 'Transfer Price', 'Billed to Distributor', 'Customer Price',
               'New Transfer Price', 'Actual Billable to Distributor',
               'Adjustment (Dist to Factory)', 'Customer Invoice Amount', 'Status']
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="166534", end_color="166534", fill_type="solid")
    summary_fill = PatternFill(start_color="ecfdf5", end_color="ecfdf5", fill_type="solid")
    summary_font = Font(bold=True, color="065f46")
    thin = Side(style="thin", color="d0d5dd")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c_idx, name in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    widths = [16, 12, 28, 16, 30, 10, 9, 12, 14, 16, 12, 14, 20, 18, 18, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    def _item_calc(item):
        qty = item.get('quantity') or 0
        customer_price = item.get('customer_selling_price') or item.get('unit_price') or 0
        commission_pct = item.get('distributor_commission_percent')
        if commission_pct is None:
            commission_pct = item.get('margin_percent')
        if commission_pct is None:
            commission_pct = 2.5
        base_price = item.get('base_price') or item.get('transfer_price') or 0
        transfer_price = base_price * (1 - commission_pct / 100) if base_price > 0 else 0
        billed = qty * transfer_price
        new_tp = customer_price * (1 - commission_pct / 100) if customer_price > 0 else 0
        actual = qty * new_tp
        return {
            'qty': qty, 'commission_pct': commission_pct, 'base_price': base_price,
            'transfer_price': transfer_price, 'billed': billed, 'customer_price': customer_price,
            'new_tp': new_tp, 'actual': actual, 'adjustment': actual - billed,
            'customer_invoice': qty * customer_price,
        }

    r = 2
    for d in deliveries:
        items = d.get('items') or []
        dev_date = d.get('delivery_date') or '-'
        if isinstance(dev_date, str):
            dev_date = dev_date[:10]
        calcs = [_item_calc(it) for it in items]
        tot = {k: sum(c[k] for c in calcs) for k in ('qty', 'billed', 'actual', 'adjustment', 'customer_invoice')} if calcs else {k: 0 for k in ('qty', 'billed', 'actual', 'adjustment', 'customer_invoice')}
        summary = [
            d.get('delivery_number') or '-', dev_date, d.get('account_name') or '-', d.get('account_city') or '-',
            f"{len(items)} SKU(s)", tot['qty'], '', '', '', round(tot['billed'], 2), '', '',
            round(tot['actual'], 2), round(tot['adjustment'], 2), round(tot['customer_invoice'], 2),
            d.get('status') or '-',
        ]
        for c_idx, val in enumerate(summary, 1):
            cell = ws.cell(row=r, column=c_idx, value=val)
            cell.border = border
            cell.fill = summary_fill
            cell.font = summary_font
            if c_idx == 6:
                cell.number_format = '#,##0'
            elif c_idx in (10, 13, 14, 15):
                cell.number_format = '#,##0.00'
        r += 1
        for it, c in zip(items, calcs):
            detail = [
                '', '', '', '', f"    {it.get('sku_name') or it.get('sku_code') or 'N/A'}",
                c['qty'], round(c['commission_pct'], 2), round(c['base_price'], 2),
                round(c['transfer_price'], 2), round(c['billed'], 2), round(c['customer_price'], 2),
                round(c['new_tp'], 2), round(c['actual'], 2), round(c['adjustment'], 2),
                round(c['customer_invoice'], 2), '',
            ]
            for c_idx, val in enumerate(detail, 1):
                cell = ws.cell(row=r, column=c_idx, value=val)
                cell.border = border
                if c_idx == 6:
                    cell.number_format = '#,##0'
                elif c_idx in (7, 8, 9, 10, 11, 12, 13, 14, 15):
                    cell.number_format = '#,##0.00'
            ws.row_dimensions[r].outline_level = 1
            r += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"stock_out_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{distributor_id}/deliveries/{delivery_id}")
async def get_delivery(
    distributor_id: str,
    delivery_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get a specific delivery with items"""
    tenant_id = get_current_tenant_id()
    
    delivery = await db.distributor_deliveries.find_one(
        {"id": delivery_id, "tenant_id": tenant_id, "distributor_id": distributor_id},
        {"_id": 0}
    )
    
    if not delivery:
        raise HTTPException(status_code=404, detail="Delivery not found")
    
    # Get delivery items
    items = await db.distributor_delivery_items.find(
        {"delivery_id": delivery_id, "tenant_id": tenant_id},
        {"_id": 0}
    ).to_list(500)
    
    delivery['items'] = items

    # Surface the account's billed_by flag on the delivery so the frontend can
    # hide Zoho invoice affordances for accounts billed by a third-party
    # distributor. This avoids an extra round-trip on the popup.
    if delivery.get('account_id'):
        acc = await db.accounts.find_one(
            {"id": delivery['account_id'], "tenant_id": tenant_id},
            {"_id": 0, "billed_by": 1}
        )
        delivery['account_billed_by'] = (acc or {}).get('billed_by') or 'company'
    else:
        delivery['account_billed_by'] = 'company'

    # Enrich applied_credit_notes with current Zoho deep-link / status — the
    # snapshot stored on the delivery doesn't carry Zoho fields, and a CN may
    # be synced to Zoho AFTER the delivery was created.
    applied = delivery.get('applied_credit_notes') or []
    cn_ids = [a.get('credit_note_id') for a in applied if a.get('credit_note_id')]
    if cn_ids:
        cns = await db.credit_notes.find(
            {'id': {'$in': cn_ids}, 'tenant_id': tenant_id},
            {'_id': 0, 'id': 1, 'zoho_creditnote_id': 1, 'zoho_creditnote_number': 1, 'zoho_creditnote_url': 1, 'status': 1}
        ).to_list(len(cn_ids))
        cn_map = {c['id']: c for c in cns}
        for a in applied:
            c = cn_map.get(a.get('credit_note_id'))
            if c:
                a['zoho_creditnote_id'] = c.get('zoho_creditnote_id')
                a['zoho_creditnote_number'] = c.get('zoho_creditnote_number')
                a['zoho_creditnote_url'] = c.get('zoho_creditnote_url')
        delivery['applied_credit_notes'] = applied
    
    return delivery


@router.get("/{distributor_id}/assigned-accounts")
async def get_assigned_accounts_for_delivery(
    distributor_id: str,
    city: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get accounts assigned to this distributor for delivery selection"""
    tenant_id = get_current_tenant_id()
    
    query = {
        "tenant_id": tenant_id,
        "distributor_id": distributor_id,
        "status": "active"
    }
    
    if city:
        query["servicing_city"] = city
    
    assignments = await db.account_distributor_assignments.find(
        query,
        {"_id": 0}
    ).to_list(500)
    
    # Get account details for each assignment
    accounts = []
    for assignment in assignments:
        account = await db.accounts.find_one(
            {"id": assignment.get('account_id')},
            {"_id": 0, "id": 1, "account_name": 1, "contact_name": 1, "city": 1, "state": 1, "delivery_address": 1, "territory": 1, "contact_number": 1, "sku_pricing": 1}
        )
        if account:
            # ── SKU pricing — pick the row that is ACTIVE today for each SKU.
            # `account.sku_pricing` may contain historical / future-dated rows;
            # we filter to the one matching today's date window. If multiple
            # rows overlap, the one with the latest `active_from` wins (most
            # recently scheduled price wins). Rows without dates are treated
            # as "always active" but lose to a dated overlapping row.
            from datetime import date as _date
            raw_pricing = account.get('sku_pricing', []) or []
            today_iso = _date.today().isoformat()

            def _in_window(row):
                af = (row.get('active_from') or '').strip()
                at = (row.get('active_to') or '').strip()
                if af and today_iso < af:
                    return False
                if at and today_iso > at:
                    return False
                return True

            by_sku: dict[str, dict] = {}
            for row in raw_pricing:
                if not _in_window(row):
                    continue
                key = (row.get('sku') or '').strip()
                if not key:
                    continue
                cur = by_sku.get(key)
                if cur is None:
                    by_sku[key] = row
                    continue
                # Prefer the row with the latest active_from (dated > undated).
                if (row.get('active_from') or '') > (cur.get('active_from') or ''):
                    by_sku[key] = row
            sku_pricing = list(by_sku.values())

            enriched_skus = []
            if sku_pricing:
                # Get all master SKUs to map names to IDs
                # Note: master_skus may not have tenant_id set (global catalog)
                master_skus = await db.master_skus.find(
                    {"$or": [{"tenant_id": tenant_id}, {"tenant_id": None}, {"tenant_id": {"$exists": False}}]},
                    {"_id": 0, "id": 1, "name": 1, "sku_name": 1}
                ).to_list(500)
                
                # Build name-to-ID mapping using sku_name (primary) or name (fallback)
                sku_name_to_id = {}
                for s in master_skus:
                    sku_key = s.get('sku_name') or s.get('name')
                    if sku_key:
                        sku_name_to_id[sku_key] = s.get('id')
                
                for sku_item in sku_pricing:
                    sku_name = sku_item.get('sku')
                    sku_id = sku_name_to_id.get(sku_name)
                    if sku_id:
                        enriched_skus.append({
                            "id": sku_id,
                            "name": sku_name,
                            "price_per_unit": sku_item.get('price_per_unit', 0),
                            "return_bottle_credit": sku_item.get('return_bottle_credit', 0)
                        })
            
            accounts.append({
                "id": account.get('id'),
                "account_name": account.get('account_name', 'Unknown Account'),
                "contact_name": account.get('contact_name', ''),
                "contact_number": account.get('contact_number', ''),
                "city": account.get('city', ''),
                "state": account.get('state', ''),
                "territory": account.get('territory', ''),
                "delivery_address": account.get('delivery_address', ''),
                "servicing_city": assignment.get('servicing_city'),
                "distributor_location_id": assignment.get('distributor_location_id'),
                "distributor_location_name": assignment.get('distributor_location_name'),
                "is_primary": assignment.get('is_primary', False),
                "sku_pricing": enriched_skus  # Include account's configured SKUs with IDs
            })
    
    return {"accounts": accounts}


# Statuses where a Stock Out order still HOLDS a reservation on stock — i.e.
# every open order that has not yet been delivered or cancelled. Stock on these
# orders is "Reserved" (committed) and must not be re-allocated to other orders.
RESERVED_DELIVERY_STATUSES = [
    "draft", "pending", "confirmed",
    "scheduled", "delivery_scheduled", "delivery_assigned",
    "on_the_way", "in_transit",
]


async def _reserved_qty_map(tenant_id, distributor_id, location_id, tracks_batches, exclude_delivery_id=None):
    """Reserved (committed) quantity per (sku_id, batch_id|None) at a source
    location, summed across all OPEN Stock Out orders. Fully derived from the
    deliveries themselves, so cancel/complete/delete need no extra bookkeeping —
    the reservation simply disappears once the order leaves the open set."""
    open_ids = await db.distributor_deliveries.distinct(
        "id",
        {
            "tenant_id": tenant_id,
            "distributor_id": distributor_id,
            "distributor_location_id": location_id,
            "status": {"$in": RESERVED_DELIVERY_STATUSES},
        },
    )
    if exclude_delivery_id:
        open_ids = [i for i in open_ids if i != exclude_delivery_id]
    if not open_ids:
        return {}
    reserved: dict = {}
    async for it in db.distributor_delivery_items.find(
        {"tenant_id": tenant_id, "delivery_id": {"$in": open_ids}},
        {"_id": 0, "sku_id": 1, "batch_id": 1, "quantity": 1},
    ):
        key = (it.get("sku_id"), it.get("batch_id") if tracks_batches else None)
        reserved[key] = reserved.get(key, 0) + int(it.get("quantity") or 0)
    return reserved


@router.post("/{distributor_id}/deliveries")
async def create_delivery(
    distributor_id: str,
    data: AccountDeliveryCreate,
    current_user: dict = Depends(get_current_user)
):
    """Create a new delivery to an account"""
    if not can_manage_distributor_data(current_user, distributor_id):
        raise HTTPException(status_code=403, detail="Not authorised to manage this distributor's deliveries")
    
    tenant_id = get_current_tenant_id()
    now = datetime.now(timezone.utc).isoformat()
    
    # Validate distributor exists
    distributor = await db.distributors.find_one(
        {"id": distributor_id, "tenant_id": tenant_id},
        {"_id": 0, "distributor_name": 1, "distributor_code": 1, "billing_approach": 1}
    )
    if not distributor:
        raise HTTPException(status_code=404, detail="Distributor not found")
    
    billing_approach = distributor.get('billing_approach', 'margin_upfront')
    
    # Validate location
    location = await db.distributor_locations.find_one(
        {"id": data.distributor_location_id, "tenant_id": tenant_id, "distributor_id": distributor_id},
        {"_id": 0, "location_name": 1, "city": 1, "track_batches": 1, "is_factory": 1}
    )
    if not location:
        raise HTTPException(status_code=400, detail="Invalid distributor location")
    source_tracks_batches = bool(location.get("track_batches"))
    source_is_factory = bool(location.get("is_factory"))
    
    # Validate account exists
    account = await db.accounts.find_one(
        {"id": data.account_id, "tenant_id": tenant_id},
        {"_id": 0, "company": 1, "name": 1, "account_name": 1, "city": 1, "state": 1, "address": 1}
    )
    if not account:
        raise HTTPException(status_code=400, detail="Account not found")
    
    # Validate items
    if not data.items or len(data.items) == 0:
        raise HTTPException(status_code=400, detail="At least one item is required")

    # Phase 2 batch tracking: when the source distributor warehouse has
    # `track_batches=True`, every line MUST carry a `batch_id`. Reject up
    # front with a clear list of missing SKUs.
    if source_tracks_batches:
        missing_batch = [
            (item.sku_name or item.sku_id)
            for item in data.items
            if not item.batch_id
        ]
        if missing_batch:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"Warehouse '{location.get('location_name')}' tracks batches — "
                    f"please pick a production batch for: {', '.join(missing_batch[:5])}"
                    + (f" and {len(missing_batch) - 5} more" if len(missing_batch) > 5 else "")
                ),
            )

    # ── Stock availability validation ─────────────────────────────────────
    # Reject the delivery up-front if any line exceeds the on-hand quantity
    # at the source warehouse. Without this, drivers could silently record
    # deliveries that drive the inventory negative (or worse, leave it
    # untouched if the deduction path fails to match a row).
    #
    # Same row-lookup convention as `complete_delivery`:
    #   • factory source       → factory_warehouse_stock (warehouse_location_id)
    #   • distributor source   → distributor_stock (distributor_location_id)
    #   • track_batches=True   → row keyed by batch_id
    #   • track_batches=False  → aggregated row
    #
    # Aggregate per (sku_id, batch_id) so multi-line deliveries of the same
    # SKU/batch are summed before checking.
    stock_demand: dict = {}
    for item in data.items:
        qty = int(item.quantity or 0)
        if qty <= 0:
            continue
        key = (item.sku_id, item.batch_id if source_tracks_batches else None)
        stock_demand[key] = stock_demand.get(key, 0) + qty

    shortages: list = []

    # ── Pre-aggregate the "derived" view of on-hand for non-factory,
    # non-batch-tracked sources. This mirrors the dashboard's formula
    # (received − delivered − pending), so the guard NEVER disagrees with the
    # dashboard the user just saw. Critical for distributors where
    # `distributor_stock` rows may be missing/stale (legacy data, partial
    # migrations) but the shipment/delivery history is intact.
    derived_on_hand_by_sku: dict = {}
    if not source_is_factory and not source_tracks_batches:
        sku_ids_needed = list({sku for (sku, _) in stock_demand.keys() if sku})
        if sku_ids_needed:
            # Single-location distributors (e.g. third-party / "not self-managed"
            # distributors with one default warehouse): scope the derived view to
            # the DISTRIBUTOR, not the location. Legacy shipments are frequently
            # missing `distributor_location_id`, so a location-scoped "received"
            # tally returns 0 while location-scoped "delivered" matches — driving
            # on-hand negative (e.g. -720) even though the dashboard (which is
            # distributor-wide) shows thousands available. Matching the dashboard
            # here keeps the guard consistent with what the user just saw.
            # (Bug hit in production for the Goa distributor, 2026-06.)
            non_factory_loc_count = await db.distributor_locations.count_documents({
                "tenant_id": tenant_id, "distributor_id": distributor_id,
                "is_factory": {"$ne": True},
            })
            single_location = non_factory_loc_count <= 1

            ship_filter = {
                "tenant_id": tenant_id,
                "distributor_id": distributor_id,
                "status": "delivered",
            }
            del_filter = {
                "tenant_id": tenant_id, "distributor_id": distributor_id,
                "status": {"$in": ["delivered", "completed", "complete"]},
            }
            if not single_location:
                ship_filter["distributor_location_id"] = data.distributor_location_id
                del_filter["distributor_location_id"] = data.distributor_location_id

            # Received = sum(delivered shipment_items for these SKUs)
            delivered_ship_ids = await db.distributor_shipments.distinct("id", ship_filter)
            for it in await db.distributor_shipment_items.find(
                {"tenant_id": tenant_id, "shipment_id": {"$in": delivered_ship_ids}, "sku_id": {"$in": sku_ids_needed}},
                {"_id": 0, "sku_id": 1, "quantity": 1, "received_quantity": 1},
            ).to_list(20000):
                # `received_quantity` is the source of truth when acknowledgement
                # was used; otherwise the dispatched `quantity`. Either way they
                # match what the dashboard tallies as "received".
                q = int(it.get("received_quantity") or it.get("quantity") or 0)
                derived_on_hand_by_sku[it["sku_id"]] = derived_on_hand_by_sku.get(it["sku_id"], 0) + q
            # Delivered out = sum(items on completed deliveries)
            done_delivery_ids = await db.distributor_deliveries.distinct("id", del_filter)
            for it in await db.distributor_delivery_items.find(
                {"tenant_id": tenant_id, "delivery_id": {"$in": done_delivery_ids}, "sku_id": {"$in": sku_ids_needed}},
                {"_id": 0, "sku_id": 1, "quantity": 1},
            ).to_list(20000):
                derived_on_hand_by_sku[it["sku_id"]] = derived_on_hand_by_sku.get(it["sku_id"], 0) - int(it.get("quantity") or 0)
            # NOTE: open/pending orders are NOT subtracted here — they are
            # handled uniformly below via the Reserved map so every source type
            # (factory / batch / distributor) nets out committed stock the same
            # way and we never double-count.

    # Reserved = stock already committed to OTHER open Stock Out orders at this
    # source location. Available = on-hand − reserved. Reserving here makes that
    # stock unavailable for allocation to any other customer/order.
    reserved_map = await _reserved_qty_map(
        tenant_id, distributor_id, data.distributor_location_id, source_tracks_batches
    )

    for (sku_id, batch_id), demand in stock_demand.items():
        if source_is_factory:
            stock_query = {
                "tenant_id": tenant_id,
                "warehouse_location_id": data.distributor_location_id,
                "sku_id": sku_id,
            }
            stock_collection = db.factory_warehouse_stock
        else:
            stock_query = {
                "tenant_id": tenant_id,
                "distributor_id": distributor_id,
                "distributor_location_id": data.distributor_location_id,
                "sku_id": sku_id,
            }
            stock_collection = db.distributor_stock
        if source_tracks_batches and batch_id:
            stock_query["batch_id"] = batch_id
        elif source_tracks_batches:
            # Legacy aggregate row (no batch_id stored).
            stock_query["batch_id"] = {"$in": [None]}

        rows = await stock_collection.find(
            stock_query, {"_id": 0, "quantity": 1, "sku_name": 1, "batch_code": 1}
        ).to_list(50)
        on_hand_rows = sum((r.get("quantity") or 0) for r in rows)

        # When source is a non-batch distributor warehouse, take the MAX of
        # the row-derived (`distributor_stock`) and the dashboard-derived
        # view. Two reasons:
        #   1. Some legacy distributor_stock rows were never populated, but
        #      shipments/deliveries are intact — the dashboard is right.
        #   2. We never want the guard to be MORE restrictive than the value
        #      the user just saw on the dashboard. Otherwise the UX is "you
        #      have 3,468 crates" and the API says "you have 0", which we
        #      hit in production on 2026-05-28.
        if not source_is_factory and not source_tracks_batches:
            on_hand = max(on_hand_rows, derived_on_hand_by_sku.get(sku_id, 0))
        else:
            on_hand = on_hand_rows

        reserved = reserved_map.get((sku_id, batch_id if source_tracks_batches else None), 0)
        available = on_hand - reserved

        if demand > available:
            label = (rows[0].get("sku_name") if rows else None) or sku_id
            batch_label = ""
            if source_tracks_batches and batch_id:
                bc = (rows[0].get("batch_code") if rows else None) or batch_id[:8]
                batch_label = f" (batch {bc})"
            committed = f", {reserved} reserved" if reserved else ""
            shortages.append(f"{label}{batch_label}: need {demand}, available {available} ({on_hand} on-hand{committed})")

    if shortages:
        raise HTTPException(
            status_code=400,
            detail=(
                f"Not enough stock at '{location.get('location_name')}' — "
                + "; ".join(shortages[:5])
                + (f"; and {len(shortages) - 5} more" if len(shortages) > 5 else "")
            ),
        )

    # Generate delivery number
    delivery_number = await generate_delivery_number(tenant_id)
    delivery_id = str(uuid.uuid4())
    
    # Build delivery address
    delivery_address = data.delivery_address or account.get('address', '')
    
    # Get margin matrix for this account/city to calculate earnings
    account_city = account.get('city', '')
    
    # VALIDATION: Check if margin matrix is configured for all SKUs in this city
    missing_margins = []
    for item_data in data.items:
        sku = await db.master_skus.find_one({"id": item_data.sku_id}, {"_id": 0, "sku_name": 1, "sku_code": 1})
        sku_name = item_data.sku_name or (sku.get('sku_name') if sku else f"SKU {item_data.sku_id}")
        
        margin = await db.distributor_margin_matrix.find_one({
            "tenant_id": tenant_id,
            "distributor_id": distributor_id,
            "city": account_city,
            "sku_id": item_data.sku_id,
            "status": "active"
        }, {"_id": 0, "id": 1})
        
        if not margin:
            missing_margins.append(sku_name)
    
    if missing_margins:
        sku_list = ", ".join(missing_margins[:3])
        if len(missing_margins) > 3:
            sku_list += f" and {len(missing_margins) - 3} more"
        raise HTTPException(
            status_code=400, 
            detail=f"Margin matrix not configured for city '{account_city}'. Please configure margins for: {sku_list}"
        )
    
    # Process items and calculate totals
    items_to_insert = []
    total_quantity = 0
    total_gross_amount = 0
    total_discount_amount = 0
    total_tax_amount = 0
    total_net_amount = 0
    total_margin_amount = 0
    total_price_premium = 0
    total_factory_adjustment = 0
    
    for item_data in data.items:
        # Get SKU info
        sku_name = item_data.sku_name
        sku_code = None
        sku = await db.master_skus.find_one({"id": item_data.sku_id}, {"_id": 0, "sku_name": 1, "sku_code": 1})
        if sku:
            sku_name = sku_name or sku.get('sku_name')
            sku_code = sku.get('sku_code')
        
        # Get margin for this SKU and city
        margin = await db.distributor_margin_matrix.find_one({
            "tenant_id": tenant_id,
            "distributor_id": distributor_id,
            "city": account_city,
            "sku_id": item_data.sku_id,
            "status": "active"
        }, {"_id": 0, "margin_type": 1, "margin_value": 1, "transfer_price": 1, "base_price": 1})
        
        margin_type = margin.get('margin_type') if margin else None
        margin_value = margin.get('margin_value') if margin else None
        transfer_price = margin.get('transfer_price') if margin else None
        base_price = margin.get('base_price') if margin else None
        
        item_dict = {
            'id': str(uuid.uuid4()),
            'tenant_id': tenant_id,
            'delivery_id': delivery_id,
            'sku_id': item_data.sku_id,
            'sku_name': sku_name,
            'sku_code': sku_code,
            'quantity': item_data.quantity,
            'unit_price': item_data.unit_price,
            'customer_selling_price': item_data.customer_selling_price or item_data.unit_price,
            'distributor_commission_percent': item_data.distributor_commission_percent or margin_value,
            'transfer_price': item_data.transfer_price or transfer_price,
            'base_price': item_data.base_price or base_price,
            'discount_percent': item_data.discount_percent or 0,
            'tax_percent': item_data.tax_percent or 0,
            'remarks': item_data.remarks,
            # Phase 2 batch identity — None when the source warehouse doesn't
            # track batches. Stored on every row so downstream code can rely
            # on the field existing.
            'batch_id': item_data.batch_id,
            'batch_code': item_data.batch_code,
            # Packaging breakdown for the delivery detail popup.
            'packaging_type_name': item_data.packaging_type_name,
            'packaging_units': item_data.packaging_units,
            'packages': item_data.packages,
        }
        
        # Calculate amounts with margin and transfer price
        item_dict = calculate_delivery_item_amounts(item_dict, margin_type, margin_value, transfer_price, base_price, billing_approach)
        items_to_insert.append(item_dict)
        
        total_quantity += item_data.quantity
        total_gross_amount += item_dict['gross_amount']
        total_discount_amount += item_dict['discount_amount']
        total_tax_amount += item_dict['tax_amount']
        total_net_amount += item_dict['net_amount']
        total_margin_amount += item_dict.get('margin_amount', 0)
        total_price_premium += item_dict.get('price_premium_payable', 0)
        total_factory_adjustment += item_dict.get('factory_distributor_adjustment', 0)
    
    # Create delivery document
    delivery_doc = {
        "id": delivery_id,
        "tenant_id": tenant_id,
        "delivery_number": delivery_number,
        "distributor_id": distributor_id,
        "distributor_name": distributor.get('distributor_name'),
        "distributor_code": distributor.get('distributor_code'),
        "distributor_location_id": data.distributor_location_id,
        "distributor_location_name": location.get('location_name'),
        "account_id": data.account_id,
        "account_name": account.get('account_name') or account.get('company') or account.get('name') or 'Unknown',
        "account_city": account.get('city'),
        "account_state": account.get('state'),
        "delivery_date": data.delivery_date,
        "reference_number": data.reference_number,
        "vehicle_number": data.vehicle_number,
        "driver_name": data.driver_name,
        "driver_contact": data.driver_contact,
        "delivery_address": delivery_address,
        "status": "draft",
        "total_quantity": total_quantity,
        "total_gross_amount": round(total_gross_amount, 2),
        "total_discount_amount": round(total_discount_amount, 2),
        "total_tax_amount": round(total_tax_amount, 2),
        "total_net_amount": round(total_net_amount, 2),
        "total_margin_amount": round(total_margin_amount, 2),
        "total_price_premium": round(total_price_premium, 2),
        "total_factory_adjustment": round(total_factory_adjustment, 2),
        "remarks": data.remarks,
        "created_at": now,
        "updated_at": now,
        "created_by": current_user.get('id'),
        "confirmed_at": None,
        "confirmed_by": None,
        "delivered_at": None,
        "delivered_by": None
    }
    
    # Handle credit notes if provided
    applied_credit_notes = []
    total_credit_applied = 0
    
    if data.credit_notes_to_apply:
        from routes.credit_notes import apply_credit_note_to_delivery
        
        for cn_request in data.credit_notes_to_apply:
            cn_id = cn_request.credit_note_id
            amount = cn_request.amount_to_apply
            
            if not cn_id or amount <= 0:
                continue
            
            try:
                result = await apply_credit_note_to_delivery(
                    tenant_id=tenant_id,
                    credit_note_id=cn_id,
                    delivery_id=delivery_id,
                    delivery_number=delivery_number,
                    amount_to_apply=amount,
                    applied_by=current_user.get('id')
                )
                
                applied_credit_notes.append({
                    'credit_note_id': cn_id,
                    'credit_note_number': result.get('credit_note_number'),
                    'amount_applied': amount,
                    'return_number': result.get('return_number')
                })
                total_credit_applied += amount
            except HTTPException as e:
                logger.warning(f"Failed to apply credit note {cn_id} during delivery creation: {e.detail}")
    
    # Calculate net customer billing
    net_customer_billing = max(0, total_net_amount - total_credit_applied)
    
    # Add credit note fields to delivery doc
    delivery_doc['applied_credit_notes'] = applied_credit_notes
    delivery_doc['total_credit_applied'] = round(total_credit_applied, 2)
    delivery_doc['net_customer_billing'] = round(net_customer_billing, 2)
    
    # Handle debit notes if provided — these ADD to the customer's billing
    # (customer owes us for missing/unreturned bottles).
    applied_debit_notes = []
    total_debit_applied = 0

    if data.debit_notes_to_apply:
        from routes.credit_notes import apply_debit_note_to_delivery

        for dn_request in data.debit_notes_to_apply:
            dn_id = dn_request.debit_note_id
            amount = dn_request.amount_to_apply

            if not dn_id or amount <= 0:
                continue

            try:
                result = await apply_debit_note_to_delivery(
                    tenant_id=tenant_id,
                    debit_note_id=dn_id,
                    delivery_id=delivery_id,
                    delivery_number=delivery_number,
                    amount_to_apply=amount,
                    applied_by=current_user.get('id')
                )

                applied_debit_notes.append({
                    'debit_note_id': dn_id,
                    'debit_note_number': result.get('debit_note_number'),
                    'amount_applied': amount,
                    'return_number': result.get('return_number')
                })
                total_debit_applied += amount
            except HTTPException as e:
                logger.warning(f"Failed to apply debit note {dn_id} during delivery creation: {e.detail}")

    # Recompute net billing including debit charges
    net_customer_billing = max(0, total_net_amount - total_credit_applied + total_debit_applied)
    delivery_doc['applied_debit_notes'] = applied_debit_notes
    delivery_doc['total_debit_applied'] = round(total_debit_applied, 2)
    delivery_doc['net_customer_billing'] = round(net_customer_billing, 2)
    
    # Insert delivery and items
    await db.distributor_deliveries.insert_one(delivery_doc)
    if items_to_insert:
        await db.distributor_delivery_items.insert_many(items_to_insert)
    
    delivery_doc.pop('_id', None)
    delivery_doc['items'] = [
        {k: v for k, v in item.items() if k not in ['_id', 'tenant_id']} 
        for item in items_to_insert
    ]
    
    logger.info(f"Delivery {delivery_number} created for account {data.account_id} with ₹{total_credit_applied} in credit notes applied by {current_user['email']}")
    
    return delivery_doc


@router.post("/{distributor_id}/deliveries/{delivery_id}/apply-credit-notes")
async def apply_credit_notes_to_delivery(
    distributor_id: str,
    delivery_id: str,
    credit_notes: List[dict],  # [{credit_note_id: str, amount_to_apply: float}]
    current_user: dict = Depends(get_current_user)
):
    """Apply credit notes to a delivery"""
    from routes.credit_notes import apply_credit_note_to_delivery
    
    if not can_manage_distributor_data(current_user, distributor_id):
        raise HTTPException(status_code=403, detail="Not authorised to manage this distributor's deliveries")
    
    tenant_id = get_current_tenant_id()
    
    delivery = await db.distributor_deliveries.find_one(
        {"id": delivery_id, "tenant_id": tenant_id, "distributor_id": distributor_id}
    )
    
    if not delivery:
        raise HTTPException(status_code=404, detail="Delivery not found")
    
    if delivery.get('status') not in ['draft', 'confirmed']:
        raise HTTPException(status_code=400, detail="Can only apply credit notes to draft or confirmed deliveries")
    
    applied_credits = delivery.get('applied_credit_notes', []) or []
    total_credit_applied = delivery.get('total_credit_applied', 0)
    
    # Apply each credit note
    applications = []
    for cn_request in credit_notes:
        cn_id = cn_request.get('credit_note_id')
        amount = cn_request.get('amount_to_apply', 0)
        
        if not cn_id or amount <= 0:
            continue
        
        try:
            result = await apply_credit_note_to_delivery(
                tenant_id=tenant_id,
                credit_note_id=cn_id,
                delivery_id=delivery_id,
                delivery_number=delivery.get('delivery_number'),
                amount_to_apply=amount,
                applied_by=current_user.get('id')
            )
            
            applied_credits.append({
                'credit_note_id': cn_id,
                'credit_note_number': result.get('credit_note_number'),
                'amount_applied': amount
            })
            total_credit_applied += amount
            applications.append(result)
        except HTTPException as e:
            # Log but continue with other credit notes
            logger.warning(f"Failed to apply credit note {cn_id}: {e.detail}")
            applications.append({'error': e.detail, 'credit_note_id': cn_id})
    
    # Calculate net customer billing
    total_net_amount = delivery.get('total_net_amount', 0)
    net_customer_billing = max(0, total_net_amount - total_credit_applied)
    
    # Update delivery
    await db.distributor_deliveries.update_one(
        {"id": delivery_id, "tenant_id": tenant_id},
        {"$set": {
            "applied_credit_notes": applied_credits,
            "total_credit_applied": round(total_credit_applied, 2),
            "net_customer_billing": round(net_customer_billing, 2),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    logger.info(f"Applied ₹{total_credit_applied} in credit notes to delivery {delivery.get('delivery_number')}")
    
    return {
        "message": f"Applied {len([a for a in applications if 'error' not in a])} credit notes",
        "applications": applications,
        "total_credit_applied": round(total_credit_applied, 2),
        "net_customer_billing": round(net_customer_billing, 2)
    }


@router.put("/{distributor_id}/deliveries/{delivery_id}")
async def update_delivery(
    distributor_id: str,
    delivery_id: str,
    data: AccountDeliveryUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Update a delivery"""
    if not can_manage_distributor_data(current_user, distributor_id):
        raise HTTPException(status_code=403, detail="Not authorised to manage this distributor's deliveries")
    
    tenant_id = get_current_tenant_id()
    
    delivery = await db.distributor_deliveries.find_one(
        {"id": delivery_id, "tenant_id": tenant_id, "distributor_id": distributor_id}
    )
    
    if not delivery:
        raise HTTPException(status_code=404, detail="Delivery not found")
    
    current_status = delivery.get('status')
    
    # Only allow limited updates for non-draft deliveries. `delivery_date` is
    # editable here so users can CORRECT the planned delivery date (e.g. when a
    # delivery was completed late) without it affecting the completion record.
    if current_status != 'draft':
        allowed_fields = ['remarks', 'vehicle_number', 'driver_name', 'driver_contact', 'delivery_date']
        update_data = {"updated_at": datetime.now(timezone.utc).isoformat()}
        
        for field in allowed_fields:
            value = getattr(data, field, None)
            if value is not None:
                update_data[field] = value
    else:
        update_data = {"updated_at": datetime.now(timezone.utc).isoformat()}
        
        for field in ['delivery_date', 'reference_number', 'vehicle_number', 
                      'driver_name', 'driver_contact', 'delivery_address', 'remarks']:
            value = getattr(data, field, None)
            if value is not None:
                update_data[field] = value
    
    await db.distributor_deliveries.update_one(
        {"id": delivery_id, "tenant_id": tenant_id},
        {"$set": update_data}
    )
    
    updated = await db.distributor_deliveries.find_one(
        {"id": delivery_id, "tenant_id": tenant_id},
        {"_id": 0}
    )
    
    return updated


@router.post("/{distributor_id}/deliveries/{delivery_id}/confirm")
async def confirm_delivery(
    distributor_id: str,
    delivery_id: str,
    background_tasks: BackgroundTasks,
    current_user: dict = Depends(get_current_user)
):
    """Confirm a draft delivery"""
    if not can_manage_distributor_data(current_user, distributor_id):
        raise HTTPException(status_code=403, detail="Not authorised to manage this distributor's deliveries")
    
    tenant_id = get_current_tenant_id()
    
    delivery = await db.distributor_deliveries.find_one(
        {"id": delivery_id, "tenant_id": tenant_id, "distributor_id": distributor_id}
    )
    
    if not delivery:
        raise HTTPException(status_code=404, detail="Delivery not found")
    
    if delivery.get('status') != 'draft':
        raise HTTPException(status_code=400, detail="Only draft deliveries can be confirmed")
    
    now = datetime.now(timezone.utc).isoformat()
    
    await db.distributor_deliveries.update_one(
        {"id": delivery_id, "tenant_id": tenant_id},
        {"$set": {
            "status": "confirmed",
            "confirmed_at": now,
            "confirmed_by": current_user.get('id'),
            "updated_at": now
        }}
    )
    
    logger.info(f"Delivery {delivery['delivery_number']} confirmed by {current_user['email']}")

    # NOTE: Zoho invoice push is intentionally NOT triggered here anymore. The
    # business flow moved to: stock-out confirm → schedule confirm (Zoho push
    # happens then, once per attached delivery). The retry endpoint below can
    # still be called manually if a particular delivery needs to be re-synced.

    return {"message": f"Delivery {delivery['delivery_number']} confirmed", "status": "confirmed"}


@router.post("/{distributor_id}/deliveries/{delivery_id}/retry-zoho-push")
async def retry_delivery_zoho_push(
    distributor_id: str,
    delivery_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Manually re-attempt the Zoho Books invoice push for a confirmed delivery.

    Surfaces specific reasons via ZohoPushSkippedError (non-factory warehouse,
    account not linked, etc.) so the user sees exactly what to fix.
    """
    if not can_manage_distributor_data(current_user, distributor_id):
        raise HTTPException(status_code=403, detail="Not authorised to manage this distributor's deliveries")

    tenant_id = get_current_tenant_id()
    delivery = await db.distributor_deliveries.find_one(
        {"id": delivery_id, "tenant_id": tenant_id, "distributor_id": distributor_id},
        {"_id": 0},
    )
    if not delivery:
        raise HTTPException(status_code=404, detail="Delivery not found")
    if delivery.get('status') not in ('confirmed', 'scheduled', 'delivered', 'complete', 'delivery_assigned', 'delivery_scheduled', 'on_the_way'):
        raise HTTPException(status_code=400, detail="Delivery must be confirmed before pushing to Zoho")

    from services.zoho_service import sync_delivery_to_zoho, ZohoPushSkippedError
    try:
        await sync_delivery_to_zoho(tenant_id, distributor_id, delivery_id)
    except ZohoPushSkippedError as skip:
        # Known reason — friendly 400 so the user sees the exact blocker
        raise HTTPException(status_code=400, detail=str(skip))
    except Exception as e:
        logger.exception(f"Manual Zoho push failed for delivery {delivery_id}")
        raise HTTPException(status_code=400, detail=f"Zoho push failed: {e}")

    # Read back to pick up zoho_invoice_url written by the sync
    fresh = await db.distributor_deliveries.find_one(
        {"id": delivery_id, "tenant_id": tenant_id},
        {"_id": 0, "zoho_invoice_url": 1, "zoho_invoice_number": 1, "zoho_invoice_id": 1},
    ) or {}
    if not fresh.get('zoho_invoice_url'):
        # Surface the exact reason recorded by the sync (e.g. warehouse not
        # mapped to a Zoho Branch, no agreed price, no SKU mapping) instead of a
        # generic catch-all, so the user knows precisely what to fix.
        failure = await db.zoho_invoice_mappings.find_one(
            {"tenant_id": tenant_id, "source_type": "distributor_delivery", "source_id": delivery_id},
            {"_id": 0, "error": 1},
            sort=[("synced_at", -1)],
        )
        reason = (failure or {}).get("error")
        raise HTTPException(
            status_code=400,
            detail=reason or (
                "Zoho push completed but no invoice URL was produced. Most common causes: account has no Zoho contact ID, line items have no agreed price, or the source warehouse is not a Factory warehouse and its distributor is not marked Self-Managed."
            ),
        )
    return {
        "message": "Zoho invoice generated.",
        "zoho_invoice_url": fresh.get('zoho_invoice_url'),
        "zoho_invoice_number": fresh.get('zoho_invoice_number'),
        "zoho_invoice_id": fresh.get('zoho_invoice_id'),
    }


_INVOICE_REGEN_ROLES = {
    "ceo", "admin", "system admin", "director", "vice president",
    "head of business", "national sales head", "regional sales manager",
}


@router.get("/{distributor_id}/deliveries/{delivery_id}/invoice-preview")
async def delivery_invoice_preview(
    distributor_id: str,
    delivery_id: str,
    current_user: dict = Depends(get_current_user),
):
    """Pre-push preview of the delivery's Zoho invoice (line items, per-line %
    discount, subtotal, total discount, net taxable). GST is computed by Zoho at
    push time. Lets reps verify discounts/quantities before generating."""
    if not can_manage_distributor_data(current_user, distributor_id):
        raise HTTPException(status_code=403, detail="Not authorised to manage this distributor's deliveries")
    tenant_id = get_current_tenant_id()
    from services.zoho_service import build_delivery_invoice_preview, ZohoPushSkippedError
    try:
        return await build_delivery_invoice_preview(tenant_id, distributor_id, delivery_id)
    except ZohoPushSkippedError as skip:
        raise HTTPException(status_code=400, detail=str(skip))


@router.post("/{distributor_id}/deliveries/{delivery_id}/regenerate-invoice")
async def regenerate_delivery_invoice_endpoint(
    distributor_id: str,
    delivery_id: str,
    current_user: dict = Depends(get_current_user),
):
    """Regenerate the Zoho invoice for a delivery (e.g. after a discount fix).
    Updates the existing invoice in place when possible (same number); otherwise
    voids it and creates a fresh one. Restricted to management roles."""
    if (current_user.get("role") or "").lower() not in _INVOICE_REGEN_ROLES:
        raise HTTPException(status_code=403, detail="Only management roles can regenerate invoices.")
    tenant_id = get_current_tenant_id()
    from services.zoho_service import (
        regenerate_delivery_invoice, ZohoPushSkippedError,
        InvoiceNotRegenerableError, MissingAgreedPriceError, ZohoBranchNotMappedError,
    )
    try:
        mapping = await regenerate_delivery_invoice(tenant_id, distributor_id, delivery_id)
    except (ZohoPushSkippedError, InvoiceNotRegenerableError, MissingAgreedPriceError, ZohoBranchNotMappedError) as known:
        raise HTTPException(status_code=400, detail=str(known))
    except Exception as e:
        logger.exception(f"Invoice regeneration failed for delivery {delivery_id}")
        raise HTTPException(status_code=400, detail=f"Invoice regeneration failed: {e}")
    mode = mapping.get("regen_mode") or "updated"
    verb = {"updated": "updated in place", "recreated": "voided and recreated", "created": "created"}.get(mode, mode)
    return {
        "message": f"Invoice {verb}.",
        "regen_mode": mode,
        "zoho_invoice_url": mapping.get("zoho_invoice_url"),
        "zoho_invoice_number": mapping.get("zoho_invoice_number"),
        "zoho_invoice_id": mapping.get("zoho_invoice_id"),
    }



@router.get("/{distributor_id}/deliveries/{delivery_id}/invoice-pdf")
async def download_delivery_invoice_pdf(
    distributor_id: str,
    delivery_id: str,
    current_user: dict = Depends(get_current_user),
):
    """Download the Zoho-generated invoice for this delivery as a PDF.

    Streams the binary back with `Content-Disposition: attachment; filename="INV-00017.pdf"`
    so the saved file matches the Zoho invoice number. Returning a server-streamed
    response (rather than a direct Zoho URL) keeps the access token server-side
    and lets us force the filename — the Zoho hosted URL is cross-origin and
    ignores the HTML `download` attribute.
    """
    if not can_manage_distributor_data(current_user, distributor_id):
        raise HTTPException(status_code=403, detail="Not authorised to download this invoice")

    tenant_id = get_current_tenant_id()
    delivery = await db.distributor_deliveries.find_one(
        {"id": delivery_id, "tenant_id": tenant_id, "distributor_id": distributor_id},
        {"_id": 0, "zoho_invoice_id": 1, "zoho_invoice_number": 1, "delivery_number": 1},
    )
    if not delivery:
        raise HTTPException(status_code=404, detail="Delivery not found")
    zoho_id = delivery.get("zoho_invoice_id")
    if not zoho_id:
        raise HTTPException(
            status_code=400,
            detail="No Zoho invoice has been generated for this delivery yet. Confirm the delivery schedule first or use 'Retry Zoho push'.",
        )

    from services.zoho_service import fetch_invoice_pdf, ZohoApiError
    try:
        pdf_bytes, invoice_number = await fetch_invoice_pdf(tenant_id, zoho_id)
    except ZohoApiError as e:
        # Try to extract Zoho's `message` field from the JSON body so the
        # caller sees the real reason (e.g. "Template not found", "Invalid
        # invoice id", "Insufficient permission to download PDF").
        zoho_msg = ""
        try:
            import json as _json
            parsed = _json.loads(e.message) if e.message else None
            if isinstance(parsed, dict):
                zoho_msg = parsed.get("message") or parsed.get("error") or ""
        except Exception:
            pass
        detail = f"Zoho returned {e.status_code} while fetching the invoice PDF."
        if zoho_msg:
            detail = f"{detail} Reason: {zoho_msg}"
        elif e.message:
            detail = f"{detail} Reason: {str(e.message)[:200]}"
        raise HTTPException(status_code=400, detail=detail)
    except Exception as e:
        logger.exception(f"Failed to download invoice PDF for delivery {delivery_id}")
        raise HTTPException(status_code=400, detail=f"Failed to download invoice: {e}")

    invoice_number = invoice_number or delivery.get("zoho_invoice_number") or zoho_id
    safe_name = re.sub(r"[^A-Za-z0-9_.-]+", "_", invoice_number).strip("_") or "invoice"
    filename = f"{safe_name}.pdf"
    from fastapi.responses import Response as _Response
    return _Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Cache-Control": "no-store",
        },
    )


@router.post("/{distributor_id}/deliveries/{delivery_id}/complete")
async def complete_delivery(
    distributor_id: str,
    delivery_id: str,
    delivery_date: Optional[str] = None,
    remarks: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Mark delivery as completed - deducts from distributor stock"""
    if not can_manage_distributor_data(current_user, distributor_id):
        raise HTTPException(status_code=403, detail="Not authorised to manage this distributor's deliveries")
    
    tenant_id = get_current_tenant_id()
    
    delivery = await db.distributor_deliveries.find_one(
        {"id": delivery_id, "tenant_id": tenant_id, "distributor_id": distributor_id}
    )
    
    if not delivery:
        raise HTTPException(status_code=404, detail="Delivery not found")
    
    if delivery.get('status') not in ['draft', 'confirmed', 'scheduled', 'in_transit', 'delivery_assigned', 'delivery_scheduled', 'on_the_way']:
        raise HTTPException(status_code=400, detail="Delivery cannot be completed in current status")
    
    now = datetime.now(timezone.utc).isoformat()

    # Group by the planned DELIVERY date, not the completion date. Completing a
    # delivery records `delivered_at` (the actual completion timestamp) but must
    # NOT clobber the original `delivery_date`. We only update delivery_date when
    # the caller explicitly passes a new one (their responsibility to correct it).
    update_data = {
        "status": "complete",
        "delivered_at": now,
        "delivered_by": current_user.get('id'),
        "updated_at": now
    }
    if delivery_date:
        update_data["delivery_date"] = delivery_date
    
    if remarks:
        update_data['remarks'] = (delivery.get('remarks', '') + '\n' + remarks).strip()
    
    await db.distributor_deliveries.update_one(
        {"id": delivery_id, "tenant_id": tenant_id},
        {"$set": update_data}
    )
    
    # Deduct from the source warehouse stock — distributor_stock for a
    # regular distributor location, factory_warehouse_stock when the source
    # is a factory warehouse (is_factory=true). Without this branch the
    # factory's at-hand never decremented after a direct factory→account
    # delivery — a silent inventory-integrity bug.
    src_loc = await db.distributor_locations.find_one(
        {"id": delivery.get("distributor_location_id"), "tenant_id": tenant_id},
        {"_id": 0, "is_factory": 1, "track_batches": 1, "location_name": 1},
    )
    src_is_factory = bool(src_loc and src_loc.get("is_factory"))
    src_tracks_batches = bool(src_loc and src_loc.get("track_batches"))

    items = await db.distributor_delivery_items.find(
        {"delivery_id": delivery_id, "tenant_id": tenant_id},
        {"_id": 0}
    ).to_list(500)

    for item in items:
        batch_id = item.get('batch_id')
        qty = item.get('quantity', 0)

        if src_is_factory:
            stock_key = {
                "tenant_id": tenant_id,
                "warehouse_location_id": delivery.get('distributor_location_id'),
                "sku_id": item.get('sku_id'),
            }
            if src_tracks_batches and batch_id:
                stock_key["batch_id"] = batch_id
            await db.factory_warehouse_stock.update_one(
                stock_key,
                {"$inc": {"quantity": -qty}, "$set": {"updated_at": now}},
            )
        else:
            stock_key = {
                "tenant_id": tenant_id,
                "distributor_id": distributor_id,
                "distributor_location_id": delivery.get('distributor_location_id'),
                "sku_id": item.get('sku_id'),
            }
            # Tracked batches deduct from the specific batch row; legacy aggregate
            # rows (no batch_id stored) match `{"$in": [None]}`.
            stock_key["batch_id"] = batch_id if batch_id else {"$in": [None]}
            await db.distributor_stock.update_one(
                stock_key,
                {
                    "$inc": {"quantity": -qty},
                    "$set": {"updated_at": now}
                }
            )

    # If the account is billed by a third-party distributor, generate an
    # External Billing Entry instead of a Zoho invoice. Idempotent — running
    # complete_delivery again won't create a duplicate.
    try:
        account_doc = await db.accounts.find_one(
            {"id": delivery.get("account_id"), "tenant_id": tenant_id},
            {"_id": 0},
        )
        if account_doc and (account_doc.get("billed_by") or "company") == "distributor":
            from services.external_billing import generate_external_billing_entry
            distributor_doc = await db.distributors.find_one(
                {"id": distributor_id, "tenant_id": tenant_id},
                {"_id": 0, "id": 1, "distributor_name": 1},
            )
            # Refresh the delivery after the status flip so EBE captures the
            # post-completion `delivered_at`.
            fresh_delivery = await db.distributor_deliveries.find_one(
                {"id": delivery_id, "tenant_id": tenant_id}, {"_id": 0},
            )
            await generate_external_billing_entry(
                tenant_id=tenant_id,
                delivery=fresh_delivery or delivery,
                account=account_doc,
                distributor=distributor_doc,
                items=items,
            )
    except Exception as e:
        # Failing the EBE side must not roll back the stock deduction — the
        # delivery is already complete. Surface in logs for admin backfill.
        logger.warning(
            f"complete_delivery succeeded but EBE generation failed for "
            f"delivery {delivery_id}: {e}"
        )
    
    logger.info(f"Delivery {delivery['delivery_number']} completed by {current_user['email']}")
    
    return {"message": f"Delivery {delivery['delivery_number']} completed", "status": "complete"}


@router.delete("/{distributor_id}/deliveries/{delivery_id}")
async def delete_delivery(
    distributor_id: str,
    delivery_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Delete a delivery - CEO/Admin can delete any status, others only draft"""
    if not can_manage_distributor_data(current_user, distributor_id):
        raise HTTPException(status_code=403, detail="Not authorised to manage this distributor's deliveries")
    
    tenant_id = get_current_tenant_id()
    user_role = current_user.get('role', '').lower()
    is_ceo_or_admin = user_role in ['ceo', 'admin']
    
    delivery = await db.distributor_deliveries.find_one(
        {"id": delivery_id, "tenant_id": tenant_id, "distributor_id": distributor_id}
    )
    
    if not delivery:
        raise HTTPException(status_code=404, detail="Delivery not found")
    
    # CEO/Admin can delete any delivery regardless of status
    # Others can only delete draft deliveries
    if not is_ceo_or_admin and delivery.get('status') != 'draft':
        raise HTTPException(status_code=400, detail="Only draft deliveries can be deleted. Contact CEO/Admin to delete non-draft deliveries.")
    
    # If delivery is settled, prevent deletion even for CEO/Admin
    if delivery.get('settlement_id'):
        raise HTTPException(status_code=400, detail="Cannot delete delivery that is part of a settlement")
    
    # Delete items first
    del_items = await db.distributor_delivery_items.find({"delivery_id": delivery_id, "tenant_id": tenant_id}).to_list(1000)
    await db.distributor_delivery_items.delete_many({"delivery_id": delivery_id, "tenant_id": tenant_id})
    
    # Record deletion audit (who/when/what) BEFORE removing the delivery
    await _record_deletion_audit(
        tenant_id, 'delivery', delivery, distributor_id, current_user,
        entity_number=delivery.get('delivery_number'),
        extra={'item_count': len(del_items), 'items_snapshot': [{k: v for k, v in i.items() if k != '_id'} for i in del_items]},
    )
    
    # Delete delivery
    await db.distributor_deliveries.delete_one({"id": delivery_id, "tenant_id": tenant_id})
    
    logger.info(f"Delivery {delivery['delivery_number']} (status: {delivery.get('status')}) deleted by {current_user['email']} (role: {user_role})")
    
    return {"message": f"Delivery {delivery['delivery_number']} deleted"}


# IMPORTANT: keep this static route BEFORE '/{distributor_id}/deletion-audit'
# below — otherwise FastAPI captures 'deletion-audit' as distributor_id.
@router.get("/deletion-audit/all")
async def get_deletion_audit_all(
    entity_type: Optional[str] = None,
    distributor_id: Optional[str] = None,
    limit: int = 500,
    current_user: dict = Depends(get_current_user)
):
    """Global deletion audit history (CEO/Admin only). Who deleted what & when."""
    user_role = (current_user.get('role') or '').lower()
    if not any(r in user_role for r in ['ceo', 'admin', 'system admin']):
        raise HTTPException(status_code=403, detail="Only CEO/Admin can view the deletion audit")
    tenant_id = get_current_tenant_id()
    q = {'tenant_id': tenant_id}
    if entity_type:
        q['entity_type'] = entity_type
    if distributor_id:
        q['distributor_id'] = distributor_id
    records = await db.deletion_audit.find(q, {'_id': 0}).sort('deleted_at', -1).limit(min(max(1, limit), 2000)).to_list(2000)
    return {'total': len(records), 'records': records}


@router.get("/{distributor_id}/deletion-audit")
async def get_deletion_audit(
    distributor_id: str,
    entity_type: Optional[str] = 'delivery',
    current_user: dict = Depends(get_current_user)
):
    """Deletion history for a distributor's deliveries/shipments (who/when/what)."""
    if not can_manage_distributor_data(current_user, distributor_id):
        raise HTTPException(status_code=403, detail="Not authorised to view this distributor's audit")
    tenant_id = get_current_tenant_id()
    q = {'tenant_id': tenant_id, 'distributor_id': distributor_id}
    if entity_type and entity_type != 'all':
        q['entity_type'] = entity_type
    records = await db.deletion_audit.find(q, {'_id': 0}).sort('deleted_at', -1).limit(1000).to_list(1000)
    return {'total': len(records), 'records': records}


@router.post("/{distributor_id}/deletion-audit/{audit_id}/restore")
async def restore_deleted_delivery(
    distributor_id: str,
    audit_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Restore a deleted delivery from its audit snapshot (CEO/Admin only).

    Re-inserts the delivery + its line items exactly as captured at deletion.
    Does NOT re-run stock deduction or Zoho pushes — deletion never reversed
    those, so the restored record is consistent with the current DB state.
    """
    user_role = (current_user.get('role') or '').lower()
    if not any(r in user_role for r in ['ceo', 'admin', 'system admin']):
        raise HTTPException(status_code=403, detail="Only CEO/Admin can restore deleted records")
    tenant_id = get_current_tenant_id()

    audit = await db.deletion_audit.find_one({'id': audit_id, 'tenant_id': tenant_id})
    if not audit:
        raise HTTPException(status_code=404, detail="Audit record not found")
    if audit.get('entity_type') != 'delivery':
        raise HTTPException(status_code=400, detail="Only deleted deliveries can be restored")
    if audit.get('restored_at'):
        raise HTTPException(status_code=400, detail="This delivery has already been restored")

    snapshot = audit.get('snapshot') or {}
    del_id = snapshot.get('id')
    if not del_id:
        raise HTTPException(status_code=400, detail="Snapshot is missing the delivery id")

    existing = await db.distributor_deliveries.find_one({'id': del_id, 'tenant_id': tenant_id})
    if existing:
        raise HTTPException(status_code=400, detail="A delivery with this id already exists — nothing to restore")

    # Re-insert delivery + items from the snapshot
    doc = {k: v for k, v in snapshot.items() if k != '_id'}
    doc['tenant_id'] = tenant_id
    doc['restored_at'] = datetime.now(timezone.utc).isoformat()
    doc['restored_by'] = current_user.get('email')
    await db.distributor_deliveries.insert_one(doc)

    items = audit.get('items_snapshot') or []
    if items:
        clean_items = []
        for it in items:
            ci = {k: v for k, v in it.items() if k != '_id'}
            ci['tenant_id'] = tenant_id
            clean_items.append(ci)
        await db.distributor_delivery_items.insert_many(clean_items)

    # Mark the audit record as restored (keep the row for history)
    await db.deletion_audit.update_one(
        {'id': audit_id, 'tenant_id': tenant_id},
        {'$set': {'restored_at': doc['restored_at'], 'restored_by': current_user.get('email')}}
    )

    logger.info(f"Delivery {snapshot.get('delivery_number')} restored by {current_user['email']} from audit {audit_id}")
    return {'success': True, 'message': f"Delivery {snapshot.get('delivery_number') or del_id} restored", 'delivery_id': del_id}




@router.post("/{distributor_id}/deliveries/{delivery_id}/cancel")
async def cancel_delivery(
    distributor_id: str,
    delivery_id: str,
    reason: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Cancel a delivery"""
    if not can_manage_distributor_data(current_user, distributor_id):
        raise HTTPException(status_code=403, detail="Not authorised to manage this distributor's deliveries")
    
    tenant_id = get_current_tenant_id()
    
    delivery = await db.distributor_deliveries.find_one(
        {"id": delivery_id, "tenant_id": tenant_id, "distributor_id": distributor_id}
    )
    
    if not delivery:
        raise HTTPException(status_code=404, detail="Delivery not found")
    
    if delivery.get('status') in ['delivered', 'cancelled']:
        raise HTTPException(status_code=400, detail="Delivery cannot be cancelled")
    
    now = datetime.now(timezone.utc).isoformat()
    
    remarks = delivery.get('remarks', '') or ''
    if reason:
        remarks = f"{remarks}\nCancelled: {reason}".strip()
    
    # Revert any applied credit notes
    reverted_notes = []
    if delivery.get('applied_credit_notes'):
        from routes.credit_notes import revert_credit_note_application
        reverted_notes = await revert_credit_note_application(
            tenant_id=tenant_id,
            delivery_id=delivery_id,
            delivery_number=delivery.get('delivery_number')
        )
        if reverted_notes:
            logger.info(f"Reverted {len(reverted_notes)} credit note(s) for cancelled delivery {delivery['delivery_number']}")
    
    await db.distributor_deliveries.update_one(
        {"id": delivery_id, "tenant_id": tenant_id},
        {"$set": {
            "status": "cancelled",
            "remarks": remarks,
            "updated_at": now
        }}
    )
    
    logger.info(f"Delivery {delivery['delivery_number']} cancelled by {current_user['email']}")
    
    return {
        "message": f"Delivery {delivery['delivery_number']} cancelled", 
        "status": "cancelled",
        "reverted_credit_notes": reverted_notes
    }


# Statuses for which a delivery is "done" and can no longer be reversed.
_NON_REVERSIBLE_DELIVERY_STATUSES = {"complete", "completed", "delivered", "cancelled", "reversed"}


async def _reverse_delivery_invoice_mirror(tenant_id: str, delivery: dict) -> None:
    """Undo the local accounting side-effects of a delivery's Zoho invoice mirror:
    remove the mirror `invoices` doc and decrement the account's running
    `outstanding_balance` by the amount that was added when the invoice synced.
    Idempotent — only decrements when the mirror was actually counted."""
    mirror = await db.invoices.find_one(
        {"tenant_id": tenant_id, "source_type": "distributor_delivery", "source_id": delivery.get("id")},
        {"_id": 0},
    )
    if not mirror:
        return
    now = datetime.now(timezone.utc).isoformat()
    if mirror.get("outstanding_counted"):
        gross = float(mirror.get("gross_invoice_value") or mirror.get("net_invoice_value") or 0)
        if gross:
            await db.accounts.update_one(
                {"tenant_id": tenant_id,
                 "$or": [{"id": delivery.get("account_id")}, {"account_id": delivery.get("account_id")}]},
                {"$inc": {"outstanding_balance": round(-gross, 2)}, "$set": {"updated_at": now}},
            )
    await db.invoices.delete_one(
        {"tenant_id": tenant_id, "source_type": "distributor_delivery", "source_id": delivery.get("id")},
    )


async def _readd_completed_delivery_stock(tenant_id: str, distributor_id: str, delivery: dict) -> None:
    """Inverse of `complete_delivery`'s deduction: add the delivered quantities
    back to the source warehouse (factory_warehouse_stock for a factory source,
    distributor_stock otherwise), using the same batch/aggregate row convention."""
    now = datetime.now(timezone.utc).isoformat()
    src_loc = await db.distributor_locations.find_one(
        {"id": delivery.get("distributor_location_id"), "tenant_id": tenant_id},
        {"_id": 0, "is_factory": 1, "track_batches": 1},
    )
    src_is_factory = bool(src_loc and src_loc.get("is_factory"))
    src_tracks_batches = bool(src_loc and src_loc.get("track_batches"))
    items = await db.distributor_delivery_items.find(
        {"delivery_id": delivery.get("id"), "tenant_id": tenant_id}, {"_id": 0},
    ).to_list(500)
    for item in items:
        batch_id = item.get("batch_id")
        qty = item.get("quantity", 0) or 0
        if not qty:
            continue
        if src_is_factory:
            stock_key = {
                "tenant_id": tenant_id,
                "warehouse_location_id": delivery.get("distributor_location_id"),
                "sku_id": item.get("sku_id"),
            }
            if src_tracks_batches and batch_id:
                stock_key["batch_id"] = batch_id
            await db.factory_warehouse_stock.update_one(
                stock_key, {"$inc": {"quantity": qty}, "$set": {"updated_at": now}}, upsert=True,
            )
        else:
            stock_key = {
                "tenant_id": tenant_id,
                "distributor_id": distributor_id,
                "distributor_location_id": delivery.get("distributor_location_id"),
                "sku_id": item.get("sku_id"),
            }
            stock_key["batch_id"] = batch_id if batch_id else {"$in": [None]}
            await db.distributor_stock.update_one(
                stock_key, {"$inc": {"quantity": qty}, "$set": {"updated_at": now}},
            )


@router.post("/{distributor_id}/deliveries/{delivery_id}/reverse")
async def reverse_delivery(
    distributor_id: str,
    delivery_id: str,
    reason: Optional[str] = None,
    acknowledge: bool = False,
    current_user: dict = Depends(get_current_user),
):
    """Reverse a regular distributor→customer delivery at ANY stage:
      • draft / open (reserved-only): release the reservation; void invoice +
        undo mirror/outstanding/credit-notes when one was already generated.
      • complete / delivered: ADD the deducted stock back to the source
        warehouse, void the Zoho invoice OR delete the External Billing Entry,
        undo the local mirror + account outstanding, revert credit notes.
    The delivery is kept and marked 'reversed' for audit.

    Non-draft reversals require `acknowledge=true` (server-side double-confirm).
    Settlement-linked deliveries cannot be reversed (detach from the settlement
    first)."""
    if not can_manage_distributor_data(current_user, distributor_id):
        raise HTTPException(status_code=403, detail="Not authorised to manage this distributor's deliveries")

    tenant_id = get_current_tenant_id()
    delivery = await db.distributor_deliveries.find_one(
        {"id": delivery_id, "tenant_id": tenant_id, "distributor_id": distributor_id}, {"_id": 0},
    )
    if not delivery:
        raise HTTPException(status_code=404, detail="Delivery not found")

    status = (delivery.get("status") or "").lower()
    if status in {"cancelled", "reversed"}:
        raise HTTPException(status_code=400, detail="This delivery is already cancelled or reversed.")
    if delivery.get("settlement_id"):
        raise HTTPException(
            status_code=400,
            detail="Cannot reverse a delivery that is part of a settlement. Remove it from the settlement first.",
        )

    # Server-side double-confirm: anything past draft must be acknowledged.
    is_draft = status == "draft"
    if not is_draft and not acknowledge:
        raise HTTPException(
            status_code=400,
            detail="This delivery is past draft. Confirm the reversal explicitly (acknowledge=true).",
        )

    was_completed = status in {"complete", "completed", "delivered"}
    now = datetime.now(timezone.utc).isoformat()

    # 1) Add stock back when it was physically deducted on completion.
    if was_completed:
        try:
            await _readd_completed_delivery_stock(tenant_id, distributor_id, delivery)
        except Exception:
            logger.exception(f"Failed to re-add stock during reverse for delivery {delivery_id}")

    # 2) Void the Zoho invoice if one exists (best-effort; retry-pending flag).
    from services.zoho_service import void_invoice
    void_pending = False
    void_error = None
    zoho_invoice_id = delivery.get("zoho_invoice_id")
    if zoho_invoice_id:
        try:
            await void_invoice(tenant_id, zoho_invoice_id)
        except Exception as exc:
            void_pending = True
            void_error = str(exc)[:500]
            logger.exception(f"Zoho invoice void failed during reverse for delivery {delivery.get('delivery_number')}")

    # 3) Undo the local invoice mirror / External Billing Entry + outstanding.
    #    `_reverse_delivery_invoice_mirror` deletes whichever invoice doc is
    #    linked to this delivery (Zoho mirror OR EBE) and only decrements the
    #    account outstanding for the counted Zoho mirror (EBE carries no
    #    outstanding by design).
    try:
        await _reverse_delivery_invoice_mirror(tenant_id, delivery)
    except Exception:
        logger.exception(f"Failed to reverse invoice mirror for delivery {delivery_id}")

    # 4) Revert applied credit notes.
    reverted_notes = []
    if delivery.get("applied_credit_notes"):
        from routes.credit_notes import revert_credit_note_application
        reverted_notes = await revert_credit_note_application(
            tenant_id=tenant_id, delivery_id=delivery_id,
            delivery_number=delivery.get("delivery_number"),
        )

    # 5) Mark reversed (drops it from the reserved set → any reservation released).
    remarks = delivery.get("remarks", "") or ""
    if reason:
        remarks = f"{remarks}\nReversed: {reason}".strip()
    set_fields = {
        "status": "reversed",
        "remarks": remarks,
        "reversed_at": now,
        "reversed_by": current_user.get("id"),
        "reversed_by_name": current_user.get("name"),
        "reversed_from_status": status,
        "stock_readded": was_completed,
        "updated_at": now,
        "applied_credit_notes": [],
        "zoho_void_pending": void_pending,
        "zoho_void_error": void_error,
    }
    if zoho_invoice_id and not void_pending:
        set_fields["zoho_invoice_voided"] = True
    await db.distributor_deliveries.update_one(
        {"id": delivery_id, "tenant_id": tenant_id}, {"$set": set_fields})

    stock_msg = "stock added back" if was_completed else "stock reservation released"
    msg = f"Delivery {delivery.get('delivery_number')} reversed — {stock_msg}"
    if zoho_invoice_id:
        msg += " (Zoho invoice void pending — will retry)" if void_pending else " and Zoho invoice voided"
    updated = await db.distributor_deliveries.find_one({"id": delivery_id, "tenant_id": tenant_id}, {"_id": 0})
    return {
        "ok": True, "message": msg, "delivery": updated,
        "zoho_void_pending": void_pending, "zoho_void_error": void_error,
        "reverted_credit_notes": reverted_notes,
    }


@router.post("/{distributor_id}/deliveries/{delivery_id}/reverse-zoho-cleanup")
async def retry_delivery_void(
    distributor_id: str,
    delivery_id: str,
    current_user: dict = Depends(get_current_user),
):
    """Re-attempt voiding the Zoho invoice for a reversed delivery whose void is
    still pending (e.g. a transient Zoho error during the original reverse)."""
    if not can_manage_distributor_data(current_user, distributor_id):
        raise HTTPException(status_code=403, detail="Not authorised to manage this distributor's deliveries")
    tenant_id = get_current_tenant_id()
    delivery = await db.distributor_deliveries.find_one(
        {"id": delivery_id, "tenant_id": tenant_id, "distributor_id": distributor_id}, {"_id": 0},
    )
    if not delivery:
        raise HTTPException(status_code=404, detail="Delivery not found")
    if delivery.get("status") != "reversed" or not delivery.get("zoho_void_pending"):
        raise HTTPException(status_code=400, detail="No pending Zoho void for this delivery.")
    now = datetime.now(timezone.utc).isoformat()
    from services.zoho_service import void_invoice
    try:
        await void_invoice(tenant_id, delivery.get("zoho_invoice_id"))
    except Exception as exc:
        err = str(exc)[:500]
        await db.distributor_deliveries.update_one(
            {"id": delivery_id, "tenant_id": tenant_id},
            {"$set": {"zoho_void_error": err, "updated_at": now}})
        raise HTTPException(status_code=400, detail=f"Zoho void failed: {err}")
    await db.distributor_deliveries.update_one(
        {"id": delivery_id, "tenant_id": tenant_id},
        {"$set": {"zoho_void_pending": False, "zoho_void_error": None,
                  "zoho_invoice_voided": True, "updated_at": now}})
    return {"ok": True, "message": "Zoho invoice voided."}



@router.get("/{distributor_id}/deliveries/{delivery_id}/customer-invoice")
async def generate_customer_invoice(
    distributor_id: str,
    delivery_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Generate a customer invoice PDF for a delivery with GST"""
    tenant_id = get_current_tenant_id()
    
    # Check access
    user_is_admin = is_distributor_admin(current_user)
    user_is_distributor = is_distributor_user(current_user) and current_user.get('distributor_id') == distributor_id
    
    if not user_is_admin and not user_is_distributor:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Get delivery
    delivery = await db.distributor_deliveries.find_one(
        {"id": delivery_id, "tenant_id": tenant_id, "distributor_id": distributor_id}
    )
    
    if not delivery:
        raise HTTPException(status_code=404, detail="Delivery not found")

    # Attach line items — they live in a separate collection. Without this the
    # PDF "Item Details" table is silently empty.
    delivery["items"] = await db.distributor_delivery_items.find(
        {"delivery_id": delivery_id, "tenant_id": tenant_id},
        {"_id": 0},
    ).to_list(500)
    
    # Get distributor
    distributor = await db.distributors.find_one({"id": distributor_id, "tenant_id": tenant_id})
    if not distributor:
        raise HTTPException(status_code=404, detail="Distributor not found")
    
    # Get account (customer)
    account = await db.accounts.find_one({"id": delivery.get('account_id'), "tenant_id": tenant_id})
    if not account:
        # Try getting minimal data from delivery
        account = {
            "account_name": delivery.get('account_name', 'Customer'),
            "city": delivery.get('account_city', ''),
            "state": "",
            "address": "",
            "gst_number": "",
            "contact_name": "",
            "contact_number": ""
        }
    
    # Get tenant for company profile and GST rate
    tenant = await db.tenants.find_one({"tenant_id": tenant_id}, {"_id": 0})
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")
    
    company_profile = tenant.get('company_profile', {})
    branding = tenant.get('branding', {})
    settings = tenant.get('settings', {})
    
    # Get GST rate from tenant settings (default 18%)
    gst_percent = settings.get('default_distributor_gst_percent', 18.0)

    # If the account is billed by a distributor, render this as an External
    # Billing Entry (no GST, banner indicating it's not a tax invoice).
    is_ebe = (account.get("billed_by") or "company") == "distributor"
    ebe_number = None
    if is_ebe:
        # Reuse / lazily create the EBE so the printed number matches what we
        # stored in the `invoices` table.
        from services.external_billing import generate_external_billing_entry, get_existing_ebe
        existing = await get_existing_ebe(tenant_id, delivery_id)
        if existing:
            ebe_number = existing.get("invoice_number")
        else:
            created = await generate_external_billing_entry(
                tenant_id=tenant_id,
                delivery=delivery,
                account=account,
                distributor=distributor,
                items=delivery["items"],
            )
            ebe_number = (created or {}).get("invoice_number")

    # Generate PDF
    try:
        pdf_bytes = generate_customer_invoice_pdf(
            delivery_data=delivery,
            company_profile=company_profile,
            account_data=account,
            distributor_data=distributor,
            gst_percent=gst_percent,
            branding=branding,
            is_external_billing=is_ebe,
            external_billing_number=ebe_number,
        )

        # Generate filename
        if is_ebe:
            invoice_number = ebe_number or f"EXT-{delivery.get('delivery_number', 'N-A')}"
            filename = f"external_billing_entry_{invoice_number}.pdf"
        else:
            invoice_number = f"INV-{delivery.get('delivery_number', 'N-A').replace('DEL-', '')}"
            filename = f"customer_invoice_{invoice_number}.pdf"

        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            }
        )
    except Exception as e:
        logger.error(f"Failed to generate customer invoice PDF: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to generate invoice: {str(e)}")


@router.post("/{distributor_id}/external-billing/backfill")
async def backfill_external_billing_entries(
    distributor_id: str,
    current_user: dict = Depends(get_current_user),
):
    """One-shot generator for historical completed deliveries whose accounts
    are billed by a third-party distributor but never had an External Billing
    Entry created (e.g., the delivery completed before this feature shipped).

    Idempotent — only creates EBEs where none exist. Returns a count summary.
    """
    if not is_distributor_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin access required")

    tenant_id = get_current_tenant_id()
    from services.external_billing import generate_external_billing_entry, get_existing_ebe

    distributor = await db.distributors.find_one(
        {"id": distributor_id, "tenant_id": tenant_id},
        {"_id": 0, "id": 1, "distributor_name": 1},
    )
    if not distributor:
        raise HTTPException(status_code=404, detail="Distributor not found")

    completed = await db.distributor_deliveries.find(
        {
            "tenant_id": tenant_id,
            "distributor_id": distributor_id,
            "status": {"$in": ["delivered", "completed", "complete"]},
        },
        {"_id": 0},
    ).to_list(10000)

    created, skipped, errors = 0, 0, 0
    for delivery in completed:
        try:
            account = await db.accounts.find_one(
                {"id": delivery.get("account_id"), "tenant_id": tenant_id},
                {"_id": 0},
            )
            if not account or (account.get("billed_by") or "company") != "distributor":
                skipped += 1
                continue
            if await get_existing_ebe(tenant_id, delivery["id"]):
                skipped += 1
                continue
            await generate_external_billing_entry(
                tenant_id=tenant_id,
                delivery=delivery,
                account=account,
                distributor=distributor,
            )
            created += 1
        except Exception as e:
            errors += 1
            logger.warning(f"EBE backfill failed for delivery {delivery.get('id')}: {e}")

    return {
        "distributor_id": distributor_id,
        "examined": len(completed),
        "created": created,
        "skipped": skipped,
        "errors": errors,
    }



# ============ Distributor Settlement CRUD ============

SETTLEMENT_STATUSES = {
    "draft": "Draft - Being prepared",
    "pending_approval": "Pending Approval",
    "approved": "Approved - Ready for payment",
    "rejected": "Rejected",
    "paid": "Paid",
    "cancelled": "Cancelled"
}


async def generate_settlement_number(tenant_id: str) -> str:
    """Generate unique settlement number like STL-2026-0001"""
    year = datetime.now().year
    count = await db.distributor_settlements.count_documents({
        "tenant_id": tenant_id,
        "settlement_number": {"$regex": f"^STL-{year}-"}
    })
    return f"STL-{year}-{count + 1:04d}"


@router.get("/settlements/all")
async def list_all_settlements(
    distributor_id: Optional[str] = None,
    status: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    current_user: dict = Depends(get_current_user)
):
    """List all settlements with filters"""
    tenant_id = get_current_tenant_id()
    
    query = {"tenant_id": tenant_id}
    
    if distributor_id:
        query["distributor_id"] = distributor_id
    if status and status != 'all':
        query["status"] = status
    if from_date:
        query["period_start"] = {"$gte": from_date}
    if to_date:
        if "period_end" in query:
            query["period_end"]["$lte"] = to_date
        else:
            query["period_end"] = {"$lte": to_date}
    
    total = await db.distributor_settlements.count_documents(query)
    
    settlements = await db.distributor_settlements.find(
        query,
        {"_id": 0}
    ).sort("created_at", -1).skip((page - 1) * page_size).limit(page_size).to_list(page_size)
    
    return {
        "settlements": settlements,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size
    }


@router.get("/settlements/summary")
async def get_settlements_summary(
    distributor_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get settlements summary stats"""
    tenant_id = get_current_tenant_id()
    
    query = {"tenant_id": tenant_id}
    if distributor_id:
        query["distributor_id"] = distributor_id
    
    total = await db.distributor_settlements.count_documents(query)
    
    # Count by status
    draft = await db.distributor_settlements.count_documents({**query, "status": "draft"})
    pending = await db.distributor_settlements.count_documents({**query, "status": "pending_approval"})
    approved = await db.distributor_settlements.count_documents({**query, "status": "approved"})
    paid = await db.distributor_settlements.count_documents({**query, "status": "paid"})
    
    # Calculate totals for paid settlements
    pipeline = [
        {"$match": {**query, "status": "paid"}},
        {"$group": {
            "_id": None,
            "total_paid": {"$sum": "$final_payout"}
        }}
    ]
    
    totals = await db.distributor_settlements.aggregate(pipeline).to_list(1)
    total_paid = totals[0]["total_paid"] if totals else 0
    
    # Pending payout (approved but not paid)
    pipeline_pending = [
        {"$match": {**query, "status": "approved"}},
        {"$group": {
            "_id": None,
            "pending_payout": {"$sum": "$final_payout"}
        }}
    ]
    
    pending_totals = await db.distributor_settlements.aggregate(pipeline_pending).to_list(1)
    pending_payout = pending_totals[0]["pending_payout"] if pending_totals else 0
    
    return {
        "total": total,
        "by_status": {
            "draft": draft,
            "pending_approval": pending,
            "approved": approved,
            "paid": paid
        },
        "total_paid": round(total_paid, 2),
        "pending_payout": round(pending_payout, 2)
    }


@router.get("/{distributor_id}/settlements")
async def list_distributor_settlements(
    distributor_id: str,
    status: Optional[str] = None,
    month: Optional[int] = None,
    year: Optional[int] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    current_user: dict = Depends(get_current_user)
):
    """List settlements for a specific distributor with month/year filters"""
    tenant_id = get_current_tenant_id()
    
    query = {"tenant_id": tenant_id, "distributor_id": distributor_id}
    
    if status and status != 'all':
        query["status"] = status
    
    if month:
        query["settlement_month"] = int(month)
    
    if year:
        query["settlement_year"] = int(year)
    
    total = await db.distributor_settlements.count_documents(query)
    
    settlements = await db.distributor_settlements.find(
        query,
        {"_id": 0}
    ).sort([("settlement_year", -1), ("settlement_month", -1), ("created_at", -1)]).skip((page - 1) * page_size).limit(page_size).to_list(page_size)

    # Enrich each settlement with stockout_totals so the parent page uses the
    # SAME math as the Settlement Detail popup (single source of truth).
    await _enrich_settlements_with_stockout_totals(tenant_id, settlements)

    return {
        "settlements": settlements,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size
    }


def _compute_delivery_stockout_view(delivery: dict, delivery_items: list) -> dict:
    """Compute the EXACT per-delivery numbers shown on the Stock Out / Deliveries
    tab so any consumer (settlement detail, reconciliation, etc.) can render the
    same single-source-of-truth values without re-deriving them.

    Mirrors the math in /app/frontend/src/components/distributor/DeliveriesTab.jsx
    (rows ~1163-1193) so settlements never drift from delivery vocabulary.
    """
    customer_order_value = 0.0     # sum(qty × customer_price × (1 - disc%))
    distributor_margin = 0.0       # sum(qty × customer_price × commission%)
    actual_billable = 0.0          # sum(qty × customer_price × (1 - commission%))
    billed_at_transfer = 0.0       # sum(qty × transfer_price)  ← what dist was already billed

    for it in delivery_items or []:
        qty = it.get('quantity') or 0
        customer_price = it.get('customer_selling_price') or it.get('unit_price') or 0
        disc_pct = it.get('discount_percent') or 0
        commission_pct = (
            it.get('distributor_commission_percent')
            or it.get('margin_percent')
            or 0
        )
        transfer_price = it.get('transfer_price') or 0
        gross = qty * customer_price
        customer_order_value += gross * (1 - disc_pct / 100)
        distributor_margin += gross * (commission_pct / 100)
        if customer_price > 0:
            new_transfer = customer_price * (1 - commission_pct / 100)
            actual_billable += qty * new_transfer
        billed_at_transfer += qty * transfer_price

    credit_applied = float(delivery.get('total_credit_applied') or 0)
    net_billable = actual_billable - credit_applied

    return {
        "customer_order_value": round(customer_order_value, 2),
        "distributor_margin": round(distributor_margin, 2),
        "actual_billable": round(actual_billable, 2),
        "credit_applied": round(credit_applied, 2),
        "net_billable": round(net_billable, 2),
        "billed_at_transfer": round(billed_at_transfer, 2),
        "applied_credit_notes": delivery.get('applied_credit_notes') or [],
    }


async def _enrich_settlements_with_stockout_totals(tenant_id: str, settlements: list) -> list:
    """Enrich a list of settlement docs in-place with `items` (each enriched with
    the per-delivery stockout view) and `stockout_totals`. Single source of truth
    shared between list and detail endpoints so the parent Settlements page and
    the popup never drift.
    """
    if not settlements:
        return settlements

    settlement_ids = [s['id'] for s in settlements if s.get('id')]
    items_by_settlement: dict[str, list] = {}
    all_items = await db.distributor_settlement_items.find(
        {"settlement_id": {"$in": settlement_ids}, "tenant_id": tenant_id},
        {"_id": 0}
    ).sort("delivery_date", 1).to_list(5000)
    for it in all_items:
        items_by_settlement.setdefault(it['settlement_id'], []).append(it)

    delivery_ids = list({it['delivery_id'] for it in all_items if it.get('delivery_id')})
    delivery_map: dict[str, dict] = {}
    delivery_items_map: dict[str, list] = {}
    if delivery_ids:
        deliveries = await db.distributor_deliveries.find(
            {"tenant_id": tenant_id, "id": {"$in": delivery_ids}},
            {"_id": 0}
        ).to_list(len(delivery_ids))
        for d in deliveries:
            delivery_map[d['id']] = d
        di_rows = await db.distributor_delivery_items.find(
            {"tenant_id": tenant_id, "delivery_id": {"$in": delivery_ids}},
            {"_id": 0}
        ).to_list(20000)
        for di in di_rows:
            delivery_items_map.setdefault(di['delivery_id'], []).append(di)

    for s in settlements:
        items = items_by_settlement.get(s['id'], [])
        for it in items:
            d = delivery_map.get(it.get('delivery_id'))
            if not d:
                continue
            view = _compute_delivery_stockout_view(d, delivery_items_map.get(d['id'], []))
            it.update(view)
            it['delivery_status'] = d.get('status')
        s['items'] = items

        # Direct credit notes paid to customer outside any delivery (standalone
        # issuances) for this account in the settlement period. Distributor paid
        # the customer out-of-pocket on the factory's behalf, so this reduces
        # what the distributor owes the factory in the settlement.
        # Only `issued` (actually handed over) issuances count.
        #
        # IMPORTANT (no-double-count): a single Credit Note can be drained via
        # TWO independent channels — (a) applied to a delivery, contributing to
        # `delivery.total_credit_applied` (which we already sum into
        # `credit_applied`) and (b) paid as cash via an issuance row. To avoid
        # counting the same money twice, per CN we cap the issuance contribution
        # at `max(0, original_amount − delivery_applied)`. Truly-standalone CNs
        # (no return_id) are unaffected by this cap, since they have no
        # delivery applications by construction.
        direct_credit_amount = 0.0
        direct_issuances: list[dict] = []
        if s.get('account_id'):
            year = s.get('settlement_year')
            month = s.get('settlement_month')
            month_match: dict = {
                "tenant_id": tenant_id,
                "distributor_id": s['distributor_id'],
                "account_id": s['account_id'],
                "status": "issued",
            }
            if year and month:
                # issued_at is stored as YYYY-MM-DD string
                start = f"{year:04d}-{month:02d}-01"
                # exclusive upper bound (next month)
                if month == 12:
                    end = f"{year + 1:04d}-01-01"
                else:
                    end = f"{year:04d}-{month + 1:02d}-01"
                month_match["issued_at"] = {"$gte": start, "$lt": end}
            iss_rows = await db.credit_note_issuances.find(
                month_match,
                {"_id": 0, "id": 1, "amount": 1, "credit_note_id": 1,
                 "credit_note_number": 1, "issuance_method": 1, "issued_at": 1,
                 "reason": 1, "return_id": 1}
            ).to_list(500)

            # Group by CN so we can apply the per-CN cap once.
            cn_ids = list({r.get('credit_note_id') for r in iss_rows if r.get('credit_note_id')})
            cn_meta: dict = {}
            if cn_ids:
                async for cn in db.credit_notes.find(
                    {"id": {"$in": cn_ids}, "tenant_id": tenant_id},
                    {"_id": 0, "id": 1, "original_amount": 1, "amount": 1, "applications": 1}
                ):
                    delivery_applied = sum(
                        (app or {}).get('amount_applied', 0)
                        for app in (cn.get('applications') or [])
                    )
                    original = cn.get('original_amount') or cn.get('amount') or 0
                    cn_meta[cn['id']] = {
                        "original": float(original),
                        "delivery_applied": float(delivery_applied),
                        "remaining_capacity": max(0.0, float(original) - float(delivery_applied)),
                    }

            # Sum issuance amount per CN, then cap at remaining_capacity.
            issuance_sum_by_cn: dict = {}
            for r in iss_rows:
                cid = r.get('credit_note_id')
                if not cid:
                    continue
                issuance_sum_by_cn[cid] = issuance_sum_by_cn.get(cid, 0.0) + float(r.get('amount') or 0)

            kept_iss_ids: set = set()
            running_per_cn: dict = {}
            for r in iss_rows:
                cid = r.get('credit_note_id')
                amount = float(r.get('amount') or 0)
                if not cid or amount <= 0:
                    continue
                meta = cn_meta.get(cid)
                if not meta:
                    # No parent CN found (orphan issuance) — count as-is.
                    direct_credit_amount += amount
                    kept_iss_ids.add(r.get('id'))
                    continue
                cap = meta["remaining_capacity"]
                used = running_per_cn.get(cid, 0.0)
                allowed = max(0.0, min(amount, cap - used))
                if allowed > 0:
                    direct_credit_amount += allowed
                    running_per_cn[cid] = used + allowed
                    kept_iss_ids.add(r.get('id'))

            direct_credit_amount = round(direct_credit_amount, 2)
            # Surface only the issuances that actually contributed (so the UI
            # popup reflects what was counted, not what was filtered out as
            # already-covered-by-delivery).
            direct_issuances = [r for r in iss_rows if r.get('id') in kept_iss_ids]

        s['direct_credit_issuances'] = direct_issuances
        s['stockout_totals'] = {
            "customer_order_value": round(sum(i.get('customer_order_value', 0) for i in items), 2),
            "distributor_margin": round(sum(i.get('distributor_margin', 0) for i in items), 2),
            "actual_billable": round(sum(i.get('actual_billable', 0) for i in items), 2),
            "credit_applied": round(sum(i.get('credit_applied', 0) for i in items), 2),
            "net_billable": round(sum(i.get('net_billable', 0) for i in items), 2),
            "billed_at_transfer": round(sum(i.get('billed_at_transfer', 0) for i in items), 2),
            "direct_credit_issued": direct_credit_amount,
        }

    return settlements


@router.get("/{distributor_id}/settlements/{settlement_id}")
async def get_settlement(
    distributor_id: str,
    settlement_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get a specific settlement with items.

    Each item in the response is enriched with the EXACT per-delivery numbers
    calculated at Stock Out time (see _compute_delivery_stockout_view) so the
    settlement screen renders the same numbers as the Delivery preview — no
    recalculation, no drift.
    """
    tenant_id = get_current_tenant_id()
    
    settlement = await db.distributor_settlements.find_one(
        {"id": settlement_id, "tenant_id": tenant_id, "distributor_id": distributor_id},
        {"_id": 0}
    )
    
    if not settlement:
        raise HTTPException(status_code=404, detail="Settlement not found")

    await _enrich_settlements_with_stockout_totals(tenant_id, [settlement])
    return settlement


@router.get("/{distributor_id}/unsettled-deliveries")
async def get_unsettled_deliveries(
    distributor_id: str,
    month: Optional[int] = None,
    year: Optional[int] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get completed deliveries that haven't been settled yet, filtered by month/year"""
    tenant_id = get_current_tenant_id()
    
    # Get all settled delivery IDs
    settled_items = await db.distributor_settlement_items.find(
        {"tenant_id": tenant_id},
        {"delivery_id": 1}
    ).to_list(10000)
    
    settled_delivery_ids = [item['delivery_id'] for item in settled_items]
    
    # Query for completed deliveries not in settled list
    query = {
        "tenant_id": tenant_id,
        "distributor_id": distributor_id,
        "status": {"$in": ["delivered", "complete"]},
        "id": {"$nin": settled_delivery_ids}
    }
    
    # Filter by month/year if provided
    if month and year:
        # Calculate date range for the month
        start_date = f"{year}-{month:02d}-01"
        if month == 12:
            end_date = f"{year + 1}-01-01"
        else:
            end_date = f"{year}-{month + 1:02d}-01"
        query["delivery_date"] = {"$gte": start_date, "$lt": end_date}
    else:
        # Use from_date/to_date if provided
        if from_date:
            query["delivery_date"] = {"$gte": from_date}
        if to_date:
            if "delivery_date" in query:
                query["delivery_date"]["$lte"] = to_date
            else:
                query["delivery_date"] = {"$lte": to_date}
    
    deliveries = await db.distributor_deliveries.find(
        query,
        {"_id": 0}
    ).sort("delivery_date", 1).to_list(500)
    
    # Fetch items for each delivery
    for delivery in deliveries:
        items = await db.distributor_delivery_items.find(
            {"delivery_id": delivery['id'], "tenant_id": tenant_id},
            {"_id": 0}
        ).to_list(500)
        delivery['items'] = items
    
    # Calculate totals
    total_quantity = sum(d.get('total_quantity', 0) for d in deliveries)
    total_amount = sum(d.get('total_net_amount', 0) for d in deliveries)
    total_margin = sum(d.get('total_margin_amount', 0) for d in deliveries)
    
    return {
        "deliveries": deliveries,
        "count": len(deliveries),
        "total_quantity": total_quantity,
        "total_amount": round(total_amount, 2),
        "total_margin": round(total_margin, 2)
    }


@router.get("/{distributor_id}/settlement-preview")
async def get_settlement_preview(
    distributor_id: str,
    month: int,
    year: int,
    current_user: dict = Depends(get_current_user)
):
    """Get a complete preview of all settlement components for a given month:
    deliveries, credit notes, and factory returns."""
    tenant_id = get_current_tenant_id()

    start_date = f"{year}-{month:02d}-01"
    end_date = f"{year + 1}-01-01" if month == 12 else f"{year}-{month + 1:02d}-01"

    # --- Settled IDs ---
    settled_items = await db.distributor_settlement_items.find(
        {"tenant_id": tenant_id}, {"delivery_id": 1}
    ).to_list(10000)
    settled_delivery_ids = [item['delivery_id'] for item in settled_items]

    existing_settlements = await db.distributor_settlements.find(
        {"tenant_id": tenant_id, "distributor_id": distributor_id},
        {"_id": 0, "credit_note_ids": 1, "factory_return_ids": 1}
    ).to_list(10000)
    settled_cn_ids = []
    settled_fr_ids = []
    for es in existing_settlements:
        settled_cn_ids.extend(es.get('credit_note_ids') or [])
        settled_fr_ids.extend(es.get('factory_return_ids') or [])

    # --- Deliveries ---
    deliveries = await db.distributor_deliveries.find({
        "tenant_id": tenant_id,
        "distributor_id": distributor_id,
        "status": {"$in": ["delivered", "complete"]},
        "id": {"$nin": settled_delivery_ids},
        "delivery_date": {"$gte": start_date, "$lt": end_date}
    }, {"_id": 0}).sort("delivery_date", 1).to_list(500)
    for d in deliveries:
        d['items'] = await db.distributor_delivery_items.find(
            {"delivery_id": d['id'], "tenant_id": tenant_id}, {"_id": 0}
        ).to_list(500)

    # --- Credit Notes ---
    from datetime import datetime as dt_cls_prev
    period_start_dt = dt_cls_prev.fromisoformat(start_date + "T00:00:00")
    period_end_dt = dt_cls_prev.fromisoformat(end_date + "T00:00:00")
    credit_notes = await db.credit_notes.find({
        "tenant_id": tenant_id,
        "distributor_id": distributor_id,
        "status": {"$in": ["pending", "partially_applied", "fully_applied"]},
        "created_at": {"$gte": period_start_dt, "$lt": period_end_dt},
        "id": {"$nin": settled_cn_ids}
    }, {"_id": 0}).to_list(500)
    # Serialize datetime fields
    for cn in credit_notes:
        for k in ['created_at', 'updated_at']:
            if hasattr(cn.get(k), 'isoformat'):
                cn[k] = cn[k].isoformat()

    # --- Factory Returns ---
    factory_returns = await db.distributor_factory_returns.find({
        "tenant_id": tenant_id,
        "distributor_id": distributor_id,
        "status": {"$in": ["confirmed", "received"]},
        "return_date": {"$gte": start_date, "$lt": end_date},
        "$or": [{"requires_settlement": True}, {"source": "warehouse"}],
        "id": {"$nin": settled_fr_ids}
    }, {"_id": 0}).to_list(500)

    total_delivery_amount = sum(d.get('total_net_amount', 0) for d in deliveries)
    total_cn_amount = sum(cn.get('original_amount', 0) or cn.get('total_amount', 0) or 0 for cn in credit_notes)
    total_fr_amount = sum(fr.get('total_credit_amount', 0) for fr in factory_returns)

    return {
        "deliveries": deliveries,
        "credit_notes": credit_notes,
        "factory_returns": factory_returns,
        "summary": {
            "total_deliveries": len(deliveries),
            "total_delivery_amount": round(total_delivery_amount, 2),
            "total_credit_notes": len(credit_notes),
            "total_credit_note_amount": round(total_cn_amount, 2),
            "total_factory_returns": len(factory_returns),
            "total_factory_return_amount": round(total_fr_amount, 2)
        }
    }


@router.post("/{distributor_id}/settlements")
async def create_settlement(
    distributor_id: str,
    data: DistributorSettlementCreate,
    current_user: dict = Depends(get_current_user)
):
    """Create a new settlement for a period"""
    if not is_distributor_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    tenant_id = get_current_tenant_id()
    now = datetime.now(timezone.utc).isoformat()
    
    # Validate distributor
    distributor = await db.distributors.find_one(
        {"id": distributor_id, "tenant_id": tenant_id},
        {"_id": 0, "distributor_name": 1, "distributor_code": 1}
    )
    if not distributor:
        raise HTTPException(status_code=404, detail="Distributor not found")
    
    # Get all settled delivery IDs to exclude
    settled_items = await db.distributor_settlement_items.find(
        {"tenant_id": tenant_id},
        {"delivery_id": 1}
    ).to_list(10000)
    
    settled_delivery_ids = [item['delivery_id'] for item in settled_items if item.get('delivery_id')]
    
    # Get already-settled credit note IDs and factory return IDs from existing settlements
    existing_settlements = await db.distributor_settlements.find(
        {"tenant_id": tenant_id, "distributor_id": distributor_id},
        {"_id": 0, "credit_note_ids": 1, "factory_return_ids": 1}
    ).to_list(10000)
    settled_cn_ids = []
    settled_fr_ids = []
    for s in existing_settlements:
        settled_cn_ids.extend(s.get('credit_note_ids') or [])
        settled_fr_ids.extend(s.get('factory_return_ids') or [])
    
    # Get completed deliveries for the period that haven't been settled
    deliveries = await db.distributor_deliveries.find({
        "tenant_id": tenant_id,
        "distributor_id": distributor_id,
        "status": {"$in": ["delivered", "complete"]},
        "delivery_date": {"$gte": data.period_start, "$lte": data.period_end},
        "id": {"$nin": settled_delivery_ids}
    }, {"_id": 0}).sort("delivery_date", 1).to_list(500)
    
    # Get credit notes issued in the period (not yet settled)
    # credit_notes.created_at is stored as datetime object, convert period dates
    from datetime import datetime as dt_cls
    period_start_dt = dt_cls.fromisoformat(data.period_start + "T00:00:00")
    period_end_dt = dt_cls.fromisoformat(data.period_end + "T23:59:59")
    credit_notes = await db.credit_notes.find({
        "tenant_id": tenant_id,
        "distributor_id": distributor_id,
        "status": {"$in": ["pending", "partially_applied", "fully_applied"]},
        "created_at": {"$gte": period_start_dt, "$lte": period_end_dt},
        "id": {"$nin": settled_cn_ids}
    }, {"_id": 0}).to_list(500)
    total_credit_notes_issued = sum(cn.get('original_amount', 0) or cn.get('total_amount', 0) or cn.get('amount', 0) or 0 for cn in credit_notes)
    
    # Get adjustable factory returns in the period (not yet settled)
    factory_returns = await db.distributor_factory_returns.find({
        "tenant_id": tenant_id,
        "distributor_id": distributor_id,
        "status": {"$in": ["confirmed", "received"]},
        "return_date": {"$gte": data.period_start, "$lte": data.period_end},
        "$or": [{"requires_settlement": True}, {"source": "warehouse"}],
        "id": {"$nin": settled_fr_ids}
    }, {"_id": 0}).to_list(500)
    total_factory_return_credit = sum(fr.get('total_credit_amount', 0) for fr in factory_returns)
    
    if not deliveries and not credit_notes and not factory_returns:
        raise HTTPException(status_code=400, detail="No unsettled deliveries, credit notes, or factory returns found for this period")
    
    # Generate settlement
    settlement_number = await generate_settlement_number(tenant_id)
    settlement_id = str(uuid.uuid4())
    
    # Calculate totals
    total_deliveries = len(deliveries)
    total_quantity = sum(d.get('total_quantity', 0) for d in deliveries)
    total_delivery_amount = sum(d.get('total_net_amount', 0) for d in deliveries)
    total_margin_amount = sum(d.get('total_margin_amount', 0) for d in deliveries)
    
    # Adjustment from distributor to factory (when customer price > base price)
    total_dist_to_factory_adjustment = sum(d.get('total_adjustment_dist_to_factory', 0) for d in deliveries)
    
    # Net adjustments: Credit notes issued + Factory return credits (factory pays distributor) - Price adjustments (distributor pays factory)
    net_adjustments = total_credit_notes_issued + total_factory_return_credit - total_dist_to_factory_adjustment
    
    # Create settlement items
    items_to_insert = []
    for delivery in deliveries:
        items_to_insert.append({
            "id": str(uuid.uuid4()),
            "tenant_id": tenant_id,
            "settlement_id": settlement_id,
            "delivery_id": delivery['id'],
            "delivery_number": delivery.get('delivery_number'),
            "delivery_date": delivery.get('delivery_date'),
            "account_id": delivery.get('account_id'),
            "account_name": delivery.get('account_name'),
            "account_city": delivery.get('account_city'),
            "total_quantity": delivery.get('total_quantity', 0),
            "total_amount": delivery.get('total_net_amount', 0),
            "margin_amount": delivery.get('total_margin_amount', 0),
            "credit_applied": delivery.get('total_credit_applied', 0),
            "adjustment_dist_to_factory": delivery.get('total_adjustment_dist_to_factory', 0)
        })
    
    # Final payout = Net Adjustments only (margin is already retained by distributor from customer collections)
    # Net = -(Dist→Factory Adj) + Credit Notes + Factory Returns
    final_payout = net_adjustments
    
    # Create settlement document
    settlement_doc = {
        "id": settlement_id,
        "tenant_id": tenant_id,
        "settlement_number": settlement_number,
        "distributor_id": distributor_id,
        "distributor_name": distributor.get('distributor_name'),
        "distributor_code": distributor.get('distributor_code'),
        "period_type": data.period_type,
        "period_start": data.period_start,
        "period_end": data.period_end,
        "total_deliveries": total_deliveries,
        "total_quantity": total_quantity,
        "total_delivery_amount": round(total_delivery_amount, 2),
        "total_margin_amount": round(total_margin_amount, 2),
        "total_credit_notes_issued": round(total_credit_notes_issued, 2),
        "total_factory_return_credit": round(total_factory_return_credit, 2),
        "total_dist_to_factory_adjustment": round(total_dist_to_factory_adjustment, 2),
        "credit_note_ids": [cn.get('id') for cn in credit_notes],
        "factory_return_ids": [fr.get('id') for fr in factory_returns],
        "total_credit_notes": len(credit_notes),
        "total_factory_returns": len(factory_returns),
        "adjustments": round(net_adjustments, 2),
        "final_payout": round(final_payout, 2),
        "status": "draft",
        "remarks": data.remarks,
        "created_at": now,
        "updated_at": now,
        "created_by": current_user.get('id'),
        "submitted_at": None,
        "submitted_by": None,
        "approved_at": None,
        "approved_by": None,
        "approved_by_name": None,
        "rejected_at": None,
        "rejected_by": None,
        "rejection_reason": None,
        "paid_at": None,
        "paid_by": None,
        "payment_reference": None
    }
    
    # Insert settlement and items
    await db.distributor_settlements.insert_one(settlement_doc)
    if items_to_insert:
        await db.distributor_settlement_items.insert_many(items_to_insert)
    
    settlement_doc.pop('_id', None)
    settlement_doc['items'] = [
        {k: v for k, v in item.items() if k not in ['_id', 'tenant_id']} 
        for item in items_to_insert
    ]
    
    logger.info(f"Settlement {settlement_number} created for distributor {distributor_id} with {total_deliveries} deliveries by {current_user['email']}")
    
    return settlement_doc


@router.post("/{distributor_id}/settlements/generate-monthly")
async def generate_monthly_settlements(
    distributor_id: str,
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Generate monthly settlements - one per account for the given month/year"""
    if not is_distributor_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    tenant_id = get_current_tenant_id()
    
    settlement_month = data.get('settlement_month')
    settlement_year = data.get('settlement_year')
    remarks = data.get('remarks', '')
    
    if not settlement_month or not settlement_year:
        raise HTTPException(status_code=400, detail="Month and year are required")
    
    # Get distributor info
    distributor = await db.distributors.find_one(
        {"id": distributor_id, "tenant_id": tenant_id}
    )
    if not distributor:
        raise HTTPException(status_code=404, detail="Distributor not found")
    
    # Get all settled delivery IDs to exclude
    settled_items = await db.distributor_settlement_items.find(
        {"tenant_id": tenant_id},
        {"delivery_id": 1}
    ).to_list(10000)
    settled_delivery_ids = [item['delivery_id'] for item in settled_items]
    
    # Calculate date range for the month
    start_date = f"{settlement_year}-{settlement_month:02d}-01"
    if settlement_month == 12:
        end_date = f"{settlement_year + 1}-01-01"
    else:
        end_date = f"{settlement_year}-{settlement_month + 1:02d}-01"
    
    # Get unsettled deliveries for the month
    query = {
        "tenant_id": tenant_id,
        "distributor_id": distributor_id,
        "status": {"$in": ["delivered", "complete"]},
        "id": {"$nin": settled_delivery_ids},
        "delivery_date": {"$gte": start_date, "$lt": end_date}
    }
    
    deliveries = await db.distributor_deliveries.find(query, {"_id": 0}).to_list(1000)
    
    # --- Independently query credit notes for the period ---
    from datetime import datetime as dt_cls_monthly
    period_start_dt = dt_cls_monthly.fromisoformat(start_date + "T00:00:00")
    period_end_dt = dt_cls_monthly.fromisoformat(end_date + "T00:00:00")  # end_date is already 1st of next month
    
    # Get already-settled credit note IDs and factory return IDs from existing settlements
    existing_monthly_settlements = await db.distributor_settlements.find(
        {"tenant_id": tenant_id, "distributor_id": distributor_id},
        {"_id": 0, "credit_note_ids": 1, "factory_return_ids": 1}
    ).to_list(10000)
    settled_cn_ids = []
    settled_fr_ids = []
    for es in existing_monthly_settlements:
        settled_cn_ids.extend(es.get('credit_note_ids') or [])
        settled_fr_ids.extend(es.get('factory_return_ids') or [])
    
    all_credit_notes = await db.credit_notes.find({
        "tenant_id": tenant_id,
        "distributor_id": distributor_id,
        "status": {"$in": ["pending", "partially_applied", "fully_applied"]},
        "created_at": {"$gte": period_start_dt, "$lt": period_end_dt},
        "id": {"$nin": settled_cn_ids}
    }, {"_id": 0}).to_list(500)
    
    # Get adjustable factory returns in the period (warehouse-sourced, not yet settled)
    all_factory_returns = await db.distributor_factory_returns.find({
        "tenant_id": tenant_id,
        "distributor_id": distributor_id,
        "status": {"$in": ["confirmed", "received"]},
        "return_date": {"$gte": start_date, "$lt": end_date},
        "$or": [{"requires_settlement": True}, {"source": "warehouse"}],
        "id": {"$nin": settled_fr_ids}
    }, {"_id": 0}).to_list(500)
    
    total_independent_factory_return_credit = sum(fr.get('total_credit_amount', 0) for fr in all_factory_returns)
    
    if not deliveries and not all_credit_notes and not all_factory_returns:
        raise HTTPException(status_code=400, detail="No unsettled deliveries, credit notes, or factory returns found for this period")
    
    # Group deliveries by account
    accounts = {}
    for delivery in deliveries:
        account_id = delivery.get('account_id', 'unknown')
        if account_id not in accounts:
            # Get account name from delivery, or fetch from accounts table
            account_name = delivery.get('account_name')
            if not account_name and account_id != 'unknown':
                account_doc = await db.accounts.find_one({"id": account_id}, {"_id": 0, "account_name": 1, "company": 1, "name": 1})
                if account_doc:
                    account_name = account_doc.get('account_name') or account_doc.get('company') or account_doc.get('name')
            accounts[account_id] = {
                'account_id': account_id,
                'account_name': account_name or 'Unknown',
                'deliveries': []
            }
        accounts[account_id]['deliveries'].append(delivery)
    
    # Group credit notes by account_id for per-account distribution
    credit_notes_by_account = {}
    for cn in all_credit_notes:
        cn_account = cn.get('account_id', 'unknown')
        if cn_account not in credit_notes_by_account:
            credit_notes_by_account[cn_account] = []
        credit_notes_by_account[cn_account].append(cn)
    
    # Ensure accounts with credit notes (but no deliveries) are also represented
    for cn_acct_id in credit_notes_by_account:
        if cn_acct_id not in accounts:
            # Fetch account name
            cn_acct_name = 'Unknown'
            if cn_acct_id != 'unknown':
                acct_doc = await db.accounts.find_one({"id": cn_acct_id}, {"_id": 0, "account_name": 1, "company": 1, "name": 1})
                if acct_doc:
                    cn_acct_name = acct_doc.get('account_name') or acct_doc.get('company') or acct_doc.get('name') or 'Unknown'
            accounts[cn_acct_id] = {
                'account_id': cn_acct_id,
                'account_name': cn_acct_name,
                'deliveries': []
            }
    
    # If still no accounts but we have factory returns, create a placeholder account
    if not accounts and all_factory_returns:
        accounts['_factory_returns_'] = {
            'account_id': '_factory_returns_',
            'account_name': 'Factory Returns Adjustment',
            'deliveries': []
        }
    
    settlements_created = []
    factory_returns_assigned = False  # Track if factory returns have been assigned to a settlement
    
    for account_id, account_data in accounts.items():
        account_deliveries = account_data['deliveries']
        
        # Fetch items for each delivery and calculate totals
        total_billing_value = 0
        distributor_earnings = 0
        margin_at_transfer_price = 0
        total_quantity = 0
        total_price_premium = 0
        total_factory_adj = 0
        total_credit_notes_applied = 0
        total_at_transfer_price = 0
        
        items_to_insert = []
        
        for delivery in account_deliveries:
            # Get delivery items
            items = await db.distributor_delivery_items.find(
                {"delivery_id": delivery['id'], "tenant_id": tenant_id},
                {"_id": 0}
            ).to_list(500)
            
            delivery_billing = 0
            delivery_earnings = 0
            delivery_margin_at_transfer = 0
            delivery_price_premium = 0
            delivery_factory_adj = 0
            delivery_credit_applied = delivery.get('total_credit_applied', 0)
            delivery_at_transfer_price = 0
            
            for item in items:
                qty = item.get('quantity', 0)
                customer_price = item.get('customer_selling_price') or item.get('unit_price') or 0
                commission_pct = item.get('distributor_commission_percent') or item.get('margin_percent') or 2.5
                base_p = item.get('base_price') or item.get('transfer_price') or 0
                
                billing_value = qty * customer_price
                earnings = billing_value * (commission_pct / 100)
                margin_transfer = qty * base_p * (commission_pct / 100)
                
                # Price premium: extra collected when customer price > base price
                price_premium = qty * (customer_price - base_p) if customer_price > base_p and base_p > 0 else 0
                
                # NEW FORMULA: Adjustment (Dist → Factory) = Actual Billable - Billed to Dist
                # Billed to Dist = qty × transfer_price = qty × base_price × (1 - margin%)
                # Actual Billable = qty × new_transfer_price = qty × customer_price × (1 - margin%)
                # Adjustment = qty × (1 - margin%) × (customer_price - base_price)
                transfer_price = base_p * (1 - commission_pct / 100) if base_p > 0 else 0
                new_transfer_price = customer_price * (1 - commission_pct / 100) if customer_price > 0 else 0
                billed_to_dist = qty * transfer_price
                actual_billable = qty * new_transfer_price
                factory_adj = actual_billable - billed_to_dist
                
                delivery_billing += billing_value
                delivery_earnings += earnings
                delivery_margin_at_transfer += margin_transfer
                delivery_price_premium += price_premium
                delivery_factory_adj += factory_adj
                delivery_at_transfer_price += billed_to_dist
            
            total_billing_value += delivery_billing
            distributor_earnings += delivery_earnings
            margin_at_transfer_price += delivery_margin_at_transfer
            total_quantity += delivery.get('total_quantity', 0)
            total_price_premium += delivery_price_premium
            total_factory_adj += delivery_factory_adj
            total_credit_notes_applied += delivery_credit_applied
            total_at_transfer_price += delivery_at_transfer_price
            
            items_to_insert.append({
                "id": str(uuid.uuid4()),
                "tenant_id": tenant_id,
                "settlement_id": None,  # Will be set below
                "delivery_id": delivery['id'],
                "delivery_number": delivery.get('delivery_number'),
                "delivery_date": delivery.get('delivery_date'),
                "account_id": account_id,
                "account_name": account_data['account_name'],
                "total_quantity": delivery.get('total_quantity', 0),
                "total_billing_value": round(delivery_billing, 2),
                "distributor_earnings": round(delivery_earnings, 2),
                "margin_at_transfer_price": round(delivery_margin_at_transfer, 2),
                "adjustment_payable": round(delivery_earnings - delivery_margin_at_transfer, 2),
                "price_premium_payable": round(delivery_price_premium, 2),
                "factory_distributor_adjustment": round(delivery_factory_adj, 2),
                "credit_notes_applied": round(delivery_credit_applied, 2)
            })
        
        # Generate settlement for this account
        settlement_number = await generate_settlement_number(tenant_id)
        settlement_id = str(uuid.uuid4())
        
        # Update items with settlement_id
        for item in items_to_insert:
            item['settlement_id'] = settlement_id
        
        adjustment_payable = distributor_earnings - margin_at_transfer_price
        
        # Per-account credit notes (independently queried)
        account_credit_notes = credit_notes_by_account.get(account_id, [])
        account_cn_total = sum(cn.get('original_amount', 0) or cn.get('total_amount', 0) or cn.get('amount', 0) or 0 for cn in account_credit_notes)
        
        # Factory returns: assign to first settlement only (they're distributor-level, not per-account)
        account_fr_total = 0
        account_factory_returns = []
        if not factory_returns_assigned and all_factory_returns:
            account_fr_total = total_independent_factory_return_credit
            account_factory_returns = all_factory_returns
            factory_returns_assigned = True
        
        # Net Payout = Earnings - ① Price Adj (Dist→Factory) + ② Credit Notes (Factory→Dist) + ③ Factory Returns (Factory→Dist)
        final_payout = distributor_earnings - total_factory_adj + account_cn_total + account_fr_total
        
        settlement_doc = {
            "id": settlement_id,
            "tenant_id": tenant_id,
            "settlement_number": settlement_number,
            "distributor_id": distributor_id,
            "distributor_name": distributor.get('distributor_name'),
            "distributor_code": distributor.get('distributor_code'),
            "account_id": account_id,
            "account_name": account_data['account_name'],
            "settlement_month": settlement_month,
            "settlement_year": settlement_year,
            "total_deliveries": len(account_deliveries),
            "total_quantity": total_quantity,
            "total_billing_value": round(total_billing_value, 2),
            "total_at_transfer_price": round(total_at_transfer_price, 2),
            "distributor_earnings": round(distributor_earnings, 2),
            "margin_at_transfer_price": round(margin_at_transfer_price, 2),
            "adjustment_payable": round(adjustment_payable, 2),
            "price_premium_payable": round(total_price_premium, 2),
            "factory_distributor_adjustment": round(total_factory_adj, 2),
            "credit_notes_applied": round(total_credit_notes_applied, 2),
            "total_credit_notes_issued": round(account_cn_total, 2),
            "total_factory_return_credit": round(account_fr_total, 2),
            "credit_note_ids": [cn.get('id') for cn in account_credit_notes],
            "factory_return_ids": [fr.get('id') for fr in account_factory_returns],
            "final_payout": round(final_payout, 2),
            "status": "draft",
            "remarks": remarks,
            "created_by": current_user['id'],
            "created_by_name": current_user.get('name', current_user.get('email')),
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        
        # Insert settlement and items
        await db.distributor_settlements.insert_one(settlement_doc)
        if items_to_insert:
            await db.distributor_settlement_items.insert_many(items_to_insert)
        
        settlement_doc.pop('_id', None)
        settlements_created.append(settlement_doc)
    
    logger.info(f"Created {len(settlements_created)} monthly settlements for distributor {distributor_id} ({settlement_month}/{settlement_year}) by {current_user['email']}")
    
    return {
        "settlements_created": len(settlements_created),
        "settlements": settlements_created
    }


@router.put("/{distributor_id}/settlements/{settlement_id}")
async def update_settlement(
    distributor_id: str,
    settlement_id: str,
    data: DistributorSettlementUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Update a settlement (remarks, adjustments)"""
    if not is_distributor_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    tenant_id = get_current_tenant_id()
    
    settlement = await db.distributor_settlements.find_one(
        {"id": settlement_id, "tenant_id": tenant_id, "distributor_id": distributor_id}
    )
    
    if not settlement:
        raise HTTPException(status_code=404, detail="Settlement not found")
    
    if settlement.get('status') not in ['draft', 'pending_approval']:
        raise HTTPException(status_code=400, detail="Cannot modify settlement in current status")
    
    update_data = {"updated_at": datetime.now(timezone.utc).isoformat()}
    
    if data.remarks is not None:
        update_data['remarks'] = data.remarks
    
    if data.adjustments is not None:
        update_data['adjustments'] = data.adjustments
        update_data['final_payout'] = round(settlement.get('total_margin_amount', 0) + data.adjustments, 2)
    
    await db.distributor_settlements.update_one(
        {"id": settlement_id, "tenant_id": tenant_id},
        {"$set": update_data}
    )
    
    updated = await db.distributor_settlements.find_one(
        {"id": settlement_id, "tenant_id": tenant_id},
        {"_id": 0}
    )
    
    return updated


@router.post("/{distributor_id}/settlements/{settlement_id}/submit")
async def submit_settlement(
    distributor_id: str,
    settlement_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Submit settlement for approval"""
    if not is_distributor_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    tenant_id = get_current_tenant_id()
    
    settlement = await db.distributor_settlements.find_one(
        {"id": settlement_id, "tenant_id": tenant_id, "distributor_id": distributor_id}
    )
    
    if not settlement:
        raise HTTPException(status_code=404, detail="Settlement not found")
    
    if settlement.get('status') != 'draft':
        raise HTTPException(status_code=400, detail="Only draft settlements can be submitted")
    
    now = datetime.now(timezone.utc).isoformat()
    
    await db.distributor_settlements.update_one(
        {"id": settlement_id, "tenant_id": tenant_id},
        {"$set": {
            "status": "pending_approval",
            "submitted_at": now,
            "submitted_by": current_user.get('id'),
            "updated_at": now
        }}
    )
    
    logger.info(f"Settlement {settlement['settlement_number']} submitted for approval by {current_user['email']}")
    
    return {"message": f"Settlement {settlement['settlement_number']} submitted for approval", "status": "pending_approval"}


@router.post("/{distributor_id}/settlements/{settlement_id}/approve")
async def approve_settlement(
    distributor_id: str,
    settlement_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Approve a settlement"""
    # Only CEO/Director can approve
    if current_user.get('role') not in ['CEO', 'Director', 'Vice President']:
        raise HTTPException(status_code=403, detail="Only senior management can approve settlements")
    
    tenant_id = get_current_tenant_id()
    
    settlement = await db.distributor_settlements.find_one(
        {"id": settlement_id, "tenant_id": tenant_id, "distributor_id": distributor_id}
    )
    
    if not settlement:
        raise HTTPException(status_code=404, detail="Settlement not found")
    
    if settlement.get('status') != 'pending_approval':
        raise HTTPException(status_code=400, detail="Settlement is not pending approval")
    
    now = datetime.now(timezone.utc).isoformat()
    
    await db.distributor_settlements.update_one(
        {"id": settlement_id, "tenant_id": tenant_id},
        {"$set": {
            "status": "approved",
            "approved_at": now,
            "approved_by": current_user.get('id'),
            "approved_by_name": current_user.get('name') or current_user.get('email'),
            "updated_at": now
        }}
    )
    
    logger.info(f"Settlement {settlement['settlement_number']} approved by {current_user['email']}")
    
    return {"message": f"Settlement {settlement['settlement_number']} approved", "status": "approved"}


@router.post("/{distributor_id}/settlements/{settlement_id}/reject")
async def reject_settlement(
    distributor_id: str,
    settlement_id: str,
    reason: str = "",
    current_user: dict = Depends(get_current_user)
):
    """Reject a settlement"""
    if current_user.get('role') not in ['CEO', 'Director', 'Vice President']:
        raise HTTPException(status_code=403, detail="Only senior management can reject settlements")
    
    tenant_id = get_current_tenant_id()
    
    settlement = await db.distributor_settlements.find_one(
        {"id": settlement_id, "tenant_id": tenant_id, "distributor_id": distributor_id}
    )
    
    if not settlement:
        raise HTTPException(status_code=404, detail="Settlement not found")
    
    if settlement.get('status') != 'pending_approval':
        raise HTTPException(status_code=400, detail="Settlement is not pending approval")
    
    now = datetime.now(timezone.utc).isoformat()
    
    await db.distributor_settlements.update_one(
        {"id": settlement_id, "tenant_id": tenant_id},
        {"$set": {
            "status": "rejected",
            "rejected_at": now,
            "rejected_by": current_user.get('id'),
            "rejection_reason": reason,
            "updated_at": now
        }}
    )
    
    logger.info(f"Settlement {settlement['settlement_number']} rejected by {current_user['email']}")
    
    return {"message": f"Settlement {settlement['settlement_number']} rejected", "status": "rejected"}


@router.post("/{distributor_id}/settlements/{settlement_id}/mark-paid")
async def mark_settlement_paid(
    distributor_id: str,
    settlement_id: str,
    payment_reference: str = "",
    current_user: dict = Depends(get_current_user)
):
    """Mark settlement as paid"""
    if not is_distributor_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    tenant_id = get_current_tenant_id()
    
    settlement = await db.distributor_settlements.find_one(
        {"id": settlement_id, "tenant_id": tenant_id, "distributor_id": distributor_id}
    )
    
    if not settlement:
        raise HTTPException(status_code=404, detail="Settlement not found")
    
    if settlement.get('status') != 'approved':
        raise HTTPException(status_code=400, detail="Only approved settlements can be marked as paid")
    
    now = datetime.now(timezone.utc).isoformat()
    
    await db.distributor_settlements.update_one(
        {"id": settlement_id, "tenant_id": tenant_id},
        {"$set": {
            "status": "paid",
            "paid_at": now,
            "paid_by": current_user.get('id'),
            "payment_reference": payment_reference,
            "updated_at": now
        }}
    )
    
    logger.info(f"Settlement {settlement['settlement_number']} marked as paid by {current_user['email']}")
    
    return {"message": f"Settlement {settlement['settlement_number']} marked as paid", "status": "paid"}


@router.delete("/{distributor_id}/settlements/{settlement_id}")
async def delete_settlement(
    distributor_id: str,
    settlement_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Delete a draft settlement"""
    if not is_distributor_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    tenant_id = get_current_tenant_id()
    
    settlement = await db.distributor_settlements.find_one(
        {"id": settlement_id, "tenant_id": tenant_id, "distributor_id": distributor_id}
    )
    
    if not settlement:
        raise HTTPException(status_code=404, detail="Settlement not found")
    
    if settlement.get('status') != 'draft':
        raise HTTPException(status_code=400, detail="Only draft settlements can be deleted")
    
    # Delete items first
    await db.distributor_settlement_items.delete_many({"settlement_id": settlement_id, "tenant_id": tenant_id})
    
    # Delete settlement
    await db.distributor_settlements.delete_one({"id": settlement_id, "tenant_id": tenant_id})
    
    logger.info(f"Settlement {settlement['settlement_number']} deleted by {current_user['email']}")
    
    return {"message": f"Settlement {settlement['settlement_number']} deleted"}


# ============ Billing Configuration CRUD ============

async def generate_billing_config_id() -> str:
    return str(uuid.uuid4())


@router.get("/{distributor_id}/billing-config")
async def get_billing_configs(
    distributor_id: str,
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get billing configurations (base prices) for a distributor"""
    tenant_id = get_current_tenant_id()
    
    query = {"tenant_id": tenant_id, "distributor_id": distributor_id}
    if status:
        query["status"] = status
    
    configs = await db.distributor_billing_config.find(
        query, {"_id": 0}
    ).sort("sku_name", 1).to_list(500)
    
    return {"configs": configs, "count": len(configs)}


@router.post("/{distributor_id}/billing-config")
async def create_billing_config(
    distributor_id: str,
    data: BillingConfigCreate,
    current_user: dict = Depends(get_current_user)
):
    """Create billing configuration for a SKU at distributor level"""
    if not is_distributor_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    tenant_id = get_current_tenant_id()
    
    # Check for existing config
    existing = await db.distributor_billing_config.find_one({
        "tenant_id": tenant_id,
        "distributor_id": distributor_id,
        "sku_id": data.sku_id,
        "status": "active"
    })
    if existing:
        raise HTTPException(status_code=400, detail="Active billing config already exists for this SKU")
    
    # Get SKU name if not provided
    sku_name = data.sku_name
    if not sku_name:
        sku = await db.master_skus.find_one({"id": data.sku_id})
        sku_name = sku.get('sku_name') or sku.get('name') if sku else None
    
    # Calculate transfer price
    transfer_price = data.base_price * (1 - data.margin_percent / 100)
    
    now = datetime.now(timezone.utc).isoformat()
    config_id = await generate_billing_config_id()
    
    config_doc = {
        "id": config_id,
        "tenant_id": tenant_id,
        "distributor_id": distributor_id,
        "sku_id": data.sku_id,
        "sku_name": sku_name,
        "base_price": data.base_price,
        "margin_percent": data.margin_percent,
        "transfer_price": transfer_price,
        "effective_from": data.effective_from or now[:10],
        "effective_to": data.effective_to,
        "remarks": data.remarks,
        "status": data.status or "active",
        "created_at": now,
        "updated_at": now,
        "created_by": current_user.get('id')
    }
    
    await db.distributor_billing_config.insert_one(config_doc)
    config_doc.pop('_id', None)
    
    return config_doc


@router.post("/{distributor_id}/billing-config/bulk")
async def bulk_create_billing_config(
    distributor_id: str,
    configs: List[BillingConfigCreate],
    current_user: dict = Depends(get_current_user)
):
    """Bulk create billing configurations"""
    if not is_distributor_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    tenant_id = get_current_tenant_id()
    now = datetime.now(timezone.utc).isoformat()
    
    # Get all SKU names
    sku_ids = [c.sku_id for c in configs]
    skus = await db.master_skus.find(
        {"$or": [{"id": {"$in": sku_ids}}, {"tenant_id": tenant_id}]}
    ).to_list(500)
    sku_map = {s.get('id'): s.get('sku_name') or s.get('name') for s in skus}
    
    created = []
    skipped = []
    
    for data in configs:
        # Check for existing
        existing = await db.distributor_billing_config.find_one({
            "tenant_id": tenant_id,
            "distributor_id": distributor_id,
            "sku_id": data.sku_id,
            "status": "active"
        })
        if existing:
            skipped.append(data.sku_id)
            continue
        
        transfer_price = data.base_price * (1 - data.margin_percent / 100)
        config_id = await generate_billing_config_id()
        
        config_doc = {
            "id": config_id,
            "tenant_id": tenant_id,
            "distributor_id": distributor_id,
            "sku_id": data.sku_id,
            "sku_name": data.sku_name or sku_map.get(data.sku_id),
            "base_price": data.base_price,
            "margin_percent": data.margin_percent,
            "transfer_price": transfer_price,
            "effective_from": data.effective_from or now[:10],
            "effective_to": data.effective_to,
            "remarks": data.remarks,
            "status": data.status or "active",
            "created_at": now,
            "updated_at": now,
            "created_by": current_user.get('id')
        }
        
        await db.distributor_billing_config.insert_one(config_doc)
        created.append(config_id)
    
    return {"created": len(created), "skipped": len(skipped), "skipped_sku_ids": skipped}


@router.put("/{distributor_id}/billing-config/{config_id}")
async def update_billing_config(
    distributor_id: str,
    config_id: str,
    data: BillingConfigUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Update billing configuration"""
    if not is_distributor_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    tenant_id = get_current_tenant_id()
    
    config = await db.distributor_billing_config.find_one({
        "id": config_id, "tenant_id": tenant_id, "distributor_id": distributor_id
    })
    if not config:
        raise HTTPException(status_code=404, detail="Billing config not found")
    
    update_data = {"updated_at": datetime.now(timezone.utc).isoformat()}
    
    if data.base_price is not None:
        update_data["base_price"] = data.base_price
    if data.margin_percent is not None:
        update_data["margin_percent"] = data.margin_percent
    
    # Recalculate transfer price if base_price or margin changed
    base_price = update_data.get("base_price", config.get("base_price"))
    margin_percent = update_data.get("margin_percent", config.get("margin_percent"))
    update_data["transfer_price"] = base_price * (1 - margin_percent / 100)
    
    if data.effective_from is not None:
        update_data["effective_from"] = data.effective_from
    if data.effective_to is not None:
        update_data["effective_to"] = data.effective_to
    if data.remarks is not None:
        update_data["remarks"] = data.remarks
    if data.status is not None:
        update_data["status"] = data.status
    
    await db.distributor_billing_config.update_one(
        {"id": config_id, "tenant_id": tenant_id},
        {"$set": update_data}
    )
    
    return {"message": "Billing config updated", "id": config_id}


@router.delete("/{distributor_id}/billing-config/{config_id}")
async def delete_billing_config(
    distributor_id: str,
    config_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Delete billing configuration"""
    if not is_distributor_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    tenant_id = get_current_tenant_id()
    
    result = await db.distributor_billing_config.delete_one({
        "id": config_id, "tenant_id": tenant_id, "distributor_id": distributor_id
    })
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Billing config not found")
    
    return {"message": "Billing config deleted"}


# ============ Provisional Invoice Generation ============

async def generate_provisional_invoice_number(tenant_id: str) -> str:
    """Generate unique provisional invoice number like PINV-2026-0001"""
    year = datetime.now().year
    count = await db.distributor_provisional_invoices.count_documents({
        "tenant_id": tenant_id,
        "invoice_number": {"$regex": f"^PINV-{year}-"}
    })
    return f"PINV-{year}-{count + 1:04d}"


@router.post("/{distributor_id}/provisional-invoices/generate")
async def generate_provisional_invoice(
    distributor_id: str,
    shipment_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Generate provisional invoice when shipment is delivered"""
    if not is_distributor_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    tenant_id = get_current_tenant_id()
    
    # Get shipment
    shipment = await db.distributor_shipments.find_one({
        "id": shipment_id, "tenant_id": tenant_id, "distributor_id": distributor_id
    })
    if not shipment:
        raise HTTPException(status_code=404, detail="Shipment not found")
    
    if shipment.get('status') != 'delivered':
        raise HTTPException(status_code=400, detail="Shipment must be delivered to generate invoice")
    
    # Check if invoice already exists
    existing = await db.distributor_provisional_invoices.find_one({
        "tenant_id": tenant_id, "shipment_id": shipment_id
    })
    if existing:
        raise HTTPException(status_code=400, detail="Provisional invoice already exists for this shipment")
    
    # Get billing configs for this distributor
    configs = await db.distributor_billing_config.find({
        "tenant_id": tenant_id,
        "distributor_id": distributor_id,
        "status": "active"
    }).to_list(500)
    config_map = {c.get('sku_id'): c for c in configs}
    
    # Get shipment items
    shipment_items = await db.distributor_shipment_items.find({
        "shipment_id": shipment_id, "tenant_id": tenant_id
    }).to_list(500)
    
    # Generate invoice
    now = datetime.now(timezone.utc).isoformat()
    invoice_id = str(uuid.uuid4())
    invoice_number = await generate_provisional_invoice_number(tenant_id)
    
    invoice_items = []
    total_quantity = 0
    total_gross_amount = 0
    total_margin_amount = 0
    total_net_amount = 0
    
    for item in shipment_items:
        sku_id = item.get('sku_id')
        quantity = item.get('quantity', 0)
        config = config_map.get(sku_id)
        
        if not config:
            # If no config, use shipment item price as base
            base_price = item.get('unit_price', 0)
            margin_percent = 2.5  # Default margin
        else:
            base_price = config.get('base_price', item.get('unit_price', 0))
            margin_percent = config.get('margin_percent', 2.5)
        
        transfer_price = base_price * (1 - margin_percent / 100)
        gross_amount = quantity * base_price
        margin_amount = gross_amount * margin_percent / 100
        net_amount = gross_amount - margin_amount
        
        item_id = str(uuid.uuid4())
        invoice_item = {
            "id": item_id,
            "invoice_id": invoice_id,
            "tenant_id": tenant_id,
            "sku_id": sku_id,
            "sku_name": item.get('sku_name'),
            "quantity": quantity,
            "base_price": base_price,
            "margin_percent": margin_percent,
            "transfer_price": transfer_price,
            "gross_amount": gross_amount,
            "margin_amount": margin_amount,
            "net_amount": net_amount
        }
        invoice_items.append(invoice_item)
        
        total_quantity += quantity
        total_gross_amount += gross_amount
        total_margin_amount += margin_amount
        total_net_amount += net_amount
    
    # Create invoice
    distributor = await db.distributors.find_one({"id": distributor_id})
    
    invoice_doc = {
        "id": invoice_id,
        "tenant_id": tenant_id,
        "invoice_number": invoice_number,
        "distributor_id": distributor_id,
        "distributor_name": distributor.get('distributor_name') if distributor else None,
        "distributor_code": distributor.get('distributor_code') if distributor else None,
        "shipment_id": shipment_id,
        "shipment_number": shipment.get('shipment_number'),
        "invoice_date": now[:10],
        "due_date": None,
        "total_quantity": total_quantity,
        "total_gross_amount": round(total_gross_amount, 2),
        "total_margin_amount": round(total_margin_amount, 2),
        "total_net_amount": round(total_net_amount, 2),
        "status": "pending",
        "reconciliation_status": "pending",
        "reconciled_quantity": 0,
        "reconciled_amount": 0,
        "remarks": f"Auto-generated from shipment {shipment.get('shipment_number')}",
        "created_at": now,
        "updated_at": now,
        "created_by": current_user.get('id')
    }
    
    await db.distributor_provisional_invoices.insert_one(invoice_doc)
    if invoice_items:
        await db.distributor_provisional_invoice_items.insert_many(invoice_items)
    
    invoice_doc.pop('_id', None)
    invoice_doc['items'] = invoice_items
    
    logger.info(f"Generated provisional invoice {invoice_number} for shipment {shipment.get('shipment_number')}")
    
    return invoice_doc


@router.get("/{distributor_id}/provisional-invoices")
async def get_provisional_invoices(
    distributor_id: str,
    status: Optional[str] = None,
    reconciliation_status: Optional[str] = None,
    page: int = 1,
    page_size: int = 20,
    current_user: dict = Depends(get_current_user)
):
    """Get provisional invoices for a distributor"""
    tenant_id = get_current_tenant_id()
    
    query = {"tenant_id": tenant_id, "distributor_id": distributor_id}
    if status:
        query["status"] = status
    if reconciliation_status:
        query["reconciliation_status"] = reconciliation_status
    
    total = await db.distributor_provisional_invoices.count_documents(query)
    invoices = await db.distributor_provisional_invoices.find(
        query, {"_id": 0}
    ).sort("created_at", -1).skip((page - 1) * page_size).limit(page_size).to_list(page_size)
    
    return {
        "invoices": invoices,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size
    }


@router.get("/{distributor_id}/provisional-invoices/{invoice_id}")
async def get_provisional_invoice(
    distributor_id: str,
    invoice_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get provisional invoice with items"""
    tenant_id = get_current_tenant_id()
    
    invoice = await db.distributor_provisional_invoices.find_one({
        "id": invoice_id, "tenant_id": tenant_id, "distributor_id": distributor_id
    }, {"_id": 0})
    
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    items = await db.distributor_provisional_invoice_items.find({
        "invoice_id": invoice_id, "tenant_id": tenant_id
    }, {"_id": 0}).to_list(500)
    
    invoice['items'] = items
    return invoice


# ============ Reconciliation Engine ============

async def generate_reconciliation_number(tenant_id: str) -> str:
    """Generate unique reconciliation number like REC-2026-0001"""
    year = datetime.now().year
    count = await db.distributor_reconciliations.count_documents({
        "tenant_id": tenant_id,
        "reconciliation_number": {"$regex": f"^REC-{year}-"}
    })
    return f"REC-{year}-{count + 1:04d}"


@router.post("/{distributor_id}/reconciliations/calculate")
async def calculate_reconciliation(
    distributor_id: str,
    data: ReconciliationCreate,
    current_user: dict = Depends(get_current_user)
):
    """Calculate reconciliation for a period (preview without saving)"""
    tenant_id = get_current_tenant_id()
    
    # Get margin matrix entries for this distributor
    # Filter by active dates - entry is valid if:
    # (active_from is null or active_from <= period_end) AND (active_to is null or active_to >= period_start)
    margin_entries = await db.distributor_margin_matrix.find({
        "tenant_id": tenant_id,
        "distributor_id": distributor_id,
        "status": "active",
        "$or": [
            {"active_from": None},
            {"active_from": {"$lte": data.period_end}}
        ]
    }).to_list(500)
    
    # Filter out entries that ended before the period started
    margin_entries = [
        m for m in margin_entries 
        if not m.get('active_to') or m.get('active_to') >= data.period_start
    ]
    
    # Build lookup map: (sku_id, city) -> margin entry
    # For entries without city match, we'll use sku_id only as fallback
    margin_map_by_city_sku = {}
    margin_map_by_sku = {}
    for m in margin_entries:
        key = (m.get('sku_id'), m.get('city'))
        margin_map_by_city_sku[key] = m
        # Also keep by sku_id only for fallback
        if m.get('sku_id') not in margin_map_by_sku:
            margin_map_by_sku[m.get('sku_id')] = m
    
    # Get deliveries in the period
    deliveries = await db.distributor_deliveries.find({
        "tenant_id": tenant_id,
        "distributor_id": distributor_id,
        "status": {"$in": ["delivered", "complete"]},
        "delivery_date": {"$gte": data.period_start, "$lte": data.period_end}
    }).to_list(1000)
    
    if not deliveries:
        return {
            "message": "No delivered items found in this period",
            "period_start": data.period_start,
            "period_end": data.period_end,
            "total_deliveries": 0,
            "items": []
        }
    
    # Get all delivery items
    delivery_ids = [d.get('id') for d in deliveries]
    delivery_items = await db.distributor_delivery_items.find({
        "delivery_id": {"$in": delivery_ids},
        "tenant_id": tenant_id
    }).to_list(5000)
    
    # Build delivery map
    delivery_map = {d.get('id'): d for d in deliveries}
    
    # Get accounts to get city info
    account_ids = list(set(d.get('account_id') for d in deliveries if d.get('account_id')))
    accounts = await db.accounts.find(
        {"id": {"$in": account_ids}},
        {"_id": 0, "id": 1, "city": 1}
    ).to_list(500)
    account_city_map = {a.get('id'): a.get('city') for a in accounts}
    
    # Calculate reconciliation items
    items = []
    total_provisional = 0
    total_actual_gross = 0
    total_entitled_margin = 0
    total_actual_net = 0
    total_quantity = 0
    
    for item in delivery_items:
        delivery = delivery_map.get(item.get('delivery_id'))
        if not delivery:
            continue
        
        sku_id = item.get('sku_id')
        quantity = item.get('quantity', 0)
        actual_selling_price = item.get('unit_price', 0)  # Price sold to customer
        
        # Get account city for lookup
        account_id = delivery.get('account_id')
        account_city = delivery.get('account_city') or account_city_map.get(account_id)
        
        # Get margin entry - first try city+sku, then sku only
        margin_entry = margin_map_by_city_sku.get((sku_id, account_city)) or margin_map_by_sku.get(sku_id)
        
        if margin_entry:
            base_price = margin_entry.get('base_price', actual_selling_price)
            margin_type = margin_entry.get('margin_type', 'percentage')
            margin_value = margin_entry.get('margin_value', 2.5)
            
            # Calculate margin based on type
            if margin_type == 'percentage':
                margin_percent = margin_value
            else:
                # For fixed margin types, convert to percentage for calculation
                margin_percent = (margin_value / base_price * 100) if base_price > 0 else 2.5
        else:
            base_price = actual_selling_price  # Fallback - no margin entry found
            margin_percent = 2.5
        
        transfer_price = base_price * (1 - margin_percent / 100)
        
        # Provisional (what distributor paid)
        provisional_amount = quantity * transfer_price
        
        # Actual (based on customer selling price)
        actual_gross_amount = quantity * actual_selling_price
        entitled_margin_amount = actual_gross_amount * margin_percent / 100
        actual_net_amount = actual_gross_amount - entitled_margin_amount
        
        # Difference
        difference_amount = actual_net_amount - provisional_amount
        
        rec_item = {
            "delivery_id": delivery.get('id'),
            "delivery_number": delivery.get('delivery_number'),
            "delivery_date": delivery.get('delivery_date'),
            "account_id": delivery.get('account_id'),
            "account_name": delivery.get('account_name'),
            "account_city": account_city,
            "sku_id": sku_id,
            "sku_name": item.get('sku_name'),
            "quantity": quantity,
            "base_price": round(base_price, 2),
            "margin_percent": round(margin_percent, 2),
            "transfer_price": round(transfer_price, 2),
            "provisional_amount": round(provisional_amount, 2),
            "actual_selling_price": round(actual_selling_price, 2),
            "actual_gross_amount": round(actual_gross_amount, 2),
            "entitled_margin_amount": round(entitled_margin_amount, 2),
            "actual_net_amount": round(actual_net_amount, 2),
            "difference_amount": round(difference_amount, 2),
            "margin_entry_found": margin_entry is not None
        }
        items.append(rec_item)
        
        total_quantity += quantity
        total_provisional += provisional_amount
        total_actual_gross += actual_gross_amount
        total_entitled_margin += entitled_margin_amount
        total_actual_net += actual_net_amount
    
    total_difference = total_actual_net - total_provisional
    settlement_type = "debit_note" if total_difference > 0 else "credit_note" if total_difference < 0 else None
    
    return {
        "period_start": data.period_start,
        "period_end": data.period_end,
        "total_deliveries": len(deliveries),
        "total_quantity": total_quantity,
        "total_provisional_amount": round(total_provisional, 2),
        "total_actual_gross_amount": round(total_actual_gross, 2),
        "total_entitled_margin": round(total_entitled_margin, 2),
        "total_actual_net_amount": round(total_actual_net, 2),
        "total_difference": round(total_difference, 2),
        "settlement_type": settlement_type,
        "items": items
    }


@router.post("/{distributor_id}/reconciliations")
async def create_reconciliation(
    distributor_id: str,
    data: ReconciliationCreate,
    current_user: dict = Depends(get_current_user)
):
    """Create and save reconciliation"""
    if not is_distributor_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    tenant_id = get_current_tenant_id()
    
    # First calculate
    calc_result = await calculate_reconciliation(distributor_id, data, current_user)
    
    if calc_result.get('total_deliveries', 0) == 0:
        raise HTTPException(status_code=400, detail="No delivered items found in this period")
    
    now = datetime.now(timezone.utc).isoformat()
    reconciliation_id = str(uuid.uuid4())
    reconciliation_number = await generate_reconciliation_number(tenant_id)
    
    # Get distributor info
    distributor = await db.distributors.find_one({"id": distributor_id})
    
    # Create reconciliation document
    rec_doc = {
        "id": reconciliation_id,
        "tenant_id": tenant_id,
        "reconciliation_number": reconciliation_number,
        "distributor_id": distributor_id,
        "distributor_name": distributor.get('distributor_name') if distributor else None,
        "distributor_code": distributor.get('distributor_code') if distributor else None,
        "period_start": data.period_start,
        "period_end": data.period_end,
        "total_deliveries": calc_result.get('total_deliveries'),
        "total_quantity": calc_result.get('total_quantity'),
        "total_provisional_amount": calc_result.get('total_provisional_amount'),
        "total_actual_gross_amount": calc_result.get('total_actual_gross_amount'),
        "total_entitled_margin": calc_result.get('total_entitled_margin'),
        "total_actual_net_amount": calc_result.get('total_actual_net_amount'),
        "total_difference": calc_result.get('total_difference'),
        "adjustments": 0,
        "final_settlement_amount": calc_result.get('total_difference'),
        "settlement_type": calc_result.get('settlement_type'),
        "status": "draft",
        "remarks": data.remarks,
        "created_at": now,
        "updated_at": now,
        "created_by": current_user.get('id')
    }
    
    await db.distributor_reconciliations.insert_one(rec_doc)
    
    # Create line items
    rec_items = []
    for item in calc_result.get('items', []):
        item_id = str(uuid.uuid4())
        rec_item = {
            "id": item_id,
            "reconciliation_id": reconciliation_id,
            "tenant_id": tenant_id,
            **item
        }
        rec_items.append(rec_item)
    
    if rec_items:
        await db.distributor_reconciliation_items.insert_many(rec_items)
    
    rec_doc.pop('_id', None)
    # Remove _id from each item after insert
    for item in rec_items:
        item.pop('_id', None)
    rec_doc['items'] = rec_items
    
    logger.info(f"Created reconciliation {reconciliation_number} for distributor {distributor_id}")
    
    return rec_doc


@router.get("/{distributor_id}/reconciliations")
async def get_reconciliations(
    distributor_id: str,
    status: Optional[str] = None,
    page: int = 1,
    page_size: int = 20,
    current_user: dict = Depends(get_current_user)
):
    """Get reconciliations for a distributor with delivery items"""
    tenant_id = get_current_tenant_id()
    
    query = {"tenant_id": tenant_id, "distributor_id": distributor_id}
    if status:
        query["status"] = status
    
    total = await db.distributor_reconciliations.count_documents(query)
    recs = await db.distributor_reconciliations.find(
        query, {"_id": 0}
    ).sort("created_at", -1).skip((page - 1) * page_size).limit(page_size).to_list(page_size)
    
    # For each reconciliation, fetch the delivery items with adjustment amounts
    for rec in recs:
        items = await db.distributor_reconciliation_items.find(
            {"reconciliation_id": rec['id'], "tenant_id": tenant_id},
            {"_id": 0}
        ).to_list(500)
        
        # Enrich items with delivery details
        delivery_items = []
        for item in items:
            delivery_id = item.get('delivery_id')
            if delivery_id:
                # Get delivery info
                delivery = await db.distributor_deliveries.find_one(
                    {"id": delivery_id, "tenant_id": tenant_id},
                    {"_id": 0, "delivery_number": 1, "account_name": 1, "account_id": 1}
                )
                if delivery:
                    item['delivery_number'] = delivery.get('delivery_number')
                    item['account_name'] = delivery.get('account_name')
            
            # Calculate adjustment if not already present
            distributor_earnings = item.get('distributor_earnings') or item.get('margin_amount') or 0
            margin_at_transfer = item.get('margin_at_transfer_price') or 0
            item['adjustment_amount'] = item.get('adjustment_amount') or item.get('adjustment_payable') or (distributor_earnings - margin_at_transfer)
            item['total_billing_value'] = item.get('total_billing_value') or item.get('gross_amount') or item.get('actual_net_amount') or 0
            
            delivery_items.append(item)
        
        rec['delivery_items'] = delivery_items
    
    return {
        "reconciliations": recs,
        "total": total,
        "page": page,
        "page_size": page_size
    }


@router.get("/{distributor_id}/reconciliations/{reconciliation_id}")
async def get_reconciliation(
    distributor_id: str,
    reconciliation_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get reconciliation with items"""
    tenant_id = get_current_tenant_id()
    
    rec = await db.distributor_reconciliations.find_one({
        "id": reconciliation_id, "tenant_id": tenant_id, "distributor_id": distributor_id
    }, {"_id": 0})
    
    if not rec:
        raise HTTPException(status_code=404, detail="Reconciliation not found")
    
    items = await db.distributor_reconciliation_items.find({
        "reconciliation_id": reconciliation_id, "tenant_id": tenant_id
    }, {"_id": 0}).to_list(5000)
    
    rec['items'] = items
    return rec


@router.post("/{distributor_id}/reconciliations/{reconciliation_id}/confirm")
async def confirm_reconciliation(
    distributor_id: str,
    reconciliation_id: str,
    adjustments: float = 0,
    current_user: dict = Depends(get_current_user)
):
    """Confirm reconciliation and generate debit/credit note"""
    if not is_distributor_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    tenant_id = get_current_tenant_id()
    now = datetime.now(timezone.utc).isoformat()
    
    rec = await db.distributor_reconciliations.find_one({
        "id": reconciliation_id, "tenant_id": tenant_id, "distributor_id": distributor_id
    })
    
    if not rec:
        raise HTTPException(status_code=404, detail="Reconciliation not found")
    
    if rec.get('status') != 'draft':
        raise HTTPException(status_code=400, detail="Only draft reconciliations can be confirmed")
    
    # Calculate final amount with adjustments
    total_difference = rec.get('total_difference', 0)
    final_amount = total_difference + adjustments
    settlement_type = "debit_note" if final_amount > 0 else "credit_note" if final_amount < 0 else None
    
    # Generate debit/credit note if there's a settlement amount
    note_id = None
    if abs(final_amount) > 0 and settlement_type:
        note_id = str(uuid.uuid4())
        year = datetime.now().year
        prefix = "DN" if settlement_type == "debit_note" else "CN"
        count = await db.distributor_debit_credit_notes.count_documents({
            "tenant_id": tenant_id,
            "note_number": {"$regex": f"^{prefix}-{year}-"}
        })
        note_number = f"{prefix}-{year}-{count + 1:04d}"
        
        distributor = await db.distributors.find_one({"id": distributor_id})
        
        note_doc = {
            "id": note_id,
            "tenant_id": tenant_id,
            "note_number": note_number,
            "note_type": "debit" if settlement_type == "debit_note" else "credit",
            "reconciliation_id": reconciliation_id,
            "reconciliation_number": rec.get('reconciliation_number'),
            "distributor_id": distributor_id,
            "distributor_name": distributor.get('distributor_name') if distributor else None,
            "distributor_code": distributor.get('distributor_code') if distributor else None,
            "amount": abs(round(final_amount, 2)),
            "status": "pending",
            "paid_amount": 0,
            "balance_amount": abs(round(final_amount, 2)),
            "remarks": f"Generated from reconciliation {rec.get('reconciliation_number')}",
            "created_at": now,
            "updated_at": now,
            "created_by": current_user.get('id')
        }
        
        await db.distributor_debit_credit_notes.insert_one(note_doc)
    
    # Update reconciliation
    await db.distributor_reconciliations.update_one(
        {"id": reconciliation_id, "tenant_id": tenant_id},
        {"$set": {
            "status": "confirmed",
            "adjustments": adjustments,
            "final_settlement_amount": round(final_amount, 2),
            "settlement_type": settlement_type,
            "debit_credit_note_id": note_id,
            "confirmed_at": now,
            "confirmed_by": current_user.get('id'),
            "updated_at": now
        }}
    )
    
    return {
        "message": "Reconciliation confirmed",
        "reconciliation_number": rec.get('reconciliation_number'),
        "final_settlement_amount": round(final_amount, 2),
        "settlement_type": settlement_type,
        "note_id": note_id
    }


@router.delete("/{distributor_id}/reconciliations/{reconciliation_id}")
async def delete_reconciliation(
    distributor_id: str,
    reconciliation_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Delete a draft reconciliation"""
    if not is_distributor_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    tenant_id = get_current_tenant_id()
    
    rec = await db.distributor_reconciliations.find_one({
        "id": reconciliation_id, "tenant_id": tenant_id, "distributor_id": distributor_id
    })
    
    if not rec:
        raise HTTPException(status_code=404, detail="Reconciliation not found")
    
    if rec.get('status') != 'draft':
        raise HTTPException(status_code=400, detail="Only draft reconciliations can be deleted")
    
    await db.distributor_reconciliation_items.delete_many({"reconciliation_id": reconciliation_id})
    await db.distributor_reconciliations.delete_one({"id": reconciliation_id, "tenant_id": tenant_id})
    
    return {"message": "Reconciliation deleted"}


# ============ Debit/Credit Notes ============

@router.get("/{distributor_id}/debit-credit-notes")
async def get_debit_credit_notes(
    distributor_id: str,
    note_type: Optional[str] = None,
    status: Optional[str] = None,
    page: int = 1,
    page_size: int = 20,
    current_user: dict = Depends(get_current_user)
):
    """Get debit/credit notes for a distributor"""
    tenant_id = get_current_tenant_id()
    
    query = {"tenant_id": tenant_id, "distributor_id": distributor_id}
    if note_type:
        query["note_type"] = note_type
    if status:
        query["status"] = status
    
    total = await db.distributor_debit_credit_notes.count_documents(query)
    notes = await db.distributor_debit_credit_notes.find(
        query, {"_id": 0}
    ).sort("created_at", -1).skip((page - 1) * page_size).limit(page_size).to_list(page_size)
    
    return {
        "notes": notes,
        "total": total,
        "page": page,
        "page_size": page_size
    }


@router.get("/{distributor_id}/debit-credit-notes/{note_id}")
async def get_debit_credit_note(
    distributor_id: str,
    note_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get debit/credit note details"""
    tenant_id = get_current_tenant_id()
    
    note = await db.distributor_debit_credit_notes.find_one({
        "id": note_id, "tenant_id": tenant_id, "distributor_id": distributor_id
    }, {"_id": 0})
    
    if not note:
        raise HTTPException(status_code=404, detail="Note not found")
    
    # Get reconciliation details if available
    if note.get('reconciliation_id'):
        rec = await db.distributor_reconciliations.find_one({
            "id": note.get('reconciliation_id'), "tenant_id": tenant_id
        }, {"_id": 0})
        note['reconciliation'] = rec
    
    return note


@router.post("/{distributor_id}/debit-credit-notes/{note_id}/record-payment")
async def record_note_payment(
    distributor_id: str,
    note_id: str,
    amount: float,
    payment_reference: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Record payment against debit/credit note"""
    if not is_distributor_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    tenant_id = get_current_tenant_id()
    now = datetime.now(timezone.utc).isoformat()
    
    note = await db.distributor_debit_credit_notes.find_one({
        "id": note_id, "tenant_id": tenant_id, "distributor_id": distributor_id
    })
    
    if not note:
        raise HTTPException(status_code=404, detail="Note not found")
    
    if note.get('status') == 'paid':
        raise HTTPException(status_code=400, detail="Note is already fully paid")
    
    if note.get('status') == 'cancelled':
        raise HTTPException(status_code=400, detail="Note is cancelled")
    
    new_paid_amount = (note.get('paid_amount', 0) or 0) + amount
    new_balance = note.get('amount', 0) - new_paid_amount
    new_status = "paid" if new_balance <= 0 else "partially_paid"
    
    update_data = {
        "paid_amount": round(new_paid_amount, 2),
        "balance_amount": round(max(0, new_balance), 2),
        "status": new_status,
        "updated_at": now
    }
    
    if new_status == "paid":
        update_data["paid_at"] = now
        update_data["paid_by"] = current_user.get('id')
    
    if payment_reference:
        update_data["payment_reference"] = payment_reference
    
    await db.distributor_debit_credit_notes.update_one(
        {"id": note_id, "tenant_id": tenant_id},
        {"$set": update_data}
    )
    
    # If note is paid, update reconciliation status
    if new_status == "paid" and note.get('reconciliation_id'):
        await db.distributor_reconciliations.update_one(
            {"id": note.get('reconciliation_id'), "tenant_id": tenant_id},
            {"$set": {
                "status": "settled",
                "settled_at": now,
                "settled_by": current_user.get('id'),
                "updated_at": now
            }}
        )
    
    return {
        "message": "Payment recorded",
        "paid_amount": round(new_paid_amount, 2),
        "balance_amount": round(max(0, new_balance), 2),
        "status": new_status
    }


@router.delete("/{distributor_id}/notes/{note_id}")
async def delete_debit_credit_note(
    distributor_id: str,
    note_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Delete a debit/credit note - CEO/Admin only"""
    tenant_id = get_current_tenant_id()
    
    # Check role - only CEO and Admin can delete notes
    user_role = current_user.get('role', '')
    if user_role not in ['CEO', 'Admin', 'System Admin']:
        raise HTTPException(
            status_code=403, 
            detail="Only CEO and Admin can delete debit/credit notes"
        )
    
    # Find the note
    note = await db.distributor_debit_credit_notes.find_one({
        "id": note_id,
        "distributor_id": distributor_id,
        "tenant_id": tenant_id
    })
    
    if not note:
        raise HTTPException(status_code=404, detail="Note not found")
    
    # Revert linked settlements back to unreconciled (for draft/pending notes)
    note_status = note.get('status', 'draft')
    if note_status in ['draft', 'pending']:
        settlement_ids = note.get('settlement_ids', [])
        if settlement_ids:
            await db.distributor_settlements.update_many(
                {"id": {"$in": settlement_ids}, "tenant_id": tenant_id},
                {"$set": {"reconciled": False}, "$unset": {"note_id": "", "note_number": ""}}
            )
            logger.info(f"Reverted {len(settlement_ids)} settlement(s) to unreconciled after deleting draft note {note.get('note_number')}")
    
    # Delete the note
    await db.distributor_debit_credit_notes.delete_one({
        "id": note_id,
        "tenant_id": tenant_id
    })
    
    # If note was linked to a reconciliation, update the reconciliation status back to draft
    if note.get('reconciliation_id'):
        await db.distributor_reconciliations.update_one(
            {"id": note.get('reconciliation_id'), "tenant_id": tenant_id},
            {"$set": {
                "status": "draft",
                "settlement_type": None,
                "updated_at": datetime.now(timezone.utc)
            }}
        )
    
    return {"message": "Note deleted successfully", "note_number": note.get('note_number')}


# ============ Stock Dashboard ============

@router.get("/{distributor_id}/stock-dashboard")
async def get_stock_dashboard(
    distributor_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Real-time stock dashboard with complete inventory picture per SKU"""
    tenant_id = get_current_tenant_id()
    
    distributor = await db.distributors.find_one(
        {"id": distributor_id, "tenant_id": tenant_id},
        {"_id": 0, "id": 1, "distributor_name": 1}
    )
    if not distributor:
        raise HTTPException(status_code=404, detail="Distributor not found")

    tdb = get_tenant_db()

    # ── Units helper ────────────────────────────────────────────────────────
    # The dashboard displays everything in CRATES. Some data sources store
    # bottles, some store crates, and master_skus often lacks bpc. We solve
    # this by:
    #   1. Reading the per-item `packaging_units` field when present (delivery
    #      / shipment items always store it).
    #   2. Falling back to `bpc_by_sku` (seeded from master_skus +
    #      factory_warehouse_stock rows below) for sources that don't store
    #      packaging_units (returns).
    # ── UNIT MODEL (updated per user requirement, Jul 2026) ─────────────────
    # The dashboard reports every quantity in the SKU's base UOM = BOTTLES
    # (the lowest unit defined in SKU Management). A line's bottle count is
    # `quantity × packaging_units` (the packaging chosen on that line). When no
    # packaging was chosen (`packaging_units` unset/≤0) the quantity is already
    # in UOM, so it is used as-is. Returns / factory-warehouse rows are already
    # stored in bottles, so no conversion is applied to them.
    bpc_by_sku: dict = {}

    def _to_crates(sku_id: str, bottles_qty, row_bpc=None) -> int:
        # Returns / factory-warehouse quantities are already stored in BOTTLES
        # (the base UOM). Report them as-is — no bottles→crates division.
        try:
            return int(round(float(bottles_qty or 0)))
        except Exception:
            return 0

    def _bottle_to_crates(item: dict) -> int:
        """Item `quantity` is in bottles → returned in bottles (base UOM)."""
        try:
            return int(item.get('quantity') or 0)
        except Exception:
            return 0

    def _item_crates(item: dict) -> int:
        """Return a line item's contribution in the base UOM (bottles).

        The reliable ground truth for base units is `packages × packaging_units`
        (crates × bottles-per-crate) whenever BOTH are known — this is immune to
        the two historical `quantity` storage inconsistencies that coexist:
          • Some rows stored `quantity` as the base-unit total already
            (quantity == packages × packaging_units) — e.g. a 10-crate line as
            quantity=150 with packaging_units=15. Multiplying quantity × pu here
            would double-count (150 → 2250).
          • Other rows stored `quantity` as the number of packages/crates while
            still carrying `packages`/`packaging_units` — e.g. a 3-crate line as
            quantity=3 with packaging_units=12. Returning quantity as-is would
            under-count (36 bottles → 3).
        Computing packages × packaging_units yields the correct base total in
        BOTH cases (10×15=150 and 3×12=36).

        Fallbacks: when `packages` is absent but packaging_units > 1, `quantity`
        is the package count (OLD model → quantity × packaging_units). Plain
        lines (loose base units, pu ≤ 1) store the base quantity directly."""
        qty = int(item.get('quantity') or 0)
        pu = int(item.get('packaging_units') or 0)
        pk = item.get('packages')
        try:
            pk = int(pk) if pk not in (None, "") else 0
        except (TypeError, ValueError):
            pk = 0
        if pk > 0 and pu > 1:
            return pk * pu        # crates × bottles-per-crate = base bottles (authoritative)
        if pk > 0:
            return qty            # loose units (pu ≤ 1): quantity is already base units
        if pu > 1:
            return qty * pu       # OLD model: quantity is number of packages
        return qty

    # === 1. STOCK IN: Shipments received (only delivered shipments) ===
    delivered_shipment_ids = await db.distributor_shipments.distinct(
        "id",
        {"tenant_id": tenant_id, "distributor_id": distributor_id, "status": "delivered"}
    )
    shipment_items = await db.distributor_shipment_items.find(
        {"tenant_id": tenant_id, "shipment_id": {"$in": delivered_shipment_ids}},
        {"_id": 0, "sku_id": 1, "sku_name": 1, "quantity": 1, "packaging_units": 1, "packages": 1}
    ).to_list(50000)
    
    stock_in_by_sku = {}
    for si in shipment_items:
        sid = si.get('sku_id', '')
        if sid not in stock_in_by_sku:
            stock_in_by_sku[sid] = {"sku_id": sid, "sku_name": si.get('sku_name', 'Unknown'), "qty": 0}
        stock_in_by_sku[sid]["qty"] += _item_crates(si)

    # === 1b. FACTORY WAREHOUSE STOCK (from production transfers) ===
    # Get factory warehouse IDs belonging to this distributor
    factory_location_ids = await db.distributor_locations.distinct(
        "id",
        {"tenant_id": tenant_id, "distributor_id": distributor_id, "is_factory": True, "status": "active"}
    )
    factory_wh_stock_by_sku = {}
    factory_wh_by_location = {}  # location_id -> {location_name, skus: [...]}
    total_factory_wh_stock = 0

    if factory_location_ids:
        fws_docs = await tdb.factory_warehouse_stock.find(
            {"tenant_id": tenant_id, "warehouse_location_id": {"$in": factory_location_ids}},
            {"_id": 0}
        ).to_list(5000)

        # Build a sku_id → bottles_per_crate map so we can convert bottles
        # (stored on the row) → crates (what the dashboard displays).
        # We prefer the value stored on the row itself (set at transfer time)
        # but fall back to master_skus.packaging_config / unit so legacy rows
        # still convert correctly.
        sku_ids_needed = list({fws.get("sku_id", "") for fws in fws_docs if fws.get("sku_id")})
        if sku_ids_needed:
            async for ms in tdb.master_skus.find(
                {"id": {"$in": sku_ids_needed}},
                {"_id": 0, "id": 1, "packaging_config": 1, "unit": 1, "bottles_per_crate": 1}
            ):
                # Order of preference: explicit bottles_per_crate field →
                # production packaging default 'units' → leave as None.
                bpc = ms.get("bottles_per_crate")
                if not bpc:
                    pc = (ms.get("packaging_config") or {}).get("production") or []
                    default = next((p for p in pc if p.get("is_default")), None)
                    if default and default.get("units"):
                        bpc = default["units"]
                # Only set when we have a real bpc — never poison the lookup
                # with the "or 1" fallback (which would override more accurate
                # row-level / item-level seeds done elsewhere).
                if bpc:
                    bpc_by_sku.setdefault(ms["id"], bpc)

        for fws in fws_docs:
            sid = fws.get("sku_id", "")
            bottles_qty = fws.get("quantity", 0)
            if bottles_qty <= 0:
                continue
            # Persist the row-level bpc back into the shared lookup so all the
            # downstream bottles→crates conversions (deliveries, returns,
            # pending out) agree with the factory_wh row's truth. Master SKU
            # often lacks `bottles_per_crate`, leaving the lookup at 1 (which
            # was the silent unit-mismatch bug on the dashboard).
            if fws.get("bottles_per_crate") and not bpc_by_sku.get(sid):
                bpc_by_sku[sid] = fws["bottles_per_crate"]
            crates_qty = _to_crates(sid, bottles_qty, fws.get("bottles_per_crate"))
            total_factory_wh_stock += crates_qty
            if sid not in factory_wh_stock_by_sku:
                factory_wh_stock_by_sku[sid] = {"sku_id": sid, "sku_name": fws.get("sku_name", "Unknown"), "qty": 0}
            factory_wh_stock_by_sku[sid]["qty"] += crates_qty
            # Per-location breakdown — aggregate per SKU so a SKU with
            # multiple `factory_warehouse_stock` rows (one per batch) shows
            # as a single line with the summed crate quantity instead of
            # one row per batch. We also keep the per-batch breakdown
            # under `batches` so the UI can disclose it on demand.
            wh_id = fws.get("warehouse_location_id", "")
            if wh_id not in factory_wh_by_location:
                factory_wh_by_location[wh_id] = {
                    "warehouse_name": fws.get("warehouse_name", ""),
                    "skus_by_id": {},
                }
            bucket = factory_wh_by_location[wh_id]["skus_by_id"].setdefault(
                sid, {"sku_id": sid, "sku_name": fws.get("sku_name", ""), "quantity": 0, "batches": []}
            )
            bucket["quantity"] += crates_qty
            bucket["batches"].append({
                "batch_id": fws.get("batch_id"),
                "batch_code": fws.get("batch_code") or "(legacy)",
                "quantity": crates_qty,
                "production_date": fws.get("production_date"),
                "received_at": fws.get("created_at"),
            })

    # Build a sku_id → list-of-batches (with warehouse_name) lookup so the
    # main Stock-by-SKU table can expose a per-batch disclosure on every row.
    fwh_batches_by_sku = {}
    for wh_id, wh_data in factory_wh_by_location.items():
        wh_name = wh_data.get("warehouse_name", "")
        for sid_key, bucket in wh_data["skus_by_id"].items():
            entry = fwh_batches_by_sku.setdefault(sid_key, [])
            for b in bucket.get("batches", []):
                entry.append({**b, "warehouse_id": wh_id, "warehouse_name": wh_name})
    # FIFO per SKU — oldest first
    for sid_key in fwh_batches_by_sku:
        fwh_batches_by_sku[sid_key].sort(
            key=lambda b: (b.get("production_date") or b.get("received_at") or "9999")
        )
    
    # === 1c. STOCK TRANSFERS in/out of this distributor's NON-FACTORY warehouses ===
    # Transfers touching a *factory* warehouse are already reflected in the live
    # factory_warehouse_stock balance above. Transfers into/out of a regular
    # (non-factory) distributor warehouse are stored only in `distributor_stock`,
    # which this dashboard does not read — so we fold them in here as explicit
    # in/out flows. `item.quantity` is in packages/crates (same unit the rest of
    # the dashboard uses), so no bottle→crate conversion is needed.
    transfer_in_by_sku = {}
    transfer_out_by_sku = {}
    xfers = await db.distributor_stock_transfers.find(
        {
            "tenant_id": tenant_id, "status": "completed",
            "$or": [{"source_distributor_id": distributor_id}, {"dest_distributor_id": distributor_id}],
        },
        {"_id": 0, "source_distributor_id": 1, "source_location_id": 1,
         "dest_distributor_id": 1, "dest_location_id": 1, "items": 1},
    ).to_list(5000)
    if xfers:
        # Resolve is_factory for every involved location in one batch query.
        _loc_ids = set()
        for xf in xfers:
            if xf.get("source_location_id"):
                _loc_ids.add(xf["source_location_id"])
            if xf.get("dest_location_id"):
                _loc_ids.add(xf["dest_location_id"])
        _factory_loc = set()
        async for _ld in db.distributor_locations.find(
            {"tenant_id": tenant_id, "id": {"$in": list(_loc_ids)}, "is_factory": True},
            {"_id": 0, "id": 1},
        ):
            _factory_loc.add(_ld["id"])
        # Fallback bottles-per-package for transfer items missing `units_per_package`
        # (older records): use the SKU's default packaging units.
        _xfer_sids = {it.get("sku_id") for xf in xfers for it in (xf.get("items") or []) if it.get("sku_id")}
        _xfer_default_upp = {}
        if _xfer_sids:
            async for ms in tdb.master_skus.find(
                {"id": {"$in": list(_xfer_sids)}},
                {"_id": 0, "id": 1, "packaging_config": 1, "bottles_per_crate": 1},
            ):
                units = ms.get("bottles_per_crate")
                if not units:
                    pc = ms.get("packaging_config") or {}
                    for key in ("stock_out", "production", "master"):
                        arr = pc.get(key) or []
                        # Prefer the default packaging (with >1 units), else any >1.
                        d = next((p for p in arr if p.get("is_default") and int(p.get("units_per_package") or 1) > 1), None) \
                            or next((p for p in arr if int(p.get("units_per_package") or 1) > 1), None)
                        if d:
                            units = int(d.get("units_per_package"))
                            break
                if units:
                    _xfer_default_upp[ms["id"]] = int(units)
        for xf in xfers:
            for it in (xf.get("items") or []):
                sid = it.get("sku_id", "")
                if not sid:
                    continue
                # Base UOM = bottles: quantity(packages) × units_per_package.
                qty = int(it.get("quantity", 0) or 0)
                upp = int(it.get("units_per_package") or 0) or _xfer_default_upp.get(sid) or 1
                bottles = qty * upp
                if bottles <= 0:
                    continue
                sku_name = it.get("sku_name", "Unknown")
                # Inbound: destination is THIS distributor & dest warehouse is non-factory
                if xf.get("dest_distributor_id") == distributor_id and xf.get("dest_location_id") not in _factory_loc:
                    b = transfer_in_by_sku.setdefault(sid, {"sku_id": sid, "sku_name": sku_name, "qty": 0})
                    b["qty"] += bottles
                # Outbound: source is THIS distributor & source warehouse is non-factory
                if xf.get("source_distributor_id") == distributor_id and xf.get("source_location_id") not in _factory_loc:
                    b = transfer_out_by_sku.setdefault(sid, {"sku_id": sid, "sku_name": sku_name, "qty": 0})
                    b["qty"] += bottles


    # === 2. STOCK OUT: Deliveries to customers (delivered/completed/complete) ===
    delivered_delivery_ids = await db.distributor_deliveries.distinct(
        "id",
        {"tenant_id": tenant_id, "distributor_id": distributor_id, "status": {"$in": ["delivered", "completed", "complete"]}}
    )
    delivery_items = await db.distributor_delivery_items.find(
        {"tenant_id": tenant_id, "delivery_id": {"$in": delivered_delivery_ids}},
        {"_id": 0, "sku_id": 1, "sku_name": 1, "quantity": 1, "packaging_units": 1, "packages": 1}
    ).to_list(50000)
    
    stock_out_by_sku = {}
    for di in delivery_items:
        sid = di.get('sku_id', '')
        if sid not in stock_out_by_sku:
            stock_out_by_sku[sid] = {"sku_id": sid, "sku_name": di.get('sku_name', 'Unknown'), "qty": 0}
        stock_out_by_sku[sid]["qty"] += _item_crates(di)

    # === 2b. RESERVED — committed to any OPEN Stock Out order, not yet delivered ===
    # Every open order (draft → in-transit) reserves stock so it can't be
    # re-allocated to another customer. On-hand still includes these units;
    # Available = On-hand − Reserved. Excludes terminal statuses
    # (delivered / cancelled).
    pending_delivery_ids = await db.distributor_deliveries.distinct(
        "id",
        {
            "tenant_id": tenant_id, "distributor_id": distributor_id,
            "status": {"$in": RESERVED_DELIVERY_STATUSES},
        },
    )
    pending_items = await db.distributor_delivery_items.find(
        {"tenant_id": tenant_id, "delivery_id": {"$in": pending_delivery_ids}},
        {"_id": 0, "sku_id": 1, "sku_name": 1, "quantity": 1, "packaging_units": 1, "packages": 1}
    ).to_list(50000)
    stock_pending_out_by_sku = {}
    for di in pending_items:
        sid = di.get('sku_id', '')
        if sid not in stock_pending_out_by_sku:
            stock_pending_out_by_sku[sid] = {"sku_id": sid, "sku_name": di.get('sku_name', 'Unknown'), "qty": 0}
        stock_pending_out_by_sku[sid]["qty"] += _item_crates(di)
    
    # Weekly delivery data (last 12 weeks) for average calculation
    from datetime import timedelta
    import calendar as cal_mod
    now = datetime.now(timezone.utc)
    twelve_weeks_ago = (now - timedelta(weeks=12)).strftime('%Y-%m-%d')
    
    recent_deliveries = await db.distributor_deliveries.find(
        {
            "tenant_id": tenant_id,
            "distributor_id": distributor_id,
            "status": {"$in": ["delivered", "completed", "complete"]},
            "delivery_date": {"$gte": twelve_weeks_ago}
        },
        {"_id": 0, "id": 1, "delivery_date": 1}
    ).to_list(10000)
    recent_delivery_ids = [d['id'] for d in recent_deliveries]
    
    recent_items = await db.distributor_delivery_items.find(
        {"tenant_id": tenant_id, "delivery_id": {"$in": recent_delivery_ids}},
        {"_id": 0, "sku_id": 1, "quantity": 1, "packaging_units": 1, "packages": 1, "delivery_id": 1}
    ).to_list(50000)
    
    # Map delivery_id -> delivery_date for weekly grouping
    del_date_map = {d['id']: d.get('delivery_date', '') for d in recent_deliveries}
    weekly_by_sku = {}
    for ri in recent_items:
        sid = ri.get('sku_id', '')
        dd = del_date_map.get(ri.get('delivery_id', ''), '')
        if not dd:
            continue
        # Get ISO week number
        try:
            dt = datetime.strptime(dd, '%Y-%m-%d')
            week_key = dt.strftime('%Y-W%W')
        except Exception:
            continue
        if sid not in weekly_by_sku:
            weekly_by_sku[sid] = {}
        if week_key not in weekly_by_sku[sid]:
            weekly_by_sku[sid][week_key] = 0
        weekly_by_sku[sid][week_key] += ri.get('quantity', 0)
    
    # === 3. CUSTOMER RETURNS (by category) ===
    customer_returns = await db.customer_returns.find(
        {
            "tenant_id": tenant_id,
            "distributor_id": distributor_id,
            "status": {"$nin": ["cancelled", "draft"]}
        },
        {"_id": 0, "items": 1, "status": 1}
    ).to_list(10000)
    
    cust_return_by_sku = {}  # {sku_id: {empty: X, damaged: X, expired: X, total: X}}
    for cr in customer_returns:
        for item in cr.get('items', []):
            sid = item.get('sku_id', '')
            qty = item.get('quantity', 0)
            cat = item.get('reason_category', 'other')
            if sid not in cust_return_by_sku:
                cust_return_by_sku[sid] = {"empty_reusable": 0, "damaged": 0, "expired": 0, "promotional": 0, "other": 0, "total": 0, "pending_factory": 0, "returned_to_factory": 0}
            bucket = cat if cat in cust_return_by_sku[sid] else "other"
            cust_return_by_sku[sid][bucket] += qty
            cust_return_by_sku[sid]["total"] += qty
            if item.get('return_to_factory'):
                if item.get('returned_to_factory'):
                    cust_return_by_sku[sid]["returned_to_factory"] += qty
                else:
                    cust_return_by_sku[sid]["pending_factory"] += qty
    
    # === 4. FACTORY RETURNS (distributor -> factory) ===
    factory_returns = await db.distributor_factory_returns.find(
        {
            "tenant_id": tenant_id,
            "distributor_id": distributor_id,
            "status": {"$nin": ["cancelled", "draft"]}
        },
        {"_id": 0, "items": 1, "reason": 1, "status": 1}
    ).to_list(10000)
    
    factory_return_by_sku = {}  # {sku_id: {qty, reason_breakdown}}
    for fr in factory_returns:
        reason = fr.get('reason', 'other')
        for item in fr.get('items', []):
            sid = item.get('sku_id', '')
            qty = item.get('quantity', 0)
            if sid not in factory_return_by_sku:
                factory_return_by_sku[sid] = {"total": 0, "empty_reusable": 0, "damaged": 0, "expired": 0, "other": 0, "sku_name": item.get('sku_name', 'Unknown')}
            bucket = reason if reason in factory_return_by_sku[sid] else "other"
            factory_return_by_sku[sid][bucket] += qty
            factory_return_by_sku[sid]["total"] += qty
    
    # === BUILD PER-SKU SUMMARY ===
    all_sku_ids = set(
        list(stock_in_by_sku.keys()) +
        list(stock_out_by_sku.keys()) +
        list(stock_pending_out_by_sku.keys()) +
        list(cust_return_by_sku.keys()) +
        list(factory_return_by_sku.keys()) +
        list(factory_wh_stock_by_sku.keys()) +
        list(transfer_in_by_sku.keys()) +
        list(transfer_out_by_sku.keys())
    )

    # bpc lookup for every SKU we touch (not just factory-warehouse ones) so
    # we can convert shipments / deliveries / returns from BOTTLES → CRATES
    # consistently with factory_warehouse_stock (which is already in crates).
    # Without this, the at-hand math mixed crates + bottles and looked wrong.
    extra_sku_ids = [sid for sid in all_sku_ids if sid and sid not in bpc_by_sku]
    if extra_sku_ids:
        async for ms in tdb.master_skus.find(
            {"id": {"$in": extra_sku_ids}},
            {"_id": 0, "id": 1, "packaging_config": 1, "unit": 1, "bottles_per_crate": 1},
        ):
            bpc = ms.get("bottles_per_crate")
            if not bpc:
                pc = (ms.get("packaging_config") or {}).get("production") or []
                default = next((p for p in pc if p.get("is_default")), None)
                if default and default.get("units"):
                    bpc = default["units"]
            if bpc:
                bpc_by_sku.setdefault(ms["id"], bpc)

    # Per-SKU base UOM + default packaging (for the "<base unit> (= N default-pkg)"
    # display). `packaging_config[*]` entries carry `units_per_package` and
    # `packaging_type_name`. The base unit is the packaging with units==1 (e.g.
    # "Bottle", "Can", "Piece") — tenant-specific, never hardcoded.
    default_pkg_by_sku: dict = {}
    base_unit_by_sku: dict = {}
    sku_ids_for_pkg = [sid for sid in all_sku_ids if sid]
    if sku_ids_for_pkg:
        async for ms in tdb.master_skus.find(
            {"id": {"$in": sku_ids_for_pkg}},
            {"_id": 0, "id": 1, "unit": 1, "base_uom": 1, "packaging_config": 1},
        ):
            pc = ms.get("packaging_config") or {}
            all_pkgs = []
            for key in ("stock_out", "production", "master", "stock_in"):
                all_pkgs.extend(pc.get(key) or [])
            # Base unit from SKU Management (`base_uom`, e.g. "Bottle"/"Can"/"Piece"),
            # falling back to a single-unit packaging name, then generic "units".
            base = next((p.get("packaging_type_name") for p in all_pkgs
                         if int(p.get("units_per_package") or 0) == 1 and p.get("packaging_type_name")), None)
            base_unit_by_sku[ms["id"]] = ms.get("base_uom") or base or "units"
            chosen = None
            for key in ("stock_out", "production", "master", "stock_in"):
                arr = pc.get(key) or []
                d = next((p for p in arr if p.get("is_default")), None)
                if d:
                    chosen = d
                    break
            if not chosen or int(chosen.get("units_per_package") or 1) <= 1:
                for key in ("stock_out", "production", "master"):
                    arr = pc.get(key) or []
                    d = next((p for p in arr if int(p.get("units_per_package") or 1) > 1), None)
                    if d:
                        chosen = d
                        break
            if chosen:
                default_pkg_by_sku[ms["id"]] = {
                    "units": int(chosen.get("units_per_package") or 1),
                    "name": chosen.get("packaging_type_name") or "",
                }

    sku_summaries = []
    total_stock_in = 0
    total_stock_out = 0
    total_stock_pending_out = 0
    total_at_hand = 0
    total_cust_returns = 0
    total_empty_bottles_returned = 0
    total_product_returns = 0
    total_factory_returns = 0
    total_transferred_out = 0

    for sid in all_sku_ids:
        si = stock_in_by_sku.get(sid, {})
        so = stock_out_by_sku.get(sid, {})
        po = stock_pending_out_by_sku.get(sid, {})
        cr_data = cust_return_by_sku.get(sid, {})
        fr_data = factory_return_by_sku.get(sid, {})
        fws_data = factory_wh_stock_by_sku.get(sid, {})
        tin_data = transfer_in_by_sku.get(sid, {})
        tout_data = transfer_out_by_sku.get(sid, {})

        # Bottle-level inputs → convert to crates. Shipment/delivery/pending
        # were already converted at the aggregation step (using each item's
        # `packaging_units`). Returns and factory-warehouse rows weren't,
        # so do them now.
        qty_in_shipments  = si.get('qty', 0)                            # already crates
        qty_out           = so.get('qty', 0)                            # already crates
        qty_pending_out   = po.get('qty', 0)                            # already crates
        qty_cust_returned = _to_crates(sid, cr_data.get('total', 0))
        qty_factory_returned = _to_crates(sid, fr_data.get('total', 0))
        qty_factory_wh    = fws_data.get('qty', 0)                       # already crates
        qty_transferred_in  = tin_data.get('qty', 0)                     # already crates
        qty_transferred_out = tout_data.get('qty', 0)                    # already crates

        # `stock_received` = everything the distributor has been *given*:
        # primary stock-in shipments + production-to-warehouse transfers +
        # stock moved IN via a stock transfer to a non-factory warehouse.
        # Returns are NEVER part of received — they're a separate flow back
        # to the factory and not deliverable (used / damaged / expired bottles).
        qty_in = qty_in_shipments + qty_factory_wh + qty_transferred_in

        # NEW formula — what's actually deliverable RIGHT NOW:
        #
        #   stock_at_hand = received
        #                   − delivered_to_customers
        #                   − scheduled_or_in_transit (committed but not yet out)
        #                   − sent_back_to_factory
        #                   − transferred_out (moved to another warehouse)
        #
        # No `+ customer_returns` add-back: returned bottles can't be re-sold
        # without going through factory QC first. The user explicitly flagged
        # this on May 27 — over-deliveries would otherwise be possible.
        stock_at_hand = qty_in - qty_out - qty_pending_out - qty_factory_returned - qty_transferred_out

        # Explicit reserve→deliver model (derived):
        #   on_hand  = physical balance still held (incl. reserved units)
        #   reserved = committed to open Stock Out orders
        #   available= on_hand − reserved  (== stock_at_hand, what's deliverable now)
        #   delivered= permanently consumed/delivered out
        stock_on_hand = qty_in - qty_out - qty_factory_returned - qty_transferred_out
        stock_reserved = qty_pending_out
        stock_available = stock_at_hand

        # Weekly average (last 12 weeks)
        weekly_data = weekly_by_sku.get(sid, {})
        weeks_with_data = len(weekly_data)
        total_recent_qty = sum(weekly_data.values())
        weekly_avg = round(total_recent_qty / max(weeks_with_data, 1), 1) if weeks_with_data > 0 else 0

        # % stock at hand — measured against received (so 100% means nothing
        # has been committed yet).
        pct_at_hand = round((stock_at_hand / qty_in * 100), 1) if qty_in > 0 else 0

        # Days of stock remaining
        daily_avg = weekly_avg / 7 if weekly_avg > 0 else 0
        days_remaining = round(stock_at_hand / daily_avg, 0) if daily_avg > 0 and stock_at_hand > 0 else None

        sku_name = si.get('sku_name') or so.get('sku_name') or po.get('sku_name') or fr_data.get('sku_name') or fws_data.get('sku_name') or tin_data.get('sku_name') or tout_data.get('sku_name') or 'Unknown'

        dp = default_pkg_by_sku.get(sid) or {}
        dp_units = dp.get('units') or 0

        sku_summaries.append({
            "sku_id": sid,
            "sku_name": sku_name,
            # All quantities below are in the SKU's base UOM (units==1 packaging).
            "unit": "units",
            "base_unit_name": base_unit_by_sku.get(sid) or "units",
            "default_packaging_name": dp.get('name') or "",
            "default_packaging_units": dp_units,
            "stock_received": qty_in,
            "stock_delivered": qty_out,
            "stock_pending_out": qty_pending_out,
            "customer_returns": qty_cust_returned,
            "customer_returns_breakdown": {
                "empty_reusable": _to_crates(sid, cr_data.get('empty_reusable', 0)),
                "damaged": _to_crates(sid, cr_data.get('damaged', 0)),
                "expired": _to_crates(sid, cr_data.get('expired', 0)),
                "promotional": _to_crates(sid, cr_data.get('promotional', 0)),
            },
            # Grouped view: Empty/Reusable + Promotional(FOC) are empty USED bottles
            # cycling back for recycling (a returnable asset, NOT lost product);
            # Damaged + Expired are genuinely unsellable finished goods.
            #
            # ⚠️ UNITS: `empty_bottles_returned` stays in **raw bottles** —
            # they are returned individually for cleaning/refill and almost
            # never in clean crate multiples. Every other column on this
            # dashboard is in **crates**.
            "empty_bottles_returned": (
                int(cr_data.get('empty_reusable', 0))
                + int(cr_data.get('promotional', 0))
            ),
            "empty_bottles_returned_unit": "bottles",
            "product_returns": (
                _to_crates(sid, cr_data.get('damaged', 0))
                + _to_crates(sid, cr_data.get('expired', 0))
            ),
            "factory_returns": qty_factory_returned,
            "factory_returns_breakdown": {
                "empty_reusable": _to_crates(sid, fr_data.get('empty_reusable', 0)),
                "damaged": _to_crates(sid, fr_data.get('damaged', 0)),
                "expired": _to_crates(sid, fr_data.get('expired', 0)),
            },
            "pending_factory_return": _to_crates(sid, cr_data.get('pending_factory', 0)),
            "factory_warehouse_stock": qty_factory_wh,
            # Per-batch breakdown of the factory warehouse stock for this SKU
            # so the Stock-by-SKU table can expose a batch-level disclosure on
            # every row. Empty list when this SKU has no factory stock.
            "factory_warehouse_batches": fwh_batches_by_sku.get(sid, []),
            "stock_at_hand": stock_at_hand,
            "stock_on_hand": stock_on_hand,
            "stock_reserved": stock_reserved,
            "stock_available": stock_available,
            "pct_stock_at_hand": pct_at_hand,
            "weekly_avg_deliveries": weekly_avg,
            "days_of_stock": days_remaining,
            "weeks_analyzed": weeks_with_data,
        })

    # Sort by stock_at_hand descending
    sku_summaries.sort(key=lambda x: x['stock_at_hand'], reverse=True)

    for s in sku_summaries:
        total_stock_in += s['stock_received']
        total_stock_out += s['stock_delivered']
        total_stock_pending_out += s['stock_pending_out']
        total_at_hand += s['stock_at_hand']
        total_cust_returns += s['customer_returns']
        total_empty_bottles_returned += s['empty_bottles_returned']
        total_product_returns += s['product_returns']
        total_factory_returns += s['factory_returns']
        total_transferred_out += transfer_out_by_sku.get(s['sku_id'], {}).get('qty', 0)
    
    # Aggregate bottle tracking — these are in BOTTLES (kept as-is so the
    # bottle counters in the UI reflect the actual number of physical bottles,
    # not crates). Don't confuse with the per-SKU breakdown above which is
    # crates for consistency with stock_at_hand math.
    total_empty = sum(cr_data.get('empty_reusable', 0) for cr_data in cust_return_by_sku.values())
    total_damaged = sum(cr_data.get('damaged', 0) for cr_data in cust_return_by_sku.values())
    total_expired = sum(cr_data.get('expired', 0) for cr_data in cust_return_by_sku.values())
    total_promotional = sum(cr_data.get('promotional', 0) for cr_data in cust_return_by_sku.values())
    total_pending_factory = sum(cr_data.get('pending_factory', 0) for cr_data in cust_return_by_sku.values())
    
    return {
        "distributor_id": distributor_id,
        "distributor_name": distributor.get('distributor_name', ''),
        "unit": "units",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "totals": {
            "stock_received": total_stock_in,
            "stock_delivered": total_stock_out,
            "stock_pending_out": total_stock_pending_out,
            "stock_reserved": total_stock_pending_out,
            "stock_on_hand": total_stock_in - total_stock_out - total_factory_returns - total_transferred_out,
            "stock_available": total_at_hand,
            "stock_at_hand": total_at_hand,
            "customer_returns": total_cust_returns,
            "empty_bottles_returned": total_empty_bottles_returned,
            "product_returns": total_product_returns,
            "factory_returns": total_factory_returns,
            "factory_warehouse_stock": total_factory_wh_stock,
            "pct_stock_at_hand": round((total_at_hand / total_stock_in * 100), 1) if total_stock_in > 0 else 0,
        },
        "bottle_tracking": {
            "empty_reusable": total_empty,
            "promotional": total_promotional,
            "damaged": total_damaged,
            "expired": total_expired,
            "pending_factory_return": total_pending_factory,
        },
        "factory_warehouses": [
            {
                "warehouse_id": wh_id,
                "warehouse_name": wh_data["warehouse_name"],
                # Convert the aggregated dict back to a list, sorted by name
                # for stable, alphabetical display in the UI. Each entry now
                # also carries a `batches` array (FIFO sorted) used by the
                # "N batches" disclosure in the dashboard card.
                "skus": sorted(
                    [
                        {
                            **s,
                            "batches": sorted(
                                s.get("batches", []),
                                key=lambda b: (b.get("production_date") or b.get("received_at") or "9999"),
                            ),
                        }
                        for s in wh_data["skus_by_id"].values()
                    ],
                    key=lambda s: (s.get("sku_name") or "").lower(),
                ),
            }
            for wh_id, wh_data in factory_wh_by_location.items()
        ],
        "sku_count": len(sku_summaries),
        "skus": sku_summaries,
    }



# ============ Real-time Reconciliation Status ============

@router.get("/{distributor_id}/billing/summary")
async def get_billing_summary(
    distributor_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get real-time billing summary for a distributor"""
    tenant_id = get_current_tenant_id()
    
    # Get distributor info
    distributor = await db.distributors.find_one(
        {"id": distributor_id, "tenant_id": tenant_id},
        {"_id": 0, "distributor_name": 1, "distributor_code": 1}
    )
    
    if not distributor:
        raise HTTPException(status_code=404, detail="Distributor not found")
    
    # Count billing configs
    config_count = await db.distributor_billing_config.count_documents({
        "tenant_id": tenant_id, "distributor_id": distributor_id, "status": "active"
    })
    
    # Provisional invoices summary
    invoice_pipeline = [
        {"$match": {"tenant_id": tenant_id, "distributor_id": distributor_id}},
        {"$group": {
            "_id": "$status",
            "count": {"$sum": 1},
            "total_amount": {"$sum": "$total_net_amount"}
        }}
    ]
    invoice_stats = await db.distributor_provisional_invoices.aggregate(invoice_pipeline).to_list(10)
    invoice_summary = {s['_id']: {"count": s['count'], "amount": s['total_amount']} for s in invoice_stats}
    
    # Reconciliation summary
    rec_pipeline = [
        {"$match": {"tenant_id": tenant_id, "distributor_id": distributor_id}},
        {"$group": {
            "_id": "$status",
            "count": {"$sum": 1},
            "total_difference": {"$sum": "$total_difference"}
        }}
    ]
    rec_stats = await db.distributor_reconciliations.aggregate(rec_pipeline).to_list(10)
    rec_summary = {s['_id']: {"count": s['count'], "total_difference": s['total_difference']} for s in rec_stats}
    
    # Debit/Credit notes summary
    note_pipeline = [
        {"$match": {"tenant_id": tenant_id, "distributor_id": distributor_id}},
        {"$group": {
            "_id": {"type": "$note_type", "status": "$status"},
            "count": {"$sum": 1},
            "total_amount": {"$sum": "$amount"},
            "total_balance": {"$sum": "$balance_amount"}
        }}
    ]
    note_stats = await db.distributor_debit_credit_notes.aggregate(note_pipeline).to_list(20)
    
    pending_debit = sum(s['total_balance'] for s in note_stats 
                       if s['_id']['type'] == 'debit' and s['_id']['status'] in ['pending', 'partially_paid'])
    pending_credit = sum(s['total_balance'] for s in note_stats 
                        if s['_id']['type'] == 'credit' and s['_id']['status'] in ['pending', 'partially_paid'])
    
    # Get unreconciled deliveries count
    # Find deliveries that are not in any reconciliation
    reconciled_delivery_ids = await db.distributor_reconciliation_items.distinct(
        "delivery_id", {"tenant_id": tenant_id}
    )
    unreconciled_count = await db.distributor_deliveries.count_documents({
        "tenant_id": tenant_id,
        "distributor_id": distributor_id,
        "status": {"$in": ["delivered", "complete"]},
        "id": {"$nin": reconciled_delivery_ids}
    })
    
    return {
        "distributor": distributor,
        "billing_configs": config_count,
        "provisional_invoices": invoice_summary,
        "reconciliations": rec_summary,
        "pending_debit_amount": round(pending_debit, 2),
        "pending_credit_amount": round(pending_credit, 2),
        "net_balance": round(pending_debit - pending_credit, 2),  # Positive = distributor owes
        "unreconciled_deliveries": unreconciled_count
    }



@router.get("/{distributor_id}/monthly-reconciliation")
async def get_monthly_reconciliation_data(
    distributor_id: str,
    month: int,
    year: int,
    current_user: dict = Depends(get_current_user)
):
    """Get approved but not yet reconciled settlements for a specific month"""
    tenant_id = get_current_tenant_id()
    
    logger.info(f"Monthly reconciliation request: distributor={distributor_id}, month={month}, year={year}, tenant={tenant_id}")
    
    # Get only APPROVED and NOT RECONCILED settlements for this month/year
    unreconciled_settlements = await db.distributor_settlements.find(
        {
            "tenant_id": tenant_id,
            "distributor_id": distributor_id,
            "settlement_month": month,
            "settlement_year": year,
            "status": "approved",
            "reconciled": {"$ne": True}  # Not reconciled yet
        },
        {"_id": 0}
    ).sort("account_name", 1).to_list(1000)
    
    # Also get already reconciled settlements for this month (for display purposes)
    reconciled_settlements = await db.distributor_settlements.find(
        {
            "tenant_id": tenant_id,
            "distributor_id": distributor_id,
            "settlement_month": month,
            "settlement_year": year,
            "reconciled": True
        },
        {"_id": 0}
    ).sort("account_name", 1).to_list(1000)
    
    logger.info(f"Found {len(unreconciled_settlements)} unreconciled and {len(reconciled_settlements)} reconciled settlements for tenant={tenant_id}")
    
    # Get all notes for this month (multiple notes allowed)
    existing_notes = await db.distributor_debit_credit_notes.find(
        {
            "tenant_id": tenant_id,
            "distributor_id": distributor_id,
            "month": month,
            "year": year
        },
        {"_id": 0, "id": 1, "note_number": 1, "note_type": 1, "amount": 1, "status": 1, "created_at": 1}
    ).sort("created_at", -1).to_list(100)
    
    # ===== Single source of truth: enrich with stockout_totals (same data the
    # Settlements tab + popup use) so Billing & Settlements ALWAYS show the same Net.
    await _enrich_settlements_with_stockout_totals(tenant_id, unreconciled_settlements)
    await _enrich_settlements_with_stockout_totals(tenant_id, reconciled_settlements)

    def _net_settlement(s: dict) -> float:
        """Mirror of the per-settlement Net Settlement formula used in
        SettlementsTab.jsx so the two tabs cannot drift.
            net_settlement = net_billable − billed_at_transfer − direct_credit_issued − factory_return_credit
        Positive ⇒ Distributor pays Supplier (Debit Note).
        Negative ⇒ Supplier pays Distributor (Credit Note).
        """
        t = s.get('stockout_totals') or {}
        net_billable = t.get('net_billable') or 0
        billed_at_transfer = t.get('billed_at_transfer') or 0
        direct_credit = t.get('direct_credit_issued') or 0
        fr_credit = s.get('total_factory_return_credit') or 0
        return net_billable - billed_at_transfer - direct_credit - fr_credit

    def _aggregate_stockout(settlements_list: list) -> dict:
        agg = {
            "customer_order_value": 0.0,
            "distributor_margin": 0.0,
            "actual_billable": 0.0,
            "credit_applied": 0.0,
            "net_billable": 0.0,
            "billed_at_transfer": 0.0,
            "direct_credit_issued": 0.0,
            "factory_return_credit": 0.0,
            "net_settlement": 0.0,
        }
        for s in settlements_list:
            t = s.get('stockout_totals') or {}
            agg["customer_order_value"] += t.get('customer_order_value') or 0
            agg["distributor_margin"] += t.get('distributor_margin') or 0
            agg["actual_billable"] += t.get('actual_billable') or 0
            agg["credit_applied"] += t.get('credit_applied') or 0
            agg["net_billable"] += t.get('net_billable') or 0
            agg["billed_at_transfer"] += t.get('billed_at_transfer') or 0
            agg["direct_credit_issued"] += t.get('direct_credit_issued') or 0
            agg["factory_return_credit"] += s.get('total_factory_return_credit') or 0
            agg["net_settlement"] += _net_settlement(s)
        return {k: round(v, 2) for k, v in agg.items()}

    stockout_unreconciled = _aggregate_stockout(unreconciled_settlements)
    stockout_reconciled = _aggregate_stockout(reconciled_settlements)

    # ===== Customer-wise reconciliation rollup (across both unreconciled + reconciled
    # settlements for the selected month). One row per Customer (account_id).
    def _build_customer_recon(all_settlements: list) -> tuple[list, dict]:
        by_account: dict = {}
        for s in all_settlements:
            aid = s.get('account_id') or 'unknown'
            t = s.get('stockout_totals') or {}
            row = by_account.setdefault(aid, {
                "account_id": aid,
                "account_name": s.get('account_name') or 'Unknown',
                "delivery_count": 0,
                "customer_order_value": 0.0,
                "credit_notes_paid": 0.0,
                "customer_invoice_value": 0.0,
                "transfer_price_value": 0.0,
            })
            row["delivery_count"] += int(s.get('total_deliveries') or 0)
            row["customer_order_value"] += t.get('customer_order_value') or 0
            row["credit_notes_paid"] += (t.get('credit_applied') or 0) + (t.get('direct_credit_issued') or 0)
            row["transfer_price_value"] += t.get('billed_at_transfer') or 0
        rows = []
        totals = {"delivery_count": 0, "customer_order_value": 0.0,
                  "credit_notes_paid": 0.0, "customer_invoice_value": 0.0,
                  "transfer_price_value": 0.0}
        for r in by_account.values():
            r["customer_invoice_value"] = round(r["customer_order_value"] - r["credit_notes_paid"], 2)
            r["customer_order_value"] = round(r["customer_order_value"], 2)
            r["credit_notes_paid"] = round(r["credit_notes_paid"], 2)
            r["transfer_price_value"] = round(r["transfer_price_value"], 2)
            rows.append(r)
            totals["delivery_count"] += r["delivery_count"]
            totals["customer_order_value"] += r["customer_order_value"]
            totals["credit_notes_paid"] += r["credit_notes_paid"]
            totals["customer_invoice_value"] += r["customer_invoice_value"]
            totals["transfer_price_value"] += r["transfer_price_value"]
        rows.sort(key=lambda x: x["account_name"].lower())
        for k, v in list(totals.items()):
            if isinstance(v, float):
                totals[k] = round(v, 2)
        return rows, totals

    customer_recon_rows, customer_recon_totals = _build_customer_recon(
        unreconciled_settlements + reconciled_settlements
    )

    # Calculate totals for unreconciled settlements only (legacy fields kept for compatibility)
    total_billing = sum(s.get('total_billing_value', 0) for s in unreconciled_settlements)
    total_earnings = sum(s.get('distributor_earnings', 0) for s in unreconciled_settlements)
    total_factory_adj = sum(s.get('factory_distributor_adjustment', 0) for s in unreconciled_settlements)
    total_credit_notes = sum(s.get('total_credit_notes_issued', 0) or s.get('credit_notes_applied', 0) for s in unreconciled_settlements)
    total_factory_return_credit = sum(s.get('total_factory_return_credit', 0) for s in unreconciled_settlements)
    
    # === ENTRY 1: Monthly Billing (at Margin Matrix Transfer Price) ===
    # Direct sum of (qty × transfer_price) from margin matrix, stored per settlement
    total_at_transfer_price = sum(s.get('total_at_transfer_price', 0) for s in unreconciled_settlements)
    # Fallback for older settlements that don't have this field
    if total_at_transfer_price == 0 and total_billing > 0:
        total_at_transfer_price = total_billing - total_earnings - total_factory_adj
    
    # --- Weekly billing breakdown (based on delivery_date) ---
    import calendar
    month_abbrevs = {1: 'Jan', 2: 'Feb', 3: 'Mar', 4: 'Apr', 5: 'May', 6: 'Jun', 7: 'Jul', 8: 'Aug', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dec'}
    days_in_month = calendar.monthrange(year, month)[1]
    week_ranges = []
    day = 1
    w = 1
    while day <= days_in_month:
        end_day = min(day + 6, days_in_month)
        week_ranges.append({"week": w, "start_day": day, "end_day": end_day, "label": f"Week {w} ({day}-{end_day} {month_abbrevs.get(month, '')})"})
        day = end_day + 1
        w += 1
    
    # Get all delivered deliveries (settled in unreconciled settlements) with their items
    all_settlement_ids = [s['id'] for s in unreconciled_settlements]
    settled_items = await db.distributor_settlement_items.find(
        {"tenant_id": tenant_id, "settlement_id": {"$in": all_settlement_ids}},
        {"_id": 0, "delivery_id": 1}
    ).to_list(10000)
    settled_delivery_ids = [si['delivery_id'] for si in settled_items]
    
    weekly_billing = []
    if settled_delivery_ids:
        settled_deliveries = await db.distributor_deliveries.find(
            {"tenant_id": tenant_id, "id": {"$in": settled_delivery_ids}},
            {"_id": 0, "id": 1, "delivery_date": 1, "account_name": 1, "delivery_number": 1}
        ).to_list(10000)
        
        # Get delivery items for transfer price calculation
        all_delivery_items = await db.distributor_delivery_items.find(
            {"tenant_id": tenant_id, "delivery_id": {"$in": settled_delivery_ids}},
            {"_id": 0}
        ).to_list(50000)
        items_by_delivery = {}
        for di in all_delivery_items:
            did = di.get('delivery_id')
            if did not in items_by_delivery:
                items_by_delivery[did] = []
            items_by_delivery[did].append(di)
        
        for wr in week_ranges:
            start_str = f"{year}-{month:02d}-{wr['start_day']:02d}"
            end_str = f"{year}-{month:02d}-{wr['end_day']:02d}"
            week_amount = 0
            week_deliveries = 0
            delivery_details = []  # Per-delivery detail for expansion
            for d in settled_deliveries:
                dd = d.get('delivery_date', '')
                if dd >= start_str and dd <= end_str:
                    week_deliveries += 1
                    del_tp_amount = 0
                    for item in items_by_delivery.get(d['id'], []):
                        qty = item.get('quantity', 0)
                        base_p = item.get('base_price') or item.get('transfer_price') or 0
                        comm = item.get('distributor_commission_percent') or item.get('margin_percent') or 0
                        tp = base_p * (1 - comm / 100)
                        del_tp_amount += qty * tp
                    week_amount += del_tp_amount
                    delivery_details.append({
                        "delivery_id": d['id'],
                        "delivery_number": d.get('delivery_number', ''),
                        "account_name": d.get('account_name', 'Unknown'),
                        "delivery_date": dd,
                        "amount_at_transfer_price": round(del_tp_amount, 2)
                    })
            weekly_billing.append({
                "week": wr['week'],
                "start_day": wr['start_day'],
                "end_day": wr['end_day'],
                "label": wr['label'],
                "amount": round(week_amount, 2),
                "deliveries": week_deliveries,
                "details": delivery_details
            })
    else:
        # No settled deliveries, return empty weeks
        for wr in week_ranges:
            weekly_billing.append({
                "week": wr['week'],
                "start_day": wr['start_day'],
                "end_day": wr['end_day'],
                "label": wr['label'],
                "amount": 0,
                "deliveries": 0
            })
    
    # === ENTRY 2: Monthly Settlement (= Sum of Net Settlement from Settlements tab) ===
    # We use the stockout-driven Net Settlement so Billing tab can never diverge
    # from Settlements tab. Positive ⇒ Debit Note (Distributor owes Supplier),
    # negative ⇒ Credit Note (Supplier owes Distributor).
    net_adjustment_amount = stockout_unreconciled["net_settlement"]
    settlement_note_type = "debit" if net_adjustment_amount > 0 else "credit" if net_adjustment_amount < 0 else "none"
    # Legacy (kept so older clients still render something sensible)
    settlement_selling_price_adj = total_factory_adj
    settlement_credits = total_credit_notes + total_factory_return_credit
    
    # === Calculate totals for already reconciled ===
    reconciled_at_tp = sum(s.get('total_at_transfer_price', 0) for s in reconciled_settlements)
    reconciled_billing = sum(s.get('total_billing_value', 0) for s in reconciled_settlements)
    reconciled_earnings = sum(s.get('distributor_earnings', 0) for s in reconciled_settlements)
    reconciled_factory_adj = sum(s.get('factory_distributor_adjustment', 0) for s in reconciled_settlements)
    if reconciled_at_tp == 0 and reconciled_billing > 0:
        reconciled_at_tp = reconciled_billing - reconciled_earnings - reconciled_factory_adj
    reconciled_credit_notes = sum(s.get('total_credit_notes_issued', 0) or s.get('credit_notes_applied', 0) for s in reconciled_settlements)
    reconciled_factory_return_credit = sum(s.get('total_factory_return_credit', 0) for s in reconciled_settlements)
    
    # --- Reconciled: Weekly billing breakdown ---
    reconciled_weekly_billing = []
    if reconciled_settlements:
        rec_settlement_ids = [s['id'] for s in reconciled_settlements]
        rec_settled_items = await db.distributor_settlement_items.find(
            {"tenant_id": tenant_id, "settlement_id": {"$in": rec_settlement_ids}},
            {"_id": 0, "delivery_id": 1}
        ).to_list(10000)
        rec_delivery_ids = list(set(si['delivery_id'] for si in rec_settled_items))
        
        if rec_delivery_ids:
            rec_deliveries = await db.distributor_deliveries.find(
                {"tenant_id": tenant_id, "id": {"$in": rec_delivery_ids}},
                {"_id": 0, "id": 1, "delivery_date": 1, "account_name": 1, "delivery_number": 1}
            ).to_list(10000)
            rec_all_items = await db.distributor_delivery_items.find(
                {"tenant_id": tenant_id, "delivery_id": {"$in": rec_delivery_ids}},
                {"_id": 0}
            ).to_list(50000)
            rec_items_by_delivery = {}
            for di in rec_all_items:
                did = di.get('delivery_id')
                if did not in rec_items_by_delivery:
                    rec_items_by_delivery[did] = []
                rec_items_by_delivery[did].append(di)
            
            for wr in week_ranges:
                start_str = f"{year}-{month:02d}-{wr['start_day']:02d}"
                end_str = f"{year}-{month:02d}-{wr['end_day']:02d}"
                week_amount = 0
                week_deliveries = 0
                delivery_details = []
                for d in rec_deliveries:
                    dd = d.get('delivery_date', '')
                    if dd >= start_str and dd <= end_str:
                        week_deliveries += 1
                        del_tp_amount = 0
                        for item in rec_items_by_delivery.get(d['id'], []):
                            qty = item.get('quantity', 0)
                            base_p = item.get('base_price') or item.get('transfer_price') or 0
                            comm = item.get('distributor_commission_percent') or item.get('margin_percent') or 0
                            tp = base_p * (1 - comm / 100)
                            del_tp_amount += qty * tp
                        week_amount += del_tp_amount
                        delivery_details.append({
                            "delivery_id": d['id'],
                            "delivery_number": d.get('delivery_number', ''),
                            "account_name": d.get('account_name', 'Unknown'),
                            "delivery_date": dd,
                            "amount_at_transfer_price": round(del_tp_amount, 2)
                        })
                reconciled_weekly_billing.append({
                    "week": wr['week'], "start_day": wr['start_day'], "end_day": wr['end_day'],
                    "label": wr['label'], "amount": round(week_amount, 2),
                    "deliveries": week_deliveries, "details": delivery_details
                })
        else:
            for wr in week_ranges:
                reconciled_weekly_billing.append({
                    "week": wr['week'], "start_day": wr['start_day'], "end_day": wr['end_day'],
                    "label": wr['label'], "amount": 0, "deliveries": 0
                })
    
    # Reconciled Entry 2: same stockout-driven Net Settlement
    reconciled_selling_price_adj = reconciled_factory_adj
    reconciled_net_adj = stockout_reconciled["net_settlement"]
    reconciled_note_type = "debit" if reconciled_net_adj > 0 else "credit" if reconciled_net_adj < 0 else "none"
    
    return {
        "unreconciled_settlements": unreconciled_settlements,
        "reconciled_settlements": reconciled_settlements,
        "total_unreconciled": len(unreconciled_settlements),
        "total_reconciled": len(reconciled_settlements),
        # Raw totals
        "total_billing_value": round(total_billing, 2),
        "total_distributor_earnings": round(total_earnings, 2),
        "total_factory_adjustment": round(total_factory_adj, 2),
        "total_credit_notes_applied": round(total_credit_notes, 2),
        "total_factory_return_credit": round(total_factory_return_credit, 2),
        # Entry 1: Billing at margin matrix transfer price
        "total_at_transfer_price": round(total_at_transfer_price, 2),
        "weekly_billing": weekly_billing,
        # Entry 2: Settlement adjustments (stockout-driven, single source of truth)
        "settlement_selling_price_adj": round(settlement_selling_price_adj, 2),
        "settlement_credits": round(settlement_credits, 2),
        "net_adjustment_amount": round(net_adjustment_amount, 2),
        "settlement_note_type": settlement_note_type,
        # NEW: stockout-driven aggregates (same source as Settlements tab)
        "stockout_aggregate": stockout_unreconciled,
        "stockout_aggregate_reconciled": stockout_reconciled,
        # NEW: customer-wise reconciliation rollup (across both reconciled + unreconciled)
        "customer_reconciliation": customer_recon_rows,
        "customer_reconciliation_totals": customer_recon_totals,
        # Reconciled - full Two-Entry data
        "reconciled_at_transfer_price": round(reconciled_at_tp, 2),
        "reconciled_weekly_billing": reconciled_weekly_billing,
        "reconciled_selling_price_adj": round(reconciled_selling_price_adj, 2),
        "reconciled_credit_notes": round(reconciled_credit_notes, 2),
        "reconciled_factory_return_credit": round(reconciled_factory_return_credit, 2),
        "reconciled_net_adjustment": round(reconciled_net_adj, 2),
        "reconciled_note_type": reconciled_note_type,
        "existing_notes": existing_notes,
        "total_notes": len(existing_notes)
    }


@router.post("/{distributor_id}/generate-monthly-note")
async def generate_monthly_note(
    distributor_id: str,
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Generate a Debit or Credit Note for approved but unreconciled settlements with PDF"""
    if not is_distributor_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    tenant_id = get_current_tenant_id()
    
    month = data.get('month')
    year = data.get('year')
    remarks = data.get('remarks', '')
    
    if not month or not year:
        raise HTTPException(status_code=400, detail="Month and year are required")
    
    # Get only APPROVED and NOT RECONCILED settlements for this month
    # Multiple notes per month are allowed - each processes unreconciled settlements
    settlements = await db.distributor_settlements.find(
        {
            "tenant_id": tenant_id,
            "distributor_id": distributor_id,
            "settlement_month": month,
            "settlement_year": year,
            "status": "approved",
            "reconciled": {"$ne": True}
        },
        {"_id": 0}
    ).to_list(1000)
    
    if not settlements:
        raise HTTPException(status_code=400, detail="No approved unreconciled settlements found for this month")
    
    # Calculate settlement adjustments (separate from billing)
    total_billing = sum(s.get('total_billing_value', 0) for s in settlements)
    total_earnings = sum(s.get('distributor_earnings', 0) for s in settlements)
    total_factory_adj = sum(s.get('factory_distributor_adjustment', 0) for s in settlements)
    total_credit_notes = sum(s.get('total_credit_notes_issued', 0) or s.get('credit_notes_applied', 0) for s in settlements)
    total_factory_return_credit = sum(s.get('total_factory_return_credit', 0) for s in settlements)
    
    # Entry 1: Billing at margin matrix transfer price
    total_at_transfer_price = sum(s.get('total_at_transfer_price', 0) for s in settlements)
    if total_at_transfer_price == 0 and total_billing > 0:
        total_at_transfer_price = total_billing - total_earnings - total_factory_adj
    
    # Entry 2: Net = stockout-driven Net Settlement (same as Settlements tab)
    await _enrich_settlements_with_stockout_totals(tenant_id, settlements)
    net_adjustment_amount = 0.0
    for s in settlements:
        t = s.get('stockout_totals') or {}
        net_billable = t.get('net_billable') or 0
        billed_at_transfer = t.get('billed_at_transfer') or 0
        direct_credit = t.get('direct_credit_issued') or 0
        fr_credit = s.get('total_factory_return_credit') or 0
        net_adjustment_amount += (net_billable - billed_at_transfer - direct_credit - fr_credit)
    net_adjustment_amount = round(net_adjustment_amount, 2)
    settlement_selling_price_adj = total_factory_adj
    settlement_credits = total_credit_notes + total_factory_return_credit
    
    if net_adjustment_amount == 0:
        raise HTTPException(status_code=400, detail="Net adjustment is zero - no note required")
    
    # Get distributor full info for PDF
    distributor = await db.distributors.find_one(
        {"id": distributor_id, "tenant_id": tenant_id},
        {"_id": 0}
    )
    
    if not distributor:
        raise HTTPException(status_code=404, detail="Distributor not found")
    
    # Get tenant settings for company profile
    tenant = await db.tenants.find_one({"tenant_id": tenant_id}, {"_id": 0})
    company_profile = tenant.get('company_profile', {}) if tenant else {}
    branding = tenant.get('branding', {}) if tenant else {}
    
    # Determine note type: positive net adjustment = distributor owes more = debit note
    note_type = "debit" if net_adjustment_amount > 0 else "credit"
    amount = abs(net_adjustment_amount)
    
    # Generate note number
    count = await db.distributor_debit_credit_notes.count_documents({"tenant_id": tenant_id})
    note_number = f"{'CN' if note_type == 'credit' else 'DN'}-{year}-{count + 1:04d}"
    
    # Create note document
    note_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    note_doc = {
        "id": note_id,
        "tenant_id": tenant_id,
        "note_number": note_number,
        "distributor_id": distributor_id,
        "distributor_name": distributor.get('distributor_name'),
        "distributor_code": distributor.get('distributor_code'),
        "note_type": note_type,
        "month": month,
        "year": year,
        "amount": round(amount, 2),
        "paid_amount": 0,
        "balance_amount": round(amount, 2),
        "settlement_ids": [s['id'] for s in settlements],
        "total_settlements": len(settlements),
        # Entry 1: Billing at margin matrix transfer price
        "total_billing_value": round(total_billing, 2),
        "total_at_transfer_price": round(total_at_transfer_price, 2),
        "total_distributor_earnings": round(total_earnings, 2),
        # Entry 2: Settlement adjustments
        "settlement_selling_price_adj": round(settlement_selling_price_adj, 2),
        "settlement_credits": round(settlement_credits, 2),
        "total_credit_notes": round(total_credit_notes, 2),
        "total_factory_return_credit": round(total_factory_return_credit, 2),
        "net_adjustment_amount": round(net_adjustment_amount, 2),
        "remarks": remarks,
        "status": "pending",
        "created_by": current_user['id'],
        "created_by_name": current_user.get('name', current_user.get('email')),
        "created_at": now,
        "updated_at": now
    }
    
    # Generate PDF
    try:
        pdf_bytes = generate_debit_credit_note_pdf(
            note_data=note_doc,
            company_profile=company_profile,
            distributor_data=distributor,
            settlements=settlements,
            branding=branding
        )
        
        # Upload PDF to object storage
        pdf_filename = f"{note_number}.pdf"
        storage_result = await upload_pdf(pdf_filename, pdf_bytes, subfolder=f"debit-credit-notes/{distributor_id}")
        
        # Store PDF reference in note document
        note_doc["pdf_path"] = storage_result.get("path")
        note_doc["pdf_size"] = storage_result.get("size")
        note_doc["pdf_generated_at"] = now
        
        logger.info(f"PDF generated and uploaded for note {note_number}: {storage_result.get('path')}")
    except Exception as e:
        logger.error(f"Failed to generate/upload PDF for note {note_number}: {e}")
        # Continue without PDF - note will still be created
        note_doc["pdf_error"] = str(e)
    
    await db.distributor_debit_credit_notes.insert_one(note_doc)
    
    # Update settlements to mark them as reconciled
    settlement_ids = [s['id'] for s in settlements]
    await db.distributor_settlements.update_many(
        {"id": {"$in": settlement_ids}, "tenant_id": tenant_id},
        {"$set": {"reconciled": True, "note_id": note_id, "note_number": note_number}}
    )
    
    note_doc.pop('_id', None)
    
    logger.info(f"Generated {note_type} note {note_number} for ₹{amount} for distributor {distributor_id} ({month}/{year}) by {current_user['email']}")
    
    return note_doc


@router.get("/{distributor_id}/notes/{note_id}/download")
async def download_note_pdf(
    distributor_id: str,
    note_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Download PDF for a debit/credit note"""
    tenant_id = get_current_tenant_id()
    
    # Find the note
    note = await db.distributor_debit_credit_notes.find_one(
        {"id": note_id, "tenant_id": tenant_id, "distributor_id": distributor_id},
        {"_id": 0}
    )
    
    if not note:
        raise HTTPException(status_code=404, detail="Note not found")
    
    pdf_path = note.get("pdf_path")
    
    if not pdf_path:
        # PDF not generated yet - generate it now
        distributor = await db.distributors.find_one(
            {"id": distributor_id, "tenant_id": tenant_id},
            {"_id": 0}
        )
        
        if not distributor:
            raise HTTPException(status_code=404, detail="Distributor not found")
        
        # Get settlements for this note
        settlements = await db.distributor_settlements.find(
            {"id": {"$in": note.get("settlement_ids", [])}, "tenant_id": tenant_id},
            {"_id": 0}
        ).to_list(1000)
        
        # Get tenant settings
        tenant = await db.tenants.find_one({"tenant_id": tenant_id}, {"_id": 0})
        company_profile = tenant.get('company_profile', {}) if tenant else {}
        branding = tenant.get('branding', {}) if tenant else {}
        
        try:
            pdf_bytes = generate_debit_credit_note_pdf(
                note_data=note,
                company_profile=company_profile,
                distributor_data=distributor,
                settlements=settlements,
                branding=branding
            )
            
            # Upload and store reference
            pdf_filename = f"{note.get('note_number', note_id)}.pdf"
            storage_result = await upload_pdf(pdf_filename, pdf_bytes, subfolder=f"debit-credit-notes/{distributor_id}")
            
            # Update note with PDF path
            await db.distributor_debit_credit_notes.update_one(
                {"id": note_id, "tenant_id": tenant_id},
                {"$set": {
                    "pdf_path": storage_result.get("path"),
                    "pdf_size": storage_result.get("size"),
                    "pdf_generated_at": datetime.now(timezone.utc).isoformat()
                }}
            )
            
            pdf_path = storage_result.get("path")
            
        except Exception as e:
            logger.error(f"Failed to generate PDF on-demand for note {note_id}: {e}")
            raise HTTPException(status_code=500, detail=f"Failed to generate PDF: {str(e)}")
    
    # Download PDF from storage
    try:
        pdf_content = await download_pdf(pdf_path)
        
        filename = f"{note.get('note_number', note_id)}.pdf"
        
        return Response(
            content=pdf_content,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            }
        )
    except Exception as e:
        logger.error(f"Failed to download PDF from storage for note {note_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to download PDF: {str(e)}")
