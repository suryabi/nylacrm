"""
Accounting Transactions — pulls bank transactions from Zoho Books (the bank
feed, the true bank-synced source) and lets users enrich each with master tags
(expense or income), a vendor, proof uploads, and — for money received — a CRM
account link that adjusts the account's outstanding balance.

De-dup: one doc per Zoho `bank_transaction_id`, enforced by a unique index on
(tenant_id, zoho_org_id, zoho_transaction_id). Sync upserts: new -> 'untagged';
existing -> only Zoho-side fields refresh, user tags/proofs/links preserved.
"""
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form, Header, Response, BackgroundTasks
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime, timezone, timedelta
from pymongo import ReturnDocument, UpdateOne, InsertOne
import asyncio
import uuid
import io
import csv
import logging

from database import db
from deps import get_current_user
from core.tenant import get_current_tenant_id
from services import zoho_service
from services.object_storage import put_object, get_object, build_path, guess_content_type

ALLOWED_PROOF_TYPES = {"image/png", "image/jpeg", "image/jpg", "image/webp", "image/gif", "application/pdf"}

router = APIRouter(prefix="/accounting/transactions", tags=["Accounting Transactions"])
logger = logging.getLogger(__name__)

COLL = "accounting_transactions"
SYNC_COLL = "accounting_txn_sync_state"
SYNC_JOB_COLL = "accounting_txn_sync_jobs"
ADMIN_ROLES = {"CEO", "Director", "System Admin", "Admin", "Vice President", "Head of Business"}

# Zoho transaction_type -> our direction. Anything else is inferred from amount sign.
_CREDIT_TYPES = {"deposit", "sales_without_invoices", "interest_income", "owner_contribution", "other_income", "customer_payment", "vendor_payment_refund"}
_DEBIT_TYPES = {"expense", "withdrawal", "transfer_fund", "card_payment", "owner_drawings", "supplier_payment", "vendor_payment"}


def _require_admin(user: dict):
    if user.get("role") not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")


def _now():
    return datetime.now(timezone.utc).isoformat()


async def ensure_indexes():
    await db[COLL].create_index(
        [("tenant_id", 1), ("zoho_org_id", 1), ("zoho_transaction_id", 1)],
        unique=True, name="uniq_zoho_txn",
    )
    await db[SYNC_JOB_COLL].create_index("created_at", name="created_at_idx")


async def _purge_old_sync_jobs(tenant_id: str, keep_days: int = 30) -> None:
    """Lazy cleanup — drop sync-job audit docs older than `keep_days`. Called
    at sync kickoff so the collection never grows unbounded."""
    cutoff = (datetime.now(timezone.utc) - timedelta(days=keep_days)).isoformat()
    try:
        await db[SYNC_JOB_COLL].delete_many({"tenant_id": tenant_id, "created_at": {"$lt": cutoff}})
    except Exception:
        # cleanup is best-effort — never fail a sync because of it
        pass


async def _next_txn_code(tenant_id: str) -> str:
    """Atomic per-tenant sequence -> human-friendly transaction id e.g. TXN-000123."""
    doc = await db["counters"].find_one_and_update(
        {"_id": f"{tenant_id}:accounting_txn"},
        {"$inc": {"seq": 1}}, upsert=True, return_document=ReturnDocument.AFTER,
    )
    return f"TXN-{str(doc['seq']).zfill(6)}"


async def _ensure_txn_codes(tenant_id: str):
    """Assign a txn_code to any transactions that don't yet have one (oldest first)."""
    cursor = db[COLL].find(
        {"tenant_id": tenant_id, "$or": [{"txn_code": {"$exists": False}}, {"txn_code": None}]},
        {"_id": 0, "id": 1},
    ).sort([("date", 1), ("created_at", 1)])
    async for d in cursor:
        code = await _next_txn_code(tenant_id)
        await db[COLL].update_one({"id": d["id"], "tenant_id": tenant_id}, {"$set": {"txn_code": code}})


def _direction_of(txn: dict) -> str:
    ttype = (txn.get("transaction_type") or "").lower()
    if ttype in _CREDIT_TYPES:
        return "credit"
    if ttype in _DEBIT_TYPES:
        return "debit"
    # fall back to debit/credit flags or amount
    if txn.get("debit_or_credit"):
        return "credit" if str(txn["debit_or_credit"]).lower().startswith("c") else "debit"
    amt = float(txn.get("amount") or 0)
    return "credit" if amt >= 0 else "debit"


def _normalize(txn: dict, tenant_id: str, org_id: str) -> dict:
    """Map a raw Zoho bank transaction to the Zoho-side fields we store."""
    txn_id = txn.get("bank_transaction_id") or txn.get("transaction_id")
    return {
        "tenant_id": tenant_id,
        "zoho_org_id": org_id,
        "zoho_transaction_id": txn_id,
        "source": "zoho_bank",
        "direction": _direction_of(txn),
        "amount": abs(float(txn.get("amount") or 0)),
        "currency": txn.get("currency_code") or txn.get("currency") or "INR",
        "date": (txn.get("date") or txn.get("transaction_date") or "")[:10],
        "zoho_transaction_type": txn.get("transaction_type"),
        "zoho_status": txn.get("status"),
        "zoho_account_id": txn.get("account_id"),
        "bank_account_name": txn.get("account_name"),
        "payee": txn.get("payee"),
        "reference_number": txn.get("reference_number"),
        "description": txn.get("description"),
        "raw": txn,
    }


# ----------------------- Sync -----------------------

async def _allocate_txn_codes(tenant_id: str, n: int) -> list:
    """Atomically reserve `n` consecutive txn codes from the per-tenant counter
    in ONE round-trip (vs N round-trips when called individually)."""
    if n <= 0:
        return []
    doc = await db["counters"].find_one_and_update(
        {"_id": f"{tenant_id}:accounting_txn"},
        {"$inc": {"seq": n}}, upsert=True, return_document=ReturnDocument.AFTER,
    )
    end = int(doc["seq"])
    start = end - n + 1
    return [f"TXN-{str(i).zfill(6)}" for i in range(start, end + 1)]


async def _run_sync(tenant_id: str, user_id: str, date_start: str, date_end: str,
                    explicit_range: bool, job_id: str) -> None:
    """Background worker — pulls Zoho pages, batches the upserts via bulk_write,
    and updates the job doc with progress / final state."""
    org_id = None
    try:
        creds = await zoho_service.get_credentials(tenant_id)
        org_id = (creds or {}).get("organization_id")
        new_count, updated_count, page = 0, 0, 1
        while page <= 50:
            res = await zoho_service.fetch_bank_transactions(tenant_id, date_start, date_end, page=page)
            txns = res.get("transactions") or []
            # Build (zoho_txn_id -> normalised doc) map, drop any without an id.
            page_docs = []
            for raw in txns:
                d = _normalize(raw, tenant_id, org_id)
                if d["zoho_transaction_id"]:
                    page_docs.append(d)
            if page_docs:
                ids = [d["zoho_transaction_id"] for d in page_docs]
                existing = set()
                async for ex in db[COLL].find(
                    {"tenant_id": tenant_id, "zoho_org_id": org_id, "zoho_transaction_id": {"$in": ids}},
                    {"_id": 0, "zoho_transaction_id": 1},
                ):
                    existing.add(ex["zoho_transaction_id"])

                new_docs = [d for d in page_docs if d["zoho_transaction_id"] not in existing]
                codes = await _allocate_txn_codes(tenant_id, len(new_docs))

                ops = []
                now = _now()
                for d in page_docs:
                    key = {"tenant_id": tenant_id, "zoho_org_id": org_id, "zoho_transaction_id": d["zoho_transaction_id"]}
                    if d["zoho_transaction_id"] in existing:
                        # refresh only Zoho-side fields; preserve user tags/proofs/links
                        ops.append(UpdateOne(key, {"$set": {
                            **{k: d[k] for k in ("amount", "currency", "date", "zoho_status",
                                                 "zoho_transaction_type", "payee", "reference_number",
                                                 "description", "bank_account_name", "zoho_account_id", "raw")},
                            "updated_at": now,
                        }}))
                        updated_count += 1
                for d, code in zip(new_docs, codes):
                    d.update({
                        "id": str(uuid.uuid4()), "status": "untagged", "txn_code": code,
                        "tags": {}, "vendor_id": None, "vendor_name": None,
                        "account_id": None, "account_name": None,
                        "account_adjustment": None, "proofs": [], "notes": None,
                        "tagged_by": None, "tagged_at": None,
                        "created_at": now, "updated_at": now,
                    })
                    ops.append(InsertOne(d))
                    new_count += 1

                if ops:
                    await db[COLL].bulk_write(ops, ordered=False)

            # update progress on the job doc so the UI can show progress
            await db[SYNC_JOB_COLL].update_one(
                {"id": job_id},
                {"$set": {"progress": {"page": page, "new": new_count, "updated": updated_count},
                          "updated_at": _now()}},
            )

            if not res.get("has_more"):
                break
            page += 1

        if not explicit_range:
            await db[SYNC_COLL].update_one(
                {"tenant_id": tenant_id},
                {"$set": {"tenant_id": tenant_id, "last_synced_date": date_end,
                          "last_synced_at": _now(), "last_synced_by": user_id}},
                upsert=True,
            )

        await db[SYNC_JOB_COLL].update_one(
            {"id": job_id},
            {"$set": {"status": "completed", "new": new_count, "updated": updated_count,
                      "finished_at": _now(), "updated_at": _now()}},
        )
    except Exception as e:
        logger.exception("Zoho bank transaction sync failed (job %s)", job_id)
        err_text = str(e)
        # Translate the common Zoho-side authorisation failure into something
        # actionable in the UI. Zoho returns 401 + code:57 when the connected
        # OAuth token is missing the ZohoBooks.banking.READ scope.
        is_scope_error = ("401" in err_text and '"code":57' in err_text) or \
                          "not authorized" in err_text.lower()
        if is_scope_error:
            friendly = (
                "Zoho rejected the request: the connected account is missing the "
                "banking access scope. Open Settings → Integrations → Zoho Books "
                "and reconnect Zoho with the 'Banking' permission, then try again."
            )
        else:
            # Don't leak stack-trace text to the UI. Full detail is in server logs.
            friendly = "Sync failed unexpectedly. Please retry; if it persists check the server logs."
        await db[SYNC_JOB_COLL].update_one(
            {"id": job_id},
            {"$set": {"status": "failed", "error": friendly,
                      "error_kind": "zoho_banking_scope" if is_scope_error else "other",
                      "finished_at": _now(), "updated_at": _now()}},
        )


@router.post("/sync")
async def sync_transactions(
    background_tasks: BackgroundTasks,
    date_start: Optional[str] = Query(None),
    date_end: Optional[str] = Query(None),
    current_user: dict = Depends(get_current_user),
):
    """Kick off a Zoho bank-transactions sync in the background and return a
    job id immediately (avoids ingress / proxy timeouts on long syncs). New
    rows land as 'untagged'; existing rows refresh Zoho-side fields only."""
    _require_admin(current_user)
    tenant_id = get_current_tenant_id()
    await ensure_indexes()
    await _purge_old_sync_jobs(tenant_id)

    creds = await zoho_service.get_credentials(tenant_id)
    if not zoho_service.is_zoho_configured() or not creds:
        raise HTTPException(status_code=400, detail="Zoho Books is not connected for this tenant. Connect it under Settings → Integrations → Zoho Books (with banking access).")

    explicit_range = bool(date_start or date_end)
    state = await db[SYNC_COLL].find_one({"tenant_id": tenant_id}, {"_id": 0})
    if not date_start and state and state.get("last_synced_date"):
        date_start = state["last_synced_date"]
    if not date_end:
        date_end = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    job_id = str(uuid.uuid4())
    job_doc = {
        "id": job_id, "tenant_id": tenant_id, "status": "running",
        "from": date_start, "to": date_end, "new": 0, "updated": 0,
        "progress": {"page": 0, "new": 0, "updated": 0},
        "started_at": _now(), "started_by": current_user.get("id"),
        "created_at": _now(), "updated_at": _now(),
    }
    await db[SYNC_JOB_COLL].insert_one(job_doc)

    # Fire-and-forget background task — runs after the HTTP response is sent.
    async def _runner():
        await _run_sync(tenant_id, current_user.get("id"), date_start, date_end, explicit_range, job_id)
    background_tasks.add_task(_runner)

    job_doc.pop("_id", None)
    return {"job_id": job_id, "status": "started", "from": date_start, "to": date_end}


@router.get("/sync/status/{job_id}")
async def sync_status(job_id: str, current_user: dict = Depends(get_current_user)):
    tenant_id = get_current_tenant_id()
    job = await db[SYNC_JOB_COLL].find_one({"id": job_id, "tenant_id": tenant_id}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="Sync job not found")
    return job


# ----------------------- List / filters -----------------------

@router.get("")
async def list_transactions(
    status: Optional[str] = Query(None),
    direction: Optional[str] = Query(None),
    bank_account_id: Optional[str] = Query(None),
    date_start: Optional[str] = Query(None),
    date_end: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    category_root: Optional[str] = Query(None),
    page: int = Query(1),
    limit: int = Query(50),
    current_user: dict = Depends(get_current_user),
):
    tenant_id = get_current_tenant_id()
    await _ensure_txn_codes(tenant_id)
    q = {"tenant_id": tenant_id}
    if status:
        q["status"] = status
    if direction:
        q["direction"] = direction
    if bank_account_id:
        q["zoho_account_id"] = bank_account_id
    if date_start or date_end:
        q["date"] = {}
        if date_start:
            q["date"]["$gte"] = date_start
        if date_end:
            q["date"]["$lte"] = date_end
    if search:
        q["$or"] = [
            {"payee": {"$regex": search, "$options": "i"}},
            {"description": {"$regex": search, "$options": "i"}},
            {"reference_number": {"$regex": search, "$options": "i"}},
        ]
    if category_root:
        ids = await _expense_category_descendants(tenant_id, category_root)
        q["tags.expense_category"] = {"$in": list(ids)}
    total = await db[COLL].count_documents(q)
    rows = await db[COLL].find(q, {"_id": 0, "raw": 0}).sort("date", -1).skip((page - 1) * limit).limit(limit).to_list(limit)
    # summary counts for the tabs
    summary = {"untagged": 0, "tagged": 0, "all": 0}
    async for r in db[COLL].aggregate([{"$match": {"tenant_id": tenant_id}}, {"$group": {"_id": "$status", "n": {"$sum": 1}}}]):
        summary[r["_id"]] = r["n"]
        summary["all"] += r["n"]
    return {"items": rows, "total": total, "page": page, "limit": limit, "summary": summary}


async def _expense_category_descendants(tenant_id: str, root_id: str) -> set:
    """Return {root_id, *descendant_ids} by walking the parent_id tree (BFS)."""
    out = {root_id}
    frontier = {root_id}
    while frontier:
        nxt = set()
        async for m in db["accounting_masters"].find(
            {"tenant_id": tenant_id, "master_type": "expense_category", "parent_id": {"$in": list(frontier)}},
            {"_id": 0, "id": 1},
        ):
            mid = m["id"]
            if mid not in out:
                out.add(mid)
                nxt.add(mid)
        frontier = nxt
    return out


@router.get("/category-summary")
async def category_summary(
    status: Optional[str] = Query(None),
    direction: Optional[str] = Query(None),
    bank_account_id: Optional[str] = Query(None),
    date_start: Optional[str] = Query(None),
    date_end: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    current_user: dict = Depends(get_current_user),
):
    """Per-root expense-category spend summary for the current filtered view.
    Aggregates each transaction's tagged expense_category to its root and returns
    {root_id, name, count, total} entries sorted by total desc."""
    tenant_id = get_current_tenant_id()
    q = {"tenant_id": tenant_id, "tags.expense_category": {"$ne": None}}
    if status:
        q["status"] = status
    if direction:
        q["direction"] = direction
    if bank_account_id:
        q["zoho_account_id"] = bank_account_id
    if date_start or date_end:
        q["date"] = {}
        if date_start:
            q["date"]["$gte"] = date_start
        if date_end:
            q["date"]["$lte"] = date_end
    if search:
        q["$or"] = [
            {"payee": {"$regex": search, "$options": "i"}},
            {"description": {"$regex": search, "$options": "i"}},
            {"reference_number": {"$regex": search, "$options": "i"}},
        ]

    # Build parent map once for expense_category masters.
    parent_map = {}
    name_map = {}
    async for m in db["accounting_masters"].find(
        {"tenant_id": tenant_id, "master_type": "expense_category"},
        {"_id": 0, "id": 1, "name": 1, "parent_id": 1, "level": 1},
    ):
        name_map[m["id"]] = m["name"]
        if m.get("parent_id"):
            parent_map[m["id"]] = m["parent_id"]

    def root_of(mid: str) -> str:
        cur = mid
        seen = set()
        while cur in parent_map and cur not in seen:
            seen.add(cur)
            cur = parent_map[cur]
        return cur

    buckets: dict = {}
    async for t in db[COLL].find(q, {"_id": 0, "tags.expense_category": 1, "amount": 1, "direction": 1}):
        leaf = (t.get("tags") or {}).get("expense_category")
        if not leaf:
            continue
        root = root_of(leaf)
        b = buckets.setdefault(root, {"root_id": root, "name": name_map.get(root, ""), "count": 0, "total": 0.0})
        b["count"] += 1
        b["total"] += float(t.get("amount") or 0)
    items = sorted(buckets.values(), key=lambda x: -x["total"])
    return {"items": items}


@router.get("/export")
async def export_transactions(
    format: str = Query("csv"),
    status: Optional[str] = Query(None),
    direction: Optional[str] = Query(None),
    bank_account_id: Optional[str] = Query(None),
    date_start: Optional[str] = Query(None),
    date_end: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    category_root: Optional[str] = Query(None),
    current_user: dict = Depends(get_current_user),
):
    """Export the filtered transactions to CSV / XLSX / PDF. No upload/proof
    information is included in the exported files."""
    tenant_id = get_current_tenant_id()
    await _ensure_txn_codes(tenant_id)

    q = {"tenant_id": tenant_id}
    if status:
        q["status"] = status
    if direction:
        q["direction"] = direction
    if bank_account_id:
        q["zoho_account_id"] = bank_account_id
    if date_start or date_end:
        q["date"] = {}
        if date_start:
            q["date"]["$gte"] = date_start
        if date_end:
            q["date"]["$lte"] = date_end
    if search:
        q["$or"] = [
            {"payee": {"$regex": search, "$options": "i"}},
            {"description": {"$regex": search, "$options": "i"}},
            {"reference_number": {"$regex": search, "$options": "i"}},
        ]
    if category_root:
        ids = await _expense_category_descendants(tenant_id, category_root)
        q["tags.expense_category"] = {"$in": list(ids)}
    rows = await db[COLL].find(q, {"_id": 0, "raw": 0, "proofs": 0}).sort("date", -1).to_list(20000)

    # id -> name maps for master tags + vendors
    name_map = {}
    parent_map = {}
    async for m in db["accounting_masters"].find({"tenant_id": tenant_id}, {"_id": 0, "id": 1, "name": 1, "parent_id": 1}):
        name_map[m["id"]] = m["name"]
        if m.get("parent_id"):
            parent_map[m["id"]] = m["parent_id"]
    vendor_map = {}
    async for v in db["accounting_vendors"].find({"tenant_id": tenant_id}, {"_id": 0, "id": 1, "name": 1}):
        vendor_map[v["id"]] = v["name"]

    def _path(mid: str) -> str:
        """Return slash-separated 'Parent / Child / Leaf' path for a master id."""
        if not mid:
            return ""
        chain = []
        cur = mid
        seen = set()
        while cur and cur not in seen:
            seen.add(cur)
            chain.append(name_map.get(cur, ""))
            cur = parent_map.get(cur)
        return " / ".join(reversed([n for n in chain if n]))

    headers = ["Transaction ID", "Date", "Direction", "Amount", "Currency", "Bank Account",
               "Payee", "Description", "Reference", "Zoho Transaction ID", "Status",
               "Expense Type", "Expense Category", "Cost Center", "Business Unit",
               "Payment Source", "Vendor", "Revenue Stream", "Linked Account", "Notes"]

    def row_values(t):
        tags = t.get("tags") or {}
        return [
            t.get("txn_code") or "", t.get("date") or "",
            "Money In" if t.get("direction") == "credit" else "Money Out",
            t.get("amount") or 0, t.get("currency") or "INR", t.get("bank_account_name") or "",
            t.get("payee") or "", t.get("description") or "", t.get("reference_number") or "",
            t.get("zoho_transaction_id") or "", t.get("status") or "",
            name_map.get(tags.get("expense_type"), ""), _path(tags.get("expense_category")),
            name_map.get(tags.get("cost_center"), ""), name_map.get(tags.get("project_business_unit"), ""),
            name_map.get(tags.get("payment_source"), ""), vendor_map.get(t.get("vendor_id"), t.get("vendor_name") or ""),
            name_map.get(tags.get("revenue_stream"), ""), t.get("account_name") or "", t.get("notes") or "",
        ]

    fname = f"transactions_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M')}"

    if format == "csv":
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(headers)
        for t in rows:
            w.writerow(row_values(t))
        return StreamingResponse(io.BytesIO(buf.getvalue().encode("utf-8-sig")), media_type="text/csv",
                                 headers={"Content-Disposition": f'attachment; filename="{fname}.csv"'})

    if format in ("xlsx", "xl", "excel"):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook()
        ws = wb.active
        ws.title = "Transactions"
        ws.append(headers)
        hf = Font(bold=True, color="FFFFFF")
        fill = PatternFill("solid", fgColor="4F46E5")
        for c in ws[1]:
            c.font = hf
            c.fill = fill
        for t in rows:
            ws.append(row_values(t))
        for i, h in enumerate(headers, 1):
            ws.column_dimensions[chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)].width = max(12, min(40, len(h) + 4))
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return StreamingResponse(out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": f'attachment; filename="{fname}.xlsx"'})

    if format == "pdf":
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        out = io.BytesIO()
        doc = SimpleDocTemplate(out, pagesize=landscape(A4), leftMargin=10 * mm, rightMargin=10 * mm, topMargin=12 * mm, bottomMargin=10 * mm)
        styles = getSampleStyleSheet()
        elems = [Paragraph("Accounting Transactions", styles["Title"]), Spacer(1, 6)]
        # compact column subset for PDF readability
        pdf_cols = ["Transaction ID", "Date", "Direction", "Amount", "Bank Account", "Payee",
                    "Status", "Expense Category", "Cost Center", "Vendor", "Revenue Stream", "Linked Account"]
        idx = [headers.index(c) for c in pdf_cols]
        data = [pdf_cols]
        for t in rows:
            rv = row_values(t)
            data.append([str(rv[i]) for i in idx])
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 6.5),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#E2E8F0")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        elems.append(table)
        doc.build(elems)
        out.seek(0)
        return StreamingResponse(out, media_type="application/pdf",
                                 headers={"Content-Disposition": f'attachment; filename="{fname}.pdf"'})

    raise HTTPException(status_code=400, detail="Unsupported format. Use csv, xlsx or pdf.")


@router.get("/bank-accounts")
async def list_bank_accounts(current_user: dict = Depends(get_current_user)):
    """Distinct bank accounts seen across synced transactions (for filtering)."""
    tenant_id = get_current_tenant_id()
    rows = await db[COLL].aggregate([
        {"$match": {"tenant_id": tenant_id, "zoho_account_id": {"$ne": None}}},
        {"$group": {"_id": "$zoho_account_id", "name": {"$first": "$bank_account_name"}}},
    ]).to_list(200)
    return {"items": [{"id": r["_id"], "name": r.get("name") or r["_id"]} for r in rows]}


# ----------------------- Tagging -----------------------

class TagPayload(BaseModel):
    tags: dict = {}          # master_type -> master_id (expense or income masters)
    vendor_id: Optional[str] = None
    vendor_name: Optional[str] = None
    notes: Optional[str] = None


@router.patch("/{txn_id}/tags")
async def tag_transaction(txn_id: str, payload: TagPayload, current_user: dict = Depends(get_current_user)):
    _require_admin(current_user)
    tenant_id = get_current_tenant_id()
    txn = await db[COLL].find_one({"id": txn_id, "tenant_id": tenant_id}, {"_id": 0})
    if not txn:
        raise HTTPException(status_code=404, detail="Transaction not found")
    clean_tags = {k: v for k, v in (payload.tags or {}).items() if v}
    has_selection = bool(clean_tags) or bool(payload.vendor_id)
    updates = {
        "tags": clean_tags,
        "vendor_id": payload.vendor_id, "vendor_name": payload.vendor_name,
        "notes": payload.notes,
        "status": "tagged" if has_selection else "untagged",
        "tagged_by": current_user.get("id"), "tagged_at": _now(), "updated_at": _now(),
    }
    await db[COLL].update_one({"id": txn_id, "tenant_id": tenant_id}, {"$set": updates})
    return await db[COLL].find_one({"id": txn_id, "tenant_id": tenant_id}, {"_id": 0, "raw": 0})


# ----------------------- Account (payment received) -----------------------

class AccountApplyPayload(BaseModel):
    account_id: str
    account_name: Optional[str] = None


@router.post("/{txn_id}/apply-account")
async def apply_account(txn_id: str, payload: AccountApplyPayload, current_user: dict = Depends(get_current_user)):
    """Tag a received payment (credit) to a CRM account and reduce that account's
    outstanding balance by the transaction amount. Reversible & idempotent."""
    _require_admin(current_user)
    tenant_id = get_current_tenant_id()
    txn = await db[COLL].find_one({"id": txn_id, "tenant_id": tenant_id}, {"_id": 0})
    if not txn:
        raise HTTPException(status_code=404, detail="Transaction not found")
    if txn.get("direction") != "credit":
        raise HTTPException(status_code=400, detail="Only money-received (credit) transactions can be applied to an account's outstanding.")
    if (txn.get("account_adjustment") or {}).get("applied"):
        raise HTTPException(status_code=400, detail="This transaction is already applied to an account. Remove it first to re-apply.")

    account = await db.accounts.find_one(
        {"tenant_id": tenant_id, "$or": [{"id": payload.account_id}, {"account_id": payload.account_id}]},
        {"_id": 0, "id": 1, "account_name": 1, "outstanding_balance": 1},
    )
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")

    amount = round(float(txn.get("amount") or 0), 2)
    await db.accounts.update_one(
        {"tenant_id": tenant_id, "$or": [{"id": payload.account_id}, {"account_id": payload.account_id}]},
        {"$inc": {"outstanding_balance": -amount}, "$set": {"updated_at": _now()}},
    )
    adjustment = {
        "applied": True, "amount": amount, "account_id": payload.account_id,
        "applied_at": _now(), "applied_by": current_user.get("id"),
    }
    await db[COLL].update_one({"id": txn_id, "tenant_id": tenant_id}, {"$set": {
        "account_id": payload.account_id,
        "account_name": payload.account_name or account.get("account_name"),
        "account_adjustment": adjustment, "updated_at": _now(),
    }})
    return {"message": f"Applied ₹{amount} to {payload.account_name or account.get('account_name')}; outstanding reduced.", "adjustment": adjustment}


@router.post("/{txn_id}/unapply-account")
async def unapply_account(txn_id: str, current_user: dict = Depends(get_current_user)):
    """Reverse a previously-applied account payment, restoring the outstanding."""
    _require_admin(current_user)
    tenant_id = get_current_tenant_id()
    txn = await db[COLL].find_one({"id": txn_id, "tenant_id": tenant_id}, {"_id": 0})
    if not txn:
        raise HTTPException(status_code=404, detail="Transaction not found")
    adj = txn.get("account_adjustment") or {}
    if not adj.get("applied"):
        raise HTTPException(status_code=400, detail="Nothing applied on this transaction.")
    await db.accounts.update_one(
        {"tenant_id": tenant_id, "$or": [{"id": adj.get("account_id")}, {"account_id": adj.get("account_id")}]},
        {"$inc": {"outstanding_balance": round(float(adj.get("amount") or 0), 2)}, "$set": {"updated_at": _now()}},
    )
    await db[COLL].update_one({"id": txn_id, "tenant_id": tenant_id}, {"$set": {
        "account_id": None, "account_name": None, "account_adjustment": None, "updated_at": _now(),
    }})
    return {"message": "Account application reversed; outstanding restored."}


# ----------------------- Proofs -----------------------

@router.post("/{txn_id}/proofs")
async def upload_proof(
    txn_id: str,
    file: UploadFile = File(...),
    proof_type: str = Form("other"),
    current_user: dict = Depends(get_current_user),
):
    _require_admin(current_user)
    tenant_id = get_current_tenant_id()
    txn = await db[COLL].find_one({"id": txn_id, "tenant_id": tenant_id}, {"_id": 0, "id": 1, "txn_code": 1, "proofs": 1})
    if not txn:
        raise HTTPException(status_code=404, detail="Transaction not found")
    content_type = (file.content_type or guess_content_type(file.filename) or "").lower()
    ext = (file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else "")
    if content_type not in ALLOWED_PROOF_TYPES and ext not in {"png", "jpg", "jpeg", "webp", "gif", "pdf"}:
        raise HTTPException(status_code=400, detail="Only image files (PNG/JPG/WEBP/GIF) or PDF are allowed.")
    if not txn.get("txn_code"):
        await db[COLL].update_one({"id": txn_id, "tenant_id": tenant_id}, {"$set": {"txn_code": await _next_txn_code(tenant_id)}})
        txn["txn_code"] = (await db[COLL].find_one({"id": txn_id, "tenant_id": tenant_id}, {"_id": 0, "txn_code": 1}))["txn_code"]

    is_image = content_type.startswith("image/") or ext in {"png", "jpg", "jpeg", "webp", "gif"}
    serial = len([p for p in (txn.get("proofs") or []) if not p.get("is_deleted")]) + 1
    display_name = f"{txn['txn_code']}-{serial}.{ext or ('pdf' if content_type == 'application/pdf' else 'bin')}"

    data = await file.read()
    path = build_path(tenant_id, "accounting-proofs", display_name)
    result = await put_object(path, data, content_type or "application/octet-stream")
    proof = {
        "id": str(uuid.uuid4()), "type": proof_type,
        "storage_path": result.get("path", path), "original_filename": file.filename,
        "display_name": display_name, "serial": serial, "is_image": is_image,
        "content_type": content_type, "size": result.get("size", len(data)),
        "uploaded_at": _now(), "uploaded_by": current_user.get("id"), "is_deleted": False,
    }
    await db[COLL].update_one({"id": txn_id, "tenant_id": tenant_id},
                              {"$push": {"proofs": proof}, "$set": {"updated_at": _now()}})
    return proof


@router.get("/{txn_id}/proofs/{proof_id}/download")
async def download_proof(
    txn_id: str, proof_id: str,
    authorization: str = Header(None), auth: str = Query(None),
    current_user: dict = Depends(get_current_user),
):
    tenant_id = get_current_tenant_id()
    txn = await db[COLL].find_one({"id": txn_id, "tenant_id": tenant_id}, {"_id": 0, "proofs": 1})
    if not txn:
        raise HTTPException(status_code=404, detail="Transaction not found")
    proof = next((p for p in (txn.get("proofs") or []) if p["id"] == proof_id and not p.get("is_deleted")), None)
    if not proof:
        raise HTTPException(status_code=404, detail="Proof not found")
    data, ctype = await get_object(proof["storage_path"])
    return Response(content=data, media_type=proof.get("content_type") or ctype,
                    headers={"Content-Disposition": f'inline; filename="{proof.get("original_filename", "proof")}"'})


@router.delete("/{txn_id}/proofs/{proof_id}")
async def delete_proof(txn_id: str, proof_id: str, current_user: dict = Depends(get_current_user)):
    _require_admin(current_user)
    tenant_id = get_current_tenant_id()
    res = await db[COLL].update_one(
        {"id": txn_id, "tenant_id": tenant_id, "proofs.id": proof_id},
        {"$set": {"proofs.$.is_deleted": True, "updated_at": _now()}},
    )
    if not res.matched_count:
        raise HTTPException(status_code=404, detail="Proof not found")
    return {"ok": True}
