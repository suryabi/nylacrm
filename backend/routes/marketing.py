"""
Marketing Module - Calendar, Post Planning, and Master Data
Provides endpoints for managing marketing content calendar and social media planning.
"""

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from datetime import datetime, timezone
from deps import get_current_user
from core.tenant import get_current_tenant_id
from database import db
import uuid
import io

router = APIRouter()

# ---- Auto-Events (Global + Indian) ----
AUTO_EVENTS = [
    {"date": "01-01", "name": "New Year", "type": "global"},
    {"date": "01-13", "name": "Lohri", "type": "indian"},
    {"date": "01-14", "name": "Makar Sankranti / Pongal", "type": "indian"},
    {"date": "01-26", "name": "Republic Day", "type": "indian"},
    {"date": "02-14", "name": "Valentine's Day", "type": "global"},
    {"date": "03-08", "name": "International Women's Day", "type": "global"},
    {"date": "03-22", "name": "World Water Day", "type": "global"},
    {"date": "03-31", "name": "Holi", "type": "indian"},
    {"date": "04-07", "name": "World Health Day", "type": "global"},
    {"date": "04-14", "name": "Ambedkar Jayanti / Baisakhi", "type": "indian"},
    {"date": "04-22", "name": "Earth Day", "type": "global"},
    {"date": "05-01", "name": "International Workers' Day", "type": "global"},
    {"date": "05-11", "name": "Mother's Day", "type": "global"},
    {"date": "06-05", "name": "World Environment Day", "type": "global"},
    {"date": "06-15", "name": "Father's Day", "type": "global"},
    {"date": "06-21", "name": "International Yoga Day", "type": "global"},
    {"date": "07-04", "name": "US Independence Day", "type": "global"},
    {"date": "08-15", "name": "Independence Day", "type": "indian"},
    {"date": "08-19", "name": "World Photography Day", "type": "global"},
    {"date": "09-05", "name": "Teachers' Day", "type": "indian"},
    {"date": "09-27", "name": "World Tourism Day", "type": "global"},
    {"date": "10-02", "name": "Gandhi Jayanti", "type": "indian"},
    {"date": "10-31", "name": "Halloween", "type": "global"},
    {"date": "11-01", "name": "Diwali", "type": "indian"},
    {"date": "11-14", "name": "Children's Day", "type": "indian"},
    {"date": "11-27", "name": "Thanksgiving", "type": "global"},
    {"date": "12-25", "name": "Christmas", "type": "global"},
    {"date": "12-31", "name": "New Year's Eve", "type": "global"},
]

DEFAULT_CATEGORIES = [
    {"name": "Health", "color": "#A8E6CF"},
    {"name": "Water", "color": "#4EA8DE"},
    {"name": "Luxury", "color": "#C084FC"},
    {"name": "Sustainability", "color": "#34D399"},
    {"name": "Lifestyle", "color": "#FB923C"},
    {"name": "Brand", "color": "#FF6B6B"},
    {"name": "Product", "color": "#FDE74C"},
    {"name": "Event", "color": "#818CF8"},
]

DEFAULT_PLATFORMS = [
    {"name": "LinkedIn", "key": "linkedin", "color": "#0A66C2", "enabled": True},
    {"name": "WhatsApp", "key": "whatsapp", "color": "#25D366", "enabled": True},
    {"name": "YouTube", "key": "youtube", "color": "#FF0000", "enabled": True},
    {"name": "Instagram", "key": "instagram", "color": "#E1306C", "enabled": True},
    {"name": "Facebook", "key": "facebook", "color": "#1877F2", "enabled": True},
]


DEFAULT_EVENT_TYPES = [
    {"name": "Conference", "color": "#8B5CF6"},
    {"name": "Trade Show", "color": "#F59E0B"},
    {"name": "Webinar", "color": "#3B82F6"},
    {"name": "Product Launch", "color": "#EC4899"},
    {"name": "Workshop", "color": "#14B8A6"},
    {"name": "Meetup", "color": "#F97316"},
    {"name": "Press Event", "color": "#6366F1"},
]

VALID_EVENT_STATUSES = ["planned", "in_progress", "completed", "cancelled"]

VALID_STATUSES = ["draft", "review", "scheduled", "published"]
VALID_CONTENT_TYPES = ["reel", "image", "video", "other"]


# ---- Events ----
@router.get("/events")
async def get_events(month: int = None, year: int = None, current_user: dict = Depends(get_current_user)):
    """Get auto-events, optionally filtered by month."""
    tenant_id = get_current_tenant_id()

    # Get custom events
    query = {"tenant_id": tenant_id}
    custom_events = await db.marketing_events.find(query, {"_id": 0}).to_list(500)

    events = list(AUTO_EVENTS)
    for ce in custom_events:
        events.append({"date": ce.get("date", ""), "name": ce.get("name", ""), "type": "custom", "id": ce.get("id")})

    if month:
        month_str = f"{month:02d}"
        events = [e for e in events if e["date"].startswith(month_str)]

    return events


@router.post("/events")
async def create_custom_event(data: dict, current_user: dict = Depends(get_current_user)):
    """Create a custom event."""
    tenant_id = get_current_tenant_id()
    event = {
        "id": str(uuid.uuid4()),
        "tenant_id": tenant_id,
        "date": data.get("date", ""),
        "name": data.get("name", ""),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.marketing_events.insert_one(event)
    event.pop("_id", None)
    return event


@router.delete("/events/{event_id}")
async def delete_custom_event(event_id: str, current_user: dict = Depends(get_current_user)):
    tenant_id = get_current_tenant_id()
    result = await db.marketing_events.delete_one({"id": event_id, "tenant_id": tenant_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Event not found")
    return {"message": "Event deleted"}


# ---- Event Types (Master Data) ----
@router.get("/event-types")
async def get_event_types(current_user: dict = Depends(get_current_user)):
    tenant_id = get_current_tenant_id()
    types = await db.marketing_event_types.find({"tenant_id": tenant_id}, {"_id": 0}).to_list(100)
    if not types:
        for et in DEFAULT_EVENT_TYPES:
            doc = {"id": str(uuid.uuid4()), "tenant_id": tenant_id, **et, "created_at": datetime.now(timezone.utc).isoformat()}
            await db.marketing_event_types.insert_one(doc)
        types = await db.marketing_event_types.find({"tenant_id": tenant_id}, {"_id": 0}).to_list(100)
    return types


@router.post("/event-types")
async def create_event_type(data: dict, current_user: dict = Depends(get_current_user)):
    tenant_id = get_current_tenant_id()
    et = {
        "id": str(uuid.uuid4()),
        "tenant_id": tenant_id,
        "name": data.get("name", ""),
        "color": data.get("color", "#8B5CF6"),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.marketing_event_types.insert_one(et)
    et.pop("_id", None)
    return et


@router.put("/event-types/{et_id}")
async def update_event_type(et_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    tenant_id = get_current_tenant_id()
    update = {k: v for k, v in data.items() if k in ("name", "color")}
    if not update:
        raise HTTPException(status_code=400, detail="Nothing to update")
    await db.marketing_event_types.update_one({"id": et_id, "tenant_id": tenant_id}, {"$set": update})
    return {"message": "Event type updated"}


@router.delete("/event-types/{et_id}")
async def delete_event_type(et_id: str, current_user: dict = Depends(get_current_user)):
    tenant_id = get_current_tenant_id()
    result = await db.marketing_event_types.delete_one({"id": et_id, "tenant_id": tenant_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Event type not found")
    return {"message": "Event type deleted"}


# ---- Calendar Events (full events with requirements & tasks) ----
@router.get("/calendar-events")
async def get_calendar_events(month: int = None, year: int = None, current_user: dict = Depends(get_current_user)):
    tenant_id = get_current_tenant_id()
    query = {"tenant_id": tenant_id}
    if month and year:
        start = f"{year}-{month:02d}-01"
        end = f"{year + 1}-01-01" if month == 12 else f"{year}-{month + 1:02d}-01"
        query["event_date"] = {"$gte": start, "$lt": end}
    events = await db.marketing_calendar_events.find(query, {"_id": 0}).sort("event_date", 1).to_list(500)
    return events


@router.get("/calendar-events/{event_id}")
async def get_calendar_event(event_id: str, current_user: dict = Depends(get_current_user)):
    tenant_id = get_current_tenant_id()
    ev = await db.marketing_calendar_events.find_one({"id": event_id, "tenant_id": tenant_id}, {"_id": 0})
    if not ev:
        raise HTTPException(status_code=404, detail="Calendar event not found")
    return ev


@router.post("/calendar-events")
async def create_calendar_event(data: dict, current_user: dict = Depends(get_current_user)):
    tenant_id = get_current_tenant_id()
    tasks = data.get("tasks", [])
    for t in tasks:
        if not t.get("id"):
            t["id"] = str(uuid.uuid4())
        t.setdefault("status", "pending")
    event = {
        "id": str(uuid.uuid4()),
        "tenant_id": tenant_id,
        "name": data.get("name", ""),
        "event_date": data.get("event_date", ""),
        "start_time": data.get("start_time", ""),
        "end_time": data.get("end_time", ""),
        "description": data.get("description", ""),
        "location": data.get("location", ""),
        "budget": data.get("budget"),
        "expected_attendees": data.get("expected_attendees"),
        "event_type": data.get("event_type", ""),
        "event_type_color": data.get("event_type_color", "#8B5CF6"),
        "requirements": data.get("requirements", []),
        "tasks": tasks,
        "status": data.get("status", "planned"),
        "created_by": current_user.get("id"),
        "created_by_name": current_user.get("name", ""),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.marketing_calendar_events.insert_one(event)
    event.pop("_id", None)
    return event


@router.put("/calendar-events/{event_id}")
async def update_calendar_event(event_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    tenant_id = get_current_tenant_id()
    existing = await db.marketing_calendar_events.find_one({"id": event_id, "tenant_id": tenant_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Calendar event not found")
    allowed = ["name", "event_date", "start_time", "end_time", "description", "location",
               "budget", "expected_attendees", "event_type", "event_type_color",
               "requirements", "tasks", "status"]
    update = {k: v for k, v in data.items() if k in allowed}
    if "tasks" in update:
        for t in update["tasks"]:
            if not t.get("id"):
                t["id"] = str(uuid.uuid4())
            t.setdefault("status", "pending")
    update["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.marketing_calendar_events.update_one({"id": event_id, "tenant_id": tenant_id}, {"$set": update})
    updated = await db.marketing_calendar_events.find_one({"id": event_id, "tenant_id": tenant_id}, {"_id": 0})
    return updated


@router.delete("/calendar-events/{event_id}")
async def delete_calendar_event(event_id: str, current_user: dict = Depends(get_current_user)):
    tenant_id = get_current_tenant_id()
    result = await db.marketing_calendar_events.delete_one({"id": event_id, "tenant_id": tenant_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Calendar event not found")
    return {"message": "Calendar event deleted"}


# ---- Comments (Posts & Events) ----

@router.get("/comments/{entity_type}/{entity_id}")
async def get_comments(entity_type: str, entity_id: str, current_user: dict = Depends(get_current_user)):
    """Get comments for a post or event"""
    if entity_type not in ("post", "event"):
        raise HTTPException(status_code=400, detail="entity_type must be 'post' or 'event'")
    tenant_id = get_current_tenant_id()
    comments = await db.marketing_comments.find(
        {"tenant_id": tenant_id, "entity_type": entity_type, "entity_id": entity_id},
        {"_id": 0}
    ).sort("created_at", 1).to_list(500)
    return comments


@router.post("/comments/{entity_type}/{entity_id}")
async def add_comment(entity_type: str, entity_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    """Add a comment to a post or event"""
    if entity_type not in ("post", "event"):
        raise HTTPException(status_code=400, detail="entity_type must be 'post' or 'event'")
    content = data.get("content", "").strip()
    if not content:
        raise HTTPException(status_code=400, detail="Comment content is required")
    tenant_id = get_current_tenant_id()
    comment = {
        "id": str(uuid.uuid4()),
        "tenant_id": tenant_id,
        "entity_type": entity_type,
        "entity_id": entity_id,
        "content": content,
        "created_by": current_user.get("id"),
        "created_by_name": current_user.get("name", ""),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.marketing_comments.insert_one(comment)
    comment.pop("_id", None)
    return comment


@router.delete("/comments/{comment_id}")
async def delete_comment(comment_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a comment (author or admin only)"""
    tenant_id = get_current_tenant_id()
    comment = await db.marketing_comments.find_one({"id": comment_id, "tenant_id": tenant_id})
    if not comment:
        raise HTTPException(status_code=404, detail="Comment not found")
    user_role = current_user.get("role", "")
    is_admin = user_role.lower() in ("ceo", "director", "system admin")
    if comment["created_by"] != current_user["id"] and not is_admin:
        raise HTTPException(status_code=403, detail="Only comment author or admin can delete")
    await db.marketing_comments.delete_one({"id": comment_id})
    return {"message": "Comment deleted"}



# ---- Categories (Master Data) ----
@router.get("/categories")
async def get_categories(current_user: dict = Depends(get_current_user)):
    tenant_id = get_current_tenant_id()
    cats = await db.marketing_categories.find({"tenant_id": tenant_id}, {"_id": 0}).to_list(100)
    if not cats:
        # Seed defaults
        for c in DEFAULT_CATEGORIES:
            doc = {"id": str(uuid.uuid4()), "tenant_id": tenant_id, **c, "created_at": datetime.now(timezone.utc).isoformat()}
            await db.marketing_categories.insert_one(doc)
        cats = await db.marketing_categories.find({"tenant_id": tenant_id}, {"_id": 0}).to_list(100)
    return cats


@router.post("/categories")
async def create_category(data: dict, current_user: dict = Depends(get_current_user)):
    tenant_id = get_current_tenant_id()
    cat = {
        "id": str(uuid.uuid4()),
        "tenant_id": tenant_id,
        "name": data.get("name", ""),
        "color": data.get("color", "#FF6B6B"),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.marketing_categories.insert_one(cat)
    cat.pop("_id", None)
    return cat


@router.put("/categories/{cat_id}")
async def update_category(cat_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    tenant_id = get_current_tenant_id()
    update = {}
    if "name" in data:
        update["name"] = data["name"]
    if "color" in data:
        update["color"] = data["color"]
    if not update:
        raise HTTPException(status_code=400, detail="Nothing to update")
    await db.marketing_categories.update_one({"id": cat_id, "tenant_id": tenant_id}, {"$set": update})
    return {"message": "Category updated"}


@router.delete("/categories/{cat_id}")
async def delete_category(cat_id: str, current_user: dict = Depends(get_current_user)):
    tenant_id = get_current_tenant_id()
    result = await db.marketing_categories.delete_one({"id": cat_id, "tenant_id": tenant_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Category not found")
    return {"message": "Category deleted"}


# ---- Platforms (Master Data) ----
@router.get("/platforms")
async def get_platforms(current_user: dict = Depends(get_current_user)):
    tenant_id = get_current_tenant_id()
    plats = await db.marketing_platforms.find({"tenant_id": tenant_id}, {"_id": 0}).to_list(50)
    if not plats:
        for p in DEFAULT_PLATFORMS:
            doc = {"id": str(uuid.uuid4()), "tenant_id": tenant_id, **p, "created_at": datetime.now(timezone.utc).isoformat()}
            await db.marketing_platforms.insert_one(doc)
        plats = await db.marketing_platforms.find({"tenant_id": tenant_id}, {"_id": 0}).to_list(50)
    return plats


@router.put("/platforms/{plat_id}")
async def update_platform(plat_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    tenant_id = get_current_tenant_id()
    update = {}
    if "enabled" in data:
        update["enabled"] = data["enabled"]
    if "name" in data:
        update["name"] = data["name"]
    if not update:
        raise HTTPException(status_code=400, detail="Nothing to update")
    await db.marketing_platforms.update_one({"id": plat_id, "tenant_id": tenant_id}, {"$set": update})
    return {"message": "Platform updated"}


# ---- Posts (CRUD + Calendar) ----
@router.get("/posts")
async def get_posts(
    month: int = None, year: int = None,
    status: str = None, category: str = None,
    current_user: dict = Depends(get_current_user)
):
    """Get posts, optionally filtered by month/year, status, or category."""
    tenant_id = get_current_tenant_id()
    query = {"tenant_id": tenant_id}

    if month and year:
        start = f"{year}-{month:02d}-01"
        if month == 12:
            end = f"{year + 1}-01-01"
        else:
            end = f"{year}-{month + 1:02d}-01"
        query["post_date"] = {"$gte": start, "$lt": end}

    if status:
        query["status"] = status
    if category:
        query["category"] = category

    posts = await db.marketing_posts.find(query, {"_id": 0}).sort("post_date", 1).to_list(500)
    return posts


@router.get("/posts/{post_id}")
async def get_post(post_id: str, current_user: dict = Depends(get_current_user)):
    tenant_id = get_current_tenant_id()
    post = await db.marketing_posts.find_one({"id": post_id, "tenant_id": tenant_id}, {"_id": 0})
    if not post:
        raise HTTPException(status_code=404, detail="Post not found")
    return post


@router.post("/posts")
async def create_post(data: dict, current_user: dict = Depends(get_current_user)):
    """Create a new marketing post."""
    tenant_id = get_current_tenant_id()

    post = {
        "id": str(uuid.uuid4()),
        "tenant_id": tenant_id,
        "post_date": data.get("post_date", ""),
        "category": data.get("category", ""),
        "content_type": data.get("content_type", "image"),
        "concept": data.get("concept", ""),
        "message": data.get("message", ""),
        "platforms": data.get("platforms", ["linkedin", "whatsapp", "youtube", "instagram", "facebook"]),
        "platform_links": {},
        "status": data.get("status", "draft"),
        "owner_id": data.get("owner_id", current_user.get("id")),
        "owner_name": data.get("owner_name", current_user.get("name", "")),
        "created_by": current_user.get("id"),
        "created_by_name": current_user.get("name", ""),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.marketing_posts.insert_one(post)
    post.pop("_id", None)
    return post


@router.put("/posts/{post_id}")
async def update_post(post_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    """Update a marketing post."""
    tenant_id = get_current_tenant_id()

    existing = await db.marketing_posts.find_one({"id": post_id, "tenant_id": tenant_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Post not found")

    allowed_fields = ["post_date", "category", "content_type", "concept", "message", "platforms", "status", "owner_id", "owner_name", "platform_links"]
    update = {k: v for k, v in data.items() if k in allowed_fields}
    update["updated_at"] = datetime.now(timezone.utc).isoformat()
    update["updated_by"] = current_user.get("id")

    await db.marketing_posts.update_one({"id": post_id, "tenant_id": tenant_id}, {"$set": update})
    updated = await db.marketing_posts.find_one({"id": post_id, "tenant_id": tenant_id}, {"_id": 0})
    return updated


@router.delete("/posts/{post_id}")
async def delete_post(post_id: str, current_user: dict = Depends(get_current_user)):
    tenant_id = get_current_tenant_id()
    result = await db.marketing_posts.delete_one({"id": post_id, "tenant_id": tenant_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Post not found")
    return {"message": "Post deleted"}


@router.put("/posts/{post_id}/status")
async def update_post_status(post_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    """Update only the status of a post (workflow transition)."""
    tenant_id = get_current_tenant_id()
    new_status = data.get("status")
    if new_status not in VALID_STATUSES:
        raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of: {VALID_STATUSES}")

    existing = await db.marketing_posts.find_one({"id": post_id, "tenant_id": tenant_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Post not found")

    await db.marketing_posts.update_one(
        {"id": post_id, "tenant_id": tenant_id},
        {"$set": {"status": new_status, "updated_at": datetime.now(timezone.utc).isoformat(), "updated_by": current_user.get("id")}}
    )
    return {"message": f"Status updated to {new_status}"}


VALID_ANALYTICS_FIELDS = ["url", "views", "likes", "comments", "shares", "subscribers_added"]


@router.put("/posts/{post_id}/links")
async def update_post_links(post_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    """Update platform links and analytics for a post.
    Expects: { "platform_links": { "linkedin": { "url": "...", "views": 100, ... }, ... } }
    """
    tenant_id = get_current_tenant_id()

    existing = await db.marketing_posts.find_one({"id": post_id, "tenant_id": tenant_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Post not found")

    incoming = data.get("platform_links", {})
    if not isinstance(incoming, dict):
        raise HTTPException(status_code=400, detail="platform_links must be an object")

    current_links = existing.get("platform_links", {})
    post_platforms = existing.get("platforms", [])

    # Only accept data for platforms assigned to this post
    for pk, metrics in incoming.items():
        if pk not in post_platforms:
            continue
        if not isinstance(metrics, dict):
            continue
        clean = {}
        for field in VALID_ANALYTICS_FIELDS:
            if field in metrics:
                if field == "url":
                    clean[field] = str(metrics[field])
                else:
                    try:
                        clean[field] = int(metrics[field])
                    except (ValueError, TypeError):
                        clean[field] = 0
        # Merge with existing
        if pk in current_links:
            current_links[pk].update(clean)
        else:
            current_links[pk] = clean

    await db.marketing_posts.update_one(
        {"id": post_id, "tenant_id": tenant_id},
        {"$set": {
            "platform_links": current_links,
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "updated_by": current_user.get("id"),
        }}
    )
    updated = await db.marketing_posts.find_one({"id": post_id, "tenant_id": tenant_id}, {"_id": 0})
    return updated


# ---- Calendar Summary ----
@router.get("/calendar")
async def get_calendar_data(month: int = Query(...), year: int = Query(...), current_user: dict = Depends(get_current_user)):
    """Get calendar view data: posts grouped by date + events for the month."""
    tenant_id = get_current_tenant_id()

    start = f"{year}-{month:02d}-01"
    if month == 12:
        end = f"{year + 1}-01-01"
    else:
        end = f"{year}-{month + 1:02d}-01"

    posts = await db.marketing_posts.find(
        {"tenant_id": tenant_id, "post_date": {"$gte": start, "$lt": end}},
        {"_id": 0}
    ).sort("post_date", 1).to_list(500)

    # Group posts by date
    posts_by_date = {}
    for p in posts:
        d = p.get("post_date", "")[:10]
        if d not in posts_by_date:
            posts_by_date[d] = []
        posts_by_date[d].append(p)

    # Get events for this month
    month_str = f"{month:02d}"
    auto = [e for e in AUTO_EVENTS if e["date"].startswith(month_str)]
    # Custom events stored as YYYY-MM-DD — match by month
    custom_query = {
        "tenant_id": tenant_id,
        "$or": [
            {"date": {"$regex": f"^{month_str}-"}},
            {"date": {"$regex": f"^{year}-{month_str}-"}},
        ]
    }
    custom = await db.marketing_events.find(custom_query, {"_id": 0}).to_list(100)
    events = auto + [{"date": c.get("date", ""), "name": c["name"], "type": "custom", "id": c.get("id")} for c in custom]

    # Get calendar events (full events with requirements & tasks) for this month
    cal_events = await db.marketing_calendar_events.find(
        {"tenant_id": tenant_id, "event_date": {"$gte": start, "$lt": end}},
        {"_id": 0}
    ).sort("event_date", 1).to_list(200)

    # Group calendar events by date
    cal_events_by_date = {}
    for ce in cal_events:
        d = ce.get("event_date", "")[:10]
        if d not in cal_events_by_date:
            cal_events_by_date[d] = []
        cal_events_by_date[d].append(ce)

    # Stats
    total = len(posts)
    by_status = {}
    by_category = {}
    by_content_type = {}
    for p in posts:
        s = p.get("status", "draft")
        by_status[s] = by_status.get(s, 0) + 1
        cat = p.get("category", "Uncategorized")
        if cat:
            by_category[cat] = by_category.get(cat, 0) + 1
        ct = p.get("content_type", "other")
        by_content_type[ct] = by_content_type.get(ct, 0) + 1

    return {
        "posts_by_date": posts_by_date,
        "events": events,
        "calendar_events_by_date": cal_events_by_date,
        "stats": {
            "total": total,
            "by_status": by_status,
            "by_category": by_category,
            "by_content_type": by_content_type,
            "events_count": len(events),
            "calendar_events_count": len(cal_events),
        },
        "month": month,
        "year": year,
    }


# ---- Spreadsheet Template, Upload, Export ----

TEMPLATE_COLUMNS = ["Post Date", "Category", "Content Type", "Concept", "Message / Caption", "Platforms", "Status"]
PLATFORM_KEYS = ["linkedin", "whatsapp", "youtube", "instagram", "facebook"]


def _build_workbook(rows=None, is_template=False):
    """Build an openpyxl Workbook. If is_template, adds instructions and a sample row."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Marketing Calendar"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="d0d5dd"),
        right=Side(style="thin", color="d0d5dd"),
        top=Side(style="thin", color="d0d5dd"),
        bottom=Side(style="thin", color="d0d5dd"),
    )

    # Header row
    for col_idx, col_name in enumerate(TEMPLATE_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Column widths
    widths = [14, 16, 14, 40, 50, 40, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    if is_template:
        # Sample row
        sample = ["2026-04-15", "Health", "reel", "World Health Day Reel", "Celebrate wellness with us!", "linkedin,instagram,youtube", "draft"]
        for col_idx, val in enumerate(sample, 1):
            cell = ws.cell(row=2, column=col_idx, value=val)
            cell.font = Font(italic=True, color="888888")
            cell.border = thin_border

        # Instructions sheet
        ins = wb.create_sheet("Instructions")
        instructions = [
            ("Column", "Description", "Valid Values"),
            ("Post Date", "Date in YYYY-MM-DD format", "e.g. 2026-04-15"),
            ("Category", "Marketing category name", "e.g. Health, Water, Luxury, Brand, Product, Event"),
            ("Content Type", "Type of content", "reel, image, video, other"),
            ("Concept", "Brief concept or idea", "Free text"),
            ("Message / Caption", "Post message or caption", "Free text (optional)"),
            ("Platforms", "Comma-separated platform keys", "linkedin, whatsapp, youtube, instagram, facebook"),
            ("Status", "Post workflow status", "draft, review, scheduled, published"),
        ]
        for r, row_data in enumerate(instructions, 1):
            for c, val in enumerate(row_data, 1):
                cell = ins.cell(row=r, column=c, value=val)
                if r == 1:
                    cell.font = Font(bold=True)
                ins.column_dimensions[chr(64 + c)].width = 30
    elif rows:
        for r_idx, row in enumerate(rows, 2):
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.border = thin_border

    return wb


def _wb_to_response(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/template")
async def download_template(current_user: dict = Depends(get_current_user)):
    """Download an empty Excel template for bulk upload."""
    wb = _build_workbook(is_template=True)
    return _wb_to_response(wb, "marketing_calendar_template.xlsx")


@router.get("/export")
async def export_posts(month: int = Query(...), year: int = Query(...), current_user: dict = Depends(get_current_user)):
    """Export posts for a given month/year as Excel."""
    tenant_id = get_current_tenant_id()
    start = f"{year}-{month:02d}-01"
    end = f"{year + 1}-01-01" if month == 12 else f"{year}-{month + 1:02d}-01"

    posts = await db.marketing_posts.find(
        {"tenant_id": tenant_id, "post_date": {"$gte": start, "$lt": end}},
        {"_id": 0}
    ).sort("post_date", 1).to_list(500)

    rows = []
    for p in posts:
        rows.append([
            p.get("post_date", ""),
            p.get("category", ""),
            p.get("content_type", ""),
            p.get("concept", ""),
            p.get("message", ""),
            ",".join(p.get("platforms", [])),
            p.get("status", "draft"),
        ])

    wb = _build_workbook(rows=rows)
    filename = f"marketing_calendar_{year}_{month:02d}.xlsx"
    return _wb_to_response(wb, filename)


@router.post("/upload-preview")
async def upload_preview(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    """Parse an uploaded Excel/CSV and return preview data without saving."""
    from openpyxl import load_workbook
    import csv as csv_mod

    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")

    content = await file.read()
    parsed_rows = []
    errors = []

    try:
        if file.filename.endswith(".csv"):
            text = content.decode("utf-8-sig")
            reader = csv_mod.DictReader(io.StringIO(text))
            for i, row in enumerate(reader, 2):
                parsed_rows.append({
                    "row_num": i,
                    "post_date": (row.get("Post Date") or "").strip(),
                    "category": (row.get("Category") or "").strip(),
                    "content_type": (row.get("Content Type") or "image").strip().lower(),
                    "concept": (row.get("Concept") or "").strip(),
                    "message": (row.get("Message / Caption") or row.get("Message") or "").strip(),
                    "platforms": [p.strip().lower() for p in (row.get("Platforms") or "").split(",") if p.strip()],
                    "status": (row.get("Status") or "draft").strip().lower(),
                })
        else:
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]

            col_map = {}
            for idx, h in enumerate(headers):
                hl = h.lower()
                if "date" in hl:
                    col_map["post_date"] = idx
                elif "categ" in hl:
                    col_map["category"] = idx
                elif "type" in hl:
                    col_map["content_type"] = idx
                elif "concept" in hl:
                    col_map["concept"] = idx
                elif "message" in hl or "caption" in hl:
                    col_map["message"] = idx
                elif "platform" in hl:
                    col_map["platforms"] = idx
                elif "status" in hl:
                    col_map["status"] = idx

            for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                if not row or all(v is None for v in row):
                    continue
                vals = list(row)

                def get_val(key, default=""):
                    idx = col_map.get(key)
                    if idx is not None and idx < len(vals):
                        v = vals[idx]
                        if v is None:
                            return default
                        if isinstance(v, datetime):
                            return v.strftime("%Y-%m-%d")
                        return str(v).strip()
                    return default

                date_str = get_val("post_date")
                plat_str = get_val("platforms", "linkedin,instagram,youtube,facebook,whatsapp")
                platforms = [p.strip().lower() for p in plat_str.split(",") if p.strip()]

                parsed_rows.append({
                    "row_num": r_idx,
                    "post_date": date_str,
                    "category": get_val("category"),
                    "content_type": get_val("content_type", "image").lower(),
                    "concept": get_val("concept"),
                    "message": get_val("message"),
                    "platforms": platforms,
                    "status": get_val("status", "draft").lower(),
                })
            wb.close()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")

    # Validate rows
    valid_rows = []
    for row in parsed_rows:
        row_errors = []
        # Date validation — accept multiple formats and normalize to YYYY-MM-DD
        if not row["post_date"]:
            row_errors.append("Missing date")
        else:
            date_str = row["post_date"].strip()
            parsed_date = None
            for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d", "%d.%m.%Y"):
                try:
                    parsed_date = datetime.strptime(date_str[:10], fmt)
                    break
                except ValueError:
                    continue
            if parsed_date:
                row["post_date"] = parsed_date.strftime("%Y-%m-%d")
            else:
                row_errors.append(f"Invalid date format: {row['post_date']} (use YYYY-MM-DD)")

        if not row["concept"]:
            row_errors.append("Missing concept")

        if row["content_type"] not in VALID_CONTENT_TYPES:
            row_errors.append(f"Invalid content type: {row['content_type']}")

        if row["status"] not in VALID_STATUSES:
            row["status"] = "draft"

        row["platforms"] = [p for p in row["platforms"] if p in PLATFORM_KEYS]
        if not row["platforms"]:
            row["platforms"] = PLATFORM_KEYS

        row["valid"] = len(row_errors) == 0
        row["errors"] = row_errors
        valid_rows.append(row)

    valid_count = sum(1 for r in valid_rows if r["valid"])
    return {
        "rows": valid_rows,
        "total": len(valid_rows),
        "valid_count": valid_count,
        "error_count": len(valid_rows) - valid_count,
    }


@router.post("/upload-confirm")
async def upload_confirm(data: dict, current_user: dict = Depends(get_current_user)):
    """Replace all posts for a month with the uploaded data.
    Expects: { "month": 4, "year": 2026, "rows": [ { post_date, category, ... }, ... ] }
    """
    tenant_id = get_current_tenant_id()
    month = data.get("month")
    year = data.get("year")
    rows = data.get("rows", [])

    if not month or not year:
        raise HTTPException(status_code=400, detail="month and year are required")
    if not rows:
        raise HTTPException(status_code=400, detail="No rows to save")

    # Delete existing posts ONLY for dates present in the upload
    upload_dates = list(set(r.get("post_date", "") for r in rows if r.get("post_date")))
    delete_result = await db.marketing_posts.delete_many({
        "tenant_id": tenant_id,
        "post_date": {"$in": upload_dates}
    })

    # Insert new posts
    now_iso = datetime.now(timezone.utc).isoformat()
    new_posts = []
    for row in rows:
        post = {
            "id": str(uuid.uuid4()),
            "tenant_id": tenant_id,
            "post_date": row.get("post_date", ""),
            "category": row.get("category", ""),
            "content_type": row.get("content_type", "image"),
            "concept": row.get("concept", ""),
            "message": row.get("message", ""),
            "platforms": row.get("platforms", PLATFORM_KEYS),
            "platform_links": {},
            "status": row.get("status", "draft"),
            "owner_id": current_user.get("id"),
            "owner_name": current_user.get("name", ""),
            "created_by": current_user.get("id"),
            "created_by_name": current_user.get("name", ""),
            "created_at": now_iso,
            "updated_at": now_iso,
        }
        new_posts.append(post)

    if new_posts:
        await db.marketing_posts.insert_many(new_posts)
        # Remove _id from response
        for p in new_posts:
            p.pop("_id", None)

    return {
        "message": f"Replaced {delete_result.deleted_count} existing posts with {len(new_posts)} new posts for {year}-{month:02d}",
        "deleted": delete_result.deleted_count,
        "inserted": len(new_posts),
    }

