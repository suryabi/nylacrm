"""
Invoices Routes - List, filter, and manage invoices
"""
from fastapi import APIRouter, HTTPException, Depends, Query, Request
from typing import Optional, List
from datetime import datetime, timezone
import logging

logger = logging.getLogger(__name__)

router = APIRouter(tags=["invoices"])

# Import dependencies
from database import get_tenant_db
from deps import get_current_user

def get_tdb():
    return get_tenant_db()


async def _build_invoice_query(
    tdb,
    *,
    search=None,
    territory=None,
    state=None,
    city=None,
    account_name=None,
    account_names=None,
    status=None,
    date_from=None,
    date_to=None,
    time_filter=None,
):
    """Build the MongoDB query dict for invoice list/export from filters.

    Returns (query, is_empty). When is_empty is True, the caller should return
    an empty result set (territory/state/city matched zero accounts).
    """
    conditions = []

    if search:
        conditions.append({
            '$or': [
                {'invoice_no': {'$regex': search, '$options': 'i'}},
                {'account_name': {'$regex': search, '$options': 'i'}},
            ]
        })

    # Multi-select exact account names (from the autocomplete filter)
    if account_names:
        names = [n for n in account_names if n]
        if names:
            conditions.append({'account_name': {'$in': names}})
    # Single free-text account name (legacy)
    elif account_name:
        conditions.append({'account_name': {'$regex': account_name, '$options': 'i'}})

    # Status is derived from real account linkage (account_uuid / account_id /
    # ca_lead_id), NOT the stale stored `status` field. An invoice that resolves
    # to an account (and shows on the account detail page) is "matched".
    _linked_or = [
        {'account_uuid': {'$nin': [None, '']}},
        {'account_id': {'$nin': [None, '']}},
        {'ca_lead_id': {'$nin': [None, '']}},
    ]
    if status == 'matched':
        conditions.append({'$or': _linked_or})
    elif status == 'unmatched':
        conditions.append({'$nor': _linked_or})

    if date_from:
        conditions.append({'invoice_date': {'$gte': date_from}})

    if date_to:
        conditions.append({'invoice_date': {'$lte': date_to}})

    if time_filter and time_filter != 'lifetime':
        from datetime import timedelta
        now = datetime.now(timezone.utc)
        date_ranges = {
            'this_week': (now - timedelta(days=now.weekday()), now),
            'last_week': (now - timedelta(days=now.weekday() + 7), now - timedelta(days=now.weekday())),
            'this_month': (now.replace(day=1), now),
            'last_month': ((now.replace(day=1) - timedelta(days=1)).replace(day=1), now.replace(day=1) - timedelta(days=1)),
            'last_3_months': (now - timedelta(days=90), now),
            'last_6_months': (now - timedelta(days=180), now),
            'this_quarter': (now.replace(month=((now.month - 1) // 3) * 3 + 1, day=1), now),
        }
        if time_filter in date_ranges:
            start, end = date_ranges[time_filter]
            if start:
                conditions.append({'invoice_date': {'$gte': start.strftime('%Y-%m-%d')}})
            if end:
                conditions.append({'invoice_date': {'$lte': end.strftime('%Y-%m-%d')}})

    if territory or state or city:
        account_query = {}
        if territory:
            account_query['territory'] = territory
        if state:
            account_query['state'] = state
        if city:
            account_query['city'] = city

        accounts = await tdb.accounts.find(account_query, {'_id': 0, 'id': 1, 'account_id': 1}).to_list(10000)

        if accounts:
            account_ids_list = [a.get('account_id') for a in accounts if a.get('account_id')]
            account_uuids_list = [a.get('id') for a in accounts if a.get('id')]
            account_filter = {'$or': []}
            if account_ids_list:
                account_filter['$or'].append({'account_id': {'$in': account_ids_list}})
                account_filter['$or'].append({'account_id_from_mq': {'$in': account_ids_list}})
            if account_uuids_list:
                account_filter['$or'].append({'account_uuid': {'$in': account_uuids_list}})
            if account_filter['$or']:
                conditions.append(account_filter)
        else:
            return ({}, True)

    query = {'$and': conditions} if conditions else {}
    return (query, False)


async def _build_credit_note_account_map(tdb):
    """Aggregate non-cancelled credit notes by account → {issued, applied, balance}.

    Credit notes are customer/account-level (generated from returns), not tied
    to individual invoices. Returns (by_id, by_name) lookup maps.
    """
    by_id = {}
    by_name = {}
    cns = await tdb.credit_notes.find(
        {'status': {'$ne': 'cancelled'}},
        {'_id': 0, 'account_id': 1, 'account_name': 1,
         'original_amount': 1, 'applied_amount': 1, 'balance_amount': 1}
    ).to_list(100000)
    for cn in cns:
        issued = cn.get('original_amount') or 0
        applied = cn.get('applied_amount') or 0
        balance = cn.get('balance_amount')
        if balance is None:
            balance = issued - applied
        aid = cn.get('account_id')
        aname = (cn.get('account_name') or '').strip().lower()
        if aid:
            agg = by_id.setdefault(aid, {'issued': 0, 'applied': 0, 'balance': 0})
            agg['issued'] += issued; agg['applied'] += applied; agg['balance'] += balance
        if aname:
            agg = by_name.setdefault(aname, {'issued': 0, 'applied': 0, 'balance': 0})
            agg['issued'] += issued; agg['applied'] += applied; agg['balance'] += balance
    return by_id, by_name


def _cn_for_invoice(inv, by_id, by_name):
    """Look up an invoice's account-level credit note totals."""
    aid = inv.get('account_id') or inv.get('account_uuid')
    if aid and aid in by_id:
        return by_id[aid]
    nm = (inv.get('account_name') or '').strip().lower()
    if nm and nm in by_name:
        return by_name[nm]
    return {'issued': 0, 'applied': 0, 'balance': 0}


def _derive_invoice_status(inv):
    """An invoice is 'matched' if it has real account linkage (same criteria
    that makes it appear on the account detail page)."""
    if inv.get('account_uuid') or inv.get('account_id') or inv.get('ca_lead_id'):
        return 'matched'
    return 'unmatched'


async def _build_invoice_applied_credit_map(tdb, invoice_nos):
    """Map invoice_no -> {credit, applications} for credit notes applied to the
    originating stock-out delivery.

    Linkage: invoice.invoice_no == zoho_invoice_mappings.zoho_invoice_number
    (source_type='distributor_delivery', status='synced') → mapping.source_id is
    the distributor_deliveries.id holding total_credit_applied + applied_credit_notes.
    """
    invoice_nos = [n for n in (invoice_nos or []) if n]
    if not invoice_nos:
        return {}
    mappings = await tdb.zoho_invoice_mappings.find(
        {'source_type': 'distributor_delivery', 'status': 'synced',
         'zoho_invoice_number': {'$in': invoice_nos}},
        {'_id': 0, 'zoho_invoice_number': 1, 'source_id': 1}
    ).to_list(100000)
    if not mappings:
        return {}
    delivery_to_invoice = {}
    for m in mappings:
        did = m.get('source_id')
        inv_no = m.get('zoho_invoice_number')
        if did and inv_no:
            delivery_to_invoice[did] = inv_no
    if not delivery_to_invoice:
        return {}
    deliveries = await tdb.distributor_deliveries.find(
        {'id': {'$in': list(delivery_to_invoice.keys())}, 'total_credit_applied': {'$gt': 0}},
        {'_id': 0, 'id': 1, 'total_credit_applied': 1, 'applied_credit_notes': 1}
    ).to_list(100000)
    out = {}
    for d in deliveries:
        inv_no = delivery_to_invoice.get(d.get('id'))
        if inv_no:
            out[inv_no] = {
                'credit': d.get('total_credit_applied') or 0,
                'applications': d.get('applied_credit_notes') or [],
            }
    return out


@router.get("/credit-notes")
async def list_invoice_credit_notes(
    current_user: dict = Depends(get_current_user),
):
    """List all customer/account-level credit notes (for the Invoices page
    'Credit Notes' tab), with issued / applied / balance totals."""
    tdb = get_tdb()
    cns = await tdb.credit_notes.find({}, {'_id': 0}).to_list(100000)
    cns.sort(key=lambda c: c.get('credit_note_date') or '', reverse=True)

    out = []
    t_issued = t_applied = t_balance = 0
    for cn in cns:
        issued = cn.get('original_amount') or 0
        applied = cn.get('applied_amount') or 0
        balance = cn.get('balance_amount')
        if balance is None:
            balance = issued - applied
        if cn.get('status') != 'cancelled':
            t_issued += issued
            t_applied += applied
            t_balance += balance
        out.append({
            'credit_note_number': cn.get('credit_note_number'),
            'account_id': cn.get('account_id'),
            'account_name': cn.get('account_name'),
            'original_amount': issued,
            'applied_amount': applied,
            'balance_amount': balance,
            'status': cn.get('status'),
            'credit_note_date': cn.get('credit_note_date'),
            'return_number': cn.get('return_number'),
            'notes': cn.get('notes'),
        })
    return {
        'credit_notes': out,
        'summary': {
            'total_issued': t_issued,
            'total_applied': t_applied,
            'total_balance': t_balance,
            'count': len(out),
        },
    }


@router.get("/account-options")
async def list_invoice_account_options(
    current_user: dict = Depends(get_current_user),
):
    """Return distinct account names present on invoices (enriched with city /
    contact from the accounts collection), sorted alphabetically.

    Used to populate the autocomplete account filter on the Invoices page —
    mirrors the Stock-Out account selector (two-line name + city/contact).
    """
    tdb = get_tdb()
    names = await tdb.invoices.distinct('account_name')
    clean_names = sorted(
        [n for n in names if n and isinstance(n, str)],
        key=lambda s: s.lower()
    )

    # Enrich with city / state / contact from accounts collection (best-effort)
    detail_map = {}
    if clean_names:
        accounts = await tdb.accounts.find(
            {'account_name': {'$in': clean_names}},
            {'_id': 0, 'account_name': 1, 'city': 1, 'state': 1, 'contact_name': 1, 'territory': 1}
        ).to_list(len(clean_names) * 2)
        for acc in accounts:
            nm = acc.get('account_name')
            if nm and nm not in detail_map:
                detail_map[nm] = {
                    'city': acc.get('city') or '',
                    'state': acc.get('state') or '',
                    'contact_name': acc.get('contact_name') or '',
                    'territory': acc.get('territory') or '',
                }

    options = [
        {
            'name': nm,
            'city': detail_map.get(nm, {}).get('city', ''),
            'state': detail_map.get(nm, {}).get('state', ''),
            'contact_name': detail_map.get(nm, {}).get('contact_name', ''),
            'territory': detail_map.get(nm, {}).get('territory', ''),
        }
        for nm in clean_names
    ]
    return {'accounts': options}


@router.get("")
async def list_invoices(
    request: Request,
    current_user: dict = Depends(get_current_user),
    search: Optional[str] = Query(None, description="Search by invoice number or account name"),
    territory: Optional[str] = Query(None),
    state: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    account_id: Optional[str] = Query(None),
    account_name: Optional[str] = Query(None),
    account_names: Optional[List[str]] = Query(None, description="Filter by exact account names (multi-select)"),
    status: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    time_filter: Optional[str] = Query(None, description="this_week, last_week, this_month, etc."),
    sort_by: Optional[str] = Query("invoice_date", description="Field to sort by"),
    sort_order: Optional[str] = Query("desc", description="asc or desc"),
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=200),
):
    """
    List all invoices with filtering and pagination.
    Similar to leads list but for invoices.
    """
    try:
        tdb = get_tdb()
        
        logger.info(f"[INVOICES] Listing invoices with filters: search={search}, territory={territory}, state={state}, city={city}, status={status}, time_filter={time_filter}, account_names={account_names}")
        
        query, is_empty = await _build_invoice_query(
            tdb, search=search, territory=territory, state=state, city=city,
            account_name=account_name, account_names=account_names, status=status,
            date_from=date_from, date_to=date_to, time_filter=time_filter,
        )
        if is_empty:
            return {
                'invoices': [], 'total': 0, 'page': page, 'limit': limit, 'pages': 0,
                'summary': {'total_gross': 0, 'total_net': 0, 'total_credit': 0}
            }

        logger.info(f"[INVOICES] Final query: {query}")
        
        # Sort configuration
        sort_direction = -1 if sort_order == 'desc' else 1
        sort_field = sort_by if sort_by else 'invoice_date'
        
        # Get total count
        total = await tdb.invoices.count_documents(query)
        logger.info(f"[INVOICES] Total count: {total}")
        
        # Calculate pagination
        skip = (page - 1) * limit
        pages = (total + limit - 1) // limit if total > 0 else 0
        
        # Fetch invoices
        cursor = tdb.invoices.find(query, {'_id': 0})
        cursor = cursor.sort(sort_field, sort_direction).skip(skip).limit(limit)
        invoices = await cursor.to_list(limit)
        
        # Enrich with account names if not present
        account_ids = list(set([inv.get('account_id') or inv.get('account_uuid') for inv in invoices if inv.get('account_id') or inv.get('account_uuid')]))
        
        if account_ids:
            accounts = await tdb.accounts.find(
                {'$or': [{'id': {'$in': account_ids}}, {'account_id': {'$in': account_ids}}]},
                {'_id': 0, 'id': 1, 'account_id': 1, 'account_name': 1, 'city': 1, 'state': 1, 'territory': 1}
            ).to_list(len(account_ids))
            
            account_map = {}
            for acc in accounts:
                if acc.get('id'):
                    account_map[acc['id']] = acc
                if acc.get('account_id'):
                    account_map[acc['account_id']] = acc
            
            # Enrich invoices
            for inv in invoices:
                acc_id = inv.get('account_id') or inv.get('account_uuid')
                if acc_id and acc_id in account_map:
                    acc = account_map[acc_id]
                    inv['account_name'] = inv.get('account_name') or acc.get('account_name')
                    inv['account_city'] = acc.get('city')
                    inv['account_state'] = acc.get('state')
                    inv['account_territory'] = acc.get('territory')
        
        # Normalize invoice fields (handle both old and new formats)
        for inv in invoices:
            # Normalize invoice number field
            if not inv.get('invoice_no') and inv.get('invoice_number'):
                inv['invoice_no'] = inv.get('invoice_number')
            # Normalize gross/net values (old format uses grand_total, new uses gross_invoice_value)
            if not inv.get('gross_invoice_value') and inv.get('grand_total'):
                inv['gross_invoice_value'] = inv.get('grand_total')
            if not inv.get('net_invoice_value'):
                # Calculate net as gross minus credit note
                gross = inv.get('gross_invoice_value') or inv.get('grand_total') or 0
                credit = inv.get('credit_note_value') or 0
                inv['net_invoice_value'] = gross - credit

        # Resolve each line item's SKU display name to the CURRENT master SKU
        # (code-first + sku_aliases) so historical lines with stale names /
        # retired codes show the current SKU everywhere.
        from services.sku_resolver import build_sku_resolver
        _resolver = await build_sku_resolver(tdb)
        for inv in invoices:
            if inv.get('items'):
                inv['items'] = _resolver.enrich_items(inv['items'])
            if inv.get('line_items'):
                inv['line_items'] = _resolver.enrich_items(inv['line_items'])

        # Enrich each row with account-level credit notes (issued/applied/balance).
        # Display-only — does NOT alter net_invoice_value.
        cn_by_id, cn_by_name = await _build_credit_note_account_map(tdb)
        for inv in invoices:
            cn = _cn_for_invoice(inv, cn_by_id, cn_by_name)
            inv['cn_issued'] = cn['issued']
            inv['cn_applied'] = cn['applied']
            inv['cn_balance'] = cn['balance']
            # Derive status from real account linkage (self-healing for stale field)
            inv['status'] = _derive_invoice_status(inv)

        # Calculate summary - handle both old and new field names
        all_invoices_cursor = tdb.invoices.find(query, {'_id': 0, 'gross_invoice_value': 1, 'grand_total': 1, 'net_invoice_value': 1, 'credit_note_value': 1, 'outstanding': 1, 'account_id': 1, 'account_uuid': 1, 'account_name': 1, 'invoice_no': 1, 'invoice_number': 1})
        all_invoices_for_summary = await all_invoices_cursor.to_list(10000)

        # Credit notes applied to specific invoices via their originating
        # stock-out delivery. Populates the per-invoice "Credit Note" value and
        # reduces Net (Net = Gross - applied credit).
        inv_nos_all = [
            (inv.get('invoice_no') or inv.get('invoice_number'))
            for inv in all_invoices_for_summary
        ]
        applied_credit_map = await _build_invoice_applied_credit_map(tdb, inv_nos_all)
        if applied_credit_map:
            for inv in all_invoices_for_summary:
                no = inv.get('invoice_no') or inv.get('invoice_number')
                ap = applied_credit_map.get(no)
                if ap and ap['credit'] > 0:
                    inv['credit_note_value'] = ap['credit']
            for inv in invoices:
                no = inv.get('invoice_no') or inv.get('invoice_number')
                ap = applied_credit_map.get(no)
                if ap and ap['credit'] > 0:
                    inv['credit_note_value'] = ap['credit']
                    inv['applied_credit_notes'] = ap['applications']
                    gross = inv.get('gross_invoice_value') or inv.get('grand_total') or 0
                    inv['net_invoice_value'] = gross - ap['credit']

        # Account-level credit note totals across the DISTINCT accounts in the
        # filtered set (count each account once to avoid duplication).
        cn_seen = set()
        cn_total_issued = cn_total_applied = cn_total_balance = 0
        for inv in all_invoices_for_summary:
            aid = inv.get('account_id') or inv.get('account_uuid')
            nm = (inv.get('account_name') or '').strip().lower()
            key = aid or nm
            if not key or key in cn_seen:
                continue
            cn_seen.add(key)
            cn = _cn_for_invoice(inv, cn_by_id, cn_by_name)
            cn_total_issued += cn['issued']
            cn_total_applied += cn['applied']
            cn_total_balance += cn['balance']

        total_gross = sum((inv.get('gross_invoice_value') or inv.get('grand_total') or 0) for inv in all_invoices_for_summary)
        total_credit = sum(inv.get('credit_note_value', 0) or 0 for inv in all_invoices_for_summary)
        summary = {
            'total_gross': total_gross,
            'total_net': total_gross - total_credit,
            'total_credit': total_credit,
            'total_outstanding': sum(inv.get('outstanding', 0) or 0 for inv in all_invoices_for_summary),
            'cn_total_issued': cn_total_issued,
            'cn_total_applied': cn_total_applied,
            'cn_total_balance': cn_total_balance,
        }
        
        logger.info(f"[INVOICES] Listed {len(invoices)} invoices (page {page}/{pages}, total {total})")
        
        return {
            'invoices': invoices,
            'total': total,
            'page': page,
            'limit': limit,
            'pages': pages,
            'summary': summary
        }
        
    except Exception as e:
        logger.error(f"[INVOICES] Error listing invoices: {e}")
        import traceback
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/export")
async def export_invoices(
    current_user: dict = Depends(get_current_user),
    search: Optional[str] = Query(None),
    territory: Optional[str] = Query(None),
    state: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    account_name: Optional[str] = Query(None),
    account_names: Optional[List[str]] = Query(None),
    status: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    time_filter: Optional[str] = Query(None),
    sort_by: Optional[str] = Query("invoice_date"),
    sort_order: Optional[str] = Query("desc"),
):
    """Export the filtered invoices list to an Excel (.xlsx) file."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse

    tdb = get_tdb()

    query, is_empty = await _build_invoice_query(
        tdb, search=search, territory=territory, state=state, city=city,
        account_name=account_name, account_names=account_names, status=status,
        date_from=date_from, date_to=date_to, time_filter=time_filter,
    )

    invoices = []
    if not is_empty:
        sort_direction = -1 if sort_order == 'desc' else 1
        sort_field = sort_by if sort_by else 'invoice_date'
        invoices = await tdb.invoices.find(query, {'_id': 0}).sort(sort_field, sort_direction).to_list(100000)

        # Enrich with account location info
        account_ids = list(set([inv.get('account_id') or inv.get('account_uuid') for inv in invoices if inv.get('account_id') or inv.get('account_uuid')]))
        if account_ids:
            accounts = await tdb.accounts.find(
                {'$or': [{'id': {'$in': account_ids}}, {'account_id': {'$in': account_ids}}]},
                {'_id': 0, 'id': 1, 'account_id': 1, 'account_name': 1, 'city': 1, 'state': 1, 'territory': 1}
            ).to_list(len(account_ids))
            account_map = {}
            for acc in accounts:
                if acc.get('id'):
                    account_map[acc['id']] = acc
                if acc.get('account_id'):
                    account_map[acc['account_id']] = acc
            for inv in invoices:
                acc_id = inv.get('account_id') or inv.get('account_uuid')
                if acc_id and acc_id in account_map:
                    acc = account_map[acc_id]
                    inv['account_name'] = inv.get('account_name') or acc.get('account_name')
                    inv['account_city'] = acc.get('city')
                    inv['account_state'] = acc.get('state')

        for inv in invoices:
            if not inv.get('invoice_no') and inv.get('invoice_number'):
                inv['invoice_no'] = inv.get('invoice_number')
            if not inv.get('gross_invoice_value') and inv.get('grand_total'):
                inv['gross_invoice_value'] = inv.get('grand_total')
            if not inv.get('net_invoice_value'):
                gross = inv.get('gross_invoice_value') or inv.get('grand_total') or 0
                credit = inv.get('credit_note_value') or 0
                inv['net_invoice_value'] = gross - credit

        # Account-level credit notes (issued/applied/balance) for export
        cn_by_id, cn_by_name = await _build_credit_note_account_map(tdb)
        for inv in invoices:
            cn = _cn_for_invoice(inv, cn_by_id, cn_by_name)
            inv['cn_issued'] = cn['issued']
            inv['cn_applied'] = cn['applied']
            inv['cn_balance'] = cn['balance']
            inv['status'] = _derive_invoice_status(inv)

        # Per-invoice applied credit (from originating stock-out delivery)
        applied_credit_map = await _build_invoice_applied_credit_map(
            tdb, [(inv.get('invoice_no') or inv.get('invoice_number')) for inv in invoices]
        )
        if applied_credit_map:
            for inv in invoices:
                no = inv.get('invoice_no') or inv.get('invoice_number')
                ap = applied_credit_map.get(no)
                if ap and ap['credit'] > 0:
                    inv['credit_note_value'] = ap['credit']
                    gross = inv.get('gross_invoice_value') or inv.get('grand_total') or 0
                    inv['net_invoice_value'] = gross - ap['credit']

    # Build workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"

    headers = ['Invoice #', 'Date', 'Account', 'Bottles', 'Gross Value', 'Credit Note',
               'Net Value', 'Outstanding', 'CN Issued', 'CN Applied', 'CN Balance',
               'Status', 'City', 'State']
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="166534", end_color="166534", fill_type="solid")
    thin = Side(style="thin", color="d0d5dd")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c_idx, name in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    widths = [18, 14, 36, 12, 16, 14, 16, 16, 14, 14, 14, 12, 18, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    def _total_bottles(inv):
        items = inv.get('items') or inv.get('line_items') or []
        if not isinstance(items, list):
            return 0
        return int(sum(float(it.get('bottles') or it.get('quantity') or 0) for it in items))

    for r_idx, inv in enumerate(invoices, 2):
        row = [
            inv.get('invoice_no') or inv.get('invoice_number') or '-',
            inv.get('invoice_date') or '-',
            inv.get('account_name') or inv.get('account_id') or '-',
            _total_bottles(inv),
            round(inv.get('gross_invoice_value') or 0, 2),
            round(inv.get('credit_note_value') or 0, 2),
            round(inv.get('net_invoice_value') or 0, 2),
            round(inv.get('outstanding') or 0, 2),
            round(inv.get('cn_issued') or 0, 2),
            round(inv.get('cn_applied') or 0, 2),
            round(inv.get('cn_balance') or 0, 2),
            inv.get('status') or '-',
            inv.get('account_city') or '-',
            inv.get('account_state') or '-',
        ]
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border
            if c_idx == 4:
                cell.number_format = '#,##0'
            elif c_idx in (5, 6, 7, 8, 9, 10, 11):
                cell.number_format = '#,##0.00'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"invoices_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{invoice_id}/pdf")
async def download_invoice_pdf(
    invoice_id: str,
    current_user: dict = Depends(get_current_user),
):
    """Generate a single-invoice PDF document for download."""
    import io
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from fastapi.responses import StreamingResponse

    tdb = get_tdb()
    inv = await tdb.invoices.find_one(
        {'$or': [{'id': invoice_id}, {'invoice_no': invoice_id}, {'invoice_number': invoice_id}]},
        {'_id': 0}
    )
    if not inv:
        raise HTTPException(status_code=404, detail='Invoice not found')

    # Normalize
    inv_no = inv.get('invoice_no') or inv.get('invoice_number') or '-'
    gross = inv.get('gross_invoice_value') or inv.get('grand_total') or 0

    # Account info
    acc = None
    acc_id = inv.get('account_uuid') or inv.get('account_id')
    if acc_id:
        acc = await tdb.accounts.find_one(
            {'$or': [{'id': acc_id}, {'account_id': acc_id}]},
            {'_id': 0, 'account_name': 1, 'city': 1, 'state': 1, 'address': 1, 'contact_name': 1, 'phone': 1, 'gstin': 1}
        )
    acc = acc or {}
    account_name = inv.get('account_name') or acc.get('account_name') or '-'

    # Applied credit (from originating stock-out delivery)
    applied_map = await _build_invoice_applied_credit_map(tdb, [inv_no])
    ap = applied_map.get(inv_no)
    credit = (ap['credit'] if ap and ap['credit'] > 0 else (inv.get('credit_note_value') or 0))
    net = gross - credit

    # Line items via SKU resolver
    from services.sku_resolver import build_sku_resolver
    _resolver = await build_sku_resolver(tdb)
    items = inv.get('items') or inv.get('line_items') or []
    if items:
        items = _resolver.enrich_items(items)

    def _money(v):
        try:
            return f"Rs. {float(v or 0):,.2f}"
        except Exception:
            return "Rs. 0.00"

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=16 * mm, rightMargin=16 * mm, topMargin=16 * mm, bottomMargin=16 * mm,
        title=f"Invoice {inv_no}"
    )
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Heading1"], fontSize=20, spaceAfter=2, textColor=colors.HexColor("#166534"))
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontSize=11, textColor=colors.grey, spaceAfter=10)
    body = ParagraphStyle("body", parent=styles["BodyText"], fontSize=10, leading=13)
    small = ParagraphStyle("small", parent=styles["BodyText"], fontSize=9, leading=11, textColor=colors.grey)
    right = ParagraphStyle("right", parent=body, alignment=2)

    story = []
    story.append(Paragraph("INVOICE", h1))
    story.append(Paragraph(f"{inv_no} &nbsp;·&nbsp; {inv.get('invoice_date') or '-'}", h2))

    bill_to = [account_name]
    addr_line = ", ".join([p for p in [acc.get('address'), acc.get('city'), acc.get('state')] if p])
    if addr_line:
        bill_to.append(addr_line)
    if acc.get('contact_name'):
        bill_to.append(f"Contact: {acc.get('contact_name')}")
    if acc.get('phone'):
        bill_to.append(f"Phone: {acc.get('phone')}")
    if acc.get('gstin'):
        bill_to.append(f"GSTIN: {acc.get('gstin')}")

    story.append(Paragraph("<b>Bill To</b>", body))
    story.append(Paragraph("<br/>".join(bill_to), body))
    story.append(Spacer(1, 8 * mm))

    # Line items table
    data = [["SKU", "Crates", "Bottles", "Line Total"]]
    for it in items:
        sku = it.get('sku_name') or it.get('sku') or 'N/A'
        crates = it.get('crates') if it.get('crates') is not None else (it.get('crateCount') if it.get('crateCount') is not None else '-')
        bottles = it.get('bottles') or it.get('quantity') or 0
        lt = it.get('lineTotal') or it.get('line_total') or it.get('net_amount') or it.get('total') or 0
        data.append([
            Paragraph(str(sku), body),
            str(crates),
            f"{int(bottles):,}",
            _money(lt),
        ])
    if len(data) == 1:
        data.append([Paragraph("No line items", small), "", "", ""])

    tbl = Table(data, colWidths=[None, 22 * mm, 24 * mm, 32 * mm])
    tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#166534")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#f1f5f9")]),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor("#d0d5dd")),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 6 * mm))

    # Totals
    totals = [
        ["Gross", _money(gross)],
        ["Credit Note", _money(credit)],
        ["Net", _money(net)],
        ["Outstanding", _money(inv.get('outstanding') or 0)],
    ]
    tot = Table(totals, colWidths=[None, 40 * mm], hAlign='RIGHT')
    tot.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LINEABOVE', (0, 2), (-1, 2), 0.5, colors.HexColor("#94a3b8")),
        ('FONTNAME', (0, 2), (-1, 2), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0, 2), (-1, 2), colors.HexColor("#6d28d9")),
    ]))
    story.append(tot)

    if ap and ap.get('applications'):
        cns = ", ".join(
            f"{a.get('credit_note_number') or ''} ({_money(a.get('amount_applied'))})"
            for a in ap['applications']
        )
        story.append(Spacer(1, 5 * mm))
        story.append(Paragraph(f"<b>Credit notes applied:</b> {cns}", small))

    story.append(Spacer(1, 10 * mm))
    story.append(Paragraph(
        f"Generated on {datetime.now(timezone.utc).strftime('%d %b %Y, %H:%M UTC')}", small
    ))

    doc.build(story)
    buf.seek(0)
    safe_no = str(inv_no).replace('/', '-').replace(' ', '_')
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="invoice_{safe_no}.pdf"'},
    )


_INVOICE_REGEN_ROLES = {
    "ceo", "admin", "system admin", "director", "vice president",
    "head of business", "national sales head", "regional sales manager",
}


@router.post("/{invoice_id}/regenerate")
async def regenerate_invoice(
    invoice_id: str,
    current_user: dict = Depends(get_current_user),
):
    """Regenerate the Zoho invoice behind a listed invoice (resolves it back to
    its originating distributor delivery). Updates in place when possible, else
    voids + recreates. Management roles only; only delivery-sourced invoices."""
    if (current_user.get("role") or "").lower() not in _INVOICE_REGEN_ROLES:
        raise HTTPException(status_code=403, detail="Only management roles can regenerate invoices.")
    from core.tenant import get_current_tenant_id
    tenant_id = get_current_tenant_id()
    tdb = get_tdb()
    inv = await tdb.invoices.find_one(
        {'$or': [{'id': invoice_id}, {'invoice_no': invoice_id}, {'invoice_number': invoice_id}]},
        {'_id': 0, 'invoice_no': 1, 'invoice_number': 1, 'zoho_invoice_number': 1},
    )
    inv_no = (inv or {}).get('zoho_invoice_number') or (inv or {}).get('invoice_no') or (inv or {}).get('invoice_number') or invoice_id
    mapping = await tdb.zoho_invoice_mappings.find_one(
        {'source_type': 'distributor_delivery', 'status': 'synced',
         '$or': [{'zoho_invoice_number': inv_no}, {'zoho_invoice_id': invoice_id}]},
        {'_id': 0, 'source_id': 1, 'distributor_id': 1, 'zoho_invoice_number': 1},
    )
    if not mapping or not mapping.get('source_id') or not mapping.get('distributor_id'):
        raise HTTPException(
            status_code=400,
            detail="This invoice can't be regenerated here — it isn't linked to a distributor delivery. Regenerate it from the delivery instead.",
        )
    from services.zoho_service import (
        regenerate_delivery_invoice, ZohoPushSkippedError,
        InvoiceNotRegenerableError, MissingAgreedPriceError, ZohoBranchNotMappedError,
    )
    try:
        m = await regenerate_delivery_invoice(tenant_id, mapping['distributor_id'], mapping['source_id'])
    except (ZohoPushSkippedError, InvoiceNotRegenerableError, MissingAgreedPriceError, ZohoBranchNotMappedError) as known:
        raise HTTPException(status_code=400, detail=str(known))
    except Exception as e:
        logger.exception(f"Invoice regeneration failed for {invoice_id}")
        raise HTTPException(status_code=400, detail=f"Invoice regeneration failed: {e}")
    mode = m.get("regen_mode") or "updated"
    verb = {"updated": "updated in place", "recreated": "voided and recreated", "created": "created"}.get(mode, mode)
    return {
        "message": f"Invoice {verb}.",
        "regen_mode": mode,
        "zoho_invoice_url": m.get("zoho_invoice_url"),
        "zoho_invoice_number": m.get("zoho_invoice_number"),
        "zoho_invoice_id": m.get("zoho_invoice_id"),
    }



@router.delete("/{invoice_id}")
async def delete_invoice(
    invoice_id: str,
    current_user: dict = Depends(get_current_user),
):
    """
    Delete an invoice. Only CEO and Admin can delete invoices.
    """
    # Check user role
    user_role = current_user.get('role', '').lower()
    allowed_roles = ['ceo', 'system admin', 'admin']
    
    if not any(role in user_role for role in allowed_roles):
        raise HTTPException(
            status_code=403, 
            detail='Only CEO and Admin can delete invoices'
        )
    
    tdb = get_tdb()
    
    # Find the invoice first
    invoice = await tdb.invoices.find_one({'$or': [{'id': invoice_id}, {'invoice_no': invoice_id}]})
    
    if not invoice:
        raise HTTPException(status_code=404, detail='Invoice not found')
    
    # Delete the invoice
    result = await tdb.invoices.delete_one({'$or': [{'id': invoice_id}, {'invoice_no': invoice_id}]})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail='Invoice not found')
    
    logger.info(f"[INVOICES] Invoice {invoice_id} deleted by {current_user.get('email')}")
    
    # Update account totals if the invoice was matched
    if invoice.get('account_uuid') or invoice.get('account_id'):
        account_id = invoice.get('account_uuid') or invoice.get('account_id')
        
        # Recalculate account totals
        remaining_invoices = await tdb.invoices.find({
            '$or': [
                {'account_uuid': account_id},
                {'account_id': account_id}
            ],
            'status': 'matched'
        }).to_list(1000)
        
        total_gross = sum(inv.get('gross_invoice_value', 0) or 0 for inv in remaining_invoices)
        total_net = sum(inv.get('net_invoice_value', 0) or 0 for inv in remaining_invoices)
        total_credit = sum(inv.get('credit_note_value', 0) or 0 for inv in remaining_invoices)
        total_outstanding = sum(inv.get('outstanding', 0) or 0 for inv in remaining_invoices)
        
        await tdb.accounts.update_one(
            {'$or': [{'id': account_id}, {'account_id': account_id}]},
            {
                '$set': {
                    'total_gross_invoice_value': total_gross,
                    'total_net_invoice_value': total_net,
                    'total_credit_note_value': total_credit,
                    'total_outstanding': total_outstanding,
                    'invoice_count': len(remaining_invoices),
                    'updated_at': datetime.now(timezone.utc).isoformat()
                }
            }
        )
    
    return {
        'success': True,
        'message': f'Invoice {invoice.get("invoice_no", invoice_id)} deleted successfully'
    }


@router.delete("")
async def bulk_delete_invoices(
    invoice_ids: List[str],
    current_user: dict = Depends(get_current_user),
):
    """
    Bulk delete invoices. Only CEO and Admin can delete invoices.
    """
    # Check user role
    user_role = current_user.get('role', '').lower()
    allowed_roles = ['ceo', 'system admin', 'admin']
    
    if not any(role in user_role for role in allowed_roles):
        raise HTTPException(
            status_code=403, 
            detail='Only CEO and Admin can delete invoices'
        )
    
    if not invoice_ids:
        raise HTTPException(status_code=400, detail='No invoice IDs provided')
    
    tdb = get_tdb()
    
    # Get all invoices to be deleted
    invoices = await tdb.invoices.find({
        '$or': [
            {'id': {'$in': invoice_ids}},
            {'invoice_no': {'$in': invoice_ids}}
        ]
    }).to_list(len(invoice_ids))
    
    if not invoices:
        raise HTTPException(status_code=404, detail='No invoices found')
    
    # Collect unique account IDs for recalculation
    account_ids = set()
    for inv in invoices:
        if inv.get('account_uuid'):
            account_ids.add(inv['account_uuid'])
        if inv.get('account_id'):
            account_ids.add(inv['account_id'])
    
    # Delete invoices (tenant-scoped)
    result = await tdb.invoices.delete_many({
        '$or': [
            {'id': {'$in': invoice_ids}},
            {'invoice_no': {'$in': invoice_ids}}
        ]
    })
    
    logger.info(f"[INVOICES] Bulk deleted {result.deleted_count} invoices by {current_user.get('email')}")
    
    # Recalculate totals for affected accounts
    for account_id in account_ids:
        remaining_invoices = await tdb.invoices.find({
            '$or': [
                {'account_uuid': account_id},
                {'account_id': account_id}
            ],
            'status': 'matched'
        }).to_list(1000)
        
        total_gross = sum(inv.get('gross_invoice_value', 0) or 0 for inv in remaining_invoices)
        total_net = sum(inv.get('net_invoice_value', 0) or 0 for inv in remaining_invoices)
        total_credit = sum(inv.get('credit_note_value', 0) or 0 for inv in remaining_invoices)
        total_outstanding = sum(inv.get('outstanding', 0) or 0 for inv in remaining_invoices)
        
        await tdb.accounts.update_one(
            {'$or': [{'id': account_id}, {'account_id': account_id}]},
            {
                '$set': {
                    'total_gross_invoice_value': total_gross,
                    'total_net_invoice_value': total_net,
                    'total_credit_note_value': total_credit,
                    'total_outstanding': total_outstanding,
                    'invoice_count': len(remaining_invoices),
                    'updated_at': datetime.now(timezone.utc).isoformat()
                }
            }
        )
    
    return {
        'success': True,
        'message': f'Deleted {result.deleted_count} invoices',
        'deleted_count': result.deleted_count
    }


@router.get("/summary")
async def get_invoice_summary(
    current_user: dict = Depends(get_current_user),
    territory: Optional[str] = Query(None),
    state: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
):
    """
    Get invoice summary statistics with optional filters.
    """
    tdb = get_tdb()
    
    query = {}
    
    # Get account IDs filtered by territory/state/city
    if territory or state or city:
        account_query = {}
        if territory:
            account_query['territory'] = territory
        if state:
            account_query['state'] = state
        if city:
            account_query['city'] = city
        
        accounts = await tdb.accounts.find(account_query, {'_id': 0, 'id': 1, 'account_id': 1}).to_list(10000)
        account_ids = [a.get('account_id') or a.get('id') for a in accounts]
        
        if account_ids:
            query['$or'] = [
                {'account_id': {'$in': account_ids}},
                {'account_uuid': {'$in': [a.get('id') for a in accounts if a.get('id')]}}
            ]
    
    # Get all invoices for summary
    invoices = await tdb.invoices.find(query, {'_id': 0}).to_list(100000)
    
    # Calculate stats
    total_count = len(invoices)
    matched_count = sum(1 for inv in invoices if inv.get('status') == 'matched')
    unmatched_count = sum(1 for inv in invoices if inv.get('status') == 'unmatched')
    
    total_gross = sum(inv.get('gross_invoice_value', 0) or 0 for inv in invoices)
    total_net = sum(inv.get('net_invoice_value', 0) or 0 for inv in invoices)
    total_credit = sum(inv.get('credit_note_value', 0) or 0 for inv in invoices)
    total_outstanding = sum(inv.get('outstanding', 0) or 0 for inv in invoices)
    
    return {
        'total_count': total_count,
        'matched_count': matched_count,
        'unmatched_count': unmatched_count,
        'total_gross': total_gross,
        'total_net': total_net,
        'total_credit': total_credit,
        'total_outstanding': total_outstanding
    }



@router.delete("/admin/nuke-all")
async def nuke_all_invoices(
    confirm: str = Query(..., description="Must be exactly 'YES-DELETE-ALL-INVOICES' to proceed"),
    current_user: dict = Depends(get_current_user),
):
    """DANGER: Wipes the entire invoices collection for the current tenant.
    Also resets invoice-derived financial rollups on every account.
    CEO / System Admin only. Requires explicit confirm token to prevent accidents.
    """
    user_role = (current_user.get('role') or '').strip()
    if user_role not in ('CEO', 'System Admin'):
        raise HTTPException(status_code=403, detail='Only CEO and System Admin can nuke invoices.')

    if confirm != 'YES-DELETE-ALL-INVOICES':
        raise HTTPException(
            status_code=400,
            detail="Pass ?confirm=YES-DELETE-ALL-INVOICES to confirm this destructive action."
        )

    tdb = get_tdb()

    # 1. Wipe every invoice in the tenant
    pre_count = await tdb.invoices.count_documents({})
    result = await tdb.invoices.delete_many({})
    deleted_count = result.deleted_count

    # 2. Wipe every Zoho invoice mapping for this tenant so future invoices push fresh
    try:
        await tdb.zoho_invoice_mappings.delete_many({'source_type': 'delivery'})
    except Exception:  # collection may not exist on every tenant
        pass

    # 3. Reset financial rollups on every account
    rollup_result = await tdb.accounts.update_many(
        {},
        {'$set': {
            'outstanding_balance': 0.0,
            'overdue_amount': 0.0,
            'total_gross_invoice_value': 0.0,
            'total_net_invoice_value': 0.0,
            'total_credit_note_value': 0.0,
            'total_outstanding': 0.0,
            'invoice_count': 0,
            'last_invoice_no': None,
            'last_invoice_date': None,
            'last_payment_amount': None,
            'last_payment_date': None,
            'updated_at': datetime.now(timezone.utc).isoformat(),
        }}
    )

    logger.warning(
        f"[INVOICES] NUKE: {deleted_count} invoices deleted "
        f"(was {pre_count}) by {current_user.get('email')}; "
        f"rollups reset on {rollup_result.modified_count} accounts."
    )

    return {
        'success': True,
        'deleted_count': deleted_count,
        'pre_count': pre_count,
        'accounts_reset': rollup_result.modified_count,
    }



@router.post("/admin/backfill-match-status")
async def backfill_invoice_match_status(
    current_user: dict = Depends(get_current_user),
):
    """One-time backfill: stamp `status` on every invoice in the tenant.

    - status='matched'   → invoice has account_uuid or account_id populated
    - status='unmatched' → no account linkage at all (legacy MQ orphans)

    Idempotent — safe to re-run. CEO / System Admin only.
    """
    user_role = (current_user.get('role') or '').strip()
    if user_role not in ('CEO', 'System Admin'):
        raise HTTPException(status_code=403, detail='Only CEO and System Admin can run this backfill.')

    tdb = get_tdb()

    # 1. Mark everything that has any account linkage as 'matched'
    matched_result = await tdb.invoices.update_many(
        {'$or': [
            {'account_uuid': {'$exists': True, '$nin': [None, '']}},
            {'account_id': {'$exists': True, '$nin': [None, '']}},
        ]},
        {'$set': {'status': 'matched'}}
    )

    # 2. Mark everything that has NO account linkage at all as 'unmatched'
    unmatched_result = await tdb.invoices.update_many(
        {
            '$and': [
                {'$or': [{'account_uuid': None}, {'account_uuid': {'$exists': False}}, {'account_uuid': ''}]},
                {'$or': [{'account_id': None}, {'account_id': {'$exists': False}}, {'account_id': ''}]},
            ]
        },
        {'$set': {'status': 'unmatched'}}
    )

    total = await tdb.invoices.count_documents({})
    matched_count = await tdb.invoices.count_documents({'status': 'matched'})
    unmatched_count = await tdb.invoices.count_documents({'status': 'unmatched'})

    logger.info(
        f"[INVOICES] backfill-match-status by {current_user.get('email')}: "
        f"matched_modified={matched_result.modified_count} "
        f"unmatched_modified={unmatched_result.modified_count}"
    )

    return {
        'success': True,
        'matched_updated': matched_result.modified_count,
        'unmatched_updated': unmatched_result.modified_count,
        'totals': {
            'total': total,
            'matched': matched_count,
            'unmatched': unmatched_count,
        },
    }
